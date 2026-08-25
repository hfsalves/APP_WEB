#!/usr/bin/env python3
"""Import the July 2026 BMSO delivery/invoice chain into HSOLS_FR.

The source workbook describes the external BMSO invoice, its delivery note and
the originating PHC purchase order.  Each imported detail line is linked with
BI.OBISTAMP: BC line -> BL line -> pre-invoice line.

The script is deliberately dry-run by default.  Use --execute only after the
validation summary is clean.  Re-running is safe: pre-invoices are identified
by their external invoice number in BO.FREF.
"""

from __future__ import annotations

import argparse
import os
import re
import shutil
import sys
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

import openpyxl
import pyodbc

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app import app
from modules.gr_subcontractor_measurements.service import (  # noqa: E402
    PHC_ZERO_DATE,
    _new_stamp,
    _phc_columns,
    _phc_conn_str,
    _phc_value,
)


PHC_DB = "HSOLS_FR"
PHC_SERVER = "10.0.1.12"
SUPPLIER_NO = 31106
BC_NDOS = 102
BL_NDOS = 130
PF_NDOS = 218
BL_NAME = "Bon Livraison Fourn."
PF_NAME = "Pré-Facture"
ZERO = Decimal("0")
QTY_STEP = Decimal("0.0001")
MONEY_STEP = Decimal("0.01")
_TABLE_COLUMNS: dict[str, set[str]] = {}


class ImportError(Exception):
    pass


def _phc_insert(cursor, table_name: str, values: dict[str, Any]) -> dict[str, Any]:
    """Cached version of the shared PHC insert helper.

    This import has thousands of lines. Looking up INFORMATION_SCHEMA for every
    individual insert makes a correct import needlessly slow and risks a client
    timeout before the transaction can finish.
    """
    columns = _TABLE_COLUMNS.get(table_name)
    if columns is None:
        columns = _phc_columns(cursor, table_name)
        _TABLE_COLUMNS[table_name] = columns
    filtered = {key: value for key, value in values.items() if key.lower() in columns}
    if not filtered:
        raise ImportError(f"Sem colunas válidas para inserir em {table_name}.")
    cursor.execute(
        f"INSERT INTO dbo.{table_name} ({', '.join(filtered)}) VALUES ({', '.join('?' for _ in filtered)})",
        tuple(filtered.values()),
    )
    return filtered


def text(value: Any) -> str:
    return str(value or "").strip()


def dec(value: Any) -> Decimal:
    if value is None or text(value) == "":
        return ZERO
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value).replace(" ", "").replace(",", "."))


def amount(value: Any) -> Decimal:
    return dec(value).quantize(MONEY_STEP, rounding=ROUND_HALF_UP)


def qty(value: Any) -> Decimal:
    return dec(value).quantize(QTY_STEP, rounding=ROUND_HALF_UP)


def norm(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", text(value).upper())


def norm_description(value: Any) -> str:
    value = text(value).upper()
    value = value.replace("CONTRIBUITION", "CONTRIBUTION")
    value = re.sub(r"[^A-Z0-9]", "", value)
    return value


def read_workbook(path: Path) -> list[dict[str, Any]]:
    book = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "Détails" not in book.sheetnames:
        raise ImportError("O ficheiro não tem a folha 'Détails'.")
    sheet = book["Détails"]
    headers = [text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = {headers[index]: values[index] for index in range(len(headers))}
        invoice = text(row.get("Nº Facture"))
        if not invoice:
            continue
        rows.append({
            "invoice": invoice,
            "invoice_date": row.get("Date Facture"),
            "delivery": text(row.get("Nº BL")),
            "delivery_date": row.get("Date BL"),
            "ref": text(row.get("Référence Produit")),
            "design": text(row.get("Article Facture/BL")),
            "qty": qty(row.get("Quantité")),
            "unit": text(row.get("Unité")),
            "price": amount(row.get("Prix Unitaire")),
            "vat": dec(row.get("TVA")) * Decimal("100"),
            "bc": int(dec(row.get("Nº BC PHC"))),
        })
    if not rows:
        raise ImportError("Não foram encontradas linhas no ficheiro.")
    return rows


def valid_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = text(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw[:10], fmt).date()
        except ValueError:
            pass
    raise ImportError(f"Data inválida no Excel: {raw!r}")


def fetch_dicts(cursor, sql: str, params: tuple[Any, ...] = ()) -> list[dict[str, Any]]:
    cursor.execute(sql, params)
    columns = [item[0] for item in cursor.description or []]
    rows: list[dict[str, Any]] = []
    for values in cursor.fetchall():
        row: dict[str, Any] = {}
        for column, value in zip(columns, values):
            row[str(column)] = value
            row[str(column).upper()] = value
        rows.append(row)
    return rows


def get_tax_code(cursor, rate: Decimal) -> int:
    rows = fetch_dicts(cursor, "SELECT CODIGO, TAXA FROM dbo.TAXASIVA")
    for row in rows:
        if dec(row["TAXA"]) == rate:
            return int(dec(row["CODIGO"]))
    raise ImportError(f"Não existe taxa de IVA {rate}% na HSOLS_FR.")


def load_source(cursor, bcs: set[int]) -> tuple[dict[int, dict[str, Any]], list[dict[str, Any]]]:
    placeholders = ",".join("?" for _ in bcs)
    headers = fetch_dicts(cursor, f"""
        SELECT B.OBRANO, B.BOSTAMP, B.BOANO, B.DATAOBRA, B.CCUSTO, B.FREF,
               B.MOEDA, B.NO, B.NOME, B.NCONT, B.MORADA, B.LOCAL, B.CODPOST, B.ESTAB,
               B.FECHADA
        FROM dbo.BO B
        WHERE B.NDOS = {BC_NDOS} AND B.NO = ? AND B.OBRANO IN ({placeholders})
    """, tuple([SUPPLIER_NO, *sorted(bcs)]))
    by_bc = {int(dec(row["OBRANO"])): row for row in headers}
    missing = sorted(bcs - set(by_bc))
    if missing:
        raise ImportError(f"Não encontrei os BC BMSO: {', '.join(map(str, missing))}.")
    if any(bool(row.get("FECHADA")) for row in by_bc.values()):
        bad = [str(number) for number, row in by_bc.items() if bool(row.get("FECHADA"))]
        raise ImportError(f"Os BC já estão fechados: {', '.join(bad)}.")

    lines = fetch_dicts(cursor, f"""
        SELECT I.*, I2.QTTCOMPRA, I2.QTTENC
        FROM dbo.BI I
        LEFT JOIN dbo.BI2 I2 ON I2.BI2STAMP = I.BISTAMP
        WHERE I.NDOS = {BC_NDOS} AND I.BOSTAMP IN ({','.join('?' for _ in by_bc)})
        ORDER BY I.BOSTAMP, I.LORDEM, I.BISTAMP
    """, tuple(row["BOSTAMP"] for row in by_bc.values()))
    for line in lines:
        line["BC"] = next(number for number, header in by_bc.items() if header["BOSTAMP"] == line["BOSTAMP"])
        line["remaining"] = qty(line.get("QTT"))
        line["price"] = amount(line.get("EDEBITO"))
        line["ref_key"] = norm(line.get("REF"))
        line["desc_key"] = norm_description(line.get("DESIGN"))
    return by_bc, lines


def allocate(rows: list[dict[str, Any]], source_lines: list[dict[str, Any]],
             credit_rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    """Map every Excel row to an exact BC source line, preserving quantities."""
    lines_by_bc: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for line in source_lines:
        lines_by_bc[int(line["BC"])].append(line)
    result: dict[str, list[dict[str, Any]]] = defaultdict(list)

    # Consume in delivery order so each repeated reference/price is assigned
    # to the corresponding source line in the BC.
    for row in sorted(rows, key=lambda item: (item["bc"], valid_date(item["delivery_date"]), item["invoice"])):
        candidates = [line for line in lines_by_bc[row["bc"]] if line["remaining"] > ZERO and line["price"] == row["price"]]
        # Exact reference has priority. Description+price is the safe fallback
        # for the gasoil lines whose Excel reference is P.02 but PHC uses S.03.
        exact_ref = [line for line in candidates if line["ref_key"] == norm(row["ref"])]
        exact_desc = [line for line in candidates if line["desc_key"] == norm_description(row["design"])]
        selected = exact_ref or exact_desc
        if not selected:
            raise ImportError(
                f"Sem linha BC para fatura {row['invoice']}, ref {row['ref']}, "
                f"preço {row['price']}, BC {row['bc']}."
            )
        remaining = row["qty"]
        for source in selected:
            if remaining <= ZERO:
                break
            part_qty = min(remaining, source["remaining"])
            source["remaining"] -= part_qty
            remaining -= part_qty
            result[row["invoice"]].append({"excel": row, "source": source, "qty": part_qty})
        if remaining > ZERO:
            raise ImportError(
                f"Quantidade insuficiente no BC {row['bc']} para a fatura {row['invoice']} "
                f"({remaining} em falta)."
            )

    # The note credit is deliberately not imported as BL/PF.  Its original
    # delivery rows may still be present in the BC; accept only those exact
    # cancelled quantities as an excluded balance when closing the BC.
    cancelled_qty: dict[tuple[int, str], Decimal] = defaultdict(lambda: ZERO)
    for row in credit_rows:
        cancelled_qty[(row["bc"], norm(row["ref"]))] += row["qty"]
    unsatisfied: list[dict[str, Any]] = []
    for line in source_lines:
        remaining = line["remaining"]
        if remaining <= QTY_STEP:
            continue
        key = (line["BC"], line["ref_key"])
        if remaining <= cancelled_qty[key] + QTY_STEP:
            cancelled_qty[key] -= remaining
            line["remaining"] = ZERO
            continue
        unsatisfied.append(line)
    if unsatisfied:
        sample = ", ".join(f"BC {line['BC']} {text(line.get('REF'))}: {line['remaining']}" for line in unsatisfied[:8])
        raise ImportError(f"Os BC não ficam totalmente satisfeitos: {sample}")
    return result


def next_number(cursor, ndos: int, year: int) -> int:
    cursor.execute("""
        SELECT ISNULL(MAX(TRY_CONVERT(int, OBRANO)), 0) + 1
        FROM dbo.BO WITH (UPDLOCK, HOLDLOCK)
        WHERE NDOS = ? AND BOANO = ?
    """, (ndos, year))
    return int(cursor.fetchone()[0] or 1)


def header_values(stamp: str, ndos: int, name: str, number: int, doc_date: date,
                  source_header: dict[str, Any], external_ref: str, net: Decimal,
                  vat: Decimal, closed: bool) -> dict[str, Any]:
    now = datetime.now()
    hour = now.strftime("%H:%M:%S")
    return {
        "bostamp": stamp, "nmdos": name, "ndos": ndos, "obrano": number,
        "boano": doc_date.year, "dataobra": doc_date, "dataopen": doc_date,
        "datafecho": doc_date if closed else PHC_ZERO_DATE,
        "nome": text(source_header.get("NOME"))[:55], "no": int(dec(source_header.get("NO"))),
        "ncont": text(source_header.get("NCONT")), "morada": text(source_header.get("MORADA")),
        "local": text(source_header.get("LOCAL")), "codpost": text(source_header.get("CODPOST")),
        "estab": int(dec(source_header.get("ESTAB"))), "moeda": text(source_header.get("MOEDA")) or "EURO",
        "ccusto": text(source_header.get("CCUSTO")), "fref": external_ref[:20],
        "totaldeb": _phc_value(net), "etotaldeb": net, "total": _phc_value(net + vat),
        "etotal": net + vat, "fechada": 1 if closed else 0,
        "ousrinis": "APP", "ousrdata": now, "ousrhora": hour,
        "usrinis": "APP", "usrdata": now, "usrhora": hour,
    }


def insert_header(cursor, stamp: str, ndos: int, name: str, number: int, doc_date: date,
                  source_header: dict[str, Any], external_ref: str, net: Decimal,
                  vat: Decimal, closed: bool) -> None:
    values = header_values(stamp, ndos, name, number, doc_date, source_header, external_ref, net, vat, closed)
    _phc_insert(cursor, "BO", values)
    now = datetime.now(); hour = now.strftime("%H:%M:%S")
    _phc_insert(cursor, "BO2", {
        "bo2stamp": stamp, "processo": text(source_header.get("CCUSTO")), "subproc": "", "area": "",
        "armazem": 1, "ousrinis": "APP", "ousrdata": now, "ousrhora": hour,
        "usrinis": "APP", "usrdata": now, "usrhora": hour,
    })
    # PHC presents BO3.DOCUMENTNUMBERORI in the "Equipe" field of these
    # dossiers. The external document number belongs in BO.FREF only.
    document_number = ""
    _phc_insert(cursor, "BO3", {
        "bo3stamp": stamp, "documentnumberori": document_number, "arquivadodigital": 0,
        "ousrinis": "APP", "ousrdata": now, "ousrhora": hour,
        "usrinis": "APP", "usrdata": now, "usrhora": hour,
    })


def insert_tax(cursor, bostamp: str, vat_code: int, net: Decimal, vat: Decimal) -> None:
    now = datetime.now(); hour = now.strftime("%H:%M:%S")
    _phc_insert(cursor, "BOT", {
        "botstamp": _new_stamp(), "bostamp": bostamp, "codigo": vat_code, "taxa": Decimal("20"),
        "ebaseinc": net, "baseinc": _phc_value(net), "evalor": vat, "valor": _phc_value(vat),
        "ousrinis": "APP", "ousrdata": now, "ousrhora": hour,
        "usrinis": "APP", "usrdata": now, "usrhora": hour,
    })


def insert_line(cursor, *, header_stamp: str, ndos: int, name: str, number: int, doc_date: date,
                source_header: dict[str, Any], source: dict[str, Any], line_qty: Decimal,
                external_design: str, parent_bistamp: str, closed: bool, lordem: int,
                vat_code: int, pf_number: int | None = None) -> str:
    now = datetime.now(); hour = now.strftime("%H:%M:%S")
    stamp = _new_stamp()
    unit_price = amount(source.get("EDEBITO"))
    total = amount(unit_price * line_qty)
    values = {
        "bistamp": stamp, "bostamp": header_stamp, "nmdos": name, "ndos": ndos,
        "obrano": number, "boano": doc_date.year, "dataobra": doc_date,
        "ref": text(source.get("REF")), "design": (external_design or text(source.get("DESIGN")))[:60],
        "qtt": line_qty, "qtt2": line_qty, "unidade": text(source.get("UNIDADE")),
        "pu": _phc_value(unit_price), "debito": _phc_value(unit_price), "edebito": unit_price,
        "ttdeb": _phc_value(total), "ettdeb": total, "pcusto": _phc_value(unit_price),
        "epcusto": unit_price, "prorc": _phc_value(unit_price), "iva": Decimal("20"),
        "tabiva": vat_code, "ivaincl": 0, "armazem": int(dec(source.get("ARMAZEM"))) or 1,
        "stipo": int(dec(source.get("STIPO"))), "no": int(dec(source_header.get("NO"))),
        "nome": text(source_header.get("NOME"))[:55], "ccusto": text(source_header.get("CCUSTO")),
        "bofref": text(source_header.get("FREF")), "bifref": text(source_header.get("FREF")),
        "familia": text(source.get("FAMILIA")), "lordem": lordem,
        "lobs": text(source.get("LOBS"))[:60], "lobs2": text(source.get("LOBS2"))[:60],
        "oobistamp": parent_bistamp, "oobostamp": source_header.get("BOSTAMP"),
        "obistamp": parent_bistamp, "fechada": 1 if closed else 0,
        "ousrinis": "APP", "ousrdata": now, "ousrhora": hour,
        "usrinis": "APP", "usrdata": now, "usrhora": hour,
    }
    if ndos == PF_NDOS:
        # Match PHC-created pre-invoices. This lets the UI recognise the
        # pending supplier-invoice movement instead of showing an orphan line.
        values["ndoc"] = 55
        values["nmdoc"] = "V/Facture"
        values["fno"] = 0
    elif pf_number is not None:
        values["ndoc"] = PF_NDOS
        values["nmdoc"] = PF_NAME
        values["fno"] = pf_number
    _phc_insert(cursor, "BI", values)
    _phc_insert(cursor, "BI2", {
        "bi2stamp": stamp, "bostamp": header_stamp, "fnstamp": "", "fodocnome": "", "foadoc": "",
        "fistamp": "", "origbistamp": "" if ndos == PF_NDOS else parent_bistamp,
        "qttenc": line_qty if ndos == BL_NDOS else ZERO,
        "ousrinis": "APP", "ousrdata": now, "ousrhora": hour,
        "usrinis": "APP", "usrdata": now, "usrhora": hour,
    })
    return stamp


def existing_invoices(cursor, invoices: set[str]) -> set[str]:
    found: set[str] = set()
    for offset in range(0, len(invoices), 200):
        batch = sorted(invoices)[offset:offset + 200]
        sql = f"""
            SELECT DISTINCT LTRIM(RTRIM(ISNULL(B.FREF, ''))) AS FREF
            FROM dbo.BO B
            LEFT JOIN dbo.BO3 B3 ON B3.BO3STAMP = B.BOSTAMP
            WHERE B.NDOS = {PF_NDOS}
              AND (LTRIM(RTRIM(ISNULL(B.FREF, ''))) IN ({','.join('?' for _ in batch)})
                   OR LTRIM(RTRIM(ISNULL(B3.DOCUMENTNUMBERORI, ''))) IN ({','.join('?' for _ in batch)}))
        """
        rows = fetch_dicts(cursor, sql, tuple(batch + batch))
        found.update(text(row["FREF"]) for row in rows)
    return found


def apply_import(conn, cursor, allocations: dict[str, list[dict[str, Any]]], headers: dict[int, dict[str, Any]], vat_code: int) -> dict[str, int]:
    grouped: dict[str, list[dict[str, Any]]] = allocations
    first_dates: dict[int, date] = {}
    for parts in grouped.values():
        for part in parts:
            first_dates.setdefault(part["excel"]["bc"], valid_date(part["excel"]["delivery_date"]))

    counters = {"bl": 0, "pf": 0, "lines": 0}
    for invoice, parts in sorted(grouped.items(), key=lambda item: (valid_date(item[1][0]["excel"]["delivery_date"]), item[0])):
        excel = parts[0]["excel"]
        delivery_date = valid_date(excel["delivery_date"])
        invoice_date = valid_date(excel["invoice_date"])
        bc = excel["bc"]
        source_header = headers[bc]
        net = amount(sum((amount(p["qty"] * p["source"]["price"]) for p in parts), ZERO))
        vat = amount(net * Decimal("0.20"))
        try:
            # Each supplier document commits independently. This keeps the
            # numbering lock very short and does not block active PHC users.
            bl_number = next_number(cursor, BL_NDOS, delivery_date.year)
            pf_number = next_number(cursor, PF_NDOS, invoice_date.year)
            bl_stamp, pf_stamp = _new_stamp(), _new_stamp()
            insert_header(cursor, bl_stamp, BL_NDOS, BL_NAME, bl_number, delivery_date, source_header,
                          excel["delivery"], net, vat, True)
            insert_header(cursor, pf_stamp, PF_NDOS, PF_NAME, pf_number, invoice_date, source_header,
                          invoice, net, vat, False)
            insert_tax(cursor, bl_stamp, vat_code, net, vat)
            insert_tax(cursor, pf_stamp, vat_code, net, vat)
            for index, part in enumerate(parts, start=1):
                source = part["source"]
                bl_line = insert_line(
                    cursor, header_stamp=bl_stamp, ndos=BL_NDOS, name=BL_NAME, number=bl_number,
                    doc_date=delivery_date, source_header=source_header, source=source, line_qty=part["qty"],
                    external_design=part["excel"]["design"], parent_bistamp=text(source["BISTAMP"]), closed=True,
                    lordem=index * 1000, vat_code=vat_code, pf_number=pf_number,
                )
                insert_line(
                    cursor, header_stamp=pf_stamp, ndos=PF_NDOS, name=PF_NAME, number=pf_number,
                    doc_date=invoice_date, source_header=source_header, source=source, line_qty=part["qty"],
                    external_design=part["excel"]["design"], parent_bistamp=bl_line, closed=False,
                    lordem=index * 1000, vat_code=vat_code,
                )
                counters["lines"] += 1
            conn.commit()
            counters["bl"] += 1; counters["pf"] += 1
        except Exception as exc:
            conn.rollback()
            raise ImportError(f"Falha ao gravar a fatura {invoice}: {exc}") from exc

    # All quantities were allocated exactly.  PHC therefore sees all three BC
    # and their lines as fully satisfied/closed.
    try:
        for number, header in headers.items():
            cursor.execute("UPDATE dbo.BI SET FECHADA = 1 WHERE BOSTAMP = ?", header["BOSTAMP"])
            cursor.execute("""
                UPDATE I2 SET QTTENC = I.QTT, QTTCOMPRA = I.QTT
                FROM dbo.BI2 I2 INNER JOIN dbo.BI I ON I.BISTAMP = I2.BI2STAMP
                WHERE I.BOSTAMP = ?
            """, header["BOSTAMP"])
            cursor.execute("UPDATE dbo.BO SET FECHADA = 1, DATAFECHO = ? WHERE BOSTAMP = ?", (date.today(), header["BOSTAMP"]))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    return counters


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, required=True)
    parser.add_argument("--execute", action="store_true", help="Grava documentos no PHC; por omissão é validação.")
    args = parser.parse_args()
    if not args.xlsx.is_file():
        raise ImportError(f"Ficheiro não encontrado: {args.xlsx}")

    rows = read_workbook(args.xlsx)
    # An avoir can contain both positive and negative rows.  It is identified
    # by a negative price in its own document and excluded as a whole.
    credit_invoices = {row["invoice"] for row in rows if row["price"] < ZERO}
    credit_rows = [row for row in rows if row["invoice"] in credit_invoices]
    usable = [row for row in rows if row["invoice"] not in credit_invoices]
    print(f"Excel: {len(rows)} linhas; {len(usable)} úteis; nota(s) de crédito excluída(s): {', '.join(sorted(credit_invoices)) or 'nenhuma'}.")
    bcs = {row["bc"] for row in usable}

    with app.app_context():
        conn = pyodbc.connect(_phc_conn_str(PHC_DB, PHC_SERVER), timeout=60)
        conn.autocommit = False
        cursor = conn.cursor()
        cursor.execute("SET LOCK_TIMEOUT 5000")
        try:
            headers, source_lines = load_source(cursor, bcs)
            allocations = allocate(usable, source_lines, credit_rows)
            invoices = set(allocations)
            prior = existing_invoices(cursor, invoices)
            if prior:
                raise ImportError("Já existem pré-faturas BMSO para: " + ", ".join(sorted(prior)[:20]))
            vat_code = get_tax_code(cursor, Decimal("20"))
            total = amount(sum((p["qty"] * p["source"]["price"] for parts in allocations.values() for p in parts), ZERO))
            print(f"Validação OK: {len(invoices)} BL + {len(invoices)} pré-faturas; {sum(len(x) for x in allocations.values())} linhas; HT {total} EUR.")
            if not args.execute:
                conn.rollback()
                print("Dry-run concluído. Nenhum registo foi alterado.")
                return 0
            counters = apply_import(conn, cursor, allocations, headers, vat_code)
            print(f"IMPORTAÇÃO CONCLUÍDA: {counters['bl']} BL fechados, {counters['pf']} pré-faturas abertas, {counters['lines']} linhas; BC 1047/1088/1137 fechados.")
            return 0
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except ImportError as exc:
        print(f"ERRO DE VALIDAÇÃO: {exc}", file=sys.stderr)
        raise SystemExit(2)

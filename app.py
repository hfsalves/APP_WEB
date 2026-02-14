import os
import tempfile
import subprocess
import shutil
import pyodbc
import json
import threading
import time
import uuid
import unicodedata
import re
import io
import importlib.util
from decimal import Decimal
from flask import Flask, render_template, request, jsonify, redirect, url_for, send_from_directory, session
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from datetime import datetime, date, timedelta, time as dtime
from sqlalchemy import text
from urllib.request import Request, urlopen
from urllib.parse import quote

# Importa a instÃ¢ncia db e modelos
from models import db, US, Menu, Acessos, Widget, UsWidget, MenuBotoes, Linhas
from models import Modais, CamposModal
#from your_app_folder import db  # ou ajusta conforme a tua estrutura

login_manager = LoginManager()

def new_stamp() -> str:
    return uuid.uuid4().hex.upper()[:25]

def create_app():
    app = Flask(__name__, static_folder='static', template_folder='templates')
    app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key')
    # Cache-busting para assets estÃ¡ticos (evita o browser usar JS antigo)
    app.config['STATIC_VERSION'] = int(time.time())
    app.config['SQLALCHEMY_DATABASE_URI'] = (
        "mssql+pyodbc://sa:enterprise@hfsalves.mooo.com,50002/GESTAO"
        "?driver=ODBC+Driver+17+for+SQL+Server"
        "&TrustServerCertificate=Yes&protocol=TCP&application_name=SZERO"
    )
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    radar_conn_str = os.environ.get('RADAR_CONN_STR') or (
        "DRIVER={ODBC Driver 17 for SQL Server};"
        "SERVER=hfsalves.mooo.com,50002;"
        "DATABASE=GESTAO;"
        "UID=sa;"
        "PWD=enterprise;"
        "TrustServerCertificate=Yes;"
        "Application Name=SZERO"
    )

    # Garantir JSON e respostas com UTF-8 para evitar problemas de acentuaÃ§Ã£o
    try:
        # Flask 2.x/3.x JSON provider
        app.json.ensure_ascii = False
    except Exception:
        app.config['JSON_AS_ASCII'] = False

    @app.after_request
    def _force_utf8(resp):
        try:
            mt = resp.mimetype or ''
            if mt.startswith('text/'):
                # acrescenta charset se nÃ£o existir
                if 'charset=' not in (resp.headers.get('Content-Type') or ''):
                    resp.headers['Content-Type'] = f"{mt}; charset=utf-8"
            elif mt == 'application/json':
                # alguns navegadores assumem latin-1 sem charset
                resp.headers['Content-Type'] = 'application/json; charset=utf-8'
        except Exception:
            pass
        return resp

    db.init_app(app)
    login_manager.init_app(app)
    login_manager.login_view = 'login'

    # Importa e regista blueprint genÃ©rico
    from blueprints.generic_crud import bp as generic_bp
    app.register_blueprint(generic_bp)

    from blueprints.anexos import bp as anexos_bp
    app.register_blueprint(anexos_bp)

    def _para_value_from_row(row):
        tipo = (row.get('TIPO') or '').strip().upper()
        if tipo == 'N':
            try:
                return float(row.get('NVALOR') or 0)
            except Exception:
                return 0.0
        if tipo == 'D':
            d = row.get('DVALOR')
            if isinstance(d, (date, datetime)):
                return d.strftime('%Y-%m-%d')
            return str(d or '')
        if tipo == 'L':
            return bool(row.get('LVALOR') or 0)
        return str(row.get('CVALOR') or '')

    def _load_para_map():
        rows = db.session.execute(text("""
            SELECT PARAMETRO, TIPO, CVALOR, DVALOR, NVALOR, LVALOR
            FROM dbo.PARA
        """)).mappings().all()
        out = {}
        for r in rows:
            key = (r.get('PARAMETRO') or '').strip()
            if not key:
                continue
            out[key] = _para_value_from_row(r)
        return out

    @app.route('/planeamento_limpezas')
    @app.route('/planeamento_limpezas/')
    @login_required
    def planeamento_limpezas_redirect():
        return redirect(url_for('generic.view_planeamento_limpezas'))

    # Favicon at root path
    @app.route('/favicon.ico')
    def favicon():
        try:
            return send_from_directory(
                os.path.join(app.root_path, 'static', 'images'),
                'favicon.ico',
                mimetype='image/x-icon'
            )
        except Exception:
            # As fallback, try /static/images/favicon.ico redirect
            return redirect(url_for('static', filename='images/favicon.ico'))

    @app.context_processor
    def inject_menu_and_access():
        page_name = None
        menu_items = []
        menu_structure = []
        perms = {}
        menu_botoes = {}
        menu_forms = {}

        if current_user.is_authenticated:
            # 1) Carrega todos os menus (filtra admin se for caso)
            if getattr(current_user, 'ADMIN', False):
                menu_items = Menu.query.order_by(Menu.ordem).all()
            else:
                menu_items = (
                    Menu.query
                        .filter_by(admin=False)
                        .order_by(Menu.ordem)
                        .all()
                )

            # 2) PermissÃµes de acesso por tabela
            rows = Acessos.query.filter_by(utilizador=current_user.LOGIN).all()
            perms = {
                a.tabela: {
                    'consultar': bool(a.consultar),
                    'inserir' : bool(a.inserir),
                    'editar'  : bool(a.editar),
                    'eliminar': bool(a.eliminar),
                }
                for a in rows
            }

            # VerificaÃ§Ã£o direta de acesso Ã  MN conforme pedido
            try:
                q_mn = db.session.execute(text(
                    "SELECT CONSULTAR FROM ACESSOS WHERE TABELA = 'MN' AND UTILIZADOR = :u"
                ), { 'u': current_user.LOGIN }).fetchone()
                can_open_mn = bool(getattr(current_user, 'ADMIN', False)) or (bool(q_mn[0]) if q_mn is not None else False)
            except Exception:
                can_open_mn = False

            try:
                q_fs = db.session.execute(text(
                    "SELECT CONSULTAR FROM ACESSOS WHERE TABELA = 'FS' AND UTILIZADOR = :u"
                ), { 'u': current_user.LOGIN }).fetchone()
                can_open_fs = bool(getattr(current_user, 'ADMIN', False)) or (bool(q_fs[0]) if q_fs is not None else False)
            except Exception:
                can_open_fs = False

            # 3) Determinar page_name e menu_botoes (igual lÃ³gica atual)
            parts = request.path.strip('/').split('/')
            if len(parts) >= 3 and parts[0] == 'generic' and parts[1] in ('view', 'form'):
                tabela_arg = parts[2]
                for m in menu_items:
                    if m.tabela == tabela_arg:
                        page_name = m.nome
                        botoes = (
                            MenuBotoes.query
                                .filter_by(TABELA=m.tabela, ATIVO=True)
                                .order_by(MenuBotoes.ORDEM)
                                .all()
                        )
                        menu_botoes = {
                            b.NOME: {
                                'icone'    : b.ICONE,
                                'texto'    : b.TEXTO,
                                'cor'      : b.COR,
                                'tipo'     : b.TIPO,
                                'acao'     : b.ACAO,
                                'condicao' : b.CONDICAO,
                                'destino'  : b.DESTINO,
                            }
                            for b in botoes
                        }
                        break

            if not page_name:
                for m in menu_items:
                    if request.path.startswith(m.url):
                        page_name = m.nome
                        break

            # 4) Montar estrutura do menu com permissÃµes
            menu_structure = []
            user_is_admin = getattr(current_user, 'ADMIN', False)

            # Lista de widgets do user (para o dashboard)
            user_widgets = set()
            if not user_is_admin:
                user_widgets = {uw.WIDGET for uw in UsWidget.query.filter_by(UTILIZADOR=current_user.LOGIN, VISIVEL=True)}

            current_group = None
            for m in menu_items:
                mostrar = False

                # Dashboard sÃ³ para quem tem widgets
                if m.tabela == "dashboard" and not user_is_admin:
                    mostrar = bool(user_widgets)
                # Monitor de Trabalho sempre visÃ­vel
                elif m.tabela == "monitor":
                    mostrar = True
                # Agrupadores (ordem % 100 == 0)
                elif m.ordem % 100 == 0:
                    mostrar = False  # SÃ³ serÃ¡ True se algum filho for permitido (mais abaixo)
                # Todos os outros: sÃ³ se tem acesso
                else:
                    mostrar = user_is_admin or perms.get(m.tabela, {}).get('consultar', False)

                # CriaÃ§Ã£o do grupo ou item
                if m.ordem % 100 == 0:
                    current_group = {
                        'name': m.nome,
                        'icon': m.icone,
                        'novo': bool(getattr(m, 'novo', False)),
                        'children': [],
                        'mostrar': False,  # SÃ³ serÃ¡ True se algum filho for mostrado
                    }
                    menu_structure.append(current_group)
                else:
                    child = {
                        'name': m.nome,
                        'url': m.url,
                        'icon': m.icone,
                        'novo': bool(getattr(m, 'novo', False))
                    }
                    if current_group:
                        if mostrar:
                            current_group['mostrar'] = True
                            current_group['children'].append(child)
                    elif mostrar:
                        # Se nÃ£o estÃ¡ num grupo, mete top-level
                        menu_structure.append(child)

            # Depois de montar, remove grupos sem filhos visÃ­veis
            menu_structure = [
                g for g in menu_structure
                if not isinstance(g, dict) or g.get('mostrar', True) or (g.get('children') and len(g['children']) > 0)
            ]

            menu_forms = { m.tabela: getattr(m, 'form', None) for m in menu_items }

        return {
            'menu_items'     : menu_items,
            'menu_structure' : menu_structure,
            'user_perms'     : perms,
            'page_name'      : page_name,
            'menu_botoes'    : menu_botoes,
            'menu_forms'     : menu_forms,
            'is_dev'         : getattr(current_user, 'DEV', False) if current_user.is_authenticated else False,
            'can_open_mn'    : can_open_mn if current_user.is_authenticated else False,
            'can_open_fs'    : can_open_fs if current_user.is_authenticated else False,
            'static_version' : app.config.get('STATIC_VERSION', 1),
            'app_params'     : session.get('APP_PARAMS', {}) if current_user.is_authenticated else {}
        }

    from sqlalchemy.sql import text

    @login_manager.user_loader
    def load_user(user_stamp):
        sql = text("""
            SELECT USSTAMP, LOGIN, NOME, EMAIL, COR, PASSWORD, ADMIN, EQUIPA, DEV, HOME, MNADMIN, LPADMIN, LSADMIN, FOTO, TEMPOS
            FROM US
            WHERE USSTAMP = :stamp
        """)
        row = db.session.execute(sql, {'stamp': user_stamp}).mappings().first()
        if not row:
            return None

        user = US()
        for k, v in row.items():
            setattr(user, k, v)
        return user

    # Rotas de autenticação
    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if request.method == 'POST':
            login_ = request.form['login']
            pwd = request.form['password']

            sql = text("""
                SELECT USSTAMP, LOGIN, NOME, EMAIL, COR, PASSWORD, ADMIN, EQUIPA, DEV, HOME, MNADMIN, LPADMIN, LSADMIN, FOTO, TEMPOS
                FROM US
                WHERE LOGIN = :login
            """)
            row = db.session.execute(sql, {'login': login_}).mappings().first()
            if row and row['PASSWORD'] == pwd:
                user = US()
                for k, v in row.items():
                    setattr(user, k, v)
                login_user(user)
                try:
                    para_map = _load_para_map()
                    session['APP_PARAMS'] = para_map
                    app.config['PARA_VALUES'] = para_map
                except Exception:
                    session['APP_PARAMS'] = {}
                return redirect(request.args.get('next') or url_for('home_page'))

            return render_template('login.html', error='Credenciais invÃ¡lidas')
        return render_template('login.html')

    @app.route('/logout')
    @login_required
    def logout():
        session.pop('APP_PARAMS', None)
        logout_user()
        return redirect(url_for('login'))

    @app.route('/')
    @login_required
    def home_page():
        home = getattr(current_user, 'HOME', '').lower().strip().lstrip('/')
        print(f"HOME do utilizador: {home}")

        # Se o utilizador regista tempos e tiver tarefas LP hoje, entre 08:00 e 20:00,
        # abre o ecrã de tempos em vez do monitor.
        try:
            tempos_enabled = int(getattr(current_user, 'TEMPOS', 0) or 0) == 1
        except Exception:
            tempos_enabled = False
        try:
            now_t = datetime.now().time()
            in_window = (dtime(8, 0) <= now_t <= dtime(20, 0))
        except Exception:
            in_window = False

        if tempos_enabled and in_window and (home == 'monitor' or not home):
            try:
                user_login = (getattr(current_user, 'LOGIN', '') or '').strip()
                if user_login:
                    has_lp_today = db.session.execute(text("""
                        SELECT TOP 1 1 AS X
                        FROM dbo.TAREFAS
                        WHERE LTRIM(RTRIM(ISNULL(ORIGEM,''))) = 'LP'
                          AND CAST(DATA AS date) = CAST(GETDATE() AS date)
                          AND ISNULL(UTILIZADOR,'') = :u
                          AND LTRIM(RTRIM(ISNULL(HORAFIM,''))) = ''
                    """), {'u': user_login}).fetchone()
                    if has_lp_today:
                        return redirect(url_for('tempos_limpeza_page'))
            except Exception:
                pass

        if home == 'dashboard':
            return redirect(url_for('dashboard_page'))
        elif home == 'monitor' or not home:
            return redirect(url_for('monitor_page'))

        return redirect(url_for('dashboard_page'))


    @app.route('/plan')
    @login_required
    def plan_page():
        return render_template('plan.html', today=date.today().isoformat())

    @app.route('/cleanings')
    @login_required
    def cleanings_page():
        return render_template('cleanings.html')

    @app.route('/cleanings/edit/<string:lpstamp>')
    @login_required
    def cleaning_edit_page(lpstamp):
        return render_template('cleaning_edit.html', lpstamp=lpstamp)

    @app.route('/dashboard')
    @login_required
    def dashboard_page():
        return render_template('dashboard.html')

    @app.route('/api/dashboard')
    @login_required
    def api_dashboard_widgets():
        user_login = current_user.LOGIN
        query = (
            db.session.query(UsWidget, Widget)
            .join(Widget, UsWidget.WIDGET == Widget.NOME)
            .filter(
                UsWidget.UTILIZADOR == user_login,
                UsWidget.VISIVEL == True,
                Widget.ATIVO == True
            )
            .order_by(UsWidget.COLUNA, UsWidget.ORDEM)
            .all()
        )
        dashboard = {1: [], 2: [], 3: []}
        for usw, widg in query:
            dashboard[usw.COLUNA].append({
                'nome': widg.NOME,
                'titulo': widg.TITULO,
                'tipo': widg.TIPO,
                'fonte': widg.FONTE,
                'config': widg.CONFIG,
                'filtros': getattr(widg, 'FILTROS', '') or '',
                'coluna': usw.COLUNA,
                'ordem': usw.ORDEM,
                'maxheight': usw.MAXHEIGHT
            })
        return jsonify(dashboard)

    @app.route('/api/widget/analise/<nome>')
    @login_required
    def widget_analise(nome):
        # 1. Carrega o widget
        widget = Widget.query.filter_by(NOME=nome, ATIVO=True).first()
        if not widget:
            return jsonify({'error': 'Widget nÃ£o encontrado'}), 404

        # 2. LÃª a query do CONFIG
        try:
            config = json.loads(widget.CONFIG)
            query = config.get('query')
            if not query:
                return jsonify({'error': 'Query nÃ£o definida no config'}), 400
        except Exception as e:
            return jsonify({'error': f'Config invÃ¡lido: {e}'}), 400

        # 3. Executa a query via SQLAlchemy
        try:
            result = db.session.execute(text(query))
        except Exception as e:
            return jsonify({'error': f'Erro ao executar query: {e}'}), 500

        # 4. ConstrÃ³i as linhas usando .mappings() para ter dicts
        mappings = result.mappings().all()
        rows = []
        for rd in mappings:
            clean = {}
            for col, val in rd.items():
                # Se for date ou datetime, passa para ISO string
                if isinstance(val, (date, datetime)):
                    clean[col] = val.strftime('%Y-%m-%d')
                else:
                    clean[col] = val
            rows.append(clean)

        # 5. Devolve colunas e linhas
        return jsonify({
            'columns': list(result.keys()),
            'rows': rows
        })

    @app.route('/api/widgets/<widget_id>/filters/options', methods=['POST'])
    @login_required
    def widget_filter_options(widget_id):
        data = request.get_json(silent=True) or {}
        options_query = (data.get('options_query') or '').strip()
        if not options_query:
            return jsonify({'error': 'options_query em falta'}), 400

        widget = Widget.query.filter((Widget.NOME == widget_id) | (Widget.WIDGETSSTAMP == widget_id)).first()
        if not widget or not widget.ATIVO:
            return jsonify({'error': 'Widget não encontrado'}), 404

        try:
            result = db.session.execute(text(options_query))
            options = []
            for row in result:
                if hasattr(row, '_mapping'):
                    m = row._mapping
                    val = m.get('value', list(m.values())[0])
                    lab = m.get('label', val)
                else:
                    val = row[0]
                    lab = row[1] if len(row) > 1 else val
                options.append({'value': val, 'label': lab})
            return jsonify({'options': options})
        except Exception as e:
            current_app.logger.exception('Erro em widget_filter_options')
            return jsonify({'error': str(e)}), 500

    @app.route('/api/widgets/<widget_id>/run', methods=['POST'])
    @login_required
    def widget_run(widget_id):
        payload = request.get_json(silent=True) or {}
        filters = payload.get('filters') or {}

        widget = Widget.query.filter((Widget.NOME == widget_id) | (Widget.WIDGETSSTAMP == widget_id)).first()
        if not widget or not widget.ATIVO:
            return jsonify({'error': 'Widget não encontrado'}), 404

        try:
            config = json.loads(widget.CONFIG or '{}')
            query = config.get('query')
            if not query:
                return jsonify({'error': 'Query não definida no config'}), 400
        except Exception as e:
            return jsonify({'error': f'Config inválido: {e}'}), 400

        # Prepara bind params apenas para as keys presentes na query (simples heurística por prefixo :)
        bind_params = {}
        for key, val in (filters.items() if isinstance(filters, dict) else []):
            bind_params[key] = val

        try:
            result = db.session.execute(text(query), bind_params)
            mappings = result.mappings().all()
            rows = []
            for rd in mappings:
                clean = {}
                for col, val in rd.items():
                    if isinstance(val, (date, datetime)):
                        clean[col] = val.strftime('%Y-%m-%d')
                    else:
                        clean[col] = val
                rows.append(clean)
            return jsonify({'columns': list(result.keys()), 'rows': rows})
        except Exception as e:
            current_app.logger.exception('Erro ao executar widget_run')
            return jsonify({'error': f'Erro ao executar query: {e}'}), 500

    # -----------------------------
    # Parâmetros
    # -----------------------------
    @app.route('/parametros')
    @login_required
    def parametros_page():
        return render_template('parametros.html', page_title='Parâmetros')

    @app.route('/api/parametros')
    @login_required
    def api_parametros():
        try:
            groups = db.session.execute(text("""
                SELECT DISTINCT LTRIM(RTRIM(ISNULL(GRUPO,''))) AS GRUPO
                FROM dbo.PARAG
                WHERE LTRIM(RTRIM(ISNULL(GRUPO,''))) <> ''
                ORDER BY LTRIM(RTRIM(ISNULL(GRUPO,'')))
            """)).fetchall()
            group_list = [str(g[0]).strip() for g in groups if g and g[0] is not None]

            rows = db.session.execute(text("""
                SELECT
                    PARASTAMP, PARAMETRO, DESCRICAO, TIPO, CVALOR, DVALOR, NVALOR, LVALOR, GRUPO
                FROM dbo.PARA
                ORDER BY LTRIM(RTRIM(ISNULL(GRUPO,''))), LTRIM(RTRIM(ISNULL(PARAMETRO,'')))
            """)).mappings().all()

            out = []
            for r in rows:
                tipo = (r.get('TIPO') or '').strip().upper()
                out.append({
                    'PARASTAMP': r.get('PARASTAMP'),
                    'PARAMETRO': (r.get('PARAMETRO') or '').strip(),
                    'DESCRICAO': (r.get('DESCRICAO') or '').strip(),
                    'TIPO': tipo,
                    'GRUPO': (r.get('GRUPO') or '').strip(),
                    'CVALOR': r.get('CVALOR') or '',
                    'DVALOR': (r.get('DVALOR').strftime('%Y-%m-%d') if isinstance(r.get('DVALOR'), (date, datetime)) else ''),
                    'NVALOR': float(r.get('NVALOR') or 0),
                    'LVALOR': 1 if int(r.get('LVALOR') or 0) else 0,
                    'VALOR': _para_value_from_row(r)
                })
            return jsonify({'groups': group_list, 'rows': out})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/parametros/save', methods=['POST'])
    @login_required
    def api_parametros_save():
        try:
            payload = request.get_json(silent=True) or {}
            rows = payload.get('rows') or []
            if not isinstance(rows, list):
                return jsonify({'error': 'Formato inválido'}), 400
            for r in rows:
                stamp = (r.get('PARASTAMP') or '').strip()
                tipo = (r.get('TIPO') or '').strip().upper()
                if not stamp:
                    continue
                if tipo == 'N':
                    db.session.execute(text("""
                        UPDATE dbo.PARA
                        SET NVALOR = :v
                        WHERE PARASTAMP = :s
                    """), {'v': float(r.get('NVALOR') or 0), 's': stamp})
                elif tipo == 'D':
                    d = (r.get('DVALOR') or '').strip()
                    db.session.execute(text("""
                        UPDATE dbo.PARA
                        SET DVALOR = CASE WHEN :d = '' THEN DVALOR ELSE CAST(:d AS date) END
                        WHERE PARASTAMP = :s
                    """), {'d': d, 's': stamp})
                elif tipo == 'L':
                    db.session.execute(text("""
                        UPDATE dbo.PARA
                        SET LVALOR = :v
                        WHERE PARASTAMP = :s
                    """), {'v': 1 if int(r.get('LVALOR') or 0) else 0, 's': stamp})
                else:
                    db.session.execute(text("""
                        UPDATE dbo.PARA
                        SET CVALOR = :v
                        WHERE PARASTAMP = :s
                    """), {'v': str(r.get('CVALOR') or '')[:200], 's': stamp})
            db.session.commit()
            para_map = _load_para_map()
            session['APP_PARAMS'] = para_map
            app.config['PARA_VALUES'] = para_map
            return jsonify({'ok': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/parametros/new', methods=['POST'])
    @login_required
    def api_parametros_new():
        try:
            payload = request.get_json(silent=True) or {}
            parametro = (payload.get('PARAMETRO') or '').strip()
            descricao = (payload.get('DESCRICAO') or '').strip()
            tipo = (payload.get('TIPO') or '').strip().upper()
            grupo = (payload.get('GRUPO') or '').strip()
            if not parametro or not descricao or tipo not in ('C', 'N', 'D', 'L') or not grupo:
                return jsonify({'error': 'Dados obrigatórios em falta.'}), 400

            exists = db.session.execute(text("""
                SELECT TOP 1 1
                FROM dbo.PARA
                WHERE UPPER(LTRIM(RTRIM(PARAMETRO))) = UPPER(LTRIM(RTRIM(:p)))
            """), {'p': parametro}).fetchone()
            if exists:
                return jsonify({'error': 'Parâmetro já existe.'}), 400

            stamp = new_stamp()
            cvalor = str(payload.get('CVALOR') or '')
            dvalor = (payload.get('DVALOR') or '').strip()
            nvalor = float(payload.get('NVALOR') or 0)
            lvalor = 1 if int(payload.get('LVALOR') or 0) else 0

            db.session.execute(text("""
                INSERT INTO dbo.PARA
                (PARASTAMP, PARAMETRO, DESCRICAO, TIPO, CVALOR, DVALOR, NVALOR, LVALOR, GRUPO)
                VALUES
                (:s, :p, :d, :t, :cv, CASE WHEN :dv = '' THEN CAST(GETDATE() AS date) ELSE CAST(:dv AS date) END, :nv, :lv, :g)
            """), {
                's': stamp,
                'p': parametro[:50],
                'd': descricao[:100],
                't': tipo,
                'cv': cvalor[:200],
                'dv': dvalor,
                'nv': nvalor,
                'lv': lvalor,
                'g': grupo[:50]
            })
            db.session.commit()
            para_map = _load_para_map()
            session['APP_PARAMS'] = para_map
            app.config['PARA_VALUES'] = para_map
            return jsonify({'ok': True, 'PARASTAMP': stamp})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/parametros/edit_meta', methods=['POST'])
    @login_required
    def api_parametros_edit_meta():
        try:
            payload = request.get_json(silent=True) or {}
            stamp = (payload.get('PARASTAMP') or '').strip()
            parametro = (payload.get('PARAMETRO') or '').strip()
            descricao = (payload.get('DESCRICAO') or '').strip()
            if not stamp or not parametro or not descricao:
                return jsonify({'error': 'Dados obrigatórios em falta.'}), 400

            dup = db.session.execute(text("""
                SELECT TOP 1 1
                FROM dbo.PARA
                WHERE UPPER(LTRIM(RTRIM(PARAMETRO))) = UPPER(LTRIM(RTRIM(:p)))
                  AND PARASTAMP <> :s
            """), {'p': parametro, 's': stamp}).fetchone()
            if dup:
                return jsonify({'error': 'Já existe outro parâmetro com esse código.'}), 400

            db.session.execute(text("""
                UPDATE dbo.PARA
                SET PARAMETRO = :p,
                    DESCRICAO = :d
                WHERE PARASTAMP = :s
            """), {'p': parametro[:50], 'd': descricao[:100], 's': stamp})
            db.session.commit()

            para_map = _load_para_map()
            session['APP_PARAMS'] = para_map
            app.config['PARA_VALUES'] = para_map
            return jsonify({'ok': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/parametros/delete', methods=['POST'])
    @login_required
    def api_parametros_delete():
        try:
            payload = request.get_json(silent=True) or {}
            stamp = (payload.get('PARASTAMP') or '').strip()
            if not stamp:
                return jsonify({'error': 'Parâmetro inválido.'}), 400
            db.session.execute(text("""
                DELETE FROM dbo.PARA
                WHERE PARASTAMP = :s
            """), {'s': stamp})
            db.session.commit()
            para_map = _load_para_map()
            session['APP_PARAMS'] = para_map
            app.config['PARA_VALUES'] = para_map
            return jsonify({'ok': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/parametros/group/new', methods=['POST'])
    @login_required
    def api_parametros_group_new():
        try:
            payload = request.get_json(silent=True) or {}
            grupo = (payload.get('GRUPO') or '').strip()
            if not grupo:
                return jsonify({'error': 'Grupo obrigatório.'}), 400
            exists = db.session.execute(text("""
                SELECT TOP 1 1
                FROM dbo.PARAG
                WHERE UPPER(LTRIM(RTRIM(GRUPO))) = UPPER(LTRIM(RTRIM(:g)))
            """), {'g': grupo}).fetchone()
            if not exists:
                db.session.execute(text("""
                    INSERT INTO dbo.PARAG (GRUPO) VALUES (:g)
                """), {'g': grupo[:50]})
                db.session.commit()
            return jsonify({'ok': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/modals/<acao>')
    @login_required
    def modal_generico(acao):
    # Exemplo de configuraÃ§Ã£o por acao
        if acao == 'agendar_tarefa':
            campos = [
            {'nome': 'PESSOA', 'label': 'Pessoa', 'tipo': 'text'},
            {'nome': 'DATA', 'label': 'Data', 'tipo': 'date'},
            {'nome': 'HORA', 'label': 'Hora', 'tipo': 'time'},
        ]
        return render_template('dynamic_modal.html', titulo="Agendar Tarefa", campos=campos, acao="gravarTarefa()")
    

    @app.route('/generic/api/modal/<modal_nome>', methods=['GET'])
    @login_required
    def api_modal_fields(modal_nome):
        from models import Modais, CamposModal
        modal = Modais.query.filter_by(NOME=modal_nome, ATIVO=True).first()
        if not modal:
            return jsonify({'success': False, 'message': 'Modal nÃ£o encontrado'}), 404

        campos = CamposModal.query.filter_by(MODAISSTAMP=modal.MODAISSTAMP).order_by(CamposModal.ORDEM).all()

        resultado = []
        for campo in campos:
            opcoes = []
            if campo.TIPO == 'COMBO' and campo.COMBO:
                try:
                    # âš ï¸ use text() aqui
                    res = db.session.execute(text(campo.COMBO))
                    for r in res.fetchall():
                        # adapte se o SELECT tiver sÃ³ uma coluna
                        opcoes.append([str(r[0]), str(r[1] if len(r)>1 else r[0])])
                except Exception as e:
                    print('Erro na query da combo:', e)

            resultado.append({
                'CAMPO': campo.CAMPO,
                'LABEL': campo.LABEL,
                'TIPO': campo.TIPO,
                'ORDEM': campo.ORDEM,
                'OPCOES': opcoes,
                'VALORDEFAULT': resolver_macros(campo.VALORDEFAULT)
            })

        return jsonify({'success': True, 'campos': resultado})

    @app.route('/generic/api/modal/gravar', methods=['POST'])
    @login_required
    def gravar_modal():
        dados = request.json
        nome_modal = dados.pop('__modal__', None)

        print(f"\nðŸ“¥ [MODAL] A gravar modal: {nome_modal}")
        print(f"ðŸ“¦ Dados recebidos: {dados}")

        modal = Modais.query.filter_by(NOME=nome_modal, ATIVO=True).first()
        if not modal:
            print(f"âŒ Modal '{nome_modal}' nÃ£o encontrado ou inativo.")
            return jsonify(success=False, error="Modal nÃ£o encontrado")

        try:
            campos = CamposModal.query.filter_by(MODAISSTAMP=modal.MODAISSTAMP).all()
            mapa = {
                c.CAMPO.upper(): (c.CAMPODESTINO or c.CAMPO)
                for c in campos
            }

            print(f"ðŸ” Mapeamento CAMPO â†’ CAMPODESTINO: {mapa}")

            dados_filtrados = {
                mapa[k.upper()]: v for k, v in dados.items() if k.upper() in mapa
            }

            print(f"âœ… Dados finais para INSERT na tabela {modal.TABELA}: {dados_filtrados}")

            if not dados_filtrados:
                print("âš ï¸ Nenhum campo corresponde ao mapping â€” abortado.")
                return jsonify(success=False, error="Nenhum dado vÃ¡lido para inserir")

            colunas = ', '.join(dados_filtrados.keys())
            marcadores = ', '.join([f':{k}' for k in dados_filtrados.keys()])
            sql = f"INSERT INTO {modal.TABELA} ({colunas}) VALUES ({marcadores})"

            print(f"ðŸ“ SQL gerado:\n{sql}")

            db.session.execute(text(sql), dados_filtrados)
            db.session.commit()
            print("âœ… InserÃ§Ã£o concluÃ­da com sucesso.")
            return jsonify(success=True)

        except Exception as e:
            db.session.rollback()
            print(f"âŒ ERRO durante gravaÃ§Ã£o:\n{e}")
            return jsonify(success=False, error=str(e)), 500
        

    from datetime import datetime
    from flask_login import current_user
    import uuid

    def resolver_macros(valor):
        def resolver_macros(valor):
            print(f"ðŸ§© A resolver macro: {valor}")

        if not isinstance(valor, str):
            return valor

        hoje = datetime.today().date()
        agora = datetime.now()

        macros = {
            '{TODAY}': hoje.strftime('%d.%m.%Y'),
            '{NOW}': agora.strftime('%H:%M'),
            '{USER}': current_user.LOGIN if current_user.is_authenticated else '',
            '{USERSTAMP}': current_user.get_id() if current_user.is_authenticated else '',
            '{UUID}': str(uuid.uuid4()),
        }

        for macro, real in macros.items():
            print(f"âž¡ï¸  Substituir {macro} por {real}")
            valor = valor.replace(macro, real)

        print(f"âœ… Resultado final: {valor}")
        return valor


    @app.route('/api/whoami')
    @login_required
    def whoami():
        from flask import jsonify

        user_data = {
            'USSTAMP': current_user.USSTAMP,
            'LOGIN': current_user.LOGIN,
            'NOME': current_user.NOME,
            'EMAIL': current_user.EMAIL,
            'COR': getattr(current_user, 'COR', None),
            'ADMIN': getattr(current_user, 'ADMIN', None),
            'DEV': getattr(current_user, 'DEV', None),
            'MNADMIN': getattr(current_user, 'MNADMIN', None),
            'LSADMIN': getattr(current_user, 'LSADMIN', None),
            'LPADMIN': getattr(current_user, 'LPADMIN', None),
            'EQUIPA': getattr(current_user, 'EQUIPA', None),
            'FOTO': getattr(current_user, 'FOTO', None)
        }

        return jsonify(user_data)


    @app.route('/analise/<usqlstamp>')
    @login_required
    def pagina_analise(usqlstamp):
        from models import Usql
        entry = Usql.query.filter_by(usqlstamp=usqlstamp).first()
        if not entry:
            return render_template('error.html', message='AnÃ¡lise nÃ£o encontrada'), 404
        return render_template('analise.html', usqlstamp=usqlstamp, titulo=entry.descricao)

    @app.route('/api/analise/<usqlstamp>')
    @login_required
    def api_analise(usqlstamp):
        from models import Usql
        entry = Usql.query.filter_by(usqlstamp=usqlstamp).first()
        if not entry:
            return jsonify({'error': 'AnÃ¡lise nÃ£o encontrada'}), 404

        try:
            result = db.session.execute(text(entry.sqlexpr))
            mappings = result.mappings().all()
            rows = []
            for rd in mappings:
                clean = {}
                for col, val in rd.items():
                    if isinstance(val, (datetime, date)):
                        clean[col] = val.strftime('%Y-%m-%d')
                    else:
                        clean[col] = val
                rows.append(clean)

            return jsonify({
                'columns': list(result.keys()),
                'rows': rows,
                'decimais': float(entry.decimais or 2),
                'totais': bool(entry.totais)
            })

        except Exception as e:
            return jsonify({'error': str(e)}), 500


    @app.route('/monitor')
    @login_required
    def monitor_page():
        return render_template('monitor.html')

    @app.route('/posicoes_pesquisas')
    @login_required
    def posicoes_pesquisas_alias():
        return redirect(url_for('pesquisas_posicoes_page'))

    @app.route('/pesquisas/posicoes')
    @login_required
    def pesquisas_posicoes_page():
        return render_template('posicoes_pesquisas.html', page_title='Posições de Pesquisas')

    @app.route('/api/pesquisas/posicoes')
    @login_required
    def api_pesquisas_posicoes():
        sql_blocos = """
DECLARE @Hoje date = CAST(GETDATE() AS date);
DECLARE @DataInicio date = @Hoje;
DECLARE @DataFim    date = DATEADD(day, 30, @Hoje);

;WITH
Aloj AS (
    SELECT
        al.nome as CCUSTO,
        al.TIPO,
        al.NMAIRBNB
    FROM dbo.AL al
    WHERE 1=1
      AND ISNULL(al.FECHADO, 0) = 0
),
Cal AS (
    SELECT @DataInicio AS [DATA]
    UNION ALL
    SELECT DATEADD(day, 1, [DATA])
    FROM Cal
    WHERE [DATA] < @DataFim
),
Occ AS (
    SELECT DISTINCT
        v.CCUSTO,
        CAST(v.[DATA] AS date) AS [DATA]
    FROM dbo.v_diario_all v
    WHERE v.[DATA] BETWEEN @DataInicio AND @DataFim
      AND ISNULL(v.VALOR, 0) <> 0  -- ignorar check-out (VALOR=0)
),
Livres AS (
    SELECT
        a.CCUSTO,
        a.TIPO,
        a.NMAIRBNB,
        c.[DATA]
    FROM Aloj a
    CROSS JOIN Cal c
    LEFT JOIN Occ o
      ON o.CCUSTO = a.CCUSTO
     AND o.[DATA] = c.[DATA]
    WHERE o.CCUSTO IS NULL
),
LivresComGrupo AS (
    SELECT
        CCUSTO,
        TIPO,
        NMAIRBNB,
        [DATA],
        DATEADD(day, -ROW_NUMBER() OVER (PARTITION BY CCUSTO ORDER BY [DATA]), [DATA]) AS grp
    FROM Livres
),
LivresValidas AS (
    SELECT
        l.CCUSTO,
        l.TIPO,
        l.NMAIRBNB,
        l.[DATA],
        l.grp
    FROM LivresComGrupo l
    JOIN (
        SELECT CCUSTO, grp
        FROM LivresComGrupo
        GROUP BY CCUSTO, grp
        HAVING COUNT(*) >= 2
    ) g
      ON g.CCUSTO = l.CCUSTO
     AND g.grp    = l.grp
    WHERE
        DATEDIFF(day, @Hoje, l.[DATA]) BETWEEN 0 AND 30
),
Blocos AS (
    SELECT
        CCUSTO,
        TIPO,
        NMAIRBNB,
        MIN([DATA]) AS DataInicio,
        MAX([DATA]) AS DataFim,
        COUNT(*) AS Noites
    FROM LivresValidas
    GROUP BY CCUSTO, TIPO, NMAIRBNB, grp
)
SELECT
    b.CCUSTO,
    b.TIPO,
    b.NMAIRBNB,
    b.DataInicio,
    b.DataFim,
    b.Noites
FROM Blocos b
ORDER BY b.CCUSTO, b.DataInicio
OPTION (MAXRECURSION 32767);
        """

        sql_pesquisas = """
DECLARE @Hoje date = CAST(GETDATE() AS date);
DECLARE @DataInicio date = @Hoje;
DECLARE @DataFim    date = DATEADD(day, 30, @Hoje);

SELECT
    PESQUISASSTAMP,
    CCUSTO,
    DATA,
    NOITES,
    HOSPEDES,
    DTPESQUISA,
    PAGINA,
    POSICAO
FROM PESQUISAS
WHERE DATA BETWEEN @DataInicio AND @DataFim
ORDER BY CCUSTO, DATA, DTPESQUISA DESC;
        """

        blocos = []
        pesquisas = []
        try:
            with pyodbc.connect(radar_conn_str) as conn:
                cur = conn.cursor()
                cur.execute(sql_blocos)
                cols = [c[0] for c in cur.description]
                for raw in cur.fetchall():
                    entry = {}
                    for col_name, val in zip(cols, raw):
                        if isinstance(val, (datetime, date)):
                            entry[col_name] = val.isoformat()
                        else:
                            entry[col_name] = val
                    blocos.append(entry)

                cur.execute(sql_pesquisas)
                cols2 = [c[0] for c in cur.description]
                for raw in cur.fetchall():
                    entry = {}
                    for col_name, val in zip(cols2, raw):
                        if isinstance(val, (datetime, date)):
                            entry[col_name] = val.isoformat()
                        else:
                            entry[col_name] = val
                    pesquisas.append(entry)
        except Exception:
            try:
                app.logger.exception("Erro ao carregar posições de pesquisas")
            except Exception:
                pass
            return jsonify({"error": "Não foi possível carregar os dados"}), 500

        return jsonify({"blocks": blocos, "pesquisas": pesquisas})

    @app.route('/api/pesquisas/posicoes/<stamp>', methods=['DELETE'])
    @login_required
    def api_pesquisas_posicoes_delete(stamp):
        if not stamp:
            return jsonify({"error": "stamp obrigatório"}), 400
        try:
            with pyodbc.connect(radar_conn_str) as conn:
                cur = conn.cursor()
                cur.execute("DELETE FROM PESQUISAS WHERE PESQUISASSTAMP = ?", stamp)
                conn.commit()
        except Exception:
            try:
                app.logger.exception("Erro ao eliminar pesquisa")
            except Exception:
                pass
            return jsonify({"error": "Não foi possível eliminar"}), 500
        return jsonify({"ok": True})

    @app.route('/api/pesquisas/posicoes/<stamp>/relaunch', methods=['POST'])
    @login_required
    def api_pesquisas_posicoes_relaunch(stamp):
        if not stamp:
            return jsonify({"error": "stamp obrigatório"}), 400
        try:
            with pyodbc.connect(radar_conn_str) as conn:
                cur = conn.cursor()
                cur.execute("""
                    SELECT CCUSTO, DATA, NOITES, HOSPEDES, NMAIRBNB
                    FROM PESQUISAS
                    WHERE PESQUISASSTAMP = ?
                """, stamp)
                row = cur.fetchone()
                if not row:
                    return jsonify({"error": "Pesquisa não encontrada"}), 404
                ccusto, data_val, noites_val, hosp_val, nmairbnb = row
                novo_stamp = datetime.now().strftime("%Y%m%d%H%M%S%f")[:25]
                cur.execute("""
                    INSERT INTO PESQUISAS
                        (PESQUISASSTAMP, CCUSTO, DATA, DTPESQUISA, NOITES, HOSPEDES, PAGINA, POSICAO, NMAIRBNB)
                    VALUES
                        (?, ?, ?, CAST(GETDATE() AS date), ?, ?, 0, 0, ?)
                """, novo_stamp, ccusto, data_val, noites_val, hosp_val, nmairbnb)
                conn.commit()
        except Exception:
            try:
                app.logger.exception("Erro ao relançar pesquisa")
            except Exception:
                pass
            return jsonify({"error": "Não foi possível relançar"}), 500
        return jsonify({"ok": True, "stamp": novo_stamp})

    @app.route('/api/pesquisas/posicoes', methods=['POST'])
    @login_required
    def api_pesquisas_posicoes_insert():
        payload = request.get_json(silent=True) or {}
        ccusto = (payload.get('ccusto') or '').strip()
        data_inicio = payload.get('data')
        noites = payload.get('noites')
        hospedes = payload.get('hospedes')
        nmairbnb = (payload.get('nmairbnb') or '').strip()

        if not ccusto or not data_inicio or not noites:
            return jsonify({"error": "ccusto, data e noites são obrigatórios"}), 400

        try:
            noites_int = int(noites)
            if noites_int < 1:
                noites_int = 1
        except Exception:
            noites_int = 1

        try:
            hospedes_int = int(hospedes or 2)
            if hospedes_int < 1:
                hospedes_int = 2
        except Exception:
            hospedes_int = 2

        try:
            data_obj = datetime.fromisoformat(data_inicio).date()
        except Exception:
            return jsonify({"error": "Data inválida"}), 400

        stamp = datetime.now().strftime("%Y%m%d%H%M%S%f")[:25]
        insert_sql = """
INSERT INTO PESQUISAS
    (PESQUISASSTAMP, CCUSTO, DATA, DTPESQUISA, NOITES, HOSPEDES, PAGINA, POSICAO, NMAIRBNB)
VALUES
    (?, ?, ?, CAST(GETDATE() AS date), ?, ?, 0, 0, ?);
        """

        try:
            with pyodbc.connect(radar_conn_str) as conn:
                cur = conn.cursor()
                cur.execute(insert_sql, stamp, ccusto, data_obj, str(noites_int), hospedes_int, nmairbnb)
                conn.commit()
        except Exception:
            try:
                app.logger.exception("Erro ao inserir pesquisa para posição")
            except Exception:
                pass
            return jsonify({"error": "Não foi possível registar a pesquisa"}), 500

        return jsonify({"ok": True, "stamp": stamp})

    @app.route('/radar')
    @login_required
    def radar_page():
        radar_rows = []
        radar_error = None
        radar_sql = """
/*  RADAR (1 linha por alojamento)
    - HOJE+1 até HOJE+30
    - Noites livres = noites que NÃO existem em v_diario_all (por CCUSTO, DATA)
    - Só conta livres vendáveis: blocos consecutivos com >= 2 noites
    - Métricas por alojamento:
        Livres_D7, Livres_D14, Livres_D30
        Pressao_D7, Pressao_D14, Pressao_D30
        ADR_60d (média VALOR>0 últimos 60 dias; fallback = média global 60d)
        Sugestão de Ação (prioriza risco D7)
*/

DECLARE @Hoje date = CAST(GETDATE() AS date);
DECLARE @DataInicio date = DATEADD(day, 1, @Hoje);
DECLARE @DataFim    date = DATEADD(day, 30, @Hoje);

;WITH
Aloj AS (
    SELECT
        al.nome as CCUSTO,
        al.TIPO
    FROM dbo.AL al
    WHERE 1=1
      -- aplica aqui os teus filtros (fechados/inativos)
      AND ISNULL(al.FECHADO, 0) = 0
      -- AND ISNULL(al.ATIVO, 1) = 1
),
Cal AS (
    SELECT @DataInicio AS [DATA]
    UNION ALL
    SELECT DATEADD(day, 1, [DATA])
    FROM Cal
    WHERE [DATA] < @DataFim
),
Occ AS (
    -- ocupadas = existem na view
    SELECT DISTINCT
        v.CCUSTO,
        CAST(v.[DATA] AS date) AS [DATA]
    FROM dbo.v_diario_all v
    WHERE v.[DATA] BETWEEN @DataInicio AND @DataFim
),
Livres AS (
    -- livres = (alojamento x calendário) sem registo em Occ
    SELECT
        a.CCUSTO,
        a.TIPO,
        c.[DATA]
    FROM Aloj a
    CROSS JOIN Cal c
    LEFT JOIN Occ o
      ON o.CCUSTO = a.CCUSTO
     AND o.[DATA] = c.[DATA]
    WHERE o.CCUSTO IS NULL
),
LivresComGrupo AS (
    -- ilhas de livres consecutivas
    SELECT
        CCUSTO,
        TIPO,
        [DATA],
        DATEADD(day, -ROW_NUMBER() OVER (PARTITION BY CCUSTO ORDER BY [DATA]), [DATA]) AS grp
    FROM Livres
),
LivresValidas AS (
    -- só blocos com >= 2 noites; e já com bucket temporal
    SELECT
        l.CCUSTO,
        l.TIPO,
        l.[DATA],
        DATEDIFF(day, @Hoje, l.[DATA]) AS Dias_Ate,
        CASE
            WHEN DATEDIFF(day, @Hoje, l.[DATA]) BETWEEN 1 AND 7  THEN 'D7'
            WHEN DATEDIFF(day, @Hoje, l.[DATA]) BETWEEN 8 AND 14 THEN 'D14'
            WHEN DATEDIFF(day, @Hoje, l.[DATA]) BETWEEN 15 AND 30 THEN 'D30'
        END AS Janela
    FROM LivresComGrupo l
    JOIN (
        SELECT CCUSTO, grp
        FROM LivresComGrupo
        GROUP BY CCUSTO, grp
        HAVING COUNT(*) >= 2
    ) g
      ON g.CCUSTO = l.CCUSTO
     AND g.grp    = l.grp
    WHERE
        DATEDIFF(day, @Hoje, l.[DATA]) BETWEEN 1 AND 30
),
LivresAgg AS (
    SELECT
        CCUSTO,
        TIPO,
        SUM(CASE WHEN Janela = 'D7'  THEN 1 ELSE 0 END)  AS Livres_D7,
        SUM(CASE WHEN Janela = 'D14' THEN 1 ELSE 0 END)  AS Livres_D14,
        SUM(CASE WHEN Janela = 'D30' THEN 1 ELSE 0 END)  AS Livres_D30
    FROM LivresValidas
    GROUP BY CCUSTO, TIPO
),
ADR_Aloj_60d AS (
    SELECT
        v.CCUSTO,
        AVG(CAST(v.VALOR AS decimal(18,6))) AS ADR_60d
    FROM dbo.v_diario_all v
    WHERE v.[DATA] >= DATEADD(day, -60, @Hoje)
      AND v.[DATA] <  @Hoje
      AND v.VALOR IS NOT NULL
      AND v.VALOR > 0
    GROUP BY v.CCUSTO
),
ADR_Portfolio_60d AS (
    SELECT
        AVG(CAST(v.VALOR AS decimal(18,6))) AS ADR_Portfolio_60d
    FROM dbo.v_diario_all v
    WHERE v.[DATA] >= DATEADD(day, -60, @Hoje)
      AND v.[DATA] <  @Hoje
      AND v.VALOR IS NOT NULL
      AND v.VALOR > 0
),
Base AS (
    SELECT
        a.CCUSTO AS Alojamento,
        a.TIPO,
        ISNULL(l.Livres_D7,  0) AS Livres_D7,
        ISNULL(l.Livres_D14, 0) AS Livres_D14,
        ISNULL(l.Livres_D30, 0) AS Livres_D30,
        CAST(1.0 * ISNULL(l.Livres_D7,  0) / 7.0  AS decimal(10,4)) AS Pressao_D7,
        CAST(1.0 * ISNULL(l.Livres_D14, 0) / 7.0  AS decimal(10,4)) AS Pressao_D14,
        CAST(1.0 * ISNULL(l.Livres_D30, 0) / 16.0 AS decimal(10,4)) AS Pressao_D30,
        COALESCE(adr.ADR_60d, p.ADR_Portfolio_60d) AS ADR_Usado_60d,
        p.ADR_Portfolio_60d
    FROM Aloj a
    LEFT JOIN LivresAgg l
      ON l.CCUSTO = a.CCUSTO
    LEFT JOIN ADR_Aloj_60d adr
      ON adr.CCUSTO = a.CCUSTO
    CROSS JOIN ADR_Portfolio_60d p
)
SELECT
    Alojamento,
    TIPO,
    Livres_D7,
    Livres_D14,
    Livres_D30,
    Pressao_D7,
    Pressao_D14,
    Pressao_D30,
    ADR_Usado_60d,
    ADR_Portfolio_60d,
    CAST((ADR_Usado_60d / NULLIF(ADR_Portfolio_60d,0)) - 1 AS decimal(10,4)) AS Desvio_ADR,
    CASE
        WHEN Pressao_D7 >= 0.40 AND ADR_Usado_60d >= ADR_Portfolio_60d
            THEN 'Urgente: baixar preco / abrir regras (D7)'
        WHEN Pressao_D7 >= 0.40
            THEN 'Urgente: mexer em regras/anuncio (D7)'
        WHEN Pressao_D14 >= 0.50 AND ADR_Usado_60d >= ADR_Portfolio_60d
            THEN 'Ajuste: preco ligeiro / min nights (D14)'
        WHEN Pressao_D14 >= 0.50
            THEN 'Ajuste: monitorizar + regras (D14)'
        WHEN Pressao_D30 >= 0.60
            THEN 'Atencao: comecar a trabalhar D30'
        ELSE 'OK'
    END AS Acao
FROM Base
ORDER BY
    CASE
        WHEN Pressao_D7  >= 0.40 THEN 1
        WHEN Pressao_D14 >= 0.50 THEN 2
        WHEN Pressao_D30 >= 0.60 THEN 3
        ELSE 4
    END,
    Pressao_D7 DESC,
    Pressao_D14 DESC,
    Pressao_D30 DESC,
    Alojamento
OPTION (MAXRECURSION 32767);
        """

        try:
            with pyodbc.connect(radar_conn_str) as conn:
                cur = conn.cursor()
                cur.execute(radar_sql)
                cols = [c[0] for c in cur.description]
                for raw in cur.fetchall():
                    row = {}
                    for col_name, val in zip(cols, raw):
                        if isinstance(val, Decimal):
                            row[col_name] = float(val)
                        elif isinstance(val, (datetime, date)):
                            row[col_name] = val.isoformat()
                        else:
                            row[col_name] = val
                    radar_rows.append(row)
        except Exception:
            radar_error = "Não foi possível carregar o radar neste momento. Tenta novamente em breve."
            try:
                app.logger.exception("Erro ao obter dados do Radar de Atenção")
            except Exception:
                pass

        status_code = 500 if radar_error else 200
        return render_template(
            'radar.html',
            page_title='Radar de Atenção',
            radar_rows=radar_rows,
            radar_error=radar_error
        ), status_code

    @app.route('/mapa-gestao')
    @login_required
    def mapa_gestao_page():
        return render_template('mapa_gestao.html', page_title='Mapa de Gestao', ano_atual=date.today().year)

    @app.route('/mapa-controlo')
    @login_required
    def mapa_controlo_page():
        hoje = date.today()
        return render_template(
            'mapa_controlo.html',
            page_title='Mapa de Controlo',
            ano_atual=hoje.year,
            mes_atual=hoje.month
        )

    @app.route('/orcamento')
    @login_required
    def orcamento_page():
        return render_template('orcamento.html', page_title='Orçamento', ano_atual=date.today().year)

    @app.route('/api/orcamento')
    @login_required
    def api_orcamento_get():
        try:
            ano = request.args.get('ano', type=int) or date.today().year
        except Exception:
            ano = date.today().year

        try:
            fam_rows = db.session.execute(text("SELECT ref, nome FROM v_stfami ORDER BY ref")).fetchall()
        except Exception as e:
            return jsonify({'error': f'Erro ao obter familias: {e}'}), 500

        def level_from_ref(ref: str) -> int:
            try:
                return ref.count('.') + 1
            except Exception:
                return 1

        def sort_key(ref: str):
            parts = []
            for p in str(ref or '').split('.'):
                try:
                    parts.append(int(p))
                except Exception:
                    parts.append(p)
            return parts

        def is_prov_ref(rf: str) -> bool:
            try:
                return str(rf or '').strip().startswith('9')
            except Exception:
                return False

        familias_map = {}
        for r in fam_rows:
            ref = str(r[0]).strip() if r and r[0] is not None else ''
            if not ref:
                continue
            if is_prov_ref(ref):
                continue  # orçamento só de custos
            nome = r[1] if len(r) > 1 else ''
            nivel = level_from_ref(ref)
            familias_map[ref] = {
                'ref': ref,
                'nome': nome,
                'nivel': nivel,
                'editable': False,
                'meses': [0.0] * 12
            }

        # Apenas famílias de movimento (folhas) são editáveis:
        # - se existir 3 níveis num ramo => nível 3 (folha)
        # - se existir 2 níveis num ramo => nível 2 (folha)
        # - se existir 1 nível => nível 1 (folha)
        parents = set()
        for ref in familias_map.keys():
            parts = ref.split('.')
            if len(parts) <= 1:
                continue
            # marca todos os prefixos como pais
            for i in range(1, len(parts)):
                parents.add('.'.join(parts[:i]))
        for ref, f in familias_map.items():
            f['editable'] = (ref not in parents)

        # carregar valores existentes
        try:
            oc_rows = db.session.execute(
                text("SELECT FAMILIA, MES, VALOR FROM OC WHERE ANO = :ano"),
                {'ano': ano}
            ).fetchall()
        except Exception as e:
            return jsonify({'error': f'Erro ao obter orçamento (OC): {e}'}), 500

        for r in oc_rows:
            fam = str(r[0]).strip() if r[0] is not None else ''
            try:
                mes = int(r[1])
            except Exception:
                mes = 0
            val = float(r[2] or 0)
            if not fam or mes < 1 or mes > 12:
                continue
            node = familias_map.get(fam)
            if node is None:
                # se existir orçamento para família não mapeada, ignora
                continue
            node['meses'][mes - 1] = round(val, 2)

        # agrega para pais (mostrar totais por nível, mas só folhas editáveis gravam)
        # regra: pais = soma dos filhos diretos (bottom-up)
        refs_sorted = sorted(familias_map.keys(), key=lambda x: (familias_map[x]['nivel'], sort_key(x)), reverse=True)
        for ref in refs_sorted:
            node = familias_map[ref]
            if node.get('editable'):
                continue
            prefix = ref + '.'
            children = [c for c in familias_map.keys() if c.startswith(prefix) and familias_map[c]['nivel'] == node['nivel'] + 1]
            if not children:
                continue
            for m in range(12):
                node['meses'][m] = round(sum(float(familias_map[c]['meses'][m] or 0) for c in children), 2)

        familias_lista = []
        for ref in sorted(familias_map.keys(), key=sort_key):
            f = familias_map[ref]
            meses = [round(float(v or 0), 2) for v in f['meses']]
            familias_lista.append({
                'ref': f['ref'],
                'nome': f.get('nome', ''),
                'nivel': int(f.get('nivel') or 1),
                'editable': bool(f.get('editable')),
                'meses': meses
            })

        return jsonify({'ano': ano, 'editable_rule': 'leaf', 'familias': familias_lista})

    @app.route('/api/orcamento/batch', methods=['POST'])
    @login_required
    def api_orcamento_batch():
        try:
            body = request.get_json(silent=True) or {}
            try:
                ano = int(body.get('ano') or date.today().year)
            except Exception:
                ano = date.today().year
            updates = body.get('updates') or []
            if not isinstance(updates, list) or not updates:
                return jsonify({'ok': True, 'updated': 0})

            updated = 0
            for u in updates[:500]:
                fam = str(u.get('familia') or '').strip()
                try:
                    mes = int(u.get('mes') or 0)
                except Exception:
                    mes = 0
                try:
                    valor = float(u.get('valor') or 0)
                except Exception:
                    valor = 0.0
                if not fam or mes < 1 or mes > 12:
                    continue

                # find existing
                row = db.session.execute(
                    text("SELECT TOP 1 OCSTAMP FROM OC WHERE ANO=:ano AND MES=:mes AND FAMILIA=:fam"),
                    {'ano': ano, 'mes': mes, 'fam': fam}
                ).fetchone()
                if row and row[0]:
                    db.session.execute(
                        text("UPDATE OC SET VALOR=:valor WHERE OCSTAMP=:id"),
                        {'valor': valor, 'id': row[0]}
                    )
                else:
                    new_id_row = db.session.execute(text("SELECT LEFT(CONVERT(varchar(36), NEWID()), 25)")).fetchone()
                    oc_id = new_id_row[0]
                    db.session.execute(
                        text("INSERT INTO OC (OCSTAMP, ANO, MES, FAMILIA, VALOR) VALUES (:id,:ano,:mes,:fam,:valor)"),
                        {'id': oc_id, 'ano': ano, 'mes': mes, 'fam': fam, 'valor': valor}
                    )
                updated += 1
            db.session.commit()
            return jsonify({'ok': True, 'updated': updated})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/mapa_controlo')
    @login_required
    def api_mapa_controlo():
        try:
            ano = request.args.get('ano', type=int) or date.today().year
        except Exception:
            ano = date.today().year
        try:
            mes = request.args.get('mes', type=int) or date.today().month
        except Exception:
            mes = date.today().month
        if mes < 1 or mes > 12:
            mes = date.today().month

        sql = text("""
            SELECT
                LTRIM(RTRIM(C.CCUSTO)) AS CCUSTO,
                SUM(CASE WHEN LEFT(LTRIM(RTRIM(C.FAMILIA)), 1) = '9' THEN ISNULL(C.TOTAL,0) ELSE 0 END) AS PROVEITO,
                SUM(CASE WHEN LEFT(LTRIM(RTRIM(C.FAMILIA)), 1) <> '9' AND LTRIM(RTRIM(C.REF)) = 'RENDA' THEN ISNULL(C.TOTAL,0) ELSE 0 END) AS RENDAS,
                SUM(CASE WHEN LEFT(LTRIM(RTRIM(C.FAMILIA)), 1) <> '9' AND LTRIM(RTRIM(C.REF)) IN ('LUZ-6','LUZ-23') THEN ISNULL(C.TOTAL,0) ELSE 0 END) AS LUZ,
                SUM(CASE WHEN LEFT(LTRIM(RTRIM(C.FAMILIA)), 1) <> '9' AND LTRIM(RTRIM(C.REF)) IN ('AGUA','SANEAMENTO') THEN ISNULL(C.TOTAL,0) ELSE 0 END) AS AGUA,
                SUM(CASE WHEN LEFT(LTRIM(RTRIM(C.FAMILIA)), 1) <> '9' AND LTRIM(RTRIM(C.REF)) = 'COMUNICACOES' THEN ISNULL(C.TOTAL,0) ELSE 0 END) AS COMUNICACOES,
                SUM(CASE
                      WHEN LEFT(LTRIM(RTRIM(C.FAMILIA)), 1) <> '9'
                       AND NOT (
                          LTRIM(RTRIM(C.REF)) = 'RENDA'
                          OR LTRIM(RTRIM(C.REF)) IN ('LUZ-6','LUZ-23')
                          OR LTRIM(RTRIM(C.REF)) IN ('AGUA','SANEAMENTO')
                          OR LTRIM(RTRIM(C.REF)) = 'COMUNICACOES'
                       )
                      THEN ISNULL(C.TOTAL,0)
                      ELSE 0
                END) AS OUTROS
            FROM v_custo C
            INNER JOIN AL a
              ON LTRIM(RTRIM(a.CCUSTO)) COLLATE SQL_Latin1_General_CP1_CI_AI
               = LTRIM(RTRIM(C.CCUSTO)) COLLATE SQL_Latin1_General_CP1_CI_AI
            WHERE YEAR(C.DATA) = :ano
              AND MONTH(C.DATA) = :mes
              AND UPPER(LTRIM(RTRIM(ISNULL(a.TIPO,'')))) = 'EXPLORACAO'
            GROUP BY LTRIM(RTRIM(C.CCUSTO))
            ORDER BY LTRIM(RTRIM(C.CCUSTO))
        """)
        try:
            rows = db.session.execute(sql, {'ano': ano, 'mes': mes}).mappings().all()
        except Exception as e:
            return jsonify({'error': f'Erro ao obter mapa controlo: {e}'}), 500

        out_rows = []
        totals = {
            'ccusto': 'TOTAL',
            'proveito': 0.0,
            'rendas': 0.0,
            'luz': 0.0,
            'agua': 0.0,
            'comunicacoes': 0.0,
            'outros': 0.0,
            'total_custos': 0.0,
            'saldo': 0.0
        }
        for r in rows:
            ccusto = (r.get('CCUSTO') or '').strip()
            proveito = float(r.get('PROVEITO') or 0)
            rendas = float(r.get('RENDAS') or 0)
            luz = float(r.get('LUZ') or 0)
            agua = float(r.get('AGUA') or 0)
            comunic = float(r.get('COMUNICACOES') or 0)
            outros = float(r.get('OUTROS') or 0)
            total_custos = rendas + luz + agua + comunic + outros
            saldo = proveito - total_custos
            out_rows.append({
                'ccusto': ccusto,
                'proveito': round(proveito, 2),
                'rendas': round(rendas, 2),
                'luz': round(luz, 2),
                'agua': round(agua, 2),
                'comunicacoes': round(comunic, 2),
                'outros': round(outros, 2),
                'total_custos': round(total_custos, 2),
                'saldo': round(saldo, 2)
            })
            totals['proveito'] += proveito
            totals['rendas'] += rendas
            totals['luz'] += luz
            totals['agua'] += agua
            totals['comunicacoes'] += comunic
            totals['outros'] += outros

        totals['total_custos'] = totals['rendas'] + totals['luz'] + totals['agua'] + totals['comunicacoes'] + totals['outros']
        totals['saldo'] = totals['proveito'] - totals['total_custos']
        for k in ('proveito','rendas','luz','agua','comunicacoes','outros','total_custos','saldo'):
            totals[k] = round(float(totals[k] or 0), 2)

        return jsonify({
            'ano': ano,
            'mes': mes,
            'rows': out_rows,
            'totals': totals
        })

    @app.route('/api/mapa_gestao/ccustos')
    @login_required
    def api_mapa_gestao_ccustos():
        try:
            rows = db.session.execute(text("""
                SELECT
                    c.CCUSTO,
                    ISNULL(a.TIPO,'') AS TIPO
                FROM (
                    SELECT DISTINCT
                        LTRIM(RTRIM(CCUSTO)) AS CCUSTO
                    FROM v_cct
                    WHERE CCUSTO IS NOT NULL
                      AND LTRIM(RTRIM(CCUSTO)) <> ''
                ) c
                LEFT JOIN (
                    SELECT
                        LTRIM(RTRIM(CCUSTO)) AS CCUSTO,
                        MAX(ISNULL(TIPO,'')) AS TIPO
                    FROM AL
                    GROUP BY LTRIM(RTRIM(CCUSTO))
                ) a
                  ON a.CCUSTO COLLATE SQL_Latin1_General_CP1_CI_AI
                   = c.CCUSTO COLLATE SQL_Latin1_General_CP1_CI_AI
                ORDER BY ISNULL(a.TIPO,''), c.CCUSTO
            """)).fetchall()
            opts = []
            for r in rows:
                cc = r[0] if len(r) > 0 else None
                tipo = r[1] if len(r) > 1 else ''
                if cc is None:
                    continue
                opts.append({'ccusto': cc, 'tipo': tipo})
            return jsonify({'options': opts})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/mapa_gestao')
    @login_required
    def api_mapa_gestao():
        try:
            ano = request.args.get('ano', type=int) or date.today().year
        except Exception:
            ano = date.today().year

        ccustos_raw = (request.args.get('ccustos') or '').strip()
        ccustos = [c.strip() for c in ccustos_raw.split(',') if c.strip()]

        try:
            fam_rows = db.session.execute(text("SELECT ref, nome FROM v_stfami ORDER BY ref")).fetchall()
        except Exception as e:
            return jsonify({'error': f'Erro ao obter familias: {e}'}), 500

        def level_from_ref(ref: str) -> int:
            try:
                return ref.count('.') + 1
            except Exception:
                return 1

        familias_map = {}
        for r in fam_rows:
            ref = str(r[0]).strip() if r and r[0] is not None else ''
            if not ref:
                continue
            nome = r[1] if len(r) > 1 else ''
            familias_map[ref] = {
                'ref': ref,
                'nome': nome,
                'nivel': level_from_ref(ref),
                'meses': [0.0] * 12,
                'total': 0.0,
                'orc_meses': [0.0] * 12,
                'orc_total': 0.0,
            }

        def ensure_node(ref: str):
            if ref not in familias_map:
                familias_map[ref] = {
                    'ref': ref,
                    'nome': ref,
                    'nivel': level_from_ref(ref),
                    'meses': [0.0] * 12,
                    'total': 0.0,
                    'orc_meses': [0.0] * 12,
                    'orc_total': 0.0,
                }

        def is_prov_ref(rf: str) -> bool:
            try:
                return str(rf or '').strip().startswith('9')
            except Exception:
                return False

        # Custos por familia/mes
        where_parts = ["YEAR(DATA) = :ano"]
        params = {'ano': ano}
        if ccustos:
            keys = []
            for idx, cc in enumerate(ccustos):
                k = f"c{idx}"
                params[k] = cc
                keys.append(f":{k}")
            where_parts.append("CCUSTO IN (" + ",".join(keys) + ")")

        where_sql = " AND ".join(where_parts)
        custos_sql = f"""
            SELECT FAMILIA, MONTH(DATA) AS MES, SUM(TOTAL) AS TOTAL
            FROM v_custo
            WHERE {where_sql}
            GROUP BY FAMILIA, MONTH(DATA)
        """

        try:
            custo_rows = db.session.execute(text(custos_sql), params).fetchall()
        except Exception as e:
            return jsonify({'error': f'Erro ao obter custos: {e}'}), 500

        for row in custo_rows:
            familia = str(row[0]).strip() if row[0] is not None else ''
            mes = int(row[1]) if row[1] is not None else None
            valor = float(row[2] or 0)
            if not familia or mes is None or mes < 1 or mes > 12:
                continue

            # acumula na familia e respetivos pais
            partes = familia.split('.')
            while partes:
                ref_atual = '.'.join(partes)
                ensure_node(ref_atual)
                node = familias_map[ref_atual]
                node['meses'][mes - 1] += valor
                node['total'] += valor
                partes = partes[:-1]

        # Orçamento (OC) por familia/mes (custos)
        try:
            oc_rows = db.session.execute(
                text("SELECT FAMILIA, MES, VALOR FROM OC WHERE ANO = :ano"),
                {'ano': ano}
            ).fetchall()
        except Exception as e:
            return jsonify({'error': f'Erro ao obter orçamento (OC): {e}'}), 500

        for row in oc_rows:
            familia = str(row[0]).strip() if row[0] is not None else ''
            try:
                mes = int(row[1]) if row[1] is not None else None
            except Exception:
                mes = None
            valor = float(row[2] or 0)
            if not familia or mes is None or mes < 1 or mes > 12:
                continue
            if is_prov_ref(familia):
                continue
            # acumula no nó e respetivos pais (para permitir evidenciar também níveis superiores)
            partes = familia.split('.')
            while partes:
                ref_atual = '.'.join(partes)
                ensure_node(ref_atual)
                node = familias_map[ref_atual]
                node['orc_meses'][mes - 1] += valor
                node['orc_total'] += valor
                partes = partes[:-1]

        # Objetivos de venda (OV) para proveitos:
        # - 9.1 => EXPLORACAO
        # - 9.2 => GESTAO
        try:
            where_ov = ["ov.ANO = :ano"]
            params_ov = {'ano': ano}
            if ccustos:
                keys = []
                for idx, cc in enumerate(ccustos):
                    k = f"cc{idx}"
                    params_ov[k] = cc
                    keys.append(f":{k}")
                where_ov.append("ov.CCUSTO IN (" + ",".join(keys) + ")")
            sql_ov = f"""
                SELECT
                    ov.CCUSTO,
                    UPPER(LTRIM(RTRIM(ISNULL(a.TIPO,'')))) AS TIPO,
                    ISNULL(ov.JANEIRO,0)   AS JANEIRO,
                    ISNULL(ov.FEVEREIRO,0) AS FEVEREIRO,
                    ISNULL(ov.MARCO,0)     AS MARCO,
                    ISNULL(ov.ABRIL,0)     AS ABRIL,
                    ISNULL(ov.MAIO,0)      AS MAIO,
                    ISNULL(ov.JUNHO,0)     AS JUNHO,
                    ISNULL(ov.JULHO,0)     AS JULHO,
                    ISNULL(ov.AGOSTO,0)    AS AGOSTO,
                    ISNULL(ov.SETEMBRO,0)  AS SETEMBRO,
                    ISNULL(ov.OUTUBRO,0)   AS OUTUBRO,
                    ISNULL(ov.NOVEMBRO,0)  AS NOVEMBRO,
                    ISNULL(ov.DEZEMBRO,0)  AS DEZEMBRO
                FROM OV ov
                LEFT JOIN AL a
                  ON LTRIM(RTRIM(a.CCUSTO)) COLLATE SQL_Latin1_General_CP1_CI_AI
                   = LTRIM(RTRIM(ov.CCUSTO)) COLLATE SQL_Latin1_General_CP1_CI_AI
                WHERE {" AND ".join(where_ov)}
            """
            ov_rows = db.session.execute(text(sql_ov), params_ov).mappings().all()
        except Exception as e:
            return jsonify({'error': f'Erro ao obter objetivos (OV): {e}'}), 500

        expl = [0.0] * 12
        gest = [0.0] * 12
        for r in ov_rows:
            tipo = (r.get('TIPO') or '').strip().upper()
            months = [
                float(r.get('JANEIRO') or 0),
                float(r.get('FEVEREIRO') or 0),
                float(r.get('MARCO') or 0),
                float(r.get('ABRIL') or 0),
                float(r.get('MAIO') or 0),
                float(r.get('JUNHO') or 0),
                float(r.get('JULHO') or 0),
                float(r.get('AGOSTO') or 0),
                float(r.get('SETEMBRO') or 0),
                float(r.get('OUTUBRO') or 0),
                float(r.get('NOVEMBRO') or 0),
                float(r.get('DEZEMBRO') or 0),
            ]
            if tipo == 'EXPLORACAO':
                for i in range(12):
                    expl[i] += months[i]
            elif tipo == 'GESTAO':
                for i in range(12):
                    gest[i] += months[i]

        ensure_node('9')
        ensure_node('9.1')
        ensure_node('9.2')
        for i in range(12):
            familias_map['9.1']['orc_meses'][i] += expl[i]
            familias_map['9.1']['orc_total'] += expl[i]
            familias_map['9.2']['orc_meses'][i] += gest[i]
            familias_map['9.2']['orc_total'] += gest[i]
            familias_map['9']['orc_meses'][i] += (expl[i] + gest[i])
            familias_map['9']['orc_total'] += (expl[i] + gest[i])

        def sort_key(ref: str):
            parts = []
            for p in ref.split('.'):
                try:
                    parts.append(int(p))
                except Exception:
                    parts.append(p)
            return parts

        total_base = sum(
            v['total']
            for v in familias_map.values()
            if v.get('nivel') == 1 and not is_prov_ref(v.get('ref'))
        )
        familias_lista = []
        for f in sorted(familias_map.values(), key=lambda x: sort_key(x['ref'])):
            total_fam = float(f['total'] or 0)
            meses_fmt = [round(float(v or 0), 2) for v in f['meses']]
            is_proveito = is_prov_ref(f.get('ref'))
            orc_meses_fmt = [round(float(v or 0), 2) for v in (f.get('orc_meses') or [0.0] * 12)]
            orc_total = round(float(f.get('orc_total') or 0), 2)
            familias_lista.append({
                'ref': f['ref'],
                'nome': f.get('nome', ''),
                'nivel': f.get('nivel', 1),
                'meses': meses_fmt,
                'total': round(total_fam, 2),
                'percent': (None if is_proveito else round((total_fam / total_base * 100) if total_base else 0, 2)),
                'orc_meses': orc_meses_fmt,
                'orc_total': orc_total,
            })

        # KPI: preco medio por noite (reservas RS) = SUM((ESTADIA+LIMPEZA) - COMISSAO) / SUM(NOITES)
        kpi_adr_valor = None
        kpi_adr_noites = 0
        kpi_adr_net = 0.0
        kpi_rs_reservas = 0
        kpi_rs_hospedes = 0
        try:
            where_rs = [
                "YEAR(ISNULL(RS.DATAOUT, RS.DATAIN)) = :ano",
                "ISNULL(RS.CANCELADA,0) = 0",
                "ISNULL(RS.NOITES,0) > 0",
                # excluir reservas do proprietario (valor medio/noite = 0)
                "ABS(((ISNULL(RS.ESTADIA,0) + ISNULL(RS.LIMPEZA,0)) - ISNULL(RS.COMISSAO,0))) > 0.005",
            ]
            params_rs = {'ano': ano}
            if ccustos:
                keys = []
                for idx, cc in enumerate(ccustos):
                    k = f"cc{idx}"
                    params_rs[k] = cc
                    keys.append(f":{k}")
                where_rs.append(
                    "LTRIM(RTRIM(ISNULL(a.CCUSTO,''))) COLLATE SQL_Latin1_General_CP1_CI_AI IN ("
                    + ",".join(keys)
                    + ")"
                )
            sql_rs = f"""
                SELECT
                    SUM((ISNULL(RS.ESTADIA,0) + ISNULL(RS.LIMPEZA,0)) - ISNULL(RS.COMISSAO,0)) AS NET_TOTAL,
                    SUM(ISNULL(RS.NOITES,0)) AS NOITES_TOTAL
                FROM RS
                LEFT JOIN AL a
                  ON LTRIM(RTRIM(a.NOME)) COLLATE SQL_Latin1_General_CP1_CI_AI
                   = LTRIM(RTRIM(ISNULL(RS.ALOJAMENTO,''))) COLLATE SQL_Latin1_General_CP1_CI_AI
                WHERE {" AND ".join(where_rs)}
            """
            r = db.session.execute(text(sql_rs), params_rs).fetchone()
            if r:
                kpi_adr_net = float((r[0] or 0) if len(r) > 0 else 0)
                kpi_adr_noites = int((r[1] or 0) if len(r) > 1 else 0)
                if kpi_adr_noites > 0:
                    kpi_adr_valor = round(kpi_adr_net / kpi_adr_noites, 2)
        except Exception:
            try:
                app.logger.exception("Erro ao calcular KPI preco medio/noite (RS)")
            except Exception:
                pass
            kpi_adr_valor = None

        # KPI: numero de reservas + numero de hospedes (adultos+criancas)
        try:
            where_rs2 = ["YEAR(ISNULL(RS.DATAOUT, RS.DATAIN)) = :ano", "ISNULL(RS.CANCELADA,0) = 0"]
            params_rs2 = {'ano': ano}
            if ccustos:
                keys = []
                for idx, cc in enumerate(ccustos):
                    k = f"ccx{idx}"
                    params_rs2[k] = cc
                    keys.append(f":{k}")
                where_rs2.append(
                    "LTRIM(RTRIM(ISNULL(a.CCUSTO,''))) COLLATE SQL_Latin1_General_CP1_CI_AI IN ("
                    + ",".join(keys)
                    + ")"
                )
            sql_rs2 = f"""
                SELECT
                    COUNT(1) AS RESERVAS,
                    SUM(ISNULL(RS.ADULTOS,0) + ISNULL(RS.CRIANCAS,0)) AS HOSPEDES
                FROM RS
                LEFT JOIN AL a
                  ON LTRIM(RTRIM(a.NOME)) COLLATE SQL_Latin1_General_CP1_CI_AI
                   = LTRIM(RTRIM(ISNULL(RS.ALOJAMENTO,''))) COLLATE SQL_Latin1_General_CP1_CI_AI
                WHERE {" AND ".join(where_rs2)}
            """
            r2 = db.session.execute(text(sql_rs2), params_rs2).fetchone()
            if r2:
                kpi_rs_reservas = int((r2[0] or 0) if len(r2) > 0 else 0)
                kpi_rs_hospedes = int((r2[1] or 0) if len(r2) > 1 else 0)
        except Exception:
            try:
                app.logger.exception("Erro ao calcular KPI reservas/hospedes (RS)")
            except Exception:
                pass

        # KPI: media por dia do ano (reservas recebidas por RDATA, mesma formula do grafico diario)
        kpi_res_ano_total = 0.0
        kpi_res_ano_media_dia = None
        kpi_fat_ano_total = 0.0
        kpi_fat_ano_media_dia = None
        try:
            start_y = date(ano, 1, 1)
            end_y = date(ano + 1, 1, 1)
            today_d = date.today()
            if ano == today_d.year:
                end_window = min(end_y, today_d + timedelta(days=1))  # até hoje (inclui hoje)
            else:
                end_window = end_y
            days_in_year = (end_window - start_y).days or 365

            where_rsy = ["RS.RDATA >= :start", "RS.RDATA < :end"]
            params_rsy = {'start': start_y, 'end': end_window}
            if ccustos:
                keys = []
                for idx, cc in enumerate(ccustos):
                    k = f"ccy{idx}"
                    params_rsy[k] = cc
                    keys.append(f":{k}")
                where_rsy.append(
                    "LTRIM(RTRIM(ISNULL(a.CCUSTO,''))) COLLATE SQL_Latin1_General_CP1_CI_AI IN ("
                    + ",".join(keys)
                    + ")"
                )

            sql_rsy = f"""
                SELECT
                    SUM(
                        CASE
                            WHEN UPPER(LTRIM(RTRIM(ISNULL(a.TIPO,'')))) = 'GESTAO' THEN
                                CASE WHEN ISNULL(a.FTLIMPEZA,0) = 0 THEN
                                    (CASE WHEN ISNULL(RS.CANCELADA,0) = 1 THEN ISNULL(RS.PCANCEL,0)
                                          ELSE ROUND(ISNULL(RS.ESTADIA,0)/1.06,2) + ROUND(ISNULL(RS.LIMPEZA,0)/1.23,2) - ISNULL(RS.COMISSAO,0)
                                     END) * (ISNULL(a.COMISSAO,0) / 100.0)
                                ELSE
                                    (CASE WHEN ISNULL(RS.CANCELADA,0) = 1 THEN ISNULL(RS.PCANCEL,0)
                                          ELSE ROUND(ISNULL(RS.ESTADIA,0)/1.06,2) - ISNULL(RS.COMISSAO,0)
                                     END) * (ISNULL(a.COMISSAO,0) / 100.0)
                                    + ROUND(ISNULL(RS.LIMPEZA,0)/1.23,2)
                                END
                            ELSE
                                CASE WHEN ISNULL(RS.CANCELADA,0) = 1 THEN ISNULL(RS.PCANCEL,0)
                                     ELSE ROUND(ISNULL(RS.ESTADIA,0)/1.06,2) + ROUND(ISNULL(RS.LIMPEZA,0)/1.23,2) - ISNULL(RS.COMISSAO,0)
                                END
                        END
                    ) AS TOTAL
                FROM RS
                LEFT JOIN AL a
                  ON LTRIM(RTRIM(a.NOME)) COLLATE SQL_Latin1_General_CP1_CI_AI
                   = LTRIM(RTRIM(ISNULL(RS.ALOJAMENTO,''))) COLLATE SQL_Latin1_General_CP1_CI_AI
                WHERE {" AND ".join(where_rsy)}
            """
            ry = db.session.execute(text(sql_rsy), params_rsy).fetchone()
            if ry:
                kpi_res_ano_total = float((ry[0] or 0) if len(ry) > 0 else 0)
                kpi_res_ano_media_dia = round(kpi_res_ano_total / float(days_in_year), 2) if days_in_year else None

            # KPI 7: Faturado líquido até hoje / dia (v_diario_all)
            where_fat = [
                "V.DATA >= :start",
                "V.DATA < :end",
                "ISNULL(V.VALOR,0) <> 0",
            ]
            params_fat = {'start': start_y, 'end': end_window}
            if ccustos:
                keys = []
                for idx, cc in enumerate(ccustos):
                    k = f"ccf{idx}"
                    params_fat[k] = cc
                    keys.append(f":{k}")
                where_fat.append(
                    "LTRIM(RTRIM(ISNULL(V.CCUSTO,''))) COLLATE SQL_Latin1_General_CP1_CI_AI IN ("
                    + ",".join(keys)
                    + ")"
                )
            sql_fat = f"""
                SELECT SUM(ISNULL(V.VALOR,0)) AS TOTAL
                FROM v_diario_all V
                WHERE {" AND ".join(where_fat)}
            """
            rf = db.session.execute(text(sql_fat), params_fat).fetchone()
            if rf:
                kpi_fat_ano_total = float((rf[0] or 0) if len(rf) > 0 else 0)
                kpi_fat_ano_media_dia = round(kpi_fat_ano_total / float(days_in_year), 2) if days_in_year else None
        except Exception:
            try:
                app.logger.exception("Erro ao calcular KPI media/dia reservas (RS)")
            except Exception:
                pass
            kpi_res_ano_total = 0.0
            kpi_res_ano_media_dia = None
            kpi_fat_ano_total = 0.0
            kpi_fat_ano_media_dia = None

        return jsonify({
            'ano': ano,
            'ccustos': ccustos,
            'total_geral': round(float(total_base or 0), 2),
            'kpis': {
                'preco_medio_noite': kpi_adr_valor,
                'preco_medio_noite_noites': kpi_adr_noites,
                'preco_medio_noite_net': round(float(kpi_adr_net or 0), 2),
                'numero_reservas': kpi_rs_reservas,
                'numero_hospedes': kpi_rs_hospedes,
                'reservas_ano_total': round(float(kpi_res_ano_total or 0), 2),
                'reservas_ano_media_dia': kpi_res_ano_media_dia,
                'faturado_ano_total': round(float(kpi_fat_ano_total or 0), 2),
                'faturado_ano_media_dia': kpi_fat_ano_media_dia,
            },
            'familias': familias_lista
        })

    @app.route('/api/mapa_gestao/detalhe')
    @login_required
    def api_mapa_gestao_detalhe():
        try:
            ano = request.args.get('ano', type=int) or date.today().year
        except Exception:
            ano = date.today().year
        familia = (request.args.get('familia') or '').strip()
        if not familia:
            return jsonify({'error': 'familia obrigatoria'}), 400
        mes = request.args.get('mes', type=int)
        include_children = request.args.get('include_children') in ('1', 'true', 'True')
        ccustos_raw = (request.args.get('ccustos') or '').strip()
        ccustos = [c.strip() for c in ccustos_raw.split(',') if c.strip()]

        where_parts = ["YEAR(data) = :ano"]
        params = {'ano': ano, 'familia': familia}
        if include_children:
            where_parts.append("(FAMILIA = :familia OR FAMILIA LIKE :familia_like)")
            params['familia_like'] = familia + ".%"
        else:
            where_parts.append("FAMILIA = :familia")
        if mes and 1 <= mes <= 12:
            where_parts.append("MONTH(data) = :mes")
            params['mes'] = mes
        if ccustos:
            keys = []
            for idx, cc in enumerate(ccustos):
                k = f"c{idx}"
                params[k] = cc
                keys.append(f":{k}")
            where_parts.append("ccusto IN (" + ",".join(keys) + ")")

        where_sql = " AND ".join(where_parts)
        sql = f"""
            SELECT
                nmdoc AS documento,
                nrdoc AS numero,
                data  AS data,
                nome  AS nome,
                ccusto AS ccusto,
                ref    AS referencia,
                design AS designacao,
                qtt    AS quantidade,
                epv    AS preco,
                total  AS total,
                CABSTAMP AS cabstamp
            FROM v_custo
            WHERE {where_sql}
            ORDER BY data, nmdoc, nrdoc
        """
        try:
            rows = db.session.execute(text(sql), params).mappings().all()
        except Exception as e:
            return jsonify({'error': f'Erro ao obter detalhe: {e}'}), 500

        # Orçamento/Objetivo:
        # - Custos => OC (por família/mês)
        # - Proveitos (famílias 9.1/9.2) => OV (por CCUSTO, mapeado via AL.TIPO)
        orc_total = 0.0
        familia_is_prov = str(familia or '').strip().startswith('9')
        if familia_is_prov:
            try:
                where_ov = ["ov.ANO = :ano"]
                params_ov = {'ano': ano}
                if ccustos:
                    keys = []
                    for idx, cc in enumerate(ccustos):
                        k = f"cc{idx}"
                        params_ov[k] = cc
                        keys.append(f":{k}")
                    where_ov.append("ov.CCUSTO IN (" + ",".join(keys) + ")")
                sql_ov = f"""
                    SELECT
                        UPPER(LTRIM(RTRIM(ISNULL(a.TIPO,'')))) AS TIPO,
                        SUM(ISNULL(ov.JANEIRO,0))   AS JANEIRO,
                        SUM(ISNULL(ov.FEVEREIRO,0)) AS FEVEREIRO,
                        SUM(ISNULL(ov.MARCO,0))     AS MARCO,
                        SUM(ISNULL(ov.ABRIL,0))     AS ABRIL,
                        SUM(ISNULL(ov.MAIO,0))      AS MAIO,
                        SUM(ISNULL(ov.JUNHO,0))     AS JUNHO,
                        SUM(ISNULL(ov.JULHO,0))     AS JULHO,
                        SUM(ISNULL(ov.AGOSTO,0))    AS AGOSTO,
                        SUM(ISNULL(ov.SETEMBRO,0))  AS SETEMBRO,
                        SUM(ISNULL(ov.OUTUBRO,0))   AS OUTUBRO,
                        SUM(ISNULL(ov.NOVEMBRO,0))  AS NOVEMBRO,
                        SUM(ISNULL(ov.DEZEMBRO,0))  AS DEZEMBRO
                    FROM OV ov
                    LEFT JOIN AL a
                      ON LTRIM(RTRIM(a.CCUSTO)) COLLATE SQL_Latin1_General_CP1_CI_AI
                       = LTRIM(RTRIM(ov.CCUSTO)) COLLATE SQL_Latin1_General_CP1_CI_AI
                    WHERE {" AND ".join(where_ov)}
                    GROUP BY UPPER(LTRIM(RTRIM(ISNULL(a.TIPO,''))))
                """
                ov_rows = db.session.execute(text(sql_ov), params_ov).mappings().all()
                by_tipo = { (r.get('TIPO') or '').strip().upper(): r for r in ov_rows }
                exp = by_tipo.get('EXPLORACAO', {}) or {}
                ges = by_tipo.get('GESTAO', {}) or {}
                month_cols = ['JANEIRO','FEVEREIRO','MARCO','ABRIL','MAIO','JUNHO','JULHO','AGOSTO','SETEMBRO','OUTUBRO','NOVEMBRO','DEZEMBRO']
                if mes and 1 <= mes <= 12:
                    col = month_cols[mes - 1]
                    val_exp = float(exp.get(col) or 0)
                    val_ges = float(ges.get(col) or 0)
                    fam = str(familia or '').strip()
                    if fam.startswith('9.1'):
                        orc_total = val_exp
                    elif fam.startswith('9.2'):
                        orc_total = val_ges
                    else:
                        # 9 (ou outras) -> soma
                        orc_total = val_exp + val_ges
                else:
                    val_exp = sum(float(exp.get(c) or 0) for c in month_cols)
                    val_ges = sum(float(ges.get(c) or 0) for c in month_cols)
                    fam = str(familia or '').strip()
                    if fam.startswith('9.1'):
                        orc_total = val_exp
                    elif fam.startswith('9.2'):
                        orc_total = val_ges
                    else:
                        orc_total = val_exp + val_ges
            except Exception as e:
                return jsonify({'error': f'Erro ao obter objetivos (OV): {e}'}), 500
        else:
            try:
                oc_where = ["ANO = :ano"]
                oc_params = {'ano': ano, 'familia': familia}
                if include_children:
                    oc_where.append("(FAMILIA = :familia OR FAMILIA LIKE :familia_like)")
                    oc_params['familia_like'] = familia + ".%"
                else:
                    oc_where.append("FAMILIA = :familia")
                if mes and 1 <= mes <= 12:
                    oc_where.append("MES = :mes")
                    oc_params['mes'] = mes
                oc_sql = "SELECT SUM(ISNULL(VALOR,0)) AS ORCAMENTO FROM OC WHERE " + " AND ".join(oc_where)
                oc_row = db.session.execute(text(oc_sql), oc_params).mappings().first() or {}
                orc_total = float(oc_row.get('ORCAMENTO') or 0)
            except Exception as e:
                return jsonify({'error': f'Erro ao obter orçamento (OC): {e}'}), 500

        out = []
        total_sum = 0.0
        cabstamps = []
        for r in rows:
            val = float(r.get('total') or 0)
            total_sum += val
            data_val = r.get('data')
            if isinstance(data_val, (datetime, date)):
                data_val = data_val.strftime('%Y-%m-%d')
            cabstamp = r.get('cabstamp')
            if cabstamp:
                cabstamps.append(str(cabstamp))
            out.append({
                'documento': r.get('documento'),
                'numero': r.get('numero'),
                'data': data_val,
                'nome': r.get('nome'),
                'ccusto': r.get('ccusto'),
                'referencia': r.get('referencia'),
                'designacao': r.get('designacao'),
                'quantidade': r.get('quantidade'),
                'preco': r.get('preco'),
                'total': round(val, 2),
                'cabstamp': cabstamp
            })

        anexo_map = {}
        if cabstamps:
            placeholders = []
            params_anx = {}
            for idx, cab in enumerate(cabstamps):
                key = f"c{idx}"
                params_anx[key] = cab
                placeholders.append(f":{key}")
            sql_anx = f"""
                SELECT RECSTAMP, CAMINHO, FICHEIRO
                FROM ANEXOS
                WHERE RECSTAMP IN ({",".join(placeholders)})
                ORDER BY RECSTAMP
            """
            try:
                rows_anx = db.session.execute(text(sql_anx), params_anx).fetchall()
                for ra in rows_anx:
                    rec = str(ra[0])
                    if rec and rec not in anexo_map:
                        caminho = ra[1] if len(ra) > 1 else None
                        ficheiro = ra[2] if len(ra) > 2 else None
                        anexo_map[rec] = caminho or ficheiro or None
            except Exception:
                anexo_map = {}

        for item in out:
            cab = item.get('cabstamp')
            url = anexo_map.get(str(cab)) if cab else None
            item['anexo_url'] = url

        desvio = total_sum - orc_total
        desvio_pct = None
        if orc_total:
            try:
                desvio_pct = (desvio / orc_total) * 100
            except Exception:
                desvio_pct = None

        return jsonify({
            'rows': out,
            'total': round(total_sum, 2),
            'orc_total': round(float(orc_total or 0), 2),
            'desvio': round(float(desvio or 0), 2),
            'desvio_pct': (None if desvio_pct is None else round(float(desvio_pct), 2))
        })

    @app.route('/api/mapa_gestao/reservas_diarias')
    @login_required
    def api_mapa_gestao_reservas_diarias():
        from datetime import timedelta

        try:
            ano = request.args.get('ano', type=int) or date.today().year
        except Exception:
            ano = date.today().year
        try:
            mes = request.args.get('mes', type=int) or date.today().month
        except Exception:
            mes = date.today().month
        if mes < 1 or mes > 12:
            return jsonify({'error': 'mes invalido'}), 400

        ccustos_raw = (request.args.get('ccustos') or '').strip()
        ccustos = [c.strip() for c in ccustos_raw.split(',') if c.strip()]

        try:
            start = date(ano, mes, 1)
            end = (date(ano + 1, 1, 1) if mes == 12 else date(ano, mes + 1, 1))
        except Exception:
            return jsonify({'error': 'ano/mes invalidos'}), 400

        where_parts = [
            "RS.RDATA >= :start",
            "RS.RDATA < :end",
        ]
        params = {'start': start, 'end': end}
        if ccustos:
            keys = []
            for idx, cc in enumerate(ccustos):
                k = f"cc{idx}"
                params[k] = cc
                keys.append(f":{k}")
            where_parts.append(
                "LTRIM(RTRIM(ISNULL(a.CCUSTO,''))) COLLATE SQL_Latin1_General_CP1_CI_AI IN ("
                + ",".join(keys)
                + ")"
            )

        sql = f"""
            SELECT
                CAST(RS.RDATA AS date) AS DIA,
                SUM(
                    CASE
                        WHEN UPPER(LTRIM(RTRIM(ISNULL(a.TIPO,'')))) = 'GESTAO' THEN
                            CASE WHEN ISNULL(a.FTLIMPEZA,0) = 0 THEN
                                (CASE WHEN ISNULL(RS.CANCELADA,0) = 1 THEN ISNULL(RS.PCANCEL,0)
                                      ELSE ROUND(ISNULL(RS.ESTADIA,0)/1.06,2) + ROUND(ISNULL(RS.LIMPEZA,0)/1.23,2) - ISNULL(RS.COMISSAO,0)
                                 END) * (ISNULL(a.COMISSAO,0) / 100.0)
                            ELSE
                                (CASE WHEN ISNULL(RS.CANCELADA,0) = 1 THEN ISNULL(RS.PCANCEL,0)
                                      ELSE ROUND(ISNULL(RS.ESTADIA,0)/1.06,2) - ISNULL(RS.COMISSAO,0)
                                 END) * (ISNULL(a.COMISSAO,0) / 100.0)
                                + ROUND(ISNULL(RS.LIMPEZA,0)/1.23,2)
                            END
                        ELSE
                            CASE WHEN ISNULL(RS.CANCELADA,0) = 1 THEN ISNULL(RS.PCANCEL,0)
                                 ELSE ROUND(ISNULL(RS.ESTADIA,0)/1.06,2) + ROUND(ISNULL(RS.LIMPEZA,0)/1.23,2) - ISNULL(RS.COMISSAO,0)
                            END
                    END
                ) AS VALOR
            FROM RS
            LEFT JOIN AL a
              ON LTRIM(RTRIM(a.NOME)) COLLATE SQL_Latin1_General_CP1_CI_AI
               = LTRIM(RTRIM(ISNULL(RS.ALOJAMENTO,''))) COLLATE SQL_Latin1_General_CP1_CI_AI
            WHERE {" AND ".join(where_parts)}
            GROUP BY CAST(RS.RDATA AS date)
            ORDER BY CAST(RS.RDATA AS date)
        """

        try:
            rows = db.session.execute(text(sql), params).fetchall()
        except Exception as e:
            return jsonify({'error': f'Erro ao obter reservas diarias: {e}'}), 500

        values_by_day = {}
        for r in rows:
            d = r[0]
            v = float(r[1] or 0)
            try:
                day_num = int(d.day)
            except Exception:
                continue
            values_by_day[day_num] = values_by_day.get(day_num, 0.0) + v

        last_day = (end - timedelta(days=1)).day
        labels = list(range(1, last_day + 1))
        values = [round(float(values_by_day.get(i, 0.0)), 2) for i in labels]
        return jsonify({
            'ano': ano,
            'mes': mes,
            'labels': labels,
            'values': values,
            'total': round(sum(values), 2)
        })

    @app.route('/api/mapa_gestao/faturacao_diaria')
    @login_required
    def api_mapa_gestao_faturacao_diaria():
        from datetime import timedelta

        try:
            ano = request.args.get('ano', type=int) or date.today().year
        except Exception:
            ano = date.today().year
        try:
            mes = request.args.get('mes', type=int) or date.today().month
        except Exception:
            mes = date.today().month
        if mes < 1 or mes > 12:
            return jsonify({'error': 'mes invalido'}), 400

        ccustos_raw = (request.args.get('ccustos') or '').strip()
        ccustos = [c.strip() for c in ccustos_raw.split(',') if c.strip()]

        try:
            start = date(ano, mes, 1)
            end = (date(ano + 1, 1, 1) if mes == 12 else date(ano, mes + 1, 1))
        except Exception:
            return jsonify({'error': 'ano/mes invalidos'}), 400

        where_parts = [
            "V.DATA >= :start",
            "V.DATA < :end",
            "ISNULL(V.VALOR,0) <> 0",
        ]
        params = {'start': start, 'end': end}
        if ccustos:
            keys = []
            for idx, cc in enumerate(ccustos):
                k = f"cc{idx}"
                params[k] = cc
                keys.append(f":{k}")
            where_parts.append(
                "LTRIM(RTRIM(ISNULL(V.CCUSTO,''))) COLLATE SQL_Latin1_General_CP1_CI_AI IN ("
                + ",".join(keys)
                + ")"
            )

        sql = f"""
            SELECT
                CAST(V.DATA AS date) AS DIA,
                SUM(ISNULL(V.VALOR,0)) AS VALOR
            FROM v_diario_all V
            WHERE {" AND ".join(where_parts)}
            GROUP BY CAST(V.DATA AS date)
            ORDER BY CAST(V.DATA AS date)
        """

        try:
            rows = db.session.execute(text(sql), params).fetchall()
        except Exception as e:
            return jsonify({'error': f'Erro ao obter faturacao diaria: {e}'}), 500

        values_by_day = {}
        for r in rows:
            d = r[0]
            v = float(r[1] or 0)
            try:
                day_num = int(d.day)
            except Exception:
                continue
            values_by_day[day_num] = values_by_day.get(day_num, 0.0) + v

        last_day = (end - timedelta(days=1)).day
        labels = list(range(1, last_day + 1))
        values = [round(float(values_by_day.get(i, 0.0)), 2) for i in labels]
        return jsonify({
            'ano': ano,
            'mes': mes,
            'labels': labels,
            'values': values,
            'total': round(sum(values), 2)
        })

    @app.route('/api/mapa_gestao/adr_diario')
    @login_required
    def api_mapa_gestao_adr_diario():
        from datetime import timedelta

        try:
            ano = request.args.get('ano', type=int) or date.today().year
        except Exception:
            ano = date.today().year
        try:
            mes = request.args.get('mes', type=int) or date.today().month
        except Exception:
            mes = date.today().month
        if mes < 1 or mes > 12:
            return jsonify({'error': 'mes invalido'}), 400

        ccustos_raw = (request.args.get('ccustos') or '').strip()
        ccustos = [c.strip() for c in ccustos_raw.split(',') if c.strip()]

        try:
            start = date(ano, mes, 1)
            end = (date(ano + 1, 1, 1) if mes == 12 else date(ano, mes + 1, 1))
        except Exception:
            return jsonify({'error': 'ano/mes invalidos'}), 400

        where_parts = [
            "ISNULL(RS.CANCELADA,0) = 0",
            "ISNULL(RS.NOITES,0) > 0",
            "RS.DATAIN IS NOT NULL",
            # excluir reservas do proprietario (valor medio/noite = 0)
            "ABS(((ISNULL(RS.ESTADIA,0) + ISNULL(RS.LIMPEZA,0)) - ISNULL(RS.COMISSAO,0))) > 0.005",
        ]
        params = {'start': start, 'end': end}
        if ccustos:
            keys = []
            for idx, cc in enumerate(ccustos):
                k = f"cc{idx}"
                params[k] = cc
                keys.append(f":{k}")
            where_parts.append(
                "LTRIM(RTRIM(ISNULL(a.CCUSTO,''))) COLLATE SQL_Latin1_General_CP1_CI_AI IN ("
                + ",".join(keys)
                + ")"
            )

        sql = f"""
            WITH nums AS (
                SELECT TOP (60) ROW_NUMBER() OVER (ORDER BY (SELECT NULL)) - 1 AS n
                FROM sys.all_objects
            ),
            exp AS (
                SELECT
                    CAST(DATEADD(day, nums.n, CAST(RS.DATAIN AS date)) AS date) AS DIA,
                    (
                        (ISNULL(RS.ESTADIA,0) + ISNULL(RS.LIMPEZA,0)) - ISNULL(RS.COMISSAO,0)
                    ) / NULLIF(CAST(ISNULL(RS.NOITES,0) AS float), 0) AS PER_NOITE
                FROM RS
                LEFT JOIN AL a
                  ON LTRIM(RTRIM(a.NOME)) COLLATE SQL_Latin1_General_CP1_CI_AI
                   = LTRIM(RTRIM(ISNULL(RS.ALOJAMENTO,''))) COLLATE SQL_Latin1_General_CP1_CI_AI
                INNER JOIN nums
                  ON nums.n < ISNULL(RS.NOITES,0)
                WHERE {" AND ".join(where_parts)}
            )
            SELECT
                DIA,
                AVG(PER_NOITE) AS ADR,
                COUNT(1) AS NOITES,
                SUM(PER_NOITE) AS NET
            FROM exp
            WHERE DIA >= :start AND DIA < :end
            GROUP BY DIA
            ORDER BY DIA
        """

        try:
            rows = db.session.execute(text(sql), params).mappings().all()
        except Exception as e:
            return jsonify({'error': f'Erro ao obter ADR diario: {e}'}), 500

        adr_by_day = {}
        net_by_day = {}
        nights_by_day = {}
        total_net = 0.0
        total_nights = 0
        for r in rows:
            d = r.get('DIA')
            adr = float(r.get('ADR') or 0)
            nights = int(r.get('NOITES') or 0)
            net = float(r.get('NET') or 0)
            try:
                day_num = int(d.day)
            except Exception:
                continue
            adr_by_day[day_num] = adr
            net_by_day[day_num] = net
            nights_by_day[day_num] = nights
            total_net += net
            total_nights += nights

        last_day = (end - timedelta(days=1)).day
        labels = list(range(1, last_day + 1))
        values = []
        noites_arr = []
        net_arr = []
        for i in labels:
            k = int(nights_by_day.get(i, 0))
            n = float(net_by_day.get(i, 0.0))
            a = adr_by_day.get(i)
            values.append(None if a is None or k <= 0 else round(float(a), 2))
            noites_arr.append(k)
            net_arr.append(round(float(n), 2))

        adr_total = (round(total_net / float(total_nights), 2) if total_nights else None)
        return jsonify({
            'ano': ano,
            'mes': mes,
            'labels': labels,
            'values': values,
            'noites_by_day': noites_arr,
            'net_by_day': net_arr,
            'noites': int(total_nights or 0),
            'net': round(float(total_net or 0), 2),
            'adr': adr_total
        })

    @app.route('/api/mapa_gestao/vendas_alojamentos')
    @login_required
    def api_mapa_gestao_vendas_alojamentos():
        try:
            ano = request.args.get('ano', type=int) or date.today().year
        except Exception:
            ano = date.today().year
        try:
            mes = request.args.get('mes', type=int) or date.today().month
        except Exception:
            mes = date.today().month
        if mes < 1 or mes > 12:
            return jsonify({'error': 'mes invalido'}), 400

        ccustos_raw = (request.args.get('ccustos') or '').strip()
        ccustos = [c.strip() for c in ccustos_raw.split(',') if c.strip()]

        where_parts = [
            "ISNULL(a.INATIVO,0) = 0",
            "UPPER(LTRIM(RTRIM(ISNULL(a.TIPO,'')))) = 'EXPLORACAO'",
        ]
        params = {'ano': ano, 'mes': mes}
        if ccustos:
            keys = []
            for idx, cc in enumerate(ccustos):
                k = f"cc{idx}"
                params[k] = cc
                keys.append(f":{k}")
            where_parts.append(
                "LTRIM(RTRIM(ISNULL(a.CCUSTO,''))) COLLATE SQL_Latin1_General_CP1_CI_AI IN ("
                + ",".join(keys)
                + ")"
            )

        sql = f"""
            WITH vc AS (
                SELECT
                    LTRIM(RTRIM(CCUSTO)) AS CCUSTO,
                    SUM(ISNULL(TOTAL,0)) AS REAL
                FROM v_custo
                WHERE YEAR(DATA) = :ano
                  AND MONTH(DATA) = :mes
                  AND FAMILIA LIKE '9%'
                GROUP BY LTRIM(RTRIM(CCUSTO))
            ),
            ovm AS (
                SELECT
                    LTRIM(RTRIM(CCUSTO)) AS CCUSTO,
                    SUM(
                        CASE :mes
                          WHEN 1 THEN ISNULL(JANEIRO,0)
                          WHEN 2 THEN ISNULL(FEVEREIRO,0)
                          WHEN 3 THEN ISNULL(MARCO,0)
                          WHEN 4 THEN ISNULL(ABRIL,0)
                          WHEN 5 THEN ISNULL(MAIO,0)
                          WHEN 6 THEN ISNULL(JUNHO,0)
                          WHEN 7 THEN ISNULL(JULHO,0)
                          WHEN 8 THEN ISNULL(AGOSTO,0)
                          WHEN 9 THEN ISNULL(SETEMBRO,0)
                          WHEN 10 THEN ISNULL(OUTUBRO,0)
                          WHEN 11 THEN ISNULL(NOVEMBRO,0)
                          WHEN 12 THEN ISNULL(DEZEMBRO,0)
                          ELSE 0
                        END
                    ) AS OBJETIVO
                FROM OV
                WHERE ANO = :ano
                GROUP BY LTRIM(RTRIM(CCUSTO))
            )
            SELECT
                LTRIM(RTRIM(ISNULL(a.NOME,''))) AS ALOJAMENTO,
                LTRIM(RTRIM(ISNULL(a.CCUSTO,''))) AS CCUSTO,
                ISNULL(vc.REAL,0) AS REAL,
                ISNULL(ovm.OBJETIVO,0) AS OBJETIVO
            FROM AL a
            LEFT JOIN vc
              ON LTRIM(RTRIM(vc.CCUSTO)) COLLATE SQL_Latin1_General_CP1_CI_AI
               = LTRIM(RTRIM(a.CCUSTO)) COLLATE SQL_Latin1_General_CP1_CI_AI
            LEFT JOIN ovm
              ON LTRIM(RTRIM(ovm.CCUSTO)) COLLATE SQL_Latin1_General_CP1_CI_AI
               = LTRIM(RTRIM(a.CCUSTO)) COLLATE SQL_Latin1_General_CP1_CI_AI
            WHERE {" AND ".join(where_parts)}
            ORDER BY LTRIM(RTRIM(ISNULL(a.NOME,'')))
        """

        try:
            rows = db.session.execute(text(sql), params).mappings().all()
        except Exception as e:
            return jsonify({'error': f'Erro ao obter vendas por alojamento: {e}'}), 500

        out = []
        total_real = 0.0
        total_obj = 0.0
        for r in rows:
            aloj = str(r.get('ALOJAMENTO') or '').strip()
            cc = str(r.get('CCUSTO') or '').strip()
            real = float(r.get('REAL') or 0)
            obj = float(r.get('OBJETIVO') or 0)
            total_real += real
            total_obj += obj
            out.append({
                'alojamento': aloj,
                'ccusto': cc,
                'real': round(real, 2),
                'objetivo': round(obj, 2),
            })

        return jsonify({
            'ano': ano,
            'mes': mes,
            'rows': out,
            'total_real': round(total_real, 2),
            'total_objetivo': round(total_obj, 2),
        })

    @app.route('/api/mapa_gestao/margem_alojamentos')
    @login_required
    def api_mapa_gestao_margem_alojamentos():
        try:
            ano = request.args.get('ano', type=int) or date.today().year
        except Exception:
            ano = date.today().year
        try:
            mes = request.args.get('mes', type=int) or date.today().month
        except Exception:
            mes = date.today().month
        if mes < 1 or mes > 12:
            return jsonify({'error': 'mes invalido'}), 400

        ccustos_raw = (request.args.get('ccustos') or '').strip()
        ccustos = [c.strip() for c in ccustos_raw.split(',') if c.strip()]

        where_parts = [
            "ISNULL(a.INATIVO,0) = 0",
            "UPPER(LTRIM(RTRIM(ISNULL(a.TIPO,'')))) = 'EXPLORACAO'",
        ]
        params = {'ano': ano, 'mes': mes}
        if ccustos:
            keys = []
            for idx, cc in enumerate(ccustos):
                k = f"cc{idx}"
                params[k] = cc
                keys.append(f":{k}")
            where_parts.append(
                "LTRIM(RTRIM(ISNULL(a.CCUSTO,''))) COLLATE SQL_Latin1_General_CP1_CI_AI IN ("
                + ",".join(keys)
                + ")"
            )

        sql = f"""
            WITH sales AS (
                SELECT
                    LTRIM(RTRIM(CCUSTO)) AS CCUSTO,
                    SUM(ISNULL(TOTAL,0)) AS VENDAS
                FROM v_custo
                WHERE YEAR(DATA) = :ano
                  AND MONTH(DATA) = :mes
                  AND FAMILIA LIKE '9%'
                GROUP BY LTRIM(RTRIM(CCUSTO))
            ),
            costs AS (
                SELECT
                    LTRIM(RTRIM(CCUSTO)) AS CCUSTO,
                    SUM(ISNULL(TOTAL,0)) AS CUSTOS
                FROM v_custo
                WHERE YEAR(DATA) = :ano
                  AND MONTH(DATA) = :mes
                  AND (FAMILIA NOT LIKE '9%')
                GROUP BY LTRIM(RTRIM(CCUSTO))
            )
            SELECT
                LTRIM(RTRIM(ISNULL(a.NOME,''))) AS ALOJAMENTO,
                LTRIM(RTRIM(ISNULL(a.CCUSTO,''))) AS CCUSTO,
                ISNULL(sales.VENDAS,0) AS VENDAS,
                ISNULL(costs.CUSTOS,0) AS CUSTOS,
                ISNULL(sales.VENDAS,0) - ISNULL(costs.CUSTOS,0) AS MARGEM
            FROM AL a
            LEFT JOIN sales
              ON LTRIM(RTRIM(sales.CCUSTO)) COLLATE SQL_Latin1_General_CP1_CI_AI
               = LTRIM(RTRIM(a.CCUSTO)) COLLATE SQL_Latin1_General_CP1_CI_AI
            LEFT JOIN costs
              ON LTRIM(RTRIM(costs.CCUSTO)) COLLATE SQL_Latin1_General_CP1_CI_AI
               = LTRIM(RTRIM(a.CCUSTO)) COLLATE SQL_Latin1_General_CP1_CI_AI
            WHERE {" AND ".join(where_parts)}
            ORDER BY LTRIM(RTRIM(ISNULL(a.NOME,'')))
        """

        try:
            rows = db.session.execute(text(sql), params).mappings().all()
        except Exception as e:
            return jsonify({'error': f'Erro ao obter margem por alojamento: {e}'}), 500

        out = []
        total_sales = 0.0
        total_costs = 0.0
        total_margin = 0.0
        for r in rows:
            aloj = str(r.get('ALOJAMENTO') or '').strip()
            cc = str(r.get('CCUSTO') or '').strip()
            vendas = float(r.get('VENDAS') or 0)
            custos = float(r.get('CUSTOS') or 0)
            margem = float(r.get('MARGEM') or 0)
            total_sales += vendas
            total_costs += custos
            total_margin += margem
            out.append({
                'alojamento': aloj,
                'ccusto': cc,
                'vendas': round(vendas, 2),
                'custos': round(custos, 2),
                'margem': round(margem, 2),
            })

        return jsonify({
            'ano': ano,
            'mes': mes,
            'rows': out,
            'total_vendas': round(total_sales, 2),
            'total_custos': round(total_costs, 2),
            'total_margem': round(total_margin, 2),
        })

    @app.route('/api/mapa_gestao/reservas_paises')
    @login_required
    def api_mapa_gestao_reservas_paises():
        from datetime import timedelta

        try:
            ano = request.args.get('ano', type=int) or date.today().year
        except Exception:
            ano = date.today().year

        ccustos_raw = (request.args.get('ccustos') or '').strip()
        ccustos = [c.strip() for c in ccustos_raw.split(',') if c.strip()]

        try:
            start = date(ano, 1, 1)
            end = date(ano + 1, 1, 1)
        except Exception:
            return jsonify({'error': 'ano invalido'}), 400

        where_parts = [
            "RS.DATAIN >= :start",
            "RS.DATAIN < :end",
            "ISNULL(RS.NOITES,0) > 0",
            "ISNULL(RS.CANCELADA,0) = 0",
        ]
        params = {'start': start, 'end': end}
        if ccustos:
            keys = []
            for idx, cc in enumerate(ccustos):
                k = f"cc{idx}"
                params[k] = cc
                keys.append(f":{k}")
            where_parts.append(
                "LTRIM(RTRIM(ISNULL(a.CCUSTO,''))) COLLATE SQL_Latin1_General_CP1_CI_AI IN ("
                + ",".join(keys)
                + ")"
            )

        sql = f"""
            SELECT
                LTRIM(RTRIM(ISNULL(RS.PAIS,''))) AS PAIS,
                COUNT(1) AS RESERVAS,
                SUM(ISNULL(RS.NOITES,0)) AS NOITES,
                SUM(ISNULL(RS.ADULTOS,0) + ISNULL(RS.CRIANCAS,0)) AS HOSPEDES,
                AVG(CAST(ISNULL(RS.NOITES,0) AS float)) AS MEDIA_NOITES,
                AVG(CAST((ISNULL(RS.ADULTOS,0) + ISNULL(RS.CRIANCAS,0)) AS float)) AS MEDIA_HOSPEDES,
                AVG(CAST(DATEDIFF(day, CAST(ISNULL(RS.RDATA, RS.DATAIN) AS date), CAST(RS.DATAIN AS date)) AS float)) AS MEDIA_ANTECIP,
                SUM(
                    CASE
                        WHEN UPPER(LTRIM(RTRIM(ISNULL(a.TIPO,'')))) = 'GESTAO' THEN
                            CASE WHEN ISNULL(a.FTLIMPEZA,0) = 0 THEN
                                (CASE WHEN ISNULL(RS.CANCELADA,0) = 1 THEN ISNULL(RS.PCANCEL,0)
                                      ELSE ROUND(ISNULL(RS.ESTADIA,0)/1.06,2) + ROUND(ISNULL(RS.LIMPEZA,0)/1.23,2) - ISNULL(RS.COMISSAO,0)
                                 END) * (ISNULL(a.COMISSAO,0) / 100.0)
                            ELSE
                                (CASE WHEN ISNULL(RS.CANCELADA,0) = 1 THEN ISNULL(RS.PCANCEL,0)
                                      ELSE ROUND(ISNULL(RS.ESTADIA,0)/1.06,2) - ISNULL(RS.COMISSAO,0)
                                 END) * (ISNULL(a.COMISSAO,0) / 100.0)
                                + ROUND(ISNULL(RS.LIMPEZA,0)/1.23,2)
                            END
                        ELSE
                            CASE WHEN ISNULL(RS.CANCELADA,0) = 1 THEN ISNULL(RS.PCANCEL,0)
                                 ELSE ROUND(ISNULL(RS.ESTADIA,0)/1.06,2) + ROUND(ISNULL(RS.LIMPEZA,0)/1.23,2) - ISNULL(RS.COMISSAO,0)
                            END
                    END
                ) AS VALOR
            FROM RS
            LEFT JOIN AL a
              ON LTRIM(RTRIM(a.NOME)) COLLATE SQL_Latin1_General_CP1_CI_AI
               = LTRIM(RTRIM(ISNULL(RS.ALOJAMENTO,''))) COLLATE SQL_Latin1_General_CP1_CI_AI
            WHERE {" AND ".join(where_parts)}
            GROUP BY LTRIM(RTRIM(ISNULL(RS.PAIS,'')))
            ORDER BY SUM(
                CASE
                    WHEN UPPER(LTRIM(RTRIM(ISNULL(a.TIPO,'')))) = 'GESTAO' THEN
                        CASE WHEN ISNULL(a.FTLIMPEZA,0) = 0 THEN
                            (CASE WHEN ISNULL(RS.CANCELADA,0) = 1 THEN ISNULL(RS.PCANCEL,0)
                                  ELSE ROUND(ISNULL(RS.ESTADIA,0)/1.06,2) + ROUND(ISNULL(RS.LIMPEZA,0)/1.23,2) - ISNULL(RS.COMISSAO,0)
                             END) * (ISNULL(a.COMISSAO,0) / 100.0)
                        ELSE
                            (CASE WHEN ISNULL(RS.CANCELADA,0) = 1 THEN ISNULL(RS.PCANCEL,0)
                                  ELSE ROUND(ISNULL(RS.ESTADIA,0)/1.06,2) - ISNULL(RS.COMISSAO,0)
                             END) * (ISNULL(a.COMISSAO,0) / 100.0)
                            + ROUND(ISNULL(RS.LIMPEZA,0)/1.23,2)
                        END
                    ELSE
                        CASE WHEN ISNULL(RS.CANCELADA,0) = 1 THEN ISNULL(RS.PCANCEL,0)
                             ELSE ROUND(ISNULL(RS.ESTADIA,0)/1.06,2) + ROUND(ISNULL(RS.LIMPEZA,0)/1.23,2) - ISNULL(RS.COMISSAO,0)
                        END
                END
            ) DESC, LTRIM(RTRIM(ISNULL(RS.PAIS,'')))
        """

        try:
            rows = db.session.execute(text(sql), params).mappings().all()
        except Exception as e:
            return jsonify({'error': f'Erro ao obter reservas por pais: {e}'}), 500

        out = []
        total_valor = 0.0
        total_res = 0
        for r in rows:
            pais = str(r.get('PAIS') or '').strip()
            valor = float(r.get('VALOR') or 0)
            noites = int(r.get('NOITES') or 0)
            reservas = int(r.get('RESERVAS') or 0)
            total_valor += valor
            total_res += reservas
            out.append({
                'pais': pais,
                'reservas': reservas,
                'valor': round(valor, 2),
                'media_noites': round(float(r.get('MEDIA_NOITES') or 0), 2),
                'media_noite': (None if noites <= 0 else round(valor / float(noites), 2)),
                'media_hospedes': round(float(r.get('MEDIA_HOSPEDES') or 0), 2),
                'media_antecip': round(float(r.get('MEDIA_ANTECIP') or 0), 1),
            })

        return jsonify({
            'ano': ano,
            'total_valor': round(total_valor, 2),
            'total_reservas': total_res,
            'rows': out
        })

    @app.route('/turnover')
    @login_required
    def turnover_page():
        return render_template('turnover.html', page_title='Turnover Diário', today=date.today().isoformat())

    @app.route('/api/turnover')
    @login_required
    def api_turnover():
        try:
            data_str = (request.args.get('data') or '').strip()
            dia = datetime.strptime(data_str, '%Y-%m-%d').date() if data_str else date.today()
            dia_iso = dia.isoformat()

            reservas_out = db.session.execute(text("""
                SELECT LTRIM(RTRIM(ALOJAMENTO)) AS ALOJAMENTO, ISNULL(HORAOUT,'') AS HORAOUT, NOITES,
                       ISNULL(ADULTOS,0) AS ADULTOS, ISNULL(CRIANCAS,0) AS CRIANCAS,
                       ISNULL(BERCO,0) AS BERCO, ISNULL(SOFACAMA,0) AS SOFACAMA
                FROM RS
                WHERE CAST(DATAOUT AS date) = :dia
                  AND ISNULL(CANCELADA,0) = 0
            """), {'dia': dia_iso}).mappings().all()

            reservas_in = db.session.execute(text("""
                SELECT LTRIM(RTRIM(RS.ALOJAMENTO)) AS ALOJAMENTO, ISNULL(RS.HORAIN,'') AS HORAIN, RS.NOITES,
                       ISNULL(RS.ADULTOS,0) AS ADULTOS, ISNULL(RS.CRIANCAS,0) AS CRIANCAS,
                       ISNULL(RS.BERCO,0) AS BERCO, ISNULL(RS.SOFACAMA,0) AS SOFACAMA,
                       ISNULL(RS.SEF,0) AS SEF, ISNULL(RS.USRSEF,'') AS USRSEF,
                       ISNULL(RS.INSTR,0) AS INSTR, ISNULL(RS.USRINSTR,'') AS USRINSTR,
                       ISNULL(RS.PRESENCIAL,0) AS PRESENCIAL,
                       ISNULL(RS.ENTROU,0) AS ENTROU,
                       ISNULL(RS.USRCHECKIN,'') AS USRCHECKIN,
                       ISNULL(uc.NOME, RS.USRCHECKIN) AS USRCHECKIN_NOME
                FROM RS
                LEFT JOIN US uc ON uc.LOGIN = RS.USRCHECKIN
                WHERE CAST(RS.DATAIN AS date) = :dia
                  AND ISNULL(RS.CANCELADA,0) = 0
            """), {'dia': dia_iso}).mappings().all()

            tarefas = db.session.execute(text("""
                SELECT LTRIM(RTRIM(t.ALOJAMENTO)) AS ALOJAMENTO,
                       t.UTILIZADOR,
                       ISNULL(t.TRATADO,0) AS TRATADO,
                       ISNULL(t.TAREFA,'') AS OBS,
                       ISNULL(u.NOME, t.UTILIZADOR) AS NOME,
                       ISNULL(t.HORA,'') AS HORA,
                       ISNULL(al.TIPOLOGIA,'') AS TIPOLOGIA
                FROM TAREFAS t
                LEFT JOIN US u ON u.LOGIN = t.UTILIZADOR
                LEFT JOIN AL al ON LTRIM(RTRIM(al.NOME)) = LTRIM(RTRIM(t.ALOJAMENTO))
                WHERE LTRIM(RTRIM(ISNULL(t.ORIGEM,''))) = 'LP'
                  AND CAST(t.DATA AS date) = :dia
            """), {'dia': dia_iso}).mappings().all()

            ultimas_limpezas = db.session.execute(text("""
                WITH cte AS (
                    SELECT LTRIM(RTRIM(t.ALOJAMENTO)) AS ALOJAMENTO,
                           CAST(t.DATA AS date) AS DIA,
                           ISNULL(t.HORA,'') AS HORA,
                           ISNULL(u.NOME, t.UTILIZADOR) AS NOME,
                           ROW_NUMBER() OVER (
                               PARTITION BY LTRIM(RTRIM(t.ALOJAMENTO))
                               ORDER BY CAST(t.DATA AS date) DESC, ISNULL(t.HORA,'') DESC
                           ) AS rn
                    FROM TAREFAS t
                    LEFT JOIN US u ON u.LOGIN = t.UTILIZADOR
                    WHERE LTRIM(RTRIM(ISNULL(t.ORIGEM,''))) = 'LP'
                      AND ISNULL(t.TRATADO,0) = 1
                      AND CAST(t.DATA AS date) <= :dia
                )
                SELECT ALOJAMENTO, DIA, HORA, NOME FROM cte WHERE rn = 1
            """), {'dia': dia_iso}).mappings().all()

            def norm_aloj(v: str) -> str:
                try:
                    return (v or '').strip().upper()
                except Exception:
                    return ''

            map_out, map_in, map_tasks, map_last = {}, {}, {}, {}
            for r in reservas_out:
                aloj = norm_aloj(r.get('ALOJAMENTO'))
                if aloj:
                    map_out.setdefault(aloj, []).append(r)
            for r in reservas_in:
                aloj = norm_aloj(r.get('ALOJAMENTO'))
                if aloj:
                    map_in.setdefault(aloj, []).append(r)
            for r in tarefas:
                aloj = norm_aloj(r.get('ALOJAMENTO'))
                if aloj:
                    map_tasks.setdefault(aloj, []).append(r)
            for r in ultimas_limpezas:
                aloj = norm_aloj(r.get('ALOJAMENTO'))
                if aloj and aloj not in map_last:
                    map_last[aloj] = r

            now_dt = datetime.now()
            all_aloj = set(map_out.keys()) | set(map_in.keys()) | set(map_tasks.keys())

            # Helpers to cache last checkout and next checkin dates
            cache_last_out = {}
            cache_next_in = {}
            cache_future_clean = {}

            def get_last_out(aloj):
                if aloj in cache_last_out:
                    return cache_last_out[aloj]
                try:
                    row = db.session.execute(text("""
                        SELECT TOP 1 CAST(DATAOUT AS date) AS DIA
                        FROM RS
                        WHERE CAST(DATAOUT AS date) < :dia
                          AND LTRIM(RTRIM(ALOJAMENTO)) = LTRIM(RTRIM(:aloj))
                          AND ISNULL(CANCELADA,0) = 0
                        ORDER BY DATAOUT DESC
                    """), {'dia': dia_iso, 'aloj': aloj}).first()
                    cache_last_out[aloj] = row[0] if row else None
                except Exception:
                    cache_last_out[aloj] = None
                return cache_last_out[aloj]

            def get_next_in(aloj):
                if aloj in cache_next_in:
                    return cache_next_in[aloj]
                try:
                    row = db.session.execute(text("""
                        SELECT TOP 1 CAST(DATAIN AS date) AS DIA
                        FROM RS
                        WHERE CAST(DATAIN AS date) > :dia
                          AND LTRIM(RTRIM(ALOJAMENTO)) = LTRIM(RTRIM(:aloj))
                          AND ISNULL(CANCELADA,0) = 0
                        ORDER BY DATAIN ASC
                    """), {'dia': dia_iso, 'aloj': aloj}).first()
                    cache_next_in[aloj] = row[0] if row else None
                except Exception:
                    cache_next_in[aloj] = None
                return cache_next_in[aloj]

            def get_future_clean(aloj, limite):
                if aloj in cache_future_clean:
                    return cache_future_clean[aloj]
                try:
                    params = {'dia': dia_iso, 'aloj': aloj}
                    sql = """
                        SELECT TOP 1 CAST(t.DATA AS date) AS DIA,
                               ISNULL(u.NOME, t.UTILIZADOR) AS NOME
                        FROM TAREFAS t
                        LEFT JOIN US u ON u.LOGIN = t.UTILIZADOR
                        WHERE LTRIM(RTRIM(t.ALOJAMENTO)) = LTRIM(RTRIM(:aloj))
                          AND LTRIM(RTRIM(ISNULL(t.ORIGEM,''))) = 'LP'
                          AND CAST(t.DATA AS date) > :dia
                          AND ISNULL(t.TRATADO,0) = 0
                    """
                    if limite:
                        sql += " AND CAST(t.DATA AS date) <= :limite"
                        params['limite'] = limite
                    sql += " ORDER BY t.DATA ASC"
                    row = db.session.execute(text(sql), params).first()
                    if row:
                        cache_future_clean[aloj] = {'dia': row[0], 'nome': row[1]}
                    else:
                        cache_future_clean[aloj] = None
                except Exception:
                    cache_future_clean[aloj] = None
                return cache_future_clean[aloj]

            dow_pt = ['SEG', 'TER', 'QUA', 'QUI', 'SEX', 'SAB', 'DOM']

            def fmt_date_parts(dval):
                try:
                    if isinstance(dval, datetime):
                        d = dval.date()
                    else:
                        d = dval
                    dow = dow_pt[d.weekday()]
                    dat = d.strftime('%d/%m')
                    return {'dow': dow, 'date': dat}
                except Exception:
                    return None

            cards = []
            for aloj in sorted(all_aloj):
                out_list = map_out.get(aloj, [])
                in_list = map_in.get(aloj, [])
                task_list = map_tasks.get(aloj, [])

                aloj_display = ''
                for src in (out_list, in_list, task_list):
                    if src:
                        aloj_display = src[0].get('ALOJAMENTO') or ''
                        break

                def first_val(lst, key):
                    try:
                        return lst[0].get(key) if lst else ''
                    except Exception:
                        return ''

                chk_out = first_val(out_list, 'HORAOUT') or ''
                chk_in = first_val(in_list, 'HORAIN') or ''
                noites = first_val(in_list, 'NOITES') or 0
                try:
                    hospedes = int(first_val(in_list, 'ADULTOS') or 0) + int(first_val(in_list, 'CRIANÇAS') or first_val(in_list, 'CRIANCAS') or 0)
                except Exception:
                    hospedes = 0
                berco = bool(first_val(in_list, 'BERCO'))
                sof = bool(first_val(in_list, 'SOFACAMA'))

                status = 'Já limpo'
                tem_planeada = False
                tem_concluida = False
                if task_list:
                    tem_planeada = any(int(t.get('TRATADO') or 0) == 0 for t in task_list)
                    tem_concluida = any(int(t.get('TRATADO') or 0) == 1 for t in task_list)
                    if tem_planeada:
                        status = 'Planeada'
                    elif tem_concluida:
                        status = 'Concluída'
                equipa = first_val(task_list, 'NOME') or ''
                hora_lp = first_val(task_list, 'HORA') or ''
                tipologia = (first_val(task_list, 'TIPOLOGIA') or '').strip().upper()
                dur_map = {'T0': 60, 'T1': 60, 'T2': 90, 'T3': 120, 'T4': 150}
                dur_min = dur_map.get(tipologia)
                hora_fim = ''
                try:
                    if hora_lp and dur_min:
                        s = str(hora_lp).strip()
                        hh = mm = 0
                        if ':' in s:
                            parts = s.split(':')
                            hh = int(parts[0])
                            mm = int(parts[1] if len(parts) > 1 else 0)
                        else:
                            digits = ''.join(ch for ch in s if ch.isdigit())
                            if len(digits) >= 3:
                                mm = int(digits[-2:])
                                hh = int(digits[:-2])
                            elif digits:
                                hh = int(digits)
                        total = hh * 60 + mm + dur_min
                        hora_fim = f"{(total // 60) % 24:02d}:{total % 60:02d}"
                except Exception:
                    hora_fim = ''
                atrasada = False
                if tem_planeada and not tem_concluida:
                    if dia < now_dt.date():
                        atrasada = True
                    elif dia == now_dt.date():
                        try:
                            if hora_fim:
                                h_end = datetime.strptime(hora_fim, '%H:%M').time()
                                if now_dt.time() > h_end:
                                    atrasada = True
                        except Exception:
                            pass
                if atrasada:
                    status = 'Atrasada'
                if out_list and not task_list:
                    status = 'Por atribuir'
                last_info = False
                if not task_list:
                    last = map_last.get(aloj)
                    if last:
                        try:
                            dia_last = last.get('DIA')
                            dia_str = ''
                            if dia_last:
                                try:
                                    dia_str = dia_last.strftime('%d/%m/%Y')
                                except Exception:
                                    dia_str = str(dia_last)[:10]
                            nome_last = last.get('NOME') or ''
                            if nome_last or dia_str:
                                equipa = f"{nome_last} ({dia_str})".strip()
                            last_info = True
                        except Exception:
                            pass
                extras_obs = []
                if berco:
                    extras_obs.append('Berço')
                if sof:
                    extras_obs.append('Sofá-cama')
                obs = ' • '.join(extras_obs)
                fallback_out = None
                if not out_list:
                    d_last = get_last_out(aloj)
                    if d_last:
                        fmt = fmt_date_parts(d_last)
                        if fmt:
                            fallback_out = fmt
                fallback_in = None
                if not in_list:
                    d_next = get_next_in(aloj)
                    if d_next:
                        fmt = fmt_date_parts(d_next)
                        if fmt:
                            fallback_in = fmt
                    else:
                        fallback_in = {'text': 'Sem Reservas'}
                # Se existe checkout e nÇõo hÇ  limpeza hoje, verificar limpeza planeada antes/prÇüx checkin
                if out_list and not task_list:
                    limite = get_next_in(aloj) or (dia + timedelta(days=14))
                    prox_limpeza = get_future_clean(aloj, limite)
                    if prox_limpeza:
                        dia_fut = prox_limpeza.get('dia')
                        nome_fut = prox_limpeza.get('nome') or ''
                        try:
                            data_fmt = dia_fut.strftime('%d/%m') if dia_fut else ''
                            status = f"Planeada {data_fmt}" if data_fmt else 'Planeada'
                        except Exception:
                            status = 'Planeada'
                        try:
                            if dia_fut:
                                data_lbl = dia_fut.strftime('%d/%m')
                            else:
                                data_lbl = ''
                            if nome_fut or data_lbl:
                                equipa = f"{nome_fut} ({data_lbl})".strip()
                                last_info = True
                        except Exception:
                            pass

                cards.append({
                    'alojamento': aloj_display or aloj,
                    'status': status,
                    'check_out': chk_out,
                    'has_check_out': bool(out_list),
                    'check_in': chk_in,
                    'has_check_in': bool(in_list),
                    'entrou': bool(first_val(in_list, 'ENTROU') or 0),
                    'equipa': equipa,
                    'last_info': last_info,
                    'hora_lp': hora_lp,
                    'hora_fim': hora_fim,
                    'noites': noites,
                    'hospedes': hospedes,
                    'berco': berco,
                    'sofacama': sof,
                    'obs': obs,
                    'sef': bool(first_val(in_list, 'SEF') or 0),
                    'instr': bool(first_val(in_list, 'INSTR') or 0),
                    'presencial': bool(first_val(in_list, 'PRESENCIAL') or 0),
                    'usrcheckin': first_val(in_list, 'USRCHECKIN') or '',
                    'usrcheckin_nome': first_val(in_list, 'USRCHECKIN_NOME') or '',
                    'fallback_out': fallback_out,
                    'fallback_in': fallback_in
                })

            return jsonify({'data': dia_iso, 'cards': cards})
        except Exception as e:
            try:
                app.logger.exception('Erro em api_turnover')
            except Exception:
                pass
            return jsonify({'error': str(e)}), 500

    @app.route('/api/turnover/checkin', methods=['GET', 'POST'])
    @login_required
    def api_turnover_checkin():
        try:
            if request.method == 'GET':
                data_str = (request.args.get('data') or '').strip()
                aloj = (request.args.get('alojamento') or '').strip()
                if not data_str or not aloj:
                    return jsonify({'error': 'Parâmetros em falta'}), 400
                try:
                    dia = datetime.strptime(data_str, '%Y-%m-%d').date()
                except Exception:
                    return jsonify({'error': 'Data inválida'}), 400

                users_rows = db.session.execute(text(
                    "SELECT LOGIN, NOME FROM US WHERE ISNULL(INATIVO,0)=0 ORDER BY NOME"
                )).fetchall()
                users = [{'value': r[0], 'label': r[1] or r[0]} for r in users_rows]

                row = db.session.execute(text("""
                    SELECT TOP 1 ISNULL(PRESENCIAL,0) AS PRESENCIAL,
                           ISNULL(ENTROU,0) AS ENTROU,
                           ISNULL(USRCHECKIN,'') AS USRCHECKIN,
                           ISNULL(SEF,0) AS SEF,
                           ISNULL(USRSEF,'') AS USRSEF,
                           ISNULL(INSTR,0) AS INSTR,
                           ISNULL(USRINSTR,'') AS USRINSTR
                    FROM RS
                    WHERE CAST(DATAIN AS date) = :dia
                      AND LTRIM(RTRIM(ALOJAMENTO)) = LTRIM(RTRIM(:aloj))
                      AND ISNULL(CANCELADA,0) = 0
                    ORDER BY DATAIN DESC
                """), {'dia': dia, 'aloj': aloj}).mappings().first()
                if not row:
                    return jsonify({'error': 'Reserva não encontrada'}), 404

                data = {
                    'presencial': bool(row.get('PRESENCIAL') or 0),
                    'entrou': bool(row.get('ENTROU') or 0),
                    'usrcheckin': row.get('USRCHECKIN') or '',
                    'sef': bool(row.get('SEF') or 0),
                    'usrsef': row.get('USRSEF') or '',
                    'instr': bool(row.get('INSTR') or 0),
                    'usrinstr': row.get('USRINSTR') or ''
                }
                return jsonify({'data': data, 'users': users, 'alojamento': aloj})

            # POST
            body = request.get_json(silent=True) or {}
            data_str = (body.get('data') or '').strip()
            aloj = (body.get('alojamento') or '').strip()
            if not data_str or not aloj:
                return jsonify({'error': 'Parâmetros em falta'}), 400
            try:
                dia = datetime.strptime(data_str, '%Y-%m-%d').date()
            except Exception:
                return jsonify({'error': 'Data inválida'}), 400
            sef = 1 if body.get('sef') else 0
            instr = 1 if body.get('instr') else 0
            presencial = 1 if body.get('presencial') else 0
            entrou = 1 if body.get('entrou') else 0
            usrcheckin = (body.get('usrcheckin') or '').strip()
            usrsef = getattr(current_user, 'LOGIN', '')
            usrinstr = getattr(current_user, 'LOGIN', '')

            upd = text("""
                UPDATE RS
                   SET PRESENCIAL = :p,
                       ENTROU = :e,
                       USRCHECKIN = :u,
                       SEF = :sef,
                       USRSEF = CASE WHEN :sef = 1 THEN :usrsef ELSE USRSEF END,
                       INSTR = :instr,
                       USRINSTR = CASE WHEN :instr = 1 THEN :usrinstr ELSE USRINSTR END
                 WHERE CAST(DATAIN AS date) = :dia
                   AND LTRIM(RTRIM(ALOJAMENTO)) = LTRIM(RTRIM(:aloj))
                   AND ISNULL(CANCELADA,0) = 0
            """)
            res = db.session.execute(upd, {
                'p': presencial,
                'e': entrou,
                'u': usrcheckin,
                'sef': sef,
                'usrsef': usrsef,
                'instr': instr,
                'usrinstr': usrinstr,
                'dia': dia,
                'aloj': aloj
            })
            db.session.commit()
            return jsonify({'ok': True, 'updated': res.rowcount})
        except Exception as e:
            db.session.rollback()
            try:
                app.logger.exception('Erro em api_turnover_checkin')
            except Exception:
                pass
            return jsonify({'error': str(e)}), 500


    # Performance dashboard page
    @app.route('/performance')
    @login_required
    def performance_page():
        return render_template('performance.html', page_title='Performance')

    # API: Alojamentos performance aggregation
    @app.route('/api/performance')
    @login_required
    def api_performance():
        try:
            # Parse dates or default to current month
            qs_ini = (request.args.get('data_inicio') or '').strip()
            qs_fim = (request.args.get('data_fim') or '').strip()

            today = date.today()
            first_day = date(today.year, today.month, 1)
            # Compute last day of month
            if today.month == 12:
                next_month_first = date(today.year + 1, 1, 1)
            else:
                next_month_first = date(today.year, today.month + 1, 1)
            from datetime import timedelta
            last_day = next_month_first - timedelta(days=1)

            def parse_d(dstr, fallback):
                try:
                    return datetime.strptime(dstr, '%Y-%m-%d').date()
                except Exception:
                    return fallback

            data_inicio = parse_d(qs_ini, first_day)
            data_fim = parse_d(qs_fim, last_day)

            # Nights total in the period (1 per day, inclusive)
            noites_totais = (data_fim - data_inicio).days + 1
            if noites_totais < 0:
                return jsonify({'error': 'Intervalo de datas inválido'}), 400

            # Future window: from max(today, start) to end (inclusive)
            future_start = first_day if data_inicio < first_day else data_inicio
            if today > future_start:
                future_start = today
            noites_futuras_totais = (data_fim - future_start).days + 1 if data_fim >= future_start else 0

            # Next 2/4/7 nights windows for alert (from today, capped to selected period)
            next_start = today if today > data_inicio else data_inicio
            next2_end_raw = today + timedelta(days=1)
            next4_end_raw = today + timedelta(days=3)
            next7_end_raw = today + timedelta(days=6)
            next2_end = next2_end_raw if next2_end_raw < data_fim else data_fim
            next4_end = next4_end_raw if next4_end_raw < data_fim else data_fim
            next7_end = next7_end_raw if next7_end_raw < data_fim else data_fim
            total_next2_days = (next2_end - next_start).days + 1 if next2_end >= next_start else 0
            total_next4_days = (next4_end - next_start).days + 1 if next4_end >= next_start else 0
            total_next7_days = (next7_end - next_start).days + 1 if next7_end >= next_start else 0

            # SQL Server aggregation with LEFT JOIN and filters in ON to preserve AL rows
            sql = text(
                """
                SELECT
                    A.NOME AS nome,
                    A.TIPOLOGIA AS tipologia,
                    A.TIPO AS tipo,
                    COUNT(DISTINCT V.data) AS noites_ocupadas,
                    SUM(ISNULL(V.valor, 0)) AS total_liquido,
                    CASE WHEN COUNT(DISTINCT V.data) > 0
                         THEN SUM(ISNULL(V.valor,0)) / NULLIF(COUNT(DISTINCT V.data), 0)
                         ELSE 0 END AS preco_medio_noite,
                    COUNT(DISTINCT CASE WHEN V.data >= :hoje THEN V.data END) AS noites_futuras_ocupadas,
                    COUNT(DISTINCT CASE WHEN V.data BETWEEN :next_start AND :next2_end THEN V.data END) AS next2_ocupadas,
                    COUNT(DISTINCT CASE WHEN V.data BETWEEN :next_start AND :next4_end THEN V.data END) AS next4_ocupadas,
                    COUNT(DISTINCT CASE WHEN V.data BETWEEN :next_start AND :next7_end THEN V.data END) AS next7_ocupadas
                FROM AL AS A
                LEFT JOIN v_diario_all AS V
                  ON V.CCUSTO = A.NOME
                 AND ISNULL(V.valor, 0) <> 0
                 AND V.data BETWEEN :data_inicio AND :data_fim
                WHERE ISNULL(A.INATIVO, 0) = 0
                GROUP BY A.NOME, A.TIPOLOGIA, A.TIPO
                ORDER BY A.NOME
                """
            )

            rows = db.session.execute(sql, {
                'data_inicio': data_inicio,
                'data_fim': data_fim,
                'hoje': today,
                'next_start': next_start,
                'next2_end': next2_end,
                'next4_end': next4_end,
                'next7_end': next7_end
            }).mappings().all()
            # Build occupancy map for next 7-day horizon to detect 2+ night gaps
            from datetime import datetime as _dt
            sql_occ = text(
                """
                SELECT V.CCUSTO AS nome, CAST(V.data AS date) AS dia
                FROM v_diario_all V
                JOIN AL A ON V.CCUSTO = A.NOME AND ISNULL(A.INATIVO,0)=0
                WHERE ISNULL(V.valor,0) <> 0
                  AND V.data BETWEEN :start AND :end
                """
            )
            occ_rows = db.session.execute(sql_occ, {
                'start': next_start,
                'end': next7_end
            }).fetchall()
            occ_map = {}
            for nome, dia in occ_rows:
                # Ensure date object
                if isinstance(dia, _dt):
                    dkey = dia.date()
                else:
                    dkey = dia
                occ_map.setdefault(nome, set()).add(dkey)

            result = []
            for r in rows:
                nome = r.get('nome')
                tipologia = r.get('tipologia')
                tipo = r.get('tipo')
                noites_ocupadas = float(r.get('noites_ocupadas') or 0)
                total_liquido = float(r.get('total_liquido') or 0)
                preco_medio_noite = float(r.get('preco_medio_noite') or 0)
                noites_futuras_ocupadas = int(r.get('noites_futuras_ocupadas') or 0)
                # Occupancy map for this lodging in next horizon
                occ_set = occ_map.get(nome, set())
                # Helper to detect if there exists a run of >=2 empty nights within first L days
                def has_two_plus_gap(L):
                    if L is None or L < 2:
                        return False
                    run = 0
                    for i in range(L):
                        d = next_start + timedelta(days=i)
                        if d in occ_set:
                            run = 0
                        else:
                            run += 1
                            if run >= 2:
                                return True
                    return False
                noites_disponiveis = max(0, noites_futuras_totais - noites_futuras_ocupadas)
                taxa = (noites_ocupadas / noites_totais * 100.0) if noites_totais > 0 else 0.0
                # Determine alert level based on first window that contains a 2+ empty-night run
                alert_level = ''
                if has_two_plus_gap(total_next2_days):
                    alert_level = 'red'
                elif has_two_plus_gap(total_next4_days):
                    alert_level = 'orange'
                elif has_two_plus_gap(total_next7_days):
                    alert_level = 'yellow'

                result.append({
                    'nome': nome,
                    'tipologia': tipologia,
                    'tipo': tipo,
                    'noites_ocupadas': round(noites_ocupadas, 2),
                    'noites_totais': noites_totais,
                    'taxa_ocupacao': round(taxa, 2),
                    'total_liquido': round(total_liquido, 2),
                    'preco_medio_noite': round(preco_medio_noite, 2),
                    'noites_futuras_ocupadas': noites_futuras_ocupadas,
                    'noites_futuras_totais': noites_futuras_totais,
                    'noites_disponiveis': noites_disponiveis,
                    'alert_level': alert_level
                })

            return jsonify({
                'rows': result,
                'data_inicio': data_inicio.strftime('%Y-%m-%d'),
                'data_fim': data_fim.strftime('%Y-%m-%d')
            })
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    # -----------------------------
    # Não disponibilidades (ND)
    # -----------------------------
    @app.route('/nd')
    @login_required
    def nd_page():
        return render_template('nd.html', page_title='Não disponibilidades')

    @app.route('/api/nd', methods=['GET', 'POST'])
    @login_required
    def api_nd():
        try:
            if request.method == 'GET':
                qs_ini = (request.args.get('data_inicio') or '').strip()
                qs_fim = (request.args.get('data_fim') or '').strip()
                def parse_d(dstr):
                    try:
                        return datetime.strptime(dstr, '%Y-%m-%d').date()
                    except Exception:
                        return None
                di = parse_d(qs_ini)
                df = parse_d(qs_fim)
                if not di or not df:
                    return jsonify({'error': 'Parâmetros de datas inválidos'}), 400
                sql = text(
                    """
                    SELECT N.NDSTAMP, N.DATA, N.TIPO, N.UTILIZADOR, U.NOME, U.COR
                    FROM ND AS N
                    INNER JOIN US AS U ON U.LOGIN = N.UTILIZADOR
                    WHERE N.DATA BETWEEN :di AND :df
                    ORDER BY N.DATA, U.NOME
                    """
                )
                rows = db.session.execute(sql, {'di': di, 'df': df}).fetchall()
                items = [
                    {
                        'id': r[0],
                        'data': r[1].isoformat(),
                        'tipo': r[2],
                        'login': r[3],
                        'nome': r[4],
                        'cor': r[5] or '#94a3b8'
                    }
                    for r in rows
                ]
                return jsonify({'rows': items})
            # POST: inserir ND (FOLGA)
            body = request.get_json(silent=True) or {}
            data_str = (body.get('data') or '').strip()
            utilizador = (body.get('utilizador') or '').strip()
            try:
                d = datetime.strptime(data_str, '%Y-%m-%d').date()
            except Exception:
                return jsonify({'error': 'Data inválida'}), 400
            if not utilizador:
                return jsonify({'error': 'Utilizador obrigatório'}), 400
            ins = text(
                """
                INSERT INTO ND (NDSTAMP, EQUIPA, DATA, TIPO, UTILIZADOR)
                SELECT LEFT(CONVERT(varchar(36), NEWID()), 25), '', :data, 'FOLGA', :util
                """
            )
            db.session.execute(ins, {'data': d, 'util': utilizador})
            db.session.commit()
            return jsonify({'ok': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/nd/move', methods=['POST'])
    @login_required
    def api_nd_move():
        try:
            body = request.get_json(silent=True) or {}
            nd_id = (body.get('id') or '').strip()
            data_str = (body.get('data') or '').strip()
            if not nd_id:
                return jsonify({'error': 'ID em falta'}), 400
            try:
                new_date = datetime.strptime(data_str, '%Y-%m-%d').date()
            except Exception:
                return jsonify({'error': 'Data inválida'}), 400
            upd = text("UPDATE ND SET DATA = :d WHERE NDSTAMP = :id")
            res = db.session.execute(upd, {'d': new_date, 'id': nd_id})
            db.session.commit()
            return jsonify({'ok': True, 'updated': res.rowcount})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/nd/delete', methods=['POST'])
    @login_required
    def api_nd_delete():
        try:
            body = request.get_json(silent=True) or {}
            nd_id = (body.get('id') or '').strip()
            if not nd_id:
                return jsonify({'error': 'ID em falta'}), 400
            del_sql = text("DELETE FROM ND WHERE NDSTAMP = :id")
            res = db.session.execute(del_sql, {'id': nd_id})
            db.session.commit()
            return jsonify({'ok': True, 'deleted': res.rowcount})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/nd/ferias', methods=['POST'])
    @login_required
    def api_nd_ferias():
        try:
            body = request.get_json(silent=True) or {}
            utilizador = (body.get('utilizador') or '').strip()
            di_str = (body.get('data_inicio') or '').strip()
            df_str = (body.get('data_fim') or '').strip()
            ano_str = (body.get('ano') or '').strip()
            if not utilizador:
                return jsonify({'error': 'Utilizador obrigatório'}), 400
            try:
                di = datetime.strptime(di_str, '%Y-%m-%d').date()
                df = datetime.strptime(df_str, '%Y-%m-%d').date()
            except Exception:
                return jsonify({'error': 'Datas inválidas'}), 400
            if df < di:
                return jsonify({'error': 'Data fim anterior à data início'}), 400
            # Determina ano
            try:
                ano = int(ano_str) if ano_str else di.year
            except Exception:
                ano = di.year
            # Nota: o ano (FA/ND.ANO) representa o direito adquirido; as datas podem ser gozadas noutro ano.
            from datetime import timedelta
            # Validação: dias disponíveis em FA vs usados em ND
            q_fa = db.session.execute(text(
                "SELECT ISNULL(DIAS,0) FROM FA WHERE UTILIZADOR = :u AND ANO = :a"
            ), { 'u': utilizador, 'a': ano }).fetchone()
            total_fa = int(q_fa[0]) if q_fa is not None else 0
            q_used = db.session.execute(text(
                "SELECT COUNT(*) FROM ND WHERE UTILIZADOR = :u AND TIPO = 'FERIAS' AND ANO = :a"
            ), { 'u': utilizador, 'a': ano }).fetchone()
            usados = int(q_used[0]) if q_used is not None else 0
            # calcula dias solicitados
            solicitados = (df - di).days + 1
            if total_fa <= 0:
                return jsonify({'error': 'Não existem dias de férias anuais configurados (FA) para este utilizador/ano'}), 400
            if usados + solicitados > total_fa:
                disponiveis = max(0, total_fa - usados)
                return jsonify({'error': f'Sem dias suficientes. Disponíveis: {disponiveis}'}), 400
            ins = text("""
                INSERT INTO ND (NDSTAMP, EQUIPA, DATA, TIPO, UTILIZADOR, ANO)
                SELECT LEFT(CONVERT(varchar(36), NEWID()), 25), '', :data, 'FERIAS', :util, :ano
            """)
            d = di
            total = 0
            while d <= df:
                db.session.execute(ins, {'data': d, 'util': utilizador, 'ano': ano})
                total += 1
                d += timedelta(days=1)
            db.session.commit()
            return jsonify({'ok': True, 'count': total})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/nd/users')
    @login_required
    def api_nd_users():
        try:
            # Apenas utilizadores ativos (INATIVO = 0)
            sql = text("SELECT LOGIN, NOME, ISNULL(COR,'') AS COR FROM US WHERE ISNULL(INATIVO,0)=0 ORDER BY NOME")
            rows = db.session.execute(sql).fetchall()
            items = [
                {
                    'login': r[0],
                    'nome': r[1],
                    'cor': r[2] or '#94a3b8'
                }
                for r in rows
            ]
            return jsonify({'users': items})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    # Tesouraria
    @app.route('/tesouraria')
    @login_required
    def tesouraria_page():
        return render_template('tesouraria.html', page_title='Tesouraria')

    # -----------------------------
    # Importação de Extrato Bancário (EXT/EL)
    # -----------------------------
    @app.route('/extratos_importar')
    @app.route('/extrato_import')
    @app.route('/extrato_importar')
    @login_required
    def extratos_importar_page():
        return render_template('extrato_import.html', page_title='Importar Extrato Bancário')

    # -----------------------------
    # Importação de Faturas de Reservas (FR)
    # -----------------------------
    @app.route('/fr_import')
    @app.route('/fr_importar')
    @login_required
    def fr_import_page():
        return render_template('fr_import.html', page_title='Importar Faturas de Reservas')

    @app.route('/reconciliacao_bancaria')
    @app.route('/reconciliacao-bancaria')
    @login_required
    def reconciliacao_bancaria_page():
        return render_template('reconciliacao_bancaria.html', page_title='Reconciliação Bancária')

    @app.route('/api/extratos')
    @login_required
    def api_extratos_list():
        try:
            cols = set(
                r[0] for r in db.session.execute(
                    text("SELECT UPPER(COLUMN_NAME) FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = 'EXT'")
                ).fetchall()
            )
            need = {'EXTSTAMP', 'NOCONTA', 'DATAINI', 'DATAFIM'}
            missing = sorted(list(need - cols))
            if missing:
                return jsonify({'error': f"Campos em falta na EXT: {', '.join(missing)}"}), 400

            rows = db.session.execute(text("""
                SELECT EXTSTAMP, NOCONTA, DATAINI, DATAFIM, ISNULL(RECONCILIADO,0) AS RECONCILIADO
                FROM dbo.EXT
                ORDER BY DATAFIM DESC, DATAINI DESC, EXTSTAMP DESC
            """)).mappings().all()
            out = []
            for r in rows:
                di = r.get('DATAINI')
                df = r.get('DATAFIM')
                out.append({
                    'EXTSTAMP': r.get('EXTSTAMP') or '',
                    'NOCONTA': int(r.get('NOCONTA') or 0),
                    'DATAINI': di.strftime('%Y-%m-%d') if isinstance(di, (date, datetime)) else (str(di) if di else ''),
                    'DATAFIM': df.strftime('%Y-%m-%d') if isinstance(df, (date, datetime)) else (str(df) if df else ''),
                    'RECONCILIADO': int(r.get('RECONCILIADO') or 0),
                })
            return jsonify(out)
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/extratos/<extstamp>')
    @login_required
    def api_extrato_detail(extstamp):
        """
        Devolve:
          - ext: cabeçalho
          - el: linhas do extrato (EL) com flag de reconciliado + groupid (se existir)
          - ba: movimentos BA no período/conta com flag de reconciliado + groupid (se existir)
        """
        try:
            extstamp = (extstamp or '').strip()
            if not extstamp:
                return jsonify({'error': 'EXTSTAMP obrigatório'}), 400

            ext = db.session.execute(text("""
                SELECT EXTSTAMP, NOCONTA, DATAINI, DATAFIM
                FROM dbo.EXT
                WHERE EXTSTAMP = :s
            """), {'s': extstamp}).mappings().first()
            if not ext:
                return jsonify({'error': 'Extrato não encontrado.'}), 404

            noconta = int(ext.get('NOCONTA') or 0)
            di = ext.get('DATAINI')
            df = ext.get('DATAFIM')
            di_s = di.strftime('%Y-%m-%d') if isinstance(di, (date, datetime)) else str(di)
            df_s = df.strftime('%Y-%m-%d') if isinstance(df, (date, datetime)) else str(df)

            el_cols = set(
                r[0] for r in db.session.execute(
                    text("SELECT UPPER(COLUMN_NAME) FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = 'EL'")
                ).fetchall()
            )
            if 'EXTSTAMP' not in el_cols:
                return jsonify({'error': 'A tabela EL precisa do campo EXTSTAMP.'}), 400
            stamp_col = 'ELSTAMP' if 'ELSTAMP' in el_cols else ('ELSATMP' if 'ELSATMP' in el_cols else None)
            if not stamp_col:
                return jsonify({'error': 'A tabela EL precisa de um campo stamp (ELSTAMP ou ELSATMP).'}), 400

            has_rec = db.session.execute(text("""
                SELECT 1 AS X
                FROM INFORMATION_SCHEMA.TABLES
                WHERE TABLE_NAME = 'REC'
            """)).fetchone()
            if not has_rec:
                return jsonify({'error': 'Tabela REC não existe.'}), 400

            # EL lines + reconciled info
            el_rows = db.session.execute(text(f"""
                SELECT
                  LTRIM(RTRIM(EL.{stamp_col})) AS STAMP,
                  CAST(EL.DATA AS date) AS DATA,
                  CAST(EL.DTVALOR AS date) AS DTVALOR,
                  ISNULL(EL.DESCRICAO,'') AS DESCRICAO,
                  CAST(EL.VALOR AS numeric(12,2)) AS VALOR,
                  ISNULL(R.GROUPID, 0) AS GROUPID
                FROM dbo.EL AS EL
                LEFT JOIN (
                  SELECT LTRIM(RTRIM(STAMP)) AS STAMP, MAX(GROUPID) AS GROUPID
                  FROM dbo.REC
                  WHERE EXTSTAMP = :s AND ORIGEM = 'EL'
                  GROUP BY LTRIM(RTRIM(STAMP))
                ) AS R ON R.STAMP = LTRIM(RTRIM(EL.{stamp_col}))
                WHERE EL.EXTSTAMP = :s
                ORDER BY CAST(EL.DATA AS date), CAST(EL.DTVALOR AS date), EL.{stamp_col}
            """), {'s': extstamp}).mappings().all()
            el_out = []
            for r in el_rows:
                d0 = r.get('DATA')
                dv = r.get('DTVALOR')
                gid = int(r.get('GROUPID') or 0)
                el_out.append({
                    'STAMP': r.get('STAMP') or '',
                    'DATA': d0.strftime('%Y-%m-%d') if isinstance(d0, (date, datetime)) else (str(d0) if d0 else ''),
                    'DTVALOR': dv.strftime('%Y-%m-%d') if isinstance(dv, (date, datetime)) else (str(dv) if dv else ''),
                    'DESCRICAO': r.get('DESCRICAO') or '',
                    'VALOR': float(r.get('VALOR') or 0),
                    'RECONCILIADO': 1 if gid > 0 else 0,
                    'GROUPID': gid,
                })

            # BA movements normalized (signed) + reconciled info
            ba_rows = db.session.execute(text("""
                SELECT
                  LTRIM(RTRIM(ISNULL(BA.BASTAMP,''))) AS BASTAMP,
                  CAST(BA.DATA AS date) AS DATA,
                  ISNULL(BA.DOCUMENTO,'') AS DOCUMENTO,
                  ISNULL(BA.DESCRICAO,'') AS DESCRICAO,
                  CAST(ISNULL(BA.EENTRADA,0) - ISNULL(BA.ESAIDA,0) AS numeric(12,2)) AS VALOR
                FROM dbo.V_BA AS BA
                WHERE CAST(BA.DATA AS date) BETWEEN :di AND :df
                  AND (ISNULL(BA.EENTRADA,0) <> 0 OR ISNULL(BA.ESAIDA,0) <> 0)
                  AND BA.NOCONTA = :acc
                  AND LTRIM(RTRIM(ISNULL(BA.BASTAMP,''))) <> ''
                ORDER BY CAST(BA.DATA AS date), ISNULL(BA.DOCUMENTO,''), ISNULL(BA.DESCRICAO,''), ISNULL(BA.BASTAMP,'')
            """), {'di': di_s, 'df': df_s, 'acc': noconta}).mappings().all()

            # groupid by BA stamp for this ext
            ba_rec = db.session.execute(text("""
                SELECT LTRIM(RTRIM(STAMP)) AS STAMP, MAX(GROUPID) AS GROUPID
                FROM dbo.REC
                WHERE EXTSTAMP = :s AND ORIGEM = 'BA'
                GROUP BY LTRIM(RTRIM(STAMP))
            """), {'s': extstamp}).mappings().all()
            ba_gid = { (r.get('STAMP') or '').strip(): int(r.get('GROUPID') or 0) for r in ba_rec }

            ba_out = []
            for r in ba_rows:
                d0 = r.get('DATA')
                bastamp = (r.get('BASTAMP') or '').strip()
                gid = ba_gid.get(bastamp, 0)
                ba_out.append({
                    'BASTAMP': bastamp,
                    'DATA': d0.strftime('%Y-%m-%d') if isinstance(d0, (date, datetime)) else (str(d0) if d0 else ''),
                    'DOCUMENTO': r.get('DOCUMENTO') or '',
                    'DESCRICAO': r.get('DESCRICAO') or '',
                    'VALOR': float(r.get('VALOR') or 0),
                    'RECONCILIADO': 1 if gid > 0 else 0,
                    'GROUPID': gid,
                })

            return jsonify({
                'ext': {
                    'EXTSTAMP': extstamp,
                    'NOCONTA': noconta,
                    'DATAINI': di_s,
                    'DATAFIM': df_s,
                },
                'el': el_out,
                'ba': ba_out,
            })
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/reconciliacao', methods=['POST'])
    @login_required
    def api_reconciliar():
        """
        Cria um grupo REC para um conjunto de linhas EL e BA.
        Regras:
          - todas as linhas ficam associadas ao mesmo EXTSTAMP + GROUPID
          - a soma EL deve igualar a soma BA (diferença ~0)
        """
        import uuid
        from decimal import Decimal
        from sqlalchemy import bindparam
        try:
            body = request.get_json(silent=True) or {}
            extstamp = (body.get('EXTSTAMP') or '').strip()
            el_list = body.get('EL') or []
            ba_list = body.get('BA') or []
            if not extstamp:
                return jsonify({'error': 'EXTSTAMP obrigatório'}), 400
            if not isinstance(el_list, list) or not isinstance(ba_list, list) or not el_list or not ba_list:
                return jsonify({'error': 'Seleciona pelo menos 1 linha EL e 1 movimento BA.'}), 400

            ext = db.session.execute(text("""
                SELECT EXTSTAMP, NOCONTA, DATAINI, DATAFIM
                FROM dbo.EXT
                WHERE EXTSTAMP = :s
            """), {'s': extstamp}).mappings().first()
            if not ext:
                return jsonify({'error': 'Extrato não encontrado.'}), 404

            noconta = int(ext.get('NOCONTA') or 0)
            di = ext.get('DATAINI')
            df = ext.get('DATAFIM')
            di_s = di.strftime('%Y-%m-%d') if isinstance(di, (date, datetime)) else str(di)
            df_s = df.strftime('%Y-%m-%d') if isinstance(df, (date, datetime)) else str(df)

            el_cols = set(
                r[0] for r in db.session.execute(
                    text("SELECT UPPER(COLUMN_NAME) FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = 'EL'")
                ).fetchall()
            )
            if 'EXTSTAMP' not in el_cols:
                return jsonify({'error': 'A tabela EL precisa do campo EXTSTAMP.'}), 400
            stamp_col = 'ELSTAMP' if 'ELSTAMP' in el_cols else ('ELSATMP' if 'ELSATMP' in el_cols else None)
            if not stamp_col:
                return jsonify({'error': 'A tabela EL precisa de um campo stamp (ELSTAMP ou ELSATMP).'}), 400

            # Carrega valores EL
            el_vals = db.session.execute(text(f"""
                SELECT LTRIM(RTRIM({stamp_col})) AS STAMP, CAST(VALOR AS numeric(12,2)) AS VALOR
                FROM dbo.EL
                WHERE EXTSTAMP = :s AND LTRIM(RTRIM({stamp_col})) IN :ids
            """).bindparams(bindparam('ids', expanding=True)), {'s': extstamp, 'ids': el_list}).mappings().all()
            if len(el_vals) != len(set(el_list)):
                return jsonify({'error': 'Uma ou mais linhas EL não foram encontradas no extrato.'}), 400

            el_sum = sum(Decimal(str(r.get('VALOR') or 0)) for r in el_vals)

            # Carrega valores BA (no período/conta)
            ba_vals = db.session.execute(text("""
                SELECT LTRIM(RTRIM(ISNULL(BA.BASTAMP,''))) AS STAMP, CAST(ISNULL(BA.EENTRADA,0) - ISNULL(BA.ESAIDA,0) AS numeric(12,2)) AS VALOR
                FROM dbo.V_BA AS BA
                WHERE CAST(BA.DATA AS date) BETWEEN :di AND :df
                  AND BA.NOCONTA = :acc
                  AND LTRIM(RTRIM(ISNULL(BA.BASTAMP,''))) IN :ids
            """).bindparams(bindparam('ids', expanding=True)), {'di': di_s, 'df': df_s, 'acc': noconta, 'ids': ba_list}).mappings().all()
            if len(ba_vals) != len(set(ba_list)):
                return jsonify({'error': 'Um ou mais movimentos BA não foram encontrados no período/conta.'}), 400

            ba_sum = sum(Decimal(str(r.get('VALOR') or 0)) for r in ba_vals)
            diff = (el_sum - ba_sum).copy_abs()
            if diff > Decimal('0.01'):
                return jsonify({'error': f'Os valores não batem. Diferença: {float(el_sum - ba_sum):.2f}'}), 400

            # Novo groupid sequencial (lock para evitar duplicados)
            gid = db.session.execute(text("""
                SELECT ISNULL(MAX(GROUPID),0) + 1 AS N
                FROM dbo.REC WITH (UPDLOCK, HOLDLOCK)
            """)).scalar()
            gid = int(gid or 1)

            # Inserir REC (uma linha por item)
            rec_rows = []
            for r in el_vals:
                rec_rows.append({
                    'RECSTAMP': uuid.uuid4().hex[:25].upper(),
                    'EXTSTAMP': extstamp,
                    'GROUPID': gid,
                    'ORIGEM': 'EL',
                    'STAMP': (r.get('STAMP') or '').strip(),
                    'VALOR': float(r.get('VALOR') or 0),
                })
            for r in ba_vals:
                rec_rows.append({
                    'RECSTAMP': uuid.uuid4().hex[:25].upper(),
                    'EXTSTAMP': extstamp,
                    'GROUPID': gid,
                    'ORIGEM': 'BA',
                    'STAMP': (r.get('STAMP') or '').strip(),
                    'VALOR': float(r.get('VALOR') or 0),
                })

            db.session.execute(text("""
                INSERT INTO dbo.REC (RECSTAMP, EXTSTAMP, GROUPID, ORIGEM, STAMP, VALOR)
                VALUES (:RECSTAMP, :EXTSTAMP, :GROUPID, :ORIGEM, :STAMP, :VALOR)
            """), rec_rows)

            # Atualiza EXT.RECONCILIADO se todas as EL estiverem conciliadas
            not_done = db.session.execute(text(f"""
                SELECT COUNT(1) AS N
                FROM dbo.EL AS EL
                LEFT JOIN dbo.REC AS R
                  ON R.EXTSTAMP = EL.EXTSTAMP AND R.ORIGEM = 'EL' AND R.STAMP = EL.{stamp_col}
                WHERE EL.EXTSTAMP = :s
                  AND R.RECSTAMP IS NULL
            """), {'s': extstamp}).scalar()
            if int(not_done or 0) == 0:
                db.session.execute(text("UPDATE dbo.EXT SET RECONCILIADO = 1 WHERE EXTSTAMP = :s"), {'s': extstamp})

            db.session.commit()
            return jsonify({'ok': True, 'GROUPID': gid})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/reconciliacao/<extstamp>/<int:groupid>', methods=['DELETE'])
    @login_required
    def api_reconciliacao_delete_group(extstamp, groupid):
        """
        Remove uma reconciliação (grupo) inteira:
          - apaga todas as linhas da REC do mesmo EXTSTAMP + GROUPID
          - recalcula EXT.RECONCILIADO (1 apenas se todas as EL estiverem reconciliadas)
        """
        try:
            extstamp = (extstamp or '').strip()
            if not extstamp:
                return jsonify({'error': 'EXTSTAMP obrigatório'}), 400
            if int(groupid or 0) <= 0:
                return jsonify({'error': 'GROUPID inválido'}), 400

            # garantir que existe
            has = db.session.execute(text("""
                SELECT TOP 1 1 AS X
                FROM dbo.REC
                WHERE EXTSTAMP = :s AND GROUPID = :g
            """), {'s': extstamp, 'g': int(groupid)}).fetchone()
            if not has:
                return jsonify({'error': 'Grupo não encontrado.'}), 404

            db.session.execute(text("""
                DELETE FROM dbo.REC
                WHERE EXTSTAMP = :s AND GROUPID = :g
            """), {'s': extstamp, 'g': int(groupid)})

            # Recalcular EXT.RECONCILIADO: 1 apenas se todas as EL tiverem reconciliação
            el_cols = set(
                r[0] for r in db.session.execute(
                    text("SELECT UPPER(COLUMN_NAME) FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = 'EL'")
                ).fetchall()
            )
            if 'EXTSTAMP' in el_cols:
                stamp_col = 'ELSTAMP' if 'ELSTAMP' in el_cols else ('ELSATMP' if 'ELSATMP' in el_cols else None)
                if stamp_col:
                    not_done = db.session.execute(text(f"""
                        SELECT COUNT(1) AS N
                        FROM dbo.EL AS EL
                        LEFT JOIN dbo.REC AS R
                          ON R.EXTSTAMP = EL.EXTSTAMP AND R.ORIGEM = 'EL' AND R.STAMP = LTRIM(RTRIM(EL.{stamp_col}))
                        WHERE EL.EXTSTAMP = :s
                          AND R.RECSTAMP IS NULL
                    """), {'s': extstamp}).scalar()
                    if int(not_done or 0) == 0:
                        db.session.execute(text("UPDATE dbo.EXT SET RECONCILIADO = 1 WHERE EXTSTAMP = :s"), {'s': extstamp})
                    else:
                        db.session.execute(text("UPDATE dbo.EXT SET RECONCILIADO = 0 WHERE EXTSTAMP = :s"), {'s': extstamp})

            db.session.commit()
            return jsonify({'ok': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/reconciliacao/ow', methods=['POST'])
    @login_required
    def api_reconciliacao_create_ow():
        """
        Gera movimentos de tesouraria (OW) no ERP a partir de entradas (VALOR > 0)
        selecionadas na grelha EL (não reconciliadas).

        Body:
          { "EXTSTAMP": "...", "EL": ["<ELSTAMP>", ...] }
        """
        try:
            body = request.get_json(silent=True) or {}
            extstamp = (body.get('EXTSTAMP') or '').strip()
            stamps = body.get('EL') or []
            if not extstamp:
                return jsonify({'error': 'EXTSTAMP obrigatório'}), 400
            if not isinstance(stamps, list) or not stamps:
                return jsonify({'error': 'Selecione movimentos EL.'}), 400

            ERP_DB = 'GUEST_SPA_TUR'

            # Detetar stamp column na EL
            el_cols = set(
                r[0] for r in db.session.execute(
                    text("SELECT UPPER(COLUMN_NAME) FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = 'EL'")
                ).fetchall()
            )
            if 'EXTSTAMP' not in el_cols:
                return jsonify({'error': 'A tabela EL precisa do campo EXTSTAMP.'}), 400
            stamp_col = 'ELSTAMP' if 'ELSTAMP' in el_cols else ('ELSATMP' if 'ELSATMP' in el_cols else None)
            if not stamp_col:
                return jsonify({'error': 'A tabela EL precisa de um campo stamp (ELSTAMP ou ELSATMP).'}), 400

            # Verificar tabela/colunas na OW (na BD do ERP)
            ow_tbl = db.session.execute(text(f"""
                SELECT 1 AS X
                FROM [{ERP_DB}].INFORMATION_SCHEMA.TABLES
                WHERE TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'OW'
            """)).fetchone()
            if not ow_tbl:
                return jsonify({'error': f"Tabela OW não existe na BD {ERP_DB}."}), 400

            ow_cols = set(
                r[0] for r in db.session.execute(
                    text(f"""
                        SELECT UPPER(COLUMN_NAME)
                        FROM [{ERP_DB}].INFORMATION_SCHEMA.COLUMNS
                        WHERE TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'OW'
                    """)
                ).fetchall()
            )
            if not ow_cols:
                return jsonify({'error': f"Não foi possível ler colunas da OW em {ERP_DB}."}), 400

            eentr_col = 'EENTR' if 'EENTR' in ow_cols else ('EENTRADA' if 'EENTRADA' in ow_cols else None)
            docno_col = 'DOCNO' if 'DOCNO' in ow_cols else ('NDOS' if 'NDOS' in ow_cols else None)

            required = {'OWSTAMP', 'DATA', 'DOCNOME', 'DESCRICAO', 'LOCAL', 'SGRUPO', 'GRUPO', 'ORIGEM', 'CONTADO', 'OLLOCAL', 'OLCODIGO'}
            missing_req = sorted([c for c in required if c not in ow_cols])
            if not eentr_col:
                missing_req.append('EENTR')
            if not docno_col:
                missing_req.append('DOCNO')
            if missing_req:
                return jsonify({'error': f"Campos em falta na OW ({ERP_DB}): " + ', '.join(missing_req)}), 400

            inserted = 0
            skipped_exists = 0
            skipped_reconciled = 0
            skipped_invalid = 0

            ins_cols = ['OWSTAMP', 'DATA', 'DOCNOME', 'DESCRICAO', eentr_col, 'LOCAL', 'SGRUPO', 'GRUPO', 'ORIGEM', 'CONTADO', 'OLLOCAL', 'OLCODIGO', docno_col]
            ins_vals = [':OWSTAMP', ':DATA', ':DOCNOME', ':DESCRICAO', ':EENTR', ':LOCAL', ':SGRUPO', ':GRUPO', ':ORIGEM', ':CONTADO', ':OLLOCAL', ':OLCODIGO', ':DOCNO']
            ins_sql = text(f"""
                INSERT INTO [{ERP_DB}].[dbo].[OW] ({", ".join(ins_cols)})
                VALUES ({", ".join(ins_vals)})
            """)

            for s in stamps[:5000]:
                stamp = (str(s or '').strip())[:25]
                if not stamp:
                    skipped_invalid += 1
                    continue

                # Já reconciliado?
                rec = db.session.execute(text("""
                    SELECT TOP 1 1 AS X
                    FROM dbo.REC
                    WHERE EXTSTAMP = :e AND ORIGEM = 'EL' AND LTRIM(RTRIM(STAMP)) = :s
                """), {'e': extstamp, 's': stamp}).fetchone()
                if rec:
                    skipped_reconciled += 1
                    continue

                # Buscar EL
                el = db.session.execute(text(f"""
                    SELECT TOP 1
                      CAST(DATA AS date) AS DATA,
                      ISNULL(DESCRICAO,'') AS DESCRICAO,
                      CAST(VALOR AS numeric(19,6)) AS VALOR
                    FROM dbo.EL
                    WHERE EXTSTAMP = :e
                      AND LTRIM(RTRIM({stamp_col})) = :s
                """), {'e': extstamp, 's': stamp}).mappings().first()
                if not el:
                    skipped_invalid += 1
                    continue

                valor = float(el.get('VALOR') or 0)
                if valor <= 0.005:
                    skipped_invalid += 1
                    continue

                # Já existe na OW?
                exists = db.session.execute(text("""
                    SELECT TOP 1 1 AS X
                    FROM [GUEST_SPA_TUR].[dbo].[OW]
                    WHERE LTRIM(RTRIM(OWSTAMP)) = :s
                """), {'s': stamp}).fetchone()
                if exists:
                    skipped_exists += 1
                    continue

                d = el.get('DATA')
                if isinstance(d, datetime):
                    d = d.date()
                if not isinstance(d, date):
                    skipped_invalid += 1
                    continue
                data_dt = datetime.combine(d, datetime.min.time())
                desc = (str(el.get('DESCRICAO') or '')).strip()

                db.session.execute(ins_sql, {
                    'OWSTAMP': stamp,
                    'DATA': data_dt,
                    'DOCNOME': 'Recebimento Airbnb',
                    'DESCRICAO': desc,
                    'EENTR': valor,
                    'LOCAL': 'B',
                    'SGRUPO': 'Recebimentos de Clientes',
                    'GRUPO': 'Actividades Operacionais',
                    'ORIGEM': 'OW',
                    'CONTADO': 1,
                    'OLLOCAL': 'Santander  DO',
                    'OLCODIGO': 'R10001',
                    'DOCNO': 15
                })
                inserted += 1

            db.session.commit()
            return jsonify({
                'ok': True,
                'inserted': inserted,
                'skipped_exists': skipped_exists,
                'skipped_reconciled': skipped_reconciled,
                'skipped_invalid': skipped_invalid
            })
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/extratos/import', methods=['POST'])
    @login_required
    def api_extratos_import():
        """
        Importa um extrato bancário a partir de um CSV (ou Excel).
        CSV (separador ';'):
          - campo 2: DATA
          - campo 3: DTVALOR
          - campo 4: DESCRICAO
          - campo 7: VALOR (com sinal +/-)
        Excel:
          Folha: "estrato"
            - B: DATA
            - C: DTVALOR
            - D: DESCRICAO
            - E: DEBITO (negativo)
            - F: CREDITO (positivo)
        Cria EXT + EL.
        """
        import io
        import uuid
        from decimal import Decimal, InvalidOperation
        import csv

        # openpyxl é opcional (só para xlsx)
        load_workbook = None
        from_excel = None
        try:
            from openpyxl import load_workbook as _lw
            from openpyxl.utils.datetime import from_excel as _fe
            load_workbook = _lw
            from_excel = _fe
        except Exception:
            load_workbook = None
            from_excel = None

        def to_date(v):
            if v is None or v == '':
                return None
            if isinstance(v, datetime):
                return v.date()
            if isinstance(v, date):
                return v
            if isinstance(v, (int, float)) and from_excel:
                try:
                    dt = from_excel(v)
                    return dt.date() if isinstance(dt, datetime) else dt
                except Exception:
                    return None
            if isinstance(v, str):
                s = v.strip()
                if not s:
                    return None
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
                    try:
                        return datetime.strptime(s, fmt).date()
                    except Exception:
                        pass
            return None

        def to_dec(v):
            if v is None or v == '':
                return Decimal('0')
            if isinstance(v, Decimal):
                return v
            if isinstance(v, (int, float)):
                return Decimal(str(v))
            if isinstance(v, str):
                s = v.strip().replace('.', '').replace(',', '.')
                if s in ('', '-', '+'):
                    return Decimal('0')
                try:
                    return Decimal(s)
                except InvalidOperation:
                    return Decimal('0')
            try:
                return Decimal(str(v))
            except Exception:
                return Decimal('0')

        try:
            noconta = (request.form.get('noconta') or '').strip()
            dataini = (request.form.get('dataini') or '').strip()
            datafim = (request.form.get('datafim') or '').strip()
            saldoini = (request.form.get('saldoini') or '0').strip()
            saldofim = (request.form.get('saldofim') or '0').strip()
            f = request.files.get('file')

            if not noconta:
                return jsonify({'error': 'Conta em falta.'}), 400
            if not dataini or not datafim:
                return jsonify({'error': 'Data início/fim em falta.'}), 400
            if not f or not getattr(f, 'filename', ''):
                return jsonify({'error': 'Ficheiro em falta.'}), 400
            filename = (getattr(f, 'filename', '') or '').lower()

            try:
                noconta_i = int(noconta)
            except Exception:
                return jsonify({'error': 'NOCONTA inválido.'}), 400

            try:
                di = datetime.strptime(dataini, '%Y-%m-%d').date()
                df = datetime.strptime(datafim, '%Y-%m-%d').date()
            except Exception:
                return jsonify({'error': 'Datas inválidas (YYYY-MM-DD).'}), 400

            if df < di:
                return jsonify({'error': 'Data fim não pode ser inferior à data início.'}), 400

            if len(saldoini) > 10 or len(saldofim) > 10:
                return jsonify({'error': 'Saldo início/fim demasiado longo (máx 10 caracteres).'}), 400

            # Verifica colunas da tabela EL
            el_cols = set(
                r[0] for r in db.session.execute(
                    text("SELECT UPPER(COLUMN_NAME) FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = 'EL'")
                ).fetchall()
            )
            ext_cols = set(
                r[0] for r in db.session.execute(
                    text("SELECT UPPER(COLUMN_NAME) FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = 'EXT'")
                ).fetchall()
            )
            need_ext = {'EXTSTAMP', 'NOCONTA', 'DATAINI', 'DATAFIM', 'SALDOINI', 'SALDOFIM', 'RECONCILIADO'}
            missing_ext = sorted(list(need_ext - ext_cols))
            if missing_ext:
                return jsonify({'error': f"Campos em falta na EXT: {', '.join(missing_ext)}"}), 400

            # Precisamos de ligar EL ao cabeçalho
            if 'EXTSTAMP' not in el_cols:
                return jsonify({'error': 'A tabela EL precisa do campo EXTSTAMP para ligar ao cabeçalho.'}), 400

            stamp_col = 'ELSTAMP' if 'ELSTAMP' in el_cols else ('ELSATMP' if 'ELSATMP' in el_cols else None)
            if not stamp_col:
                return jsonify({'error': 'A tabela EL precisa de um campo stamp (ELSTAMP ou ELSATMP).'}), 400

            rows_in = []
            skipped = 0
            file_bytes = f.read()

            if filename.endswith('.csv'):
                # decode best-effort (PT banks often cp1252/latin1)
                text_data = None
                for enc in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1'):
                    try:
                        text_data = file_bytes.decode(enc)
                        break
                    except Exception:
                        continue
                if text_data is None:
                    return jsonify({'error': 'Não foi possível ler o CSV (encoding).'}), 400

                rdr = csv.reader(io.StringIO(text_data), delimiter=';')
                for row in rdr:
                    if not row or len(row) < 7:
                        continue
                    data0 = to_date(row[1])
                    dtval0 = to_date(row[2])
                    desc0 = (row[3] or '').strip()

                    # Alguns CSVs trazem débito/crédito em colunas separadas (débitos já negativos).
                    # Heurística:
                    #   - tenta 6ª e 7ª colunas (1-based) => indexes 5 e 6 (0-based)
                    #   - se der zero, tenta o primeiro número não-zero a seguir à descrição
                    deb = to_dec(row[5]) if len(row) > 5 else Decimal('0')
                    cred = to_dec(row[6]) if len(row) > 6 else Decimal('0')
                    valor = (deb or Decimal('0')) + (cred or Decimal('0'))
                    if abs(valor) < Decimal('0.005'):
                        for cell in row[4:]:
                            v2 = to_dec(cell)
                            if abs(v2) >= Decimal('0.005'):
                                valor = v2
                                break

                    if data0 is None:
                        continue
                    # filtra pelo intervalo escolhido
                    if data0 < di or data0 > df:
                        skipped += 1
                        continue
                    if dtval0 is None:
                        dtval0 = data0
                    if abs(valor) < Decimal('0.005'):
                        skipped += 1
                        continue
                    rows_in.append({
                        'DATA': data0,
                        'DTVALOR': dtval0,
                        'DESCRICAO': (desc0 or '')[:200],
                        'VALOR': valor.quantize(Decimal('0.01')),
                    })
            else:
                if not load_workbook:
                    return jsonify({'error': 'Para importar Excel (.xlsx) é necessário instalar openpyxl.'}), 400

                wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
                ws = None
                for name in wb.sheetnames:
                    if (name or '').strip().lower() == 'estrato':
                        ws = wb[name]
                        break
                if ws is None:
                    return jsonify({'error': 'Folha "estrato" não encontrada no Excel.'}), 400

                for row in ws.iter_rows(values_only=True):
                    # Colunas B..F => indexes 1..5 (0-based)
                    if not row or len(row) < 6:
                        continue
                    data0 = to_date(row[1])
                    dtval0 = to_date(row[2])
                    desc0 = (row[3] or '').strip() if isinstance(row[3], str) else (str(row[3]).strip() if row[3] is not None else '')
                    deb = to_dec(row[4])
                    cred = to_dec(row[5])
                    valor = (cred or Decimal('0')) + (deb or Decimal('0'))

                    if data0 is None:
                        continue
                    if data0 < di or data0 > df:
                        skipped += 1
                        continue
                    if dtval0 is None:
                        dtval0 = data0
                    if abs(valor) < Decimal('0.005'):
                        skipped += 1
                        continue
                    rows_in.append({
                        'DATA': data0,
                        'DTVALOR': dtval0,
                        'DESCRICAO': (desc0 or '')[:200],
                        'VALOR': valor.quantize(Decimal('0.01')),
                    })

            if not rows_in:
                return jsonify({'error': 'Não foram encontradas linhas válidas para importar.'}), 400

            extstamp = uuid.uuid4().hex[:25].upper()

            db.session.execute(text("""
                INSERT INTO dbo.EXT (EXTSTAMP, NOCONTA, DATAINI, DATAFIM, SALDOINI, SALDOFIM, RECONCILIADO)
                VALUES (:s, :n, :di, :df, :si, :sf, 0)
            """), {'s': extstamp, 'n': noconta_i, 'di': di, 'df': df, 'si': saldoini, 'sf': saldofim})

            # Inserção de linhas
            el_rows = []
            for r0 in rows_in:
                el_rows.append({
                    'EXTSTAMP': extstamp,
                    'STAMP': uuid.uuid4().hex[:25].upper(),
                    'DATA': r0['DATA'],
                    'DTVALOR': r0['DTVALOR'],
                    'DESCRICAO': r0['DESCRICAO'],
                    'VALOR': r0['VALOR'],
                })

            ins = text(f"""
                INSERT INTO dbo.EL ({stamp_col}, EXTSTAMP, DATA, DTVALOR, DESCRICAO, VALOR)
                VALUES (:STAMP, :EXTSTAMP, :DATA, :DTVALOR, :DESCRICAO, :VALOR)
            """)
            db.session.execute(ins, el_rows)
            db.session.commit()

            sample = [
                {
                    'DATA': r['DATA'].strftime('%Y-%m-%d') if isinstance(r['DATA'], date) else str(r['DATA']),
                    'DTVALOR': r['DTVALOR'].strftime('%Y-%m-%d') if isinstance(r['DTVALOR'], date) else str(r['DTVALOR']),
                    'DESCRICAO': r['DESCRICAO'],
                    'VALOR': float(r['VALOR']),
                }
                for r in rows_in[:50]
            ]
            return jsonify({'ok': True, 'EXTSTAMP': extstamp, 'inserted': len(rows_in), 'skipped': skipped, 'sample': sample})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/fr/import', methods=['POST'])
    @login_required
    def api_fr_import():
        """
        Importa linhas para dbo.FR a partir de Excel.

        Excel:
          - 1ª linha: títulos (ignorada)
          - Colunas:
            A DOCUMENTO, B DATA, C CLIENTE, D CCUSTO, I BASE, J TAXAIVA, K IVA, L ANULADO, M ARTIGO

        Regras:
          - só importa ARTIGO in ('Estadia', 'Taxa de Limpeza')
          - upsert por (DOCUMENTO, ARTIGO, CCUSTO)
        """
        import io
        import uuid
        from decimal import Decimal, InvalidOperation

        # openpyxl é necessário para xlsx
        try:
            from openpyxl import load_workbook
            from openpyxl.utils.datetime import from_excel as from_excel_date
        except Exception:
            load_workbook = None
            from_excel_date = None

        def new_stamp():
            return uuid.uuid4().hex.upper()[:25]

        def norm_txt(v, max_len=None):
            s = ('' if v is None else str(v)).strip()
            if max_len:
                s = s[:max_len]
            return s

        def norm_artigo(v):
            s = norm_txt(v, 50)
            s2 = ' '.join(s.split()).strip().casefold()
            return s2

        def to_date(v):
            if v is None or v == '':
                return None
            if isinstance(v, datetime):
                return v.date()
            if isinstance(v, date):
                return v
            if isinstance(v, (int, float)) and from_excel_date:
                try:
                    dt = from_excel_date(v)
                    return dt.date() if isinstance(dt, datetime) else dt
                except Exception:
                    return None
            if isinstance(v, str):
                s = v.strip()
                if not s:
                    return None
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
                    try:
                        return datetime.strptime(s, fmt).date()
                    except Exception:
                        pass
            return None

        def to_dec(v):
            if v is None or v == '':
                return Decimal('0')
            if isinstance(v, Decimal):
                return v
            if isinstance(v, (int, float)):
                return Decimal(str(v))
            if isinstance(v, str):
                s = v.strip().replace('.', '').replace(',', '.')
                if s in ('', '-', '+'):
                    return Decimal('0')
                try:
                    return Decimal(s)
                except InvalidOperation:
                    return Decimal('0')
            try:
                return Decimal(str(v))
            except Exception:
                return Decimal('0')

        def to_bit(v):
            if v is None or v == '':
                return 0
            if isinstance(v, bool):
                return 1 if v else 0
            if isinstance(v, (int, float)):
                try:
                    return 1 if float(v) != 0 else 0
                except Exception:
                    return 0
            s = str(v).strip().casefold()
            return 1 if s in ('1', 'true', 'sim', 's', 'yes', 'y', 'anulado') else 0

        try:
            f = request.files.get('file')
            if not f or not getattr(f, 'filename', ''):
                return jsonify({'error': 'Ficheiro em falta.'}), 400
            filename = (getattr(f, 'filename', '') or '').lower()
            if not filename.endswith('.xlsx'):
                return jsonify({'error': 'Formato inválido. Use .xlsx'}), 400
            if not load_workbook:
                return jsonify({'error': 'openpyxl não está disponível no servidor.'}), 500

            # validar tabela/colunas mínimas
            cols = set(
                r[0] for r in db.session.execute(
                    text("SELECT UPPER(COLUMN_NAME) FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = 'FR'")
                ).fetchall()
            )
            need = {'FRSTAMP', 'DOCUMENTO', 'DATA', 'CLIENTE', 'CCUSTO', 'ARTIGO', 'BASE', 'TAXAIVA', 'IVA', 'ANULADO'}
            missing = sorted(list(need - cols))
            if missing:
                return jsonify({'error': f"Campos em falta na FR: {', '.join(missing)}"}), 400

            raw = f.read()
            wb = load_workbook(io.BytesIO(raw), data_only=True)
            ws = wb.active

            allowed = {'estadia', 'taxa de limpeza'}
            inserted = 0
            updated = 0
            skipped = 0
            sample = []

            # cache de existentes por key (DOCUMENTO, ARTIGO, CCUSTO) -> FRSTAMP
            existing_rows = db.session.execute(text("""
                SELECT FRSTAMP, DOCUMENTO, ARTIGO, CCUSTO
                FROM dbo.FR
            """)).mappings().all()
            existing = {}
            for r in existing_rows:
                k = (
                    norm_txt(r.get('DOCUMENTO'), 100),
                    norm_txt(r.get('ARTIGO'), 25),
                    norm_txt(r.get('CCUSTO'), 50),
                )
                if all(k):
                    existing[k] = norm_txt(r.get('FRSTAMP'), 25)

            ins_sql = text("""
                INSERT INTO dbo.FR
                (FRSTAMP, DOCUMENTO, DATA, CLIENTE, CCUSTO, ARTIGO, BASE, TAXAIVA, IVA, ANULADO)
                VALUES
                (:FRSTAMP, :DOCUMENTO, :DATA, :CLIENTE, :CCUSTO, :ARTIGO, :BASE, :TAXAIVA, :IVA, :ANULADO)
            """)
            upd_sql = text("""
                UPDATE dbo.FR
                SET DATA = :DATA,
                    CLIENTE = :CLIENTE,
                    BASE = :BASE,
                    TAXAIVA = :TAXAIVA,
                    IVA = :IVA,
                    ANULADO = :ANULADO
                WHERE FRSTAMP = :FRSTAMP
            """)

            # iterar a partir da linha 2 (1 = títulos)
            for row_idx in range(2, ws.max_row + 1):
                doc = norm_txt(ws.cell(row=row_idx, column=1).value, 100)   # A
                dt = to_date(ws.cell(row=row_idx, column=2).value)          # B
                cliente = norm_txt(ws.cell(row=row_idx, column=3).value, 100)  # C
                ccusto = norm_txt(ws.cell(row=row_idx, column=4).value, 50) # D
                base = to_dec(ws.cell(row=row_idx, column=9).value)         # I
                taxaiva = to_dec(ws.cell(row=row_idx, column=10).value)     # J
                iva = to_dec(ws.cell(row=row_idx, column=11).value)         # K
                anulado = to_bit(ws.cell(row=row_idx, column=12).value)     # L
                artigo = norm_txt(ws.cell(row=row_idx, column=13).value, 25)  # M

                if not doc or not artigo or not ccusto or not dt:
                    skipped += 1
                    continue
                if norm_artigo(artigo) not in allowed:
                    skipped += 1
                    continue

                k = (doc, artigo, ccusto)
                frstamp = existing.get(k)
                params = {
                    'FRSTAMP': frstamp or new_stamp(),
                    'DOCUMENTO': doc,
                    'DATA': dt,
                    'CLIENTE': cliente,
                    'CCUSTO': ccusto,
                    'ARTIGO': artigo,
                    'BASE': base,
                    'TAXAIVA': taxaiva,
                    'IVA': iva,
                    'ANULADO': anulado,
                }

                if frstamp:
                    db.session.execute(upd_sql, params)
                    updated += 1
                else:
                    db.session.execute(ins_sql, params)
                    inserted += 1
                    existing[k] = params['FRSTAMP']

                if len(sample) < 50:
                    sample.append({
                        'DOCUMENTO': doc,
                        'DATA': dt.strftime('%Y-%m-%d') if dt else '',
                        'CLIENTE': cliente,
                        'CCUSTO': ccusto,
                        'ARTIGO': artigo,
                        'BASE': float(base),
                        'TAXAIVA': float(taxaiva),
                        'IVA': float(iva),
                        'ANULADO': int(anulado),
                        'MODE': 'U' if frstamp else 'I',
                    })

            db.session.commit()
            return jsonify({'ok': True, 'inserted': inserted, 'updated': updated, 'skipped': skipped, 'sample': sample})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/generic/api/tesouraria/contas')
    @login_required
    def api_tesouraria_contas():
        try:
            sql = text("""
                SELECT BANCO, CONTA, ORDEM, NOCONTA
                FROM V_BL
                ORDER BY ORDEM, BANCO, CONTA
            """)
            rows = db.session.execute(sql).mappings().all()
            return jsonify([dict(r) for r in rows])
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/generic/api/tesouraria')
    @login_required
    def api_tesouraria():
        try:
            start = request.args.get('start')
            end = request.args.get('end')
            account = (request.args.get('account') or '').strip()
            if not start or not end:
                return jsonify({'error': 'Parâmetros start/end obrigatórios'}), 400
            base_sql = """
                SELECT DATAENTRADA AS DATA, TIPOLINHA AS TIPO, SUM(VALOR) AS VALOR, ISNULL(NoConta, 1) AS NOCONTA
                FROM DBO.V_ENTRADAS_TESOURARIA
                WHERE DATAENTRADA BETWEEN :start AND :end
            """
            params = {'start': start, 'end': end}
            if account:
                # As previsões podem vir sem NoConta; por defeito, assumimos que pertencem à conta 1.
                base_sql += " AND ISNULL(NoConta, 1) = :account"
                params['account'] = account
            base_sql += " GROUP BY DATAENTRADA, TIPOLINHA, ISNULL(NoConta, 1) ORDER BY DATAENTRADA, TIPOLINHA"
            sql = text(base_sql)
            rows = db.session.execute(sql, params).mappings().all()
            return jsonify([dict(r) for r in rows])
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/generic/api/tesouraria/ba')
    @login_required
    def api_tesouraria_ba():
        """
        Movimentos reais de tesouraria (BA) por dia:
        - Entradas: soma EENTRADA
        - Saídas: soma ESAIDA
        Fonte: dbo.V_BA (na BD GESTAO)
        """
        try:
            start = request.args.get('start')
            end = request.args.get('end')
            account = (request.args.get('account') or '').strip()
            if not start or not end:
                return jsonify({'error': 'Parâmetros start/end obrigatórios'}), 400

            where_account = ""
            params = {'start': start, 'end': end}
            if account:
                where_account = " AND BA.NOCONTA = :account"
                params['account'] = account

            sql = text(f"""
                SELECT
                    CAST(BA.DATA AS date) AS DATA,
                    SUM(ISNULL(BA.EENTRADA, 0)) AS EENTRADA,
                    SUM(ISNULL(BA.ESAIDA, 0)) AS ESAIDA
                FROM dbo.V_BA AS BA
                WHERE CAST(BA.DATA AS date) BETWEEN :start AND :end
                  AND (ISNULL(BA.EENTRADA, 0) <> 0 OR ISNULL(BA.ESAIDA, 0) <> 0)
                  {where_account}
                GROUP BY CAST(BA.DATA AS date)
                ORDER BY CAST(BA.DATA AS date)
            """)
            rows = db.session.execute(sql, params).mappings().all()
            out = []
            for r in rows:
                d = r.get('DATA')
                if isinstance(d, (datetime, date)):
                    d = d.strftime('%Y-%m-%d')
                else:
                    d = str(d) if d is not None else ''
                out.append({
                    'DATA': d,
                    'EENTRADA': float(r.get('EENTRADA') or 0),
                    'ESAIDA': float(r.get('ESAIDA') or 0)
                })
            return jsonify(out)
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/generic/api/tesouraria/ba/saldo_base')
    @login_required
    def api_tesouraria_ba_saldo_base():
        """
        Saldo base (acumulado) antes de uma data (exclusive).
        Query: ?before=YYYY-MM-DD
        Retorna: { "base": <float> }
        """
        try:
            before = (request.args.get('before') or '').strip()
            account = (request.args.get('account') or '').strip()
            if not before:
                return jsonify({'error': 'Parâmetro before obrigatório'}), 400

            where_account = ""
            params = {'before': before}
            if account:
                where_account = " AND BA.NOCONTA = :account"
                params['account'] = account

            sql = text(f"""
                SELECT
                    SUM(ISNULL(BA.EENTRADA, 0) - ISNULL(BA.ESAIDA, 0)) AS BASE
                FROM dbo.V_BA AS BA
                WHERE CAST(BA.DATA AS date) < :before
                  AND (ISNULL(BA.EENTRADA, 0) <> 0 OR ISNULL(BA.ESAIDA, 0) <> 0)
                  {where_account}
            """)
            row = db.session.execute(sql, params).mappings().first() or {}
            base = float(row.get('BASE') or 0)
            return jsonify({'base': base})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/generic/api/tesouraria/ba/detalhe')
    @login_required
    def api_tesouraria_ba_detalhe():
        """
        Detalhe de movimentos reais (V_BA) num dia.
        Query: ?date=YYYY-MM-DD&kind=in|out
        """
        try:
            d = (request.args.get('date') or '').strip()
            kind = (request.args.get('kind') or '').strip().lower()
            account = (request.args.get('account') or '').strip()
            if not d:
                return jsonify({'error': 'Parâmetro date obrigatório'}), 400
            if kind not in ('in', 'out'):
                return jsonify({'error': 'Parâmetro kind inválido (in|out)'}), 400

            if kind == 'in':
                where_kind = "ISNULL(BA.EENTRADA,0) <> 0"
                value_col = "ISNULL(BA.EENTRADA,0)"
            else:
                where_kind = "ISNULL(BA.ESAIDA,0) <> 0"
                value_col = "ISNULL(BA.ESAIDA,0)"

            where_account = ""
            params = {'d': d}
            if account:
                where_account = " AND BA.NOCONTA = :account"
                params['account'] = account

            sql = text(f"""
                SELECT
                    CAST(BA.DATA AS date) AS DATA,
                    ISNULL(BA.BASTAMP,'') AS BASTAMP,
                    ISNULL(BA.DOCUMENTO,'') AS DOCUMENTO,
                    ISNULL(BA.DESCRICAO,'') AS DESCRICAO,
                    {value_col} AS VALOR
                FROM dbo.V_BA AS BA
                WHERE CAST(BA.DATA AS date) = :d
                  AND {where_kind}
                  {where_account}
                ORDER BY ISNULL(BA.DOCUMENTO,''), ISNULL(BA.DESCRICAO,'')
            """)
            rows = db.session.execute(sql, params).mappings().all()
            out = []
            total = 0.0
            for r in rows:
                val = float(r.get('VALOR') or 0)
                total += val
                data_val = r.get('DATA')
                if isinstance(data_val, (datetime, date)):
                    data_str = data_val.strftime('%Y-%m-%d')
                else:
                    data_str = str(data_val) if data_val is not None else ''
                out.append({
                    'data': data_str,
                    'bastamp': r.get('BASTAMP') or '',
                    'documento': r.get('DOCUMENTO') or '',
                    'descricao': r.get('DESCRICAO') or '',
                    'valor': round(val, 2)
                })

            return jsonify({'date': d, 'kind': kind, 'rows': out, 'total': round(total, 2)})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/generic/api/tesouraria/ba/extrato')
    @login_required
    def api_tesouraria_ba_extrato():
        """
        Extrato de movimentos reais (V_BA) num intervalo.
        Query: ?start=YYYY-MM-DD&end=YYYY-MM-DD
        Retorna linhas normalizadas:
          { DATA, KIND: in|out, DOCUMENTO, DESCRICAO, VALOR }
        """
        try:
            start = (request.args.get('start') or '').strip()
            end = (request.args.get('end') or '').strip()
            account = (request.args.get('account') or '').strip()
            if not start or not end:
                return jsonify({'error': 'Parâmetros start/end obrigatórios'}), 400

            where_account = ""
            params = {'start': start, 'end': end}
            if account:
                where_account = " AND BA.NOCONTA = :account"
                params['account'] = account

            sql = text("""
                SELECT
                    CAST(BA.DATA AS date) AS DATA,
                    'in' AS KIND,
                    ISNULL(BA.DOCUMENTO,'') AS DOCUMENTO,
                    ISNULL(BA.DESCRICAO,'') AS DESCRICAO,
                    ISNULL(BA.EENTRADA,0) AS VALOR
                FROM dbo.V_BA AS BA
                WHERE CAST(BA.DATA AS date) BETWEEN :start AND :end
                  AND ISNULL(BA.EENTRADA,0) <> 0
            """ + where_account + """

                UNION ALL

                SELECT
                    CAST(BA.DATA AS date) AS DATA,
                    'out' AS KIND,
                    ISNULL(BA.DOCUMENTO,'') AS DOCUMENTO,
                    ISNULL(BA.DESCRICAO,'') AS DESCRICAO,
                    ISNULL(BA.ESAIDA,0) AS VALOR
                FROM dbo.V_BA AS BA
                WHERE CAST(BA.DATA AS date) BETWEEN :start AND :end
                  AND ISNULL(BA.ESAIDA,0) <> 0
            """ + where_account + """

                ORDER BY DATA, KIND, DOCUMENTO, DESCRICAO
            """)
            rows = db.session.execute(sql, params).mappings().all()
            out = []
            for r in rows:
                d = r.get('DATA')
                if isinstance(d, (datetime, date)):
                    d = d.strftime('%Y-%m-%d')
                else:
                    d = str(d) if d is not None else ''
                out.append({
                    'DATA': d,
                    'KIND': (r.get('KIND') or '').strip(),
                    'DOCUMENTO': r.get('DOCUMENTO') or '',
                    'DESCRICAO': r.get('DESCRICAO') or '',
                    'VALOR': round(float(r.get('VALOR') or 0), 2)
                })
            return jsonify(out)
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    # -----------------------------
    # Contas Correntes - Fornecedores
    # -----------------------------
    @app.route('/cc_fornecedores')
    @app.route('/contas-correntes-fornecedores')
    @login_required
    def cc_fornecedores_page():
        return render_template('cc_fornecedores.html', page_title='Contas Correntes - Fornecedores')

    # Alias para abrir FO sem o prefixo /generic (usado em links do ecrã de CC)
    @app.route('/fo_compras_form/', defaults={'record_stamp': None})
    @app.route('/fo_compras_form/<record_stamp>')
    @login_required
    def fo_compras_form_alias(record_stamp):
        if record_stamp:
            return redirect(url_for('generic.fo_compras_form', record_stamp=record_stamp))
        return redirect(url_for('generic.fo_compras_form'))

    fo_doc_extract_cache = {}

    def _fo_pick_pdf_anexo(fostamp, anexo_id=None):
        params = {'fostamp': (fostamp or '').strip()}
        if anexo_id:
            row = db.session.execute(text("""
                SELECT TOP 1 ANEXOSSTAMP, RECSTAMP, FICHEIRO, CAMINHO, TIPO, DATA
                FROM dbo.ANEXOS
                WHERE LTRIM(RTRIM(ANEXOSSTAMP)) = :anx
                  AND UPPER(LTRIM(RTRIM(TABELA))) = 'FO'
                  AND LTRIM(RTRIM(RECSTAMP)) = :fostamp
            """), {'anx': anexo_id.strip(), **params}).mappings().first()
            if row:
                return dict(row)

        row = db.session.execute(text("""
            SELECT TOP 1 ANEXOSSTAMP, RECSTAMP, FICHEIRO, CAMINHO, TIPO, DATA
            FROM dbo.ANEXOS
            WHERE UPPER(LTRIM(RTRIM(TABELA))) = 'FO'
              AND LTRIM(RTRIM(RECSTAMP)) = :fostamp
              AND (
                    LOWER(LTRIM(RTRIM(ISNULL(TIPO,'')))) = 'pdf'
                 OR LOWER(ISNULL(FICHEIRO,'')) LIKE '%.pdf'
              )
            ORDER BY DATA DESC, ANEXOSSTAMP DESC
        """), params).mappings().first()
        return dict(row) if row else None

    def _fo_resolve_local_path(caminho: str):
        p = (caminho or '').strip()
        if not p:
            return None
        if p.startswith('http://') or p.startswith('https://'):
            return None
        if p.startswith('/'):
            rel = p.lstrip('/').replace('/', os.sep)
            return os.path.join(app.root_path, rel)
        if os.path.isabs(p):
            return p
        return os.path.join(app.root_path, p.replace('/', os.sep))

    def _fo_decode_qr_from_png_bytes(png_bytes: bytes, use_opencv=True, aggressive=False):
        def _qr_score(raw_val):
            s = (raw_val or '').strip()
            if not s:
                return -999
            su = s.upper()
            score = 0
            # QR fiscal AT costuma conter pares tipo A:, B:, H:... ou ATCUD explícito
            if 'ATCUD' in su:
                score += 50
            if re.search(r'(^|[\*\|;])H:', su):
                score += 40
            if re.search(r'(^|[\*\|;])[A-Z]:', su):
                score += 25
            if '*' in s or ';' in s or '|' in s:
                score += 10
            # URL pura tende a não ser QR fiscal da AT
            if su.startswith('HTTP://') or su.startswith('HTTPS://'):
                score -= 30
            # dimensão mínima útil
            score += min(len(s) // 20, 10)
            return score

        best_raw = None
        best_method = None
        best_conf = None
        best_score = -9999

        # 0) OpenCV first (mais rápido no servidor)
        if use_opencv:
            try:
                import cv2
                import numpy as np
                try:
                    cv2.setLogLevel(0)
                except Exception:
                    try:
                        cv2.utils.logging.setLogLevel(cv2.utils.logging.LOG_LEVEL_ERROR)
                    except Exception:
                        pass
                arr = np.frombuffer(png_bytes, dtype=np.uint8)
                img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
                if img is not None:
                    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                    detector = cv2.QRCodeDetector()
                    variants0 = [img, cv2.cvtColor(gray, cv2.COLOR_GRAY2BGR)]
                    for v in variants0:
                        for k in range(2):
                            vv = np.rot90(v, k).copy() if k else v
                            val, _, _ = detector.detectAndDecode(vv)
                            val = (val or '').strip()
                            if val:
                                sc = _qr_score(val)
                                if sc >= 25:
                                    return val, 'opencv', 0.80, sc
                                if sc > best_score:
                                    best_score = sc
                                    best_raw = val
                                    best_method = 'opencv'
                                    best_conf = 0.80
            except Exception:
                pass

        # 1) pyzbar (principal) com pré-processamento e rotações
        try:
            from PIL import Image, ImageOps, ImageEnhance
            from pyzbar.pyzbar import decode as zdecode, ZBarSymbol
            img0 = Image.open(io.BytesIO(png_bytes)).convert('L')
            variants = [img0, ImageOps.autocontrast(img0)]
            if aggressive:
                variants.append(ImageEnhance.Contrast(img0).enhance(1.8))
                for base in list(variants):
                    variants.append(base.resize((int(base.width * 1.5), int(base.height * 1.5))))
            tried = []
            angles = (0, 90, 180, 270) if aggressive else (0, 180)
            for v in variants:
                for ang in angles:
                    vv = v.rotate(ang, expand=True) if ang else v
                    key = (vv.width, vv.height, ang)
                    if key in tried:
                        continue
                    tried.append(key)
                    dec = zdecode(vv, symbols=[ZBarSymbol.QRCODE])
                    for d in dec or []:
                        raw = (d.data or b'').decode('utf-8', errors='ignore').strip()
                        if raw:
                            sc = _qr_score(raw)
                            if sc > best_score:
                                best_score = sc
                                best_raw = raw
                                best_method = 'pyzbar'
                                best_conf = 0.95
        except Exception:
            pass
        # 2) OpenCV fallback com grayscale/threshold e rotações
        try:
            if not use_opencv:
                return best_raw, best_method, best_conf, best_score
            import cv2
            import numpy as np
            try:
                cv2.setLogLevel(0)
            except Exception:
                try:
                    cv2.utils.logging.setLogLevel(cv2.utils.logging.LOG_LEVEL_ERROR)
                except Exception:
                    pass
            arr = np.frombuffer(png_bytes, dtype=np.uint8)
            img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
            if img is not None:
                gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                _, th = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                variants = [img, cv2.cvtColor(gray, cv2.COLOR_GRAY2BGR)]
                if aggressive:
                    variants.append(cv2.cvtColor(th, cv2.COLOR_GRAY2BGR))
                detector = cv2.QRCodeDetector()
                for v in variants:
                    max_rot = 4 if aggressive else 2
                    for k in range(max_rot):
                        vv = np.rot90(v, k).copy() if k else v
                        val, _, _ = detector.detectAndDecode(vv)
                        val = (val or '').strip()
                        if val:
                            sc = _qr_score(val)
                            if sc > best_score:
                                best_score = sc
                                best_raw = val
                                best_method = 'opencv'
                                best_conf = 0.75
        except Exception:
            pass
        return best_raw, best_method, best_conf, best_score

    def _fo_parse_found_fields(text_all, qr_raw=None):
        found = {}
        src = '\n'.join([qr_raw or '', text_all or ''])

        def _parse_pt_num(s):
            try:
                raw = (s or '').strip().replace(' ', '')
                if not raw:
                    return None
                # Suporta 137.80, 137,80, 1.234,56 e 1,234.56
                if ',' in raw and '.' in raw:
                    if raw.rfind(',') > raw.rfind('.'):
                        # decimal = vírgula
                        raw = raw.replace('.', '').replace(',', '.')
                    else:
                        # decimal = ponto
                        raw = raw.replace(',', '')
                elif ',' in raw:
                    raw = raw.replace('.', '').replace(',', '.')
                else:
                    raw = raw.replace(',', '')
                return float(raw)
            except Exception:
                return None

        qr_map = {}
        if qr_raw:
            # Formato AT típico: A:...*B:...*C:... (ou ; como separador)
            parts = re.split(r'[\*;]', qr_raw)
            for p in parts:
                if ':' not in p:
                    continue
                k, v = p.split(':', 1)
                k = (k or '').strip().upper()
                v = (v or '').strip()
                if k:
                    qr_map[k] = v

        if qr_raw:
            found['qr_raw'] = qr_raw

        m_atcud = re.search(r'\bATCUD\s*[:=]?\s*([A-Z0-9\-/\.]+)', src, re.IGNORECASE)
        if not m_atcud and qr_raw:
            m_atcud = re.search(r'(?:^|[\*\|;])H:([^*\|;\r\n]+)', qr_raw, re.IGNORECASE)
        if m_atcud:
            found['atcud'] = (m_atcud.group(1) or '').strip()
        elif qr_map.get('H'):
            found['atcud'] = qr_map.get('H')

        m_doc = re.search(r'(?:N[ºO]\s*DOC(?:UMENTO)?|DOC(?:UMENTO)?|FATURA|FACTURA|RECIBO)\s*[:#]?\s*([A-Z0-9][A-Z0-9/\-\. ]{2,50})', text_all or '', re.IGNORECASE)
        if m_doc:
            doc_candidate = (m_doc.group(1) or '').strip()
            # Evitar capturar datas por engano (ex.: "11 fev 2026")
            if not re.search(r'\b(?:JAN|FEV|MAR|ABR|MAI|JUN|JUL|AGO|SET|OUT|NOV|DEZ)\b', doc_candidate.upper()):
                found['doc_no'] = doc_candidate

        # No QR AT: A = NIF emitente, B = NIF adquirente/receptor
        if qr_map.get('A'):
            found['nif_emitente'] = qr_map.get('A')
        if qr_map.get('B'):
            found['nif_receptor'] = qr_map.get('B')
        if not found.get('nif_emitente'):
            m_nif = re.search(r'NIF(?:\s+EMITENTE)?\s*[:#]?\s*(\d{9})', src, re.IGNORECASE)
            if not m_nif:
                m_nif = re.search(r'\b(\d{9})\b', src or '')
            if m_nif:
                found['nif_emitente'] = (m_nif.group(1) or '').strip()

        m_date = re.search(r'\b(\d{2}[\/\-.]\d{2}[\/\-.]\d{4})\b', src or '')
        if m_date:
            d = (m_date.group(1) or '').replace('.', '-').replace('/', '-')
            parts = d.split('-')
            if len(parts) == 3:
                found['data'] = f"{parts[2]}-{parts[1]}-{parts[0]}"
        if qr_map.get('F') and re.match(r'^\d{8}$', qr_map.get('F') or ''):
            f = qr_map.get('F')
            found['data'] = f"{f[0:4]}-{f[4:6]}-{f[6:8]}"

        m_total = re.search(r'(?:TOTAL(?:\s+A\s+PAGAR)?|VALOR\s+TOTAL)\s*[:=]?\s*([0-9\.\s]+,[0-9]{2})', text_all or '', re.IGNORECASE)
        if m_total:
            n = _parse_pt_num(m_total.group(1))
            if n is not None:
                found['total'] = round(n, 2)
        if qr_map.get('O'):
            n = _parse_pt_num(qr_map.get('O'))
            if n is not None:
                found['total'] = round(n, 2)

        m_iva = re.search(r'\bIVA\b\s*[:=]?\s*([0-9\.\s]+,[0-9]{2})', text_all or '', re.IGNORECASE)
        if m_iva:
            n = _parse_pt_num(m_iva.group(1))
            if n is not None:
                found['iva'] = round(n, 2)
        if qr_map.get('N'):
            n = _parse_pt_num(qr_map.get('N'))
            if n is not None:
                found['iva'] = round(n, 2)

        if qr_map.get('G'):
            found['doc_no'] = qr_map.get('G')

        # CPE (Código de Ponto de Entrega) - ex.: PT 0002 0000 3165 4577 TW
        cpe_match = re.search(
            r'\b(PT\s*\d{4}\s*\d{4}\s*\d{4}\s*\d{4}\s*[A-Z]{2})\b',
            src or '',
            re.IGNORECASE
        )
        if cpe_match:
            cpe_raw = (cpe_match.group(1) or '').upper()
            cpe_norm = re.sub(r'\s+', ' ', cpe_raw).strip()
            found['cpe'] = cpe_norm

        # Água - "Local Consumo: 79699"
        m_local_consumo = re.search(r'\bLOCAL\s+CONSUMO\s*[:\-]?\s*([0-9]{3,20})\b', src or '', re.IGNORECASE)
        if m_local_consumo:
            found['local_consumo'] = (m_local_consumo.group(1) or '').strip()

        # Desdobramento AT (principais variáveis)
        if qr_map:
            known = {
                'A': 'at_a_nif_emitente',
                'B': 'at_b_nif_receptor',
                'C': 'at_c_pais_receptor',
                'D': 'at_d_tipo_documento',
                'E': 'at_e_estado_documento',
                'F': 'at_f_data_documento',
                'G': 'at_g_numero_documento',
                'H': 'at_h_atcud',
                'N': 'at_n_iva_total',
                'O': 'at_o_total_documento',
                'Q': 'at_q_hash',
                'R': 'at_r_software_certificado'
            }
            for k, out_key in known.items():
                if qr_map.get(k):
                    found[out_key] = qr_map.get(k)
            for k in sorted(qr_map.keys()):
                if re.match(r'^I\d+$', k):
                    found[f'at_{k.lower()}'] = qr_map.get(k)

        return found

    def _fo_analyze_pdf_document(local_pdf_path: str, fast_mode=True, allow_opencv_fallback=False):
        # Resultado base
        t0 = time.time()
        out = {
            'status': 'not_found',
            'found': {},
            'message': 'Não foi possível encontrar QR/ATCUD no documento.',
            'debug': {
                'pages_total': 0,
                'render_attempts': 0,
                'text_chars': 0,
                'file_exists': False,
                'file_size': 0,
                'deps': {},
                'fast_mode': bool(fast_mode),
                'allow_opencv_fallback': bool(allow_opencv_fallback),
                'elapsed_ms': 0,
                'pages_scanned': 0
            }
        }
        out['debug']['file_exists'] = bool(local_pdf_path and os.path.isfile(local_pdf_path))
        if local_pdf_path and os.path.isfile(local_pdf_path):
            try:
                out['debug']['file_size'] = int(os.path.getsize(local_pdf_path) or 0)
            except Exception:
                out['debug']['file_size'] = 0
        if not local_pdf_path or not os.path.isfile(local_pdf_path):
            out['status'] = 'error'
            out['message'] = 'Ficheiro não encontrado no servidor.'
            return out

        deps = {
            'fitz': bool(importlib.util.find_spec('fitz')),
            'pypdf': bool(importlib.util.find_spec('pypdf')),
            'PIL': bool(importlib.util.find_spec('PIL')),
            'pyzbar': bool(importlib.util.find_spec('pyzbar')),
            'cv2': bool(importlib.util.find_spec('cv2')),
        }
        out['debug']['deps'] = deps
        if not deps.get('fitz') and not deps.get('pypdf'):
            out['status'] = 'error'
            out['message'] = 'Dependências em falta para ler PDF (fitz/pypdf).'
            return out

        text_pages = []
        first_text = ''
        first_text_page = None
        qr_raw = None
        qr_page = None
        qr_method = None
        qr_conf = None
        qr_score = -9999
        render_attempts = 0

        # PyMuPDF path (preferencial, por render+texto por página)
        try:
            import fitz
            doc = fitz.open(local_pdf_path)
            total_pages = int(doc.page_count or 0)
            out['debug']['pages_total'] = total_pages
            max_pages_scan = min(total_pages, 4 if fast_mode else total_pages)
            page_indexes = list(range(max_pages_scan))
            out['debug']['pages_scanned'] = len(page_indexes)

            text_pages = []
            for idx_txt in page_indexes:
                t = (doc[idx_txt].get_text('text') or '')
                text_pages.append(t)
                s = ' '.join((t or '').split())
                if s:
                    first_text = s[:600]
                    first_text_page = idx_txt + 1
                    break
            fast_passes = [
                {'dpi': 170, 'crop_mode': 'br'},
                {'dpi': 220, 'crop_mode': 'br'},
                {'dpi': 240, 'crop_mode': 'full'},
            ]
            fallback_passes = [
                {'dpi': 320, 'crop_mode': 'full'},
                {'dpi': 300, 'crop_mode': 'bl'},
                {'dpi': 300, 'crop_mode': 'tr'},
            ]
            max_attempts = 18 if fast_mode else 42
            for cfg in fast_passes:
                for i in page_indexes:
                    if render_attempts >= max_attempts:
                        break
                    page = doc[i]
                    rect = page.rect
                    clip = None
                    mode = (cfg.get('crop_mode') or 'full')
                    if mode == 'br':
                        clip = fitz.Rect(rect.width * 0.45, rect.height * 0.45, rect.width, rect.height)
                    elif mode == 'bl':
                        clip = fitz.Rect(0, rect.height * 0.45, rect.width * 0.55, rect.height)
                    elif mode == 'tr':
                        clip = fitz.Rect(rect.width * 0.45, 0, rect.width, rect.height * 0.55)
                    render_attempts += 1
                    pix = page.get_pixmap(dpi=cfg['dpi'], clip=clip, alpha=False)
                    raw, method, conf, sc = _fo_decode_qr_from_png_bytes(
                        pix.tobytes('png'),
                        use_opencv=True,
                        aggressive=False
                    )
                    if raw and sc > qr_score:
                        qr_raw = raw
                        qr_page = i + 1
                        qr_method = method
                        qr_conf = conf
                        qr_score = sc
                        if qr_score >= 70:
                            break
                if render_attempts >= max_attempts:
                    break
                if qr_score >= 70:
                    break

            # Recovery pass: em fast_mode, se não encontrou QR nas primeiras páginas,
            # varre as restantes com poucos passes pyzbar para não perder QR em páginas mais à frente.
            if fast_mode and (not qr_raw) and total_pages > len(page_indexes):
                recovery_indexes = list(range(len(page_indexes), total_pages))
                recovery_passes = [
                    {'dpi': 170, 'crop_mode': 'br'},
                    {'dpi': 220, 'crop_mode': 'full'},
                ]
                recovery_attempts_cap = max_attempts + 24
                for cfg in recovery_passes:
                    for i in recovery_indexes:
                        if render_attempts >= recovery_attempts_cap:
                            break
                        page = doc[i]
                        rect = page.rect
                        clip = None
                        mode = (cfg.get('crop_mode') or 'full')
                        if mode == 'br':
                            clip = fitz.Rect(rect.width * 0.45, rect.height * 0.45, rect.width, rect.height)
                        elif mode == 'bl':
                            clip = fitz.Rect(0, rect.height * 0.45, rect.width * 0.55, rect.height)
                        elif mode == 'tr':
                            clip = fitz.Rect(rect.width * 0.45, 0, rect.width, rect.height * 0.55)
                        render_attempts += 1
                        pix = page.get_pixmap(dpi=cfg['dpi'], clip=clip, alpha=False)
                        raw, method, conf, sc = _fo_decode_qr_from_png_bytes(
                            pix.tobytes('png'),
                            use_opencv=True,
                            aggressive=False
                        )
                        if raw and sc > qr_score:
                            qr_raw = raw
                            qr_page = i + 1
                            qr_method = method
                            qr_conf = conf
                            qr_score = sc
                            if qr_score >= 70:
                                break
                    if render_attempts >= recovery_attempts_cap:
                        break
                    if qr_score >= 70:
                        break
            # fallback com OpenCV apenas se pyzbar não encontrou nada útil
            if allow_opencv_fallback and (not qr_raw or qr_score < 60):
                for cfg in fallback_passes:
                    for i in page_indexes:
                        if render_attempts >= max_attempts:
                            break
                        page = doc[i]
                        rect = page.rect
                        clip = None
                        mode = (cfg.get('crop_mode') or 'full')
                        if mode == 'br':
                            clip = fitz.Rect(rect.width * 0.45, rect.height * 0.45, rect.width, rect.height)
                        elif mode == 'bl':
                            clip = fitz.Rect(0, rect.height * 0.45, rect.width * 0.55, rect.height)
                        elif mode == 'tr':
                            clip = fitz.Rect(rect.width * 0.45, 0, rect.width, rect.height * 0.55)
                        render_attempts += 1
                        pix = page.get_pixmap(dpi=cfg['dpi'], clip=clip, alpha=False)
                        raw, method, conf, sc = _fo_decode_qr_from_png_bytes(
                            pix.tobytes('png'),
                            use_opencv=True,
                            aggressive=True
                        )
                        if raw and sc > qr_score:
                            qr_raw = raw
                            qr_page = i + 1
                            qr_method = method
                            qr_conf = conf
                            qr_score = sc
                            if qr_score >= 70:
                                break
                    if render_attempts >= max_attempts:
                        break
                    if qr_score >= 70:
                        break
            doc.close()
        except Exception:
            text_pages = []

        # fallback texto sem fitz
        if not text_pages:
            try:
                from pypdf import PdfReader
                reader = PdfReader(local_pdf_path)
                total = len(reader.pages)
                max_pages_scan = min(total, 4 if fast_mode else total)
                out['debug']['pages_total'] = max(int(out['debug'].get('pages_total') or 0), total)
                out['debug']['pages_scanned'] = max(int(out['debug'].get('pages_scanned') or 0), max_pages_scan)
                text_pages = [((reader.pages[i].extract_text() or '')) for i in range(max_pages_scan)]
                out['debug']['pages_total'] = max(int(out['debug'].get('pages_total') or 0), len(text_pages))
                for idx_txt, t in enumerate(text_pages):
                    s = ' '.join((t or '').split())
                    if s:
                        first_text = s[:600]
                        first_text_page = idx_txt + 1
                        break
            except Exception:
                text_pages = []

        text_all = '\n'.join(text_pages or [])
        out['debug']['render_attempts'] = int(render_attempts)
        out['debug']['text_chars'] = int(len(text_all or ''))
        out['debug']['attempts_limited'] = bool(render_attempts >= 80)

        # Fallback textual: alguns PDFs trazem o payload AT como texto extraível.
        if not qr_raw and text_all:
            txt_norm = re.sub(r'[\r\n\t]+', ' ', text_all)
            txt_norm = re.sub(r'\s+', ' ', txt_norm).strip()
            m_qr_text = re.search(
                r'(A:\d{9}\*B:\d{9}\*C:[A-Z]{2}\*D:[A-Z]{1,3}\*E:[A-Z0-9]+\*F:\d{8}\*G:[^*]{3,120}\*H:[^*]{3,80}(?:\*[A-Z]\d?:[^*]{0,60})*)',
                txt_norm,
                re.IGNORECASE
            )
            if m_qr_text:
                qr_raw = (m_qr_text.group(1) or '').strip()
                qr_method = 'regex'
                qr_conf = 0.55
                for pi, pt in enumerate(text_pages):
                    pnorm = re.sub(r'[\r\n\t]+', ' ', (pt or ''))
                    if qr_raw[:24] in pnorm:
                        qr_page = pi + 1
                        break

        found = _fo_parse_found_fields(text_all=text_all, qr_raw=qr_raw)
        if qr_page:
            found['page'] = qr_page
        if qr_method:
            found['method'] = qr_method
        if qr_conf is not None:
            found['confidence'] = round(float(qr_conf), 2)
        if qr_score > -9999:
            found['qr_score'] = int(qr_score)
        if first_text:
            found['first_text'] = first_text
            if first_text_page:
                found['first_text_page'] = first_text_page

        fiscal_keys = ('qr_raw', 'atcud', 'at_h_atcud', 'doc_no', 'nif_emitente', 'nif_receptor', 'total', 'iva', 'at_g_numero_documento')
        has_fiscal = any(found.get(k) not in (None, '') for k in fiscal_keys)

        if has_fiscal:
            out['status'] = 'ok'
            out['found'] = found
            out['message'] = 'Documento analisado com sucesso.'
        else:
            out['status'] = 'not_found'
            out['found'] = found
            # mensagem de fallback mais explícita quando falta stack de QR
            if (not deps.get('pyzbar')) and (not deps.get('cv2')):
                out['message'] = 'PDF lido, mas sem motor de QR instalado (pyzbar/cv2).'
            else:
                out['message'] = 'Documento lido, mas sem QR/ATCUD fiscal identificado.'
        out['debug']['elapsed_ms'] = int((time.time() - t0) * 1000)
        return out

    @app.route('/api/fo_compras/<fostamp>/analisar_documento', methods=['GET'])
    @login_required
    def api_fo_compras_analisar_documento_get(fostamp):
        key = (fostamp or '').strip()
        if not key:
            return jsonify({'status': 'error', 'message': 'FOSTAMP inválido.'}), 400
        cached = fo_doc_extract_cache.get(key)
        if not cached:
            return jsonify({'status': 'not_found', 'message': 'Sem resultado em cache.'})
        return jsonify(cached)

    @app.route('/api/fo_compras/<fostamp>/analisar_documento', methods=['POST'])
    @login_required
    def api_fo_compras_analisar_documento(fostamp):
        try:
            key = (fostamp or '').strip()
            if not key:
                return jsonify({'status': 'error', 'message': 'FOSTAMP inválido.'}), 400
            payload = request.get_json(silent=True) or {}
            anexo_id = (payload.get('anexo_id') or '').strip() or None

            anx = _fo_pick_pdf_anexo(key, anexo_id=anexo_id)
            if not anx:
                result = {
                    'status': 'not_found',
                    'file': None,
                    'found': {},
                    'message': 'Não existe anexo PDF associado a esta compra.'
                }
                fo_doc_extract_cache[key] = result
                return jsonify(result)

            local_path = _fo_resolve_local_path(anx.get('CAMINHO'))
            try:
                file_mtime = os.path.getmtime(local_path) if local_path and os.path.isfile(local_path) else None
                file_size = os.path.getsize(local_path) if local_path and os.path.isfile(local_path) else None
            except Exception:
                file_mtime = None
                file_size = None
            cache_sig = {
                'anexo_id': (anx.get('ANEXOSSTAMP') or '').strip(),
                'path': local_path or '',
                'mtime': file_mtime,
                'size': file_size
            }
            cached = fo_doc_extract_cache.get(key)
            if cached and isinstance(cached, dict) and (cached.get('_sig') == cache_sig):
                return jsonify({k: v for k, v in cached.items() if k != '_sig'})
            file_obj = {
                'name': (anx.get('FICHEIRO') or '').strip(),
                'path': local_path or (anx.get('CAMINHO') or '').strip()
            }
            fast_mode = True
            allow_opencv_fallback = False
            try:
                fast_mode = bool((payload.get('fast_mode', True)))
            except Exception:
                fast_mode = True
            try:
                allow_opencv_fallback = bool((payload.get('allow_opencv_fallback', False)))
            except Exception:
                allow_opencv_fallback = False
            # defaults por env/config para servidor
            env_fast = os.environ.get('FO_QR_FAST_DEFAULT')
            if env_fast is not None:
                fast_mode = str(env_fast).strip() not in ('0', 'false', 'False', '')
            env_cv = os.environ.get('FO_QR_ENABLE_OPENCV_FALLBACK')
            if env_cv is not None:
                allow_opencv_fallback = str(env_cv).strip() in ('1', 'true', 'True')
            # Em ambientes não-Windows (tipicamente servidor Linux), forçar fallback OpenCV
            # para compensar diferenças do zbar/pyzbar entre hosts.
            if not allow_opencv_fallback:
                force_server_cv = os.environ.get('FO_QR_SERVER_FORCE_OPENCV', '1')
                if os.name != 'nt' and str(force_server_cv).strip() in ('1', 'true', 'True'):
                    allow_opencv_fallback = True

            ana = _fo_analyze_pdf_document(
                local_path,
                fast_mode=fast_mode,
                allow_opencv_fallback=allow_opencv_fallback
            )
            result = {
                'status': ana.get('status', 'error'),
                'file': file_obj,
                'found': ana.get('found') or {},
                'message': ana.get('message') or '',
                'debug': ana.get('debug') or {}
            }
            fo_doc_extract_cache[key] = {**result, '_sig': cache_sig}
            return jsonify(result)
        except Exception as e:
            return jsonify({
                'status': 'error',
                'file': None,
                'found': {},
                'message': str(e)
            }), 500

    @app.route('/api/fo_compras/fornecedor_por_nif', methods=['GET'])
    @login_required
    def api_fo_compras_fornecedor_por_nif():
        try:
            nif = (request.args.get('nif') or '').strip()
            nif_digits = re.sub(r'\D+', '', nif or '')
            if len(nif_digits) < 9:
                return jsonify({'ok': False, 'error': 'NIF inválido.'}), 400
            row = db.session.execute(text("""
                SELECT TOP 1
                    NO,
                    NOME,
                    NCONT,
                    ISNULL(MORADA,'') AS MORADA,
                    ISNULL(LOCAL,'') AS LOCAL,
                    ISNULL(CODPOST,'') AS CODPOST
                FROM dbo.V_FL
                WHERE REPLACE(REPLACE(REPLACE(LTRIM(RTRIM(ISNULL(NCONT,''))), ' ', ''), '-', ''), '.', '') = :nif
                ORDER BY NOME
            """), {'nif': nif_digits}).mappings().first()
            if not row:
                return jsonify({'ok': False, 'found': None, 'message': 'Fornecedor não encontrado para o NIF indicado.'})
            return jsonify({'ok': True, 'found': dict(row)})
        except Exception as e:
            return jsonify({'ok': False, 'error': str(e)}), 500

    @app.route('/api/fo_compras/refs_por_fornecedor', methods=['GET'])
    @login_required
    def api_fo_compras_refs_por_fornecedor():
        try:
            no_raw = (request.args.get('no') or '').strip()
            try:
                no_val = int(float(no_raw))
            except Exception:
                return jsonify({'ok': False, 'error': 'NO inválido.'}), 400

            rows = db.session.execute(text("""
                SELECT
                    CAST(NO AS int) AS NO,
                    LTRIM(RTRIM(ISNULL(REF,''))) AS REF,
                    LTRIM(RTRIM(CAST(ISNULL(TABIVA,'') AS varchar(20)))) AS TABIVA
                FROM dbo.V_REFS
                WHERE CAST(NO AS int) = :no
            """), {'no': no_val}).mappings().all()

            out = []
            for r in rows:
                ref = (r.get('REF') or '').strip()
                tab = (r.get('TABIVA') or '').strip()
                if not ref or not tab:
                    continue
                out.append({
                    'NO': int(r.get('NO') or no_val),
                    'REF': ref,
                    'TABIVA': tab
                })
            return jsonify({'ok': True, 'items': out})
        except Exception as e:
            return jsonify({'ok': False, 'error': str(e)}), 500

    @app.route('/api/fo_compras/ccusto_por_consumo', methods=['GET'])
    @login_required
    def api_fo_compras_ccusto_por_consumo():
        try:
            cpe = (request.args.get('cpe') or '').strip()
            lconsumo = (request.args.get('lconsumo') or '').strip()
            cpe_norm = re.sub(r'[\s\-]+', '', cpe.upper())
            lconsumo_norm = re.sub(r'\s+', '', lconsumo)
            if not cpe_norm and not lconsumo_norm:
                return jsonify({'ok': False, 'error': 'Parâmetros em falta (cpe/lconsumo).'}), 400

            row = db.session.execute(text("""
                SELECT TOP 1
                    LTRIM(RTRIM(ISNULL(CCUSTO,''))) AS CCUSTO,
                    LTRIM(RTRIM(ISNULL(NOME,''))) AS NOME,
                    LTRIM(RTRIM(ISNULL(CPE,''))) AS CPE,
                    LTRIM(RTRIM(ISNULL(LCONSUMO,''))) AS LCONSUMO
                FROM dbo.AL
                WHERE (
                        :cpe_norm <> ''
                        AND REPLACE(REPLACE(UPPER(LTRIM(RTRIM(ISNULL(CPE,'')))), ' ', ''), '-', '') = :cpe_norm
                      )
                   OR (
                        :lconsumo_norm <> ''
                        AND REPLACE(LTRIM(RTRIM(ISNULL(LCONSUMO,''))), ' ', '') = :lconsumo_norm
                      )
                ORDER BY
                    CASE
                        WHEN :cpe_norm <> '' AND REPLACE(REPLACE(UPPER(LTRIM(RTRIM(ISNULL(CPE,'')))), ' ', ''), '-', '') = :cpe_norm
                             AND :lconsumo_norm <> '' AND REPLACE(LTRIM(RTRIM(ISNULL(LCONSUMO,''))), ' ', '') = :lconsumo_norm
                            THEN 0
                        WHEN :cpe_norm <> '' AND REPLACE(REPLACE(UPPER(LTRIM(RTRIM(ISNULL(CPE,'')))), ' ', ''), '-', '') = :cpe_norm
                            THEN 1
                        WHEN :lconsumo_norm <> '' AND REPLACE(LTRIM(RTRIM(ISNULL(LCONSUMO,''))), ' ', '') = :lconsumo_norm
                            THEN 2
                        ELSE 9
                    END,
                    NOME
            """), {
                'cpe_norm': cpe_norm,
                'lconsumo_norm': lconsumo_norm
            }).mappings().first()

            if not row:
                return jsonify({'ok': False, 'found': None, 'message': 'Sem alojamento para CPE/Local Consumo.'})
            return jsonify({'ok': True, 'found': dict(row)})
        except Exception as e:
            return jsonify({'ok': False, 'error': str(e)}), 500

    @app.route('/api/cc_fornecedores/resumo')
    @login_required
    def api_cc_fornecedores_resumo():
        """
        Resumo por fornecedor (saldo em aberto).
        Query:
          - q (opcional): pesquisa por NO/NOME
          - pendentes=1: apenas fornecedores com saldo em aberto != 0
        """
        try:
            q = (request.args.get('q') or '').strip()
            pendentes = (request.args.get('pendentes') or '').strip() in ('1', 'true', 'True')

            # Fonte: dbo.V_FC (BD GESTAO).
            # Fornecedores: a dÃ­vida em aberto deve ficar positiva (normalmente vem do lado do crÃ©dito).
            #   (ECRED - ECREDF) - (EDEB - EDEBF)
            open_expr = "(ISNULL(ECRED,0) - ISNULL(ECREDF,0)) - (ISNULL(EDEB,0) - ISNULL(EDEBF,0))"

            where = []
            params = {}
            if q:
                where.append("(CAST(NO AS varchar(50)) LIKE :q OR NOME LIKE :q)")
                params['q'] = f"%{q}%"

            where_sql = ("WHERE " + " AND ".join(where)) if where else ""
            having_sql = f"HAVING ABS(SUM({open_expr})) > 0.005" if pendentes else ""

            sql = text(f"""
                SELECT
                    NO,
                    MAX(NOME) AS NOME,
                    SUM({open_expr}) AS SALDO_ABERTO
                FROM dbo.V_FC
                {where_sql}
                GROUP BY NO
                {having_sql}
                ORDER BY SUM({open_expr}) DESC, MAX(NOME)
            """)
            rows = db.session.execute(sql, params).mappings().all()

            items = []
            total_divida = 0.0
            total_saldo = 0.0
            for r in rows:
                no = r.get('NO')
                nome = str(r.get('NOME') or '').strip()
                saldo = float(r.get('SALDO_ABERTO') or 0)
                total_saldo += saldo
                divida = saldo if saldo > 0 else 0.0
                total_divida += divida
                items.append({
                    'NO': int(no) if str(no).isdigit() else no,
                    'NOME': nome,
                    'SALDO_ABERTO': round(saldo, 2),
                    'DIVIDA': round(divida, 2)
                })

            return jsonify({
                'total_divida': round(total_divida, 2),
                'total_saldo': round(total_saldo, 2),
                'count_fornecedores': len(items),
                'items': items
            })
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/cc_fornecedores/detalhe')
    @login_required
    def api_cc_fornecedores_detalhe():
        """
        Movimentos de conta corrente por fornecedor.
        Query:
          - no (obrigatório): fornecedor
          - pendentes=1: apenas linhas com aberto != 0
        """
        try:
            no = request.args.get('no', type=int)
            if not no:
                return jsonify({'error': 'Parâmetro no obrigatório'}), 400
            pendentes = (request.args.get('pendentes') or '').strip() in ('1', 'true', 'True')

            open_expr = "(ISNULL(ECRED,0) - ISNULL(ECREDF,0)) - (ISNULL(EDEB,0) - ISNULL(EDEBF,0))"

            where_pend = f"AND ABS({open_expr}) > 0.005" if pendentes else ""
            sql = text(f"""
                SELECT
                    FCSTAMP,
                    CAST(DATALC AS date) AS DATALC,
                    CAST(DATAVEN AS date) AS DATAVEN,
                    ISNULL(CMDESC,'') AS CMDESC,
                    ISNULL(ADOC,'') AS ADOC,
                    ISNULL(EDEB,0) AS EDEB,
                    ISNULL(ECRED,0) AS ECRED,
                    ISNULL(EDEBF,0) AS EDEBF,
                    ISNULL(ECREDF,0) AS ECREDF,
                    ISNULL(MOEDA,'') AS MOEDA,
                    ISNULL(CCUSTO,'') AS CCUSTO,
                    ISNULL(FOSTAMP,'') AS FOSTAMP,
                    {open_expr} AS ABERTO
                FROM dbo.V_FC
                WHERE NO = :no
                {where_pend}
                ORDER BY CAST(DATALC AS date), FCSTAMP
            """)
            rows = db.session.execute(sql, {'no': no}).mappings().all()

            out = []
            total_aberto = 0.0
            for r in rows:
                aberto = float(r.get('ABERTO') or 0)
                total_aberto += aberto
                dlc = r.get('DATALC')
                dven = r.get('DATAVEN')
                out.append({
                    'FCSTAMP': r.get('FCSTAMP') or '',
                    'DATALC': dlc.strftime('%Y-%m-%d') if isinstance(dlc, (datetime, date)) else (str(dlc) if dlc is not None else ''),
                    'DATAVEN': dven.strftime('%Y-%m-%d') if isinstance(dven, (datetime, date)) else (str(dven) if dven is not None else ''),
                    'CMDESC': r.get('CMDESC') or '',
                    'ADOC': r.get('ADOC') or '',
                    'EDEB': round(float(r.get('EDEB') or 0), 2),
                    'ECRED': round(float(r.get('ECRED') or 0), 2),
                    'EDEBF': round(float(r.get('EDEBF') or 0), 2),
                    'ECREDF': round(float(r.get('ECREDF') or 0), 2),
                    'ABERTO': round(aberto, 2),
                    'MOEDA': r.get('MOEDA') or '',
                    'CCUSTO': r.get('CCUSTO') or '',
                    'FOSTAMP': r.get('FOSTAMP') or ''
                })

            return jsonify({'no': no, 'rows': out, 'total_aberto': round(total_aberto, 2)})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/cc_fornecedores/pendentes')
    @login_required
    def api_cc_fornecedores_pendentes():
        """
        Lista movimentos pendentes (em aberto) para assistente de pagamentos.
        Query:
          - q (opcional): pesquisa por NO/NOME/ADOC/CMDESC

        Nota: inclui também ABERTO < 0 (créditos) para permitir compensação no pagamento.
        """
        try:
            q = (request.args.get('q') or '').strip()

            open_expr = "(ISNULL(ECRED,0) - ISNULL(ECREDF,0)) - (ISNULL(EDEB,0) - ISNULL(EDEBF,0))"
            where = [f"ABS(({open_expr})) > 0.005"]
            params = {}
            if q:
                where.append("(CAST(NO AS varchar(50)) LIKE :q OR NOME LIKE :q OR ADOC LIKE :q OR CMDESC LIKE :q)")
                params['q'] = f"%{q}%"

            where_sql = "WHERE " + " AND ".join(where)

            def run_query(include_cm: bool):
                cm_sel = "ISNULL(CM,0) AS CM," if include_cm else "CAST(0 AS int) AS CM,"
                sql = text(f"""
                    SELECT
                        FCSTAMP,
                        CAST(DATALC AS date) AS DATALC,
                        CAST(DATAVEN AS date) AS DATAVEN,
                        ISNULL(CMDESC,'') AS CMDESC,
                        ISNULL(ADOC,'') AS ADOC,
                        {cm_sel}
                        ISNULL(NO,0) AS NO,
                        ISNULL(NOME,'') AS NOME,
                        ISNULL(MOEDA,'') AS MOEDA,
                        ISNULL(CCUSTO,'') AS CCUSTO,
                        ISNULL(FOSTAMP,'') AS FOSTAMP,
                        {open_expr} AS ABERTO
                    FROM dbo.V_FC
                    {where_sql}
                    ORDER BY ISNULL(NOME,''), CAST(DATAVEN AS date), CAST(DATALC AS date), FCSTAMP
                """)
                return db.session.execute(sql, params).mappings().all()

            try:
                rows = run_query(include_cm=True)
            except Exception as e:
                # compat: a view V_FC pode nÃ£o ter a coluna CM ainda
                msg = str(e)
                if "Invalid column name 'CM'" in msg or "Invalid column name \"CM\"" in msg:
                    rows = run_query(include_cm=False)
                else:
                    raise

            out = []
            total = 0.0
            for r in rows:
                dlc = r.get('DATALC')
                dven = r.get('DATAVEN')
                aberto = float(r.get('ABERTO') or 0)
                total += aberto
                out.append({
                    'FCSTAMP': (r.get('FCSTAMP') or '').strip(),
                    'DATALC': dlc.strftime('%Y-%m-%d') if isinstance(dlc, (datetime, date)) else (str(dlc) if dlc is not None else ''),
                    'DATAVEN': dven.strftime('%Y-%m-%d') if isinstance(dven, (datetime, date)) else (str(dven) if dven is not None else ''),
                    'CMDESC': r.get('CMDESC') or '',
                    'ADOC': r.get('ADOC') or '',
                    'CM': int(r.get('CM') or 0),
                    'NO': int(r.get('NO') or 0),
                    'NOME': r.get('NOME') or '',
                    'MOEDA': r.get('MOEDA') or '',
                    'CCUSTO': r.get('CCUSTO') or '',
                    'FOSTAMP': r.get('FOSTAMP') or '',
                    'ABERTO': round(aberto, 2)
                })

            return jsonify({'rows': out, 'total': round(total, 2), 'count': len(out)})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    # -----------------------------
    # Contas Correntes - Clientes
    # -----------------------------
    @app.route('/cc_clientes')
    @app.route('/contas-correntes-clientes')
    @login_required
    def cc_clientes_page():
        return render_template('cc_clientes.html', page_title='Contas Correntes - Clientes')

    @app.route('/api/cc_clientes/resumo')
    @login_required
    def api_cc_clientes_resumo():
        """
        Resumo por cliente (saldo em aberto).
        Query:
          - q (opcional): pesquisa por NO/NOME
          - pendentes=1: apenas clientes com saldo em aberto != 0
        """
        try:
            q = (request.args.get('q') or '').strip()
            pendentes = (request.args.get('pendentes') or '').strip() in ('1', 'true', 'True')

            # Fonte: dbo.V_CC (BD GESTAO).
            # Clientes: o valor em aberto deve ficar positivo quando o cliente nos deve (normalmente vem do lado do débito).
            #   (EDEB - EDEBF) - (ECRED - ECREDF)
            open_expr = "(ISNULL(EDEB,0) - ISNULL(EDEBF,0)) - (ISNULL(ECRED,0) - ISNULL(ECREDF,0))"

            where = []
            params = {}
            if q:
                where.append("(CAST(NO AS varchar(50)) LIKE :q OR NOME LIKE :q)")
                params['q'] = f"%{q}%"

            where_sql = ("WHERE " + " AND ".join(where)) if where else ""
            having_sql = f"HAVING ABS(SUM({open_expr})) > 0.005" if pendentes else ""

            sql = text(f"""
                SELECT
                    NO,
                    MAX(NOME) AS NOME,
                    SUM({open_expr}) AS SALDO_ABERTO
                FROM dbo.V_CC
                {where_sql}
                GROUP BY NO
                {having_sql}
                ORDER BY SUM({open_expr}) DESC, MAX(NOME)
            """)
            rows = db.session.execute(sql, params).mappings().all()

            items = []
            total_aberto = 0.0
            total_saldo = 0.0
            for r in rows:
                no = r.get('NO')
                nome = str(r.get('NOME') or '').strip()
                saldo = float(r.get('SALDO_ABERTO') or 0)
                total_saldo += saldo
                aberto = saldo if saldo > 0 else 0.0
                total_aberto += aberto
                items.append({
                    'NO': int(no) if str(no).isdigit() else no,
                    'NOME': nome,
                    'SALDO_ABERTO': round(saldo, 2),
                    'ABERTO': round(aberto, 2)
                })

            return jsonify({
                'total_aberto': round(total_aberto, 2),
                'total_saldo': round(total_saldo, 2),
                'count_clientes': len(items),
                'items': items
            })
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/cc_clientes/detalhe')
    @login_required
    def api_cc_clientes_detalhe():
        """
        Movimentos de conta corrente por cliente.
        Query:
          - no (obrigatório): cliente
          - pendentes=1: apenas linhas com aberto != 0
        """
        try:
            no = request.args.get('no', type=int)
            if not no:
                return jsonify({'error': 'Parâmetro no obrigatório'}), 400
            pendentes = (request.args.get('pendentes') or '').strip() in ('1', 'true', 'True')

            open_expr = "(ISNULL(EDEB,0) - ISNULL(EDEBF,0)) - (ISNULL(ECRED,0) - ISNULL(ECREDF,0))"
            where_pend = f"AND ABS({open_expr}) > 0.005" if pendentes else ""

            sql = text(f"""
                SELECT
                    CCSTAMP,
                    CAST(DATALC AS date) AS DATALC,
                    CAST(DATAVEN AS date) AS DATAVEN,
                    ISNULL(CMDESC,'') AS CMDESC,
                    ISNULL(NRDOC,'') AS NRDOC,
                    ISNULL(EDEB,0) AS EDEB,
                    ISNULL(ECRED,0) AS ECRED,
                    ISNULL(EDEBF,0) AS EDEBF,
                    ISNULL(ECREDF,0) AS ECREDF,
                    ISNULL(MOEDA,'') AS MOEDA,
                    ISNULL(CCUSTO,'') AS CCUSTO,
                    ISNULL(FTSTAMP,'') AS FTSTAMP,
                    ISNULL(CM,0) AS CM,
                    {open_expr} AS ABERTO
                FROM dbo.V_CC
                WHERE NO = :no
                {where_pend}
                ORDER BY CAST(DATALC AS date), CCSTAMP
            """)
            rows = db.session.execute(sql, {'no': no}).mappings().all()

            out = []
            total_aberto = 0.0
            for r in rows:
                aberto = float(r.get('ABERTO') or 0)
                total_aberto += aberto
                dlc = r.get('DATALC')
                dven = r.get('DATAVEN')
                out.append({
                    'CCSTAMP': r.get('CCSTAMP') or '',
                    'DATALC': dlc.strftime('%Y-%m-%d') if isinstance(dlc, (datetime, date)) else (str(dlc) if dlc is not None else ''),
                    'DATAVEN': dven.strftime('%Y-%m-%d') if isinstance(dven, (datetime, date)) else (str(dven) if dven is not None else ''),
                    'CMDESC': r.get('CMDESC') or '',
                    'NRDOC': r.get('NRDOC') or '',
                    'EDEB': round(float(r.get('EDEB') or 0), 2),
                    'ECRED': round(float(r.get('ECRED') or 0), 2),
                    'EDEBF': round(float(r.get('EDEBF') or 0), 2),
                    'ECREDF': round(float(r.get('ECREDF') or 0), 2),
                    'ABERTO': round(aberto, 2),
                    'MOEDA': r.get('MOEDA') or '',
                    'CCUSTO': r.get('CCUSTO') or '',
                    'FTSTAMP': r.get('FTSTAMP') or '',
                    'CM': int(r.get('CM') or 0),
                })

            return jsonify({'no': no, 'rows': out, 'total_aberto': round(total_aberto, 2)})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/cc_clientes/pendentes')
    @login_required
    def api_cc_clientes_pendentes():
        """
        Lista movimentos pendentes (em aberto) para assistente de recebimentos.
        Query:
          - q (opcional): pesquisa por NO/NOME/NRDOC/CMDESC

        Nota: considera pendente quando ABS(ABERTO) > 0 (a receber / a crédito).
        """
        try:
            q = (request.args.get('q') or '').strip()

            open_expr = "(ISNULL(EDEB,0) - ISNULL(EDEBF,0)) - (ISNULL(ECRED,0) - ISNULL(ECREDF,0))"
            where = [f"ABS(({open_expr})) > 0.005"]
            params = {}
            if q:
                where.append("(CAST(NO AS varchar(50)) LIKE :q OR NOME LIKE :q OR CAST(NRDOC AS varchar(50)) LIKE :q OR CMDESC LIKE :q)")
                params['q'] = f"%{q}%"

            where_sql = "WHERE " + " AND ".join(where)

            sql = text(f"""
                SELECT
                    CCSTAMP,
                    CAST(DATALC AS date) AS DATALC,
                    CAST(DATAVEN AS date) AS DATAVEN,
                    ISNULL(CMDESC,'') AS CMDESC,
                    ISNULL(NRDOC,0) AS NRDOC,
                    ISNULL(CM,0) AS CM,
                    ISNULL(NO,0) AS NO,
                    ISNULL(NOME,'') AS NOME,
                    ISNULL(MOEDA,'') AS MOEDA,
                    ISNULL(CCUSTO,'') AS CCUSTO,
                    ISNULL(FTSTAMP,'') AS FTSTAMP,
                    {open_expr} AS ABERTO
                FROM dbo.V_CC
                {where_sql}
                ORDER BY ISNULL(NOME,''), CAST(DATAVEN AS date), CAST(DATALC AS date), CCSTAMP
            """)
            rows = db.session.execute(sql, params).mappings().all()

            out = []
            total = 0.0
            for r in rows:
                dlc = r.get('DATALC')
                dven = r.get('DATAVEN')
                aberto = float(r.get('ABERTO') or 0)
                total += aberto
                out.append({
                    'CCSTAMP': (r.get('CCSTAMP') or '').strip(),
                    'DATALC': dlc.strftime('%Y-%m-%d') if isinstance(dlc, (datetime, date)) else (str(dlc) if dlc is not None else ''),
                    'DATAVEN': dven.strftime('%Y-%m-%d') if isinstance(dven, (datetime, date)) else (str(dven) if dven is not None else ''),
                    'CMDESC': r.get('CMDESC') or '',
                    'NRDOC': int(r.get('NRDOC') or 0),
                    'CM': int(r.get('CM') or 0),
                    'NO': int(r.get('NO') or 0),
                    'NOME': r.get('NOME') or '',
                    'MOEDA': r.get('MOEDA') or '',
                    'CCUSTO': r.get('CCUSTO') or '',
                    'FTSTAMP': r.get('FTSTAMP') or '',
                    'ABERTO': round(aberto, 2)
                })

            return jsonify({'rows': out, 'total': round(total, 2), 'count': len(out)})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/cc_clientes/recebimentos/criar', methods=['POST'])
    @login_required
    def api_cc_clientes_recebimentos_criar():
        """
        Cria registos RE/RL a partir de movimentos pendentes selecionados.
        Body:
          { "rec_date": "YYYY-MM-DD", "items": [ { CCSTAMP, NO, NOME, CMDESC, NRDOC, DATALC, DATAVEN, ABERTO, PAYVAL, MOEDA } ... ] }
        Cria 1 RE por cliente e 1 RL por movimento.
        """
        try:
            body = request.get_json(silent=True) or {}
            rec_date = str(body.get('rec_date') or '').strip()
            items = body.get('items') or []
            if not rec_date:
                return jsonify({'error': 'rec_date obrigatório (YYYY-MM-DD)'}), 400
            if not isinstance(items, list) or not items:
                return jsonify({'error': 'items obrigatório'}), 400

            try:
                d = date.fromisoformat(rec_date)
            except Exception:
                return jsonify({'error': 'rec_date inválido (YYYY-MM-DD)'}), 400
            base_dt = datetime.combine(d, datetime.min.time())
            ano = d.year

            # agrupar por cliente
            by_no = {}
            for it in items:
                no = int(it.get('NO') or 0)
                if not no:
                    continue
                nome = str(it.get('NOME') or '').strip()
                ccstamp = str(it.get('CCSTAMP') or '').strip()
                if not ccstamp:
                    continue
                aberto = float(it.get('ABERTO') or 0)
                payval = it.get('PAYVAL', None)
                pay = float(payval) if payval is not None else aberto

                # Permitir também créditos (ABERTO < 0): aceita valores positivos e negativos,
                # mas força o sinal e limita à faixa válida do "aberto".
                if not (abs(aberto) > 0.005):
                    continue
                if aberto > 0:
                    # entre 0 e aberto
                    if pay < 0:
                        pay = 0.0
                    if pay > aberto:
                        pay = aberto
                else:
                    # aberto < 0: entre aberto e 0 (valores negativos)
                    if pay > 0:
                        pay = 0.0
                    if pay < aberto:
                        pay = aberto
                if abs(pay) <= 0.00001:
                    continue
                if no not in by_no:
                    by_no[no] = {'NO': no, 'NOME': nome, 'items': []}
                by_no[no]['items'].append({**it, 'PAYVAL': pay})

            if not by_no:
                return jsonify({'error': 'Nenhum movimento válido selecionado.'}), 400

            created = []
            for grp in by_no.values():
                no = grp['NO']
                nome = grp['NOME'] or ''
                movs = grp['items']
                total = sum(float(x.get('PAYVAL') or 0) for x in movs)

                # Colunas disponíveis (para compatibilidade entre bases)
                re_cols = set(
                    r['COLUMN_NAME'].upper()
                    for r in (
                        db.session.execute(
                            text("""
                                SELECT COLUMN_NAME
                                FROM INFORMATION_SCHEMA.COLUMNS
                                WHERE TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'RE'
                            """)
                        ).mappings().all()
                        or []
                    )
                )
                rl_cols = set(
                    r['COLUMN_NAME'].upper()
                    for r in (
                        db.session.execute(
                            text("""
                                SELECT COLUMN_NAME
                                FROM INFORMATION_SCHEMA.COLUMNS
                                WHERE TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'RL'
                            """)
                        ).mappings().all()
                        or []
                    )
                )

                # RNO sequencial simples
                rno_row = db.session.execute(text("SELECT ISNULL(MAX(RNO),0) + 1 AS N FROM dbo.RE")).mappings().first() or {}
                rno = int(rno_row.get('N') or 1)

                restamp = new_stamp()
                moeda = (movs[0].get('MOEDA') or 'EUR').strip()[:11]

                re_cols_sql = [
                    'RESTAMP', 'NMDOC', 'RNO', 'RDATA', 'NOME', 'TOTAL', 'ETOTAL',
                    'NO', 'REANO',
                    'OLCODIGO', 'TELOCAL', 'MOEDA', 'CONTADO', 'PROCESS', 'PROCDATA', 'OLLOCAL', 'PLANO', 'TIPO', 'PAIS', 'SYNC'
                ]
                re_vals_sql = [
                    ':RESTAMP', ':NMDOC', ':RNO', ':RDATA', ':NOME', ':TOTAL', ':ETOTAL',
                    ':NO', ':REANO',
                    ':OLCODIGO', ':TELOCAL', ':MOEDA', ':CONTADO', ':PROCESS', ':PROCDATA', ':OLLOCAL', ':PLANO', ':TIPO', ':PAIS', ':SYNC'
                ]
                # Algumas bases usam NDOC, outras NDOS.
                if 'NDOC' in re_cols:
                    re_cols_sql.insert(7, 'NDOC')
                    re_vals_sql.insert(7, ':NDOC')
                if 'NDOS' in re_cols:
                    re_cols_sql.insert(7, 'NDOS')
                    re_vals_sql.insert(7, ':NDOS')

                ins_re = text(f"""
                    INSERT INTO dbo.RE
                    ({", ".join(re_cols_sql)})
                    VALUES
                    ({", ".join(re_vals_sql)})
                """)
                re_params = {
                    'RESTAMP': restamp,
                    'NMDOC': 'Normal',
                    'RNO': rno,
                    'RDATA': base_dt,
                    'NOME': nome[:55],
                    'TOTAL': total,
                    'ETOTAL': total,
                    'NO': no,
                    'REANO': ano,
                    'OLCODIGO': 'R10001',
                    'TELOCAL': 'B',
                    'MOEDA': moeda,
                    'CONTADO': 1,
                    'PROCESS': 1,
                    'PROCDATA': base_dt,
                    'OLLOCAL': 'Santander  DO',
                    'PLANO': 0,
                    'TIPO': '',
                    'PAIS': 1,
                    'SYNC': 0
                }
                if 'NDOC' in re_cols:
                    re_params['NDOC'] = 1
                if 'NDOS' in re_cols:
                    re_params['NDOS'] = 1
                db.session.execute(ins_re, re_params)

                # RL: algumas bases podem ter VAL (em vez de EVAL). Queremos:
                #   - VAL/EVAL = valor pendente (ABERTO)
                #   - EREC = valor recebido (PAYVAL)
                rl_value_col = 'EVAL' if 'EVAL' in rl_cols else ('VAL' if 'VAL' in rl_cols else 'EVAL')
                ins_rl = text(f"""
                    INSERT INTO dbo.RL
                    (RLSTAMP, NDOC, RNO, CDESC, NRDOC, DATALC, DATAVEN, RESTAMP, CCSTAMP, CM, {rl_value_col}, EREC, PROCESS, MOEDA, RDATA)
                    VALUES
                    (:RLSTAMP, :NDOC, :RNO, :CDESC, :NRDOC, :DATALC, :DATAVEN, :RESTAMP, :CCSTAMP, :CM, :VALPEND, :EREC, :PROCESS, :MOEDA, :RDATA)
                """)
                for m in movs:
                    rlstamp = new_stamp()
                    dlc = str(m.get('DATALC') or '').strip()
                    dven = str(m.get('DATAVEN') or '').strip()
                    try:
                        dlc_dt = datetime.combine(date.fromisoformat(dlc), datetime.min.time()) if dlc else base_dt
                    except Exception:
                        dlc_dt = base_dt
                    try:
                        dven_dt = datetime.combine(date.fromisoformat(dven), datetime.min.time()) if dven else base_dt
                    except Exception:
                        dven_dt = base_dt

                    aberto0 = float(m.get('ABERTO') or 0)
                    recebido = float(m.get('PAYVAL') or 0)

                    db.session.execute(ins_rl, {
                        'RLSTAMP': rlstamp,
                        'NDOC': 1,
                        'RNO': rno,
                        'CDESC': (str(m.get('CMDESC') or '')[:20]),
                        'NRDOC': int(m.get('NRDOC') or 0),
                        'DATALC': dlc_dt,
                        'DATAVEN': dven_dt,
                        'RESTAMP': restamp,
                        'CCSTAMP': (str(m.get('CCSTAMP') or '')[:25]),
                        'CM': int(m.get('CM') or 0),
                        'VALPEND': aberto0,
                        'EREC': recebido,
                        'PROCESS': 1,
                        'MOEDA': (str(m.get('MOEDA') or moeda)[:11]),
                        'RDATA': base_dt
                    })

                created.append({'RESTAMP': restamp, 'RNO': rno, 'NO': no, 'TOTAL': round(total, 2)})

            db.session.commit()
            return jsonify({'ok': True, 'created': created})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/cc_fornecedores/pagamentos/criar', methods=['POST'])
    @login_required
    def api_cc_fornecedores_pagamentos_criar():
        """
        Cria registos PO/PL a partir de movimentos pendentes selecionados.
        Body:
          { "items": [ { FCSTAMP, NO, NOME, CMDESC, ADOC, DATALC, DATAVEN, ABERTO, MOEDA, CCUSTO } ... ] }
        Cria 1 PO por fornecedor e 1 PL por movimento.
        """
        try:
            body = request.get_json(silent=True) or {}
            pay_date_str = str(body.get('pay_date') or '').strip()
            items = body.get('items') or []
            if not isinstance(items, list) or not items:
                return jsonify({'error': 'Sem movimentos selecionados.'}), 400

            # normalizar e validar
            cleaned = []
            for it in items:
                if not isinstance(it, dict):
                    continue
                no = int(it.get('NO') or 0)
                fcstamp = str(it.get('FCSTAMP') or '').strip()
                nome = str(it.get('NOME') or '').strip()
                if not no or not fcstamp:
                    continue
                try:
                    aberto = float(it.get('ABERTO') or 0)
                except Exception:
                    aberto = 0.0
                if abs(aberto) <= 0.005:
                    continue
                try:
                    payval = float(it.get('PAYVAL') if it.get('PAYVAL') is not None else it.get('EVAL') or aberto)
                except Exception:
                    payval = aberto
                # permitir valores negativos (créditos), mas sempre com o mesmo sinal do "aberto"
                if abs(payval) <= 0.005:
                    continue
                if aberto > 0:
                    if payval < 0:
                        continue
                    if payval > aberto:
                        payval = aberto
                else:
                    # aberto < 0: payval deve ser negativo e entre [aberto, 0]
                    if payval > 0:
                        continue
                    if payval < aberto:
                        payval = aberto
                cleaned.append({
                    'NO': no,
                    'NOME': nome,
                    'FCSTAMP': fcstamp[:25],
                    'CMDESC': str(it.get('CMDESC') or '').strip()[:20],
                    'ADOC': str(it.get('ADOC') or '').strip()[:50],
                    'DATALC': str(it.get('DATALC') or '').strip(),
                    'DATAVEN': str(it.get('DATAVEN') or '').strip(),
                    'CM': int(it.get('CM') or 0),
                    'EVAL': payval,
                    'ABERTO': aberto,
                    'MOEDA': str(it.get('MOEDA') or '').strip()[:11],
                    'CCUSTO': str(it.get('CCUSTO') or '').strip()[:20],
                })
                if len(cleaned) >= 2000:
                    break
            if not cleaned:
                return jsonify({'error': 'Sem movimentos válidos.'}), 400

            # agrupar por fornecedor
            by_no = {}
            for it in cleaned:
                by_no.setdefault(it['NO'], {'NO': it['NO'], 'NOME': it['NOME'], 'items': []})
                by_no[it['NO']]['items'].append(it)

            def new_stamp():
                import uuid
                return uuid.uuid4().hex.upper()[:25]

            # Data do pagamento (sem hora)
            try:
                base_date = date.fromisoformat(pay_date_str) if pay_date_str else date.today()
            except Exception:
                base_date = date.today()
            base_dt = datetime.combine(base_date, datetime.min.time())
            ano = int(base_date.strftime('%Y'))

            # Lote sequencial (1 por execução do assistente)
            lote_row = db.session.execute(text("SELECT ISNULL(MAX(LOTE),0) + 1 AS N FROM dbo.PO")).mappings().first() or {}
            lote = int(lote_row.get('N') or 1)

            created = []
            # Algumas BD usam DVALOR, outras DTVALOR (ou ambas). Detecta e usa as que existirem.
            po_cols = set(
                r['COLUMN_NAME'].upper()
                for r in (
                    db.session.execute(
                        text("""
                            SELECT COLUMN_NAME
                            FROM INFORMATION_SCHEMA.COLUMNS
                            WHERE TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'PO'
                        """)
                    ).mappings().all()
                    or []
                )
            )
            has_dvalor = 'DVALOR' in po_cols
            has_dtvalor = 'DTVALOR' in po_cols

            for grp in by_no.values():
                no = grp['NO']
                nome = grp['NOME'] or ''
                movs = grp['items']
                total = sum(float(x['EVAL'] or 0) for x in movs)

                # RNO sequencial simples
                rno_row = db.session.execute(text("SELECT ISNULL(MAX(RNO),0) + 1 AS N FROM dbo.PO")).mappings().first() or {}
                rno = int(rno_row.get('N') or 1)

                postamp = new_stamp()
                moeda = (movs[0].get('MOEDA') or 'EUR').strip()[:11]
                ccusto = (movs[0].get('CCUSTO') or '').strip()[:20]
                fcstamp_head = (movs[0].get('FCSTAMP') or '').strip()[:25]
                cmdesc_head = (movs[0].get('CMDESC') or 'PAGAMENTO').strip()[:20]
                adoc_head = (movs[0].get('ADOC') or '').strip()[:50]

                po_cols_sql = [
                    'POSTAMP', 'RNO', 'RDATA', 'NOME', 'TOTAL', 'ETOTAL', 'PAIS', 'NO', 'POANO',
                    'OLCODIGO', 'TELOCAL', 'MOEDA', 'CONTADO',
                    'PROCESS', 'PROCDATA', 'OLLOCAL', 'CCUSTO', 'PLANO', 'OLSTAMP', 'FCSTAMP', 'CM', 'CMDESC', 'ADOC', 'LOTE'
                ]
                po_vals_sql = [
                    ':POSTAMP', ':RNO', ':RDATA', ':NOME', ':TOTAL', ':ETOTAL', ':PAIS', ':NO', ':POANO',
                    ':OLCODIGO', ':TELOCAL', ':MOEDA', ':CONTADO',
                    ':PROCESS', ':PROCDATA', ':OLLOCAL', ':CCUSTO', ':PLANO', ':OLSTAMP', ':FCSTAMP', ':CM', ':CMDESC', ':ADOC', ':LOTE'
                ]
                if has_dvalor:
                    po_cols_sql.insert(9, 'DVALOR')
                    po_vals_sql.insert(9, ':DVALOR')
                if has_dtvalor:
                    po_cols_sql.insert(9, 'DTVALOR')
                    po_vals_sql.insert(9, ':DTVALOR')

                ins_po = text(f"""
                    INSERT INTO dbo.PO
                    ({", ".join(po_cols_sql)})
                    VALUES
                    ({", ".join(po_vals_sql)})
                """)

                po_params = {
                    'POSTAMP': postamp,
                    'RNO': rno,
                    'RDATA': base_dt,
                    'NOME': nome[:55],
                    'TOTAL': total,
                    'ETOTAL': total,
                    'PAIS': 0,
                    'NO': no,
                    'POANO': ano,
                    'OLCODIGO': 'P10001',
                    'TELOCAL': 'B',
                    'MOEDA': moeda,
                    'CONTADO': 1,
                    'PROCESS': 1,
                    'PROCDATA': base_dt,
                    'OLLOCAL': 'Santander  DO',
                    'CCUSTO': ccusto,
                    'PLANO': 0,
                    'OLSTAMP': '',
                    'FCSTAMP': fcstamp_head,
                    'CM': 104,
                    'CMDESC': 'N/Trfa.',
                    'ADOC': str(rno),
                    'LOTE': lote
                }
                if has_dvalor:
                    po_params['DVALOR'] = base_dt
                if has_dtvalor:
                    po_params['DTVALOR'] = base_dt

                db.session.execute(ins_po, po_params)

                ins_pl = text("""
                    INSERT INTO dbo.PL
                    (POSTAMP, PLSTAMP, RNO, CDESC, ADOC, DATALC, DATAVEN, FCSTAMP, CM, EVAL, EREC, PROCESS, MOEDA, RDATA)
                    VALUES
                    (:POSTAMP, :PLSTAMP, :RNO, :CDESC, :ADOC, :DATALC, :DATAVEN, :FCSTAMP, :CM, :EVAL, :EREC, :PROCESS, :MOEDA, :RDATA)
                """)
                for m in movs:
                    plstamp = new_stamp()
                    # datas como datetime (YYYY-MM-DD)
                    try:
                        dlc_dt = datetime.fromisoformat(m['DATALC'])
                    except Exception:
                        dlc_dt = base_dt
                    try:
                        dven_dt = datetime.fromisoformat(m['DATAVEN'])
                    except Exception:
                        dven_dt = dlc_dt

                    dlc = datetime.combine(dlc_dt.date(), datetime.min.time())
                    dven = datetime.combine(dven_dt.date(), datetime.min.time())
                    cm_line = int(m.get('CM') or 0)

                    db.session.execute(ins_pl, {
                        'POSTAMP': postamp,
                        'PLSTAMP': plstamp,
                        'RNO': rno,
                        'CDESC': (m.get('CMDESC') or '')[:20],
                        'ADOC': (m.get('ADOC') or '')[:50],
                        'DATALC': dlc,
                        'DATAVEN': dven,
                        'FCSTAMP': (m.get('FCSTAMP') or '')[:25],
                        'CM': cm_line,
                        'EVAL': float(m.get('EVAL') or 0),
                        'EREC': float(m.get('EVAL') or 0),
                        'PROCESS': 1,
                        'MOEDA': (m.get('MOEDA') or moeda)[:11],
                        'RDATA': base_dt
                    })

                created.append({'NO': no, 'NOME': nome, 'POSTAMP': postamp, 'RNO': rno, 'TOTAL': round(total, 2), 'COUNT': len(movs), 'LOTE': lote})

            db.session.commit()
            return jsonify({'ok': True, 'lote': lote, 'created': created})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    # -----------------------------
    # Pagamentos (PO)
    # -----------------------------
    @app.route('/pagamentos')
    @login_required
    def pagamentos_page():
        return render_template('pagamentos.html', page_title='Pagamentos')

    @app.route('/api/pagamentos')
    @login_required
    def api_pagamentos_list():
        """
        Lista pagamentos (PO) com filtros simples.
        Query:
          - de=YYYY-MM-DD
          - ate=YYYY-MM-DD
          - lote=int
          - q=texto (NO, NOME, ADOC, CMDESC, OLCODIGO, OLLOCAL)
        """
        try:
            de = (request.args.get('de') or '').strip()
            ate = (request.args.get('ate') or '').strip()
            q = (request.args.get('q') or '').strip()
            try:
                lote = int(request.args.get('lote')) if request.args.get('lote') not in (None, '') else None
            except Exception:
                lote = None

            where = []
            params = {}
            if de:
                where.append("CAST(PO.RDATA AS date) >= :de")
                params['de'] = de
            if ate:
                where.append("CAST(PO.RDATA AS date) <= :ate")
                params['ate'] = ate
            if lote is not None:
                where.append("PO.LOTE = :lote")
                params['lote'] = lote
            if q:
                where.append("""(
                    CAST(PO.NO AS varchar(50)) LIKE :q OR
                    PO.NOME LIKE :q OR
                    PO.ADOC LIKE :q OR
                    PO.CMDESC LIKE :q OR
                    PO.OLCODIGO LIKE :q OR
                    PO.OLLOCAL LIKE :q
                )""")
                params['q'] = f"%{q}%"

            where_sql = ("WHERE " + " AND ".join(where)) if where else ""

            sql = text(f"""
                SELECT TOP 2000
                    PO.POSTAMP,
                    PO.RNO,
                    CAST(PO.RDATA AS date) AS RDATA,
                    ISNULL(PO.LOTE,0) AS LOTE,
                    ISNULL(PO.NO,0) AS NO,
                    ISNULL(PO.NOME,'') AS NOME,
                    ISNULL(PO.MOEDA,'') AS MOEDA,
                    ISNULL(PO.ETOTAL,0) AS ETOTAL,
                    ISNULL(PO.OLLOCAL,'') AS OLLOCAL,
                    ISNULL(PO.OLCODIGO,'') AS OLCODIGO,
                    ISNULL(PO.TELOCAL,'') AS TELOCAL,
                    ISNULL(PO.CMDESC,'') AS CMDESC,
                    ISNULL(PO.ADOC,'') AS ADOC
                FROM dbo.PO AS PO
                {where_sql}
                ORDER BY CAST(PO.RDATA AS date) DESC, ISNULL(PO.LOTE,0) DESC, PO.RNO DESC
            """)
            rows = db.session.execute(sql, params).mappings().all()

            out = []
            total = 0.0
            for r in rows:
                d = r.get('RDATA')
                if isinstance(d, (datetime, date)):
                    d = d.strftime('%Y-%m-%d')
                else:
                    d = str(d) if d is not None else ''
                et = float(r.get('ETOTAL') or 0)
                total += et
                out.append({
                    'POSTAMP': r.get('POSTAMP') or '',
                    'RNO': int(r.get('RNO') or 0),
                    'RDATA': d,
                    'LOTE': int(r.get('LOTE') or 0),
                    'NO': int(r.get('NO') or 0),
                    'NOME': r.get('NOME') or '',
                    'MOEDA': r.get('MOEDA') or '',
                    'ETOTAL': round(et, 2),
                    'OLLOCAL': r.get('OLLOCAL') or '',
                    'OLCODIGO': r.get('OLCODIGO') or '',
                    'TELOCAL': r.get('TELOCAL') or '',
                    'CMDESC': r.get('CMDESC') or '',
                    'ADOC': r.get('ADOC') or ''
                })

            return jsonify({'rows': out, 'count': len(out), 'total': round(total, 2)})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/pagamentos/<postamp>')
    @login_required
    def api_pagamentos_detail(postamp):
        """
        Detalhe de um pagamento:
          - Cabeçalho: PO
          - Linhas: PL
        """
        try:
            postamp = (postamp or '').strip()
            if not postamp:
                return jsonify({'error': 'POSTAMP obrigatório'}), 400

            head_sql = text("""
                SELECT
                    PO.POSTAMP,
                    PO.RNO,
                    CAST(PO.RDATA AS date) AS RDATA,
                    ISNULL(PO.LOTE,0) AS LOTE,
                    ISNULL(PO.NO,0) AS NO,
                    ISNULL(PO.NOME,'') AS NOME,
                    ISNULL(PO.ETOTAL,0) AS ETOTAL,
                    ISNULL(PO.MOEDA,'') AS MOEDA,
                    ISNULL(PO.OLCODIGO,'') AS OLCODIGO,
                    ISNULL(PO.OLLOCAL,'') AS OLLOCAL,
                    ISNULL(PO.TELOCAL,'') AS TELOCAL,
                    ISNULL(PO.CM,0) AS CM,
                    ISNULL(PO.CMDESC,'') AS CMDESC,
                    ISNULL(PO.ADOC,'') AS ADOC,
                    ISNULL(PO.PROCESS,0) AS PROCESS
                FROM dbo.PO AS PO
                WHERE PO.POSTAMP = :postamp
            """)
            head = db.session.execute(head_sql, {'postamp': postamp}).mappings().first()
            if not head:
                return jsonify({'error': 'Pagamento não encontrado.'}), 404

            d = head.get('RDATA')
            if isinstance(d, (datetime, date)):
                d = d.strftime('%Y-%m-%d')
            else:
                d = str(d) if d is not None else ''

            header = {
                'POSTAMP': head.get('POSTAMP') or '',
                'RNO': int(head.get('RNO') or 0),
                'RDATA': d,
                'LOTE': int(head.get('LOTE') or 0),
                'NO': int(head.get('NO') or 0),
                'NOME': head.get('NOME') or '',
                'ETOTAL': round(float(head.get('ETOTAL') or 0), 2),
                'MOEDA': head.get('MOEDA') or '',
                'OLCODIGO': head.get('OLCODIGO') or '',
                'OLLOCAL': head.get('OLLOCAL') or '',
                'TELOCAL': head.get('TELOCAL') or '',
                'CM': int(head.get('CM') or 0),
                'CMDESC': head.get('CMDESC') or '',
                'ADOC': head.get('ADOC') or '',
                'PROCESS': int(head.get('PROCESS') or 0),
            }

            lines_sql = text("""
                SELECT
                    PL.PLSTAMP,
                    PL.RNO,
                    CAST(PL.RDATA AS date) AS RDATA,
                    ISNULL(PL.CDESC,'') AS CDESC,
                    ISNULL(PL.ADOC,'') AS ADOC,
                    CAST(PL.DATALC AS date) AS DATALC,
                    CAST(PL.DATAVEN AS date) AS DATAVEN,
                    ISNULL(PL.FCSTAMP,'') AS FCSTAMP,
                    ISNULL(PL.CM,0) AS CM,
                    ISNULL(PL.EVAL,0) AS EVAL,
                    ISNULL(PL.EREC,0) AS EREC,
                    ISNULL(PL.PROCESS,0) AS PROCESS,
                    ISNULL(PL.MOEDA,'') AS MOEDA
                FROM dbo.PL AS PL
                WHERE PL.POSTAMP = :postamp
                ORDER BY CAST(PL.DATAVEN AS date), CAST(PL.DATALC AS date), PL.PLSTAMP
            """)
            rows = db.session.execute(lines_sql, {'postamp': postamp}).mappings().all()

            lines = []
            for r in rows:
                dlc = r.get('DATALC')
                dven = r.get('DATAVEN')
                lines.append({
                    'PLSTAMP': r.get('PLSTAMP') or '',
                    'RNO': int(r.get('RNO') or 0),
                    'RDATA': (r.get('RDATA').strftime('%Y-%m-%d') if isinstance(r.get('RDATA'), (datetime, date)) else str(r.get('RDATA') or '')),
                    'CDESC': r.get('CDESC') or '',
                    'ADOC': r.get('ADOC') or '',
                    'DATALC': dlc.strftime('%Y-%m-%d') if isinstance(dlc, (datetime, date)) else (str(dlc) if dlc is not None else ''),
                    'DATAVEN': dven.strftime('%Y-%m-%d') if isinstance(dven, (datetime, date)) else (str(dven) if dven is not None else ''),
                    'FCSTAMP': r.get('FCSTAMP') or '',
                    'CM': int(r.get('CM') or 0),
                    'EVAL': round(float(r.get('EVAL') or 0), 2),
                    'EREC': round(float(r.get('EREC') or 0), 2),
                    'PROCESS': int(r.get('PROCESS') or 0),
                    'MOEDA': r.get('MOEDA') or ''
                })

            return jsonify({'header': header, 'lines': lines})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/pagamentos/<postamp>', methods=['PUT'])
    @login_required
    def api_pagamentos_update(postamp):
        """
        Atualiza campos do pagamento (PO). Por agora:
          - RDATA/(DVALOR|DTVALOR)/PROCDATA (data sem hora)
          - PL.RDATA (para manter consistência)
        Body: { "RDATA": "YYYY-MM-DD" }
        """
        try:
            postamp = (postamp or '').strip()
            if not postamp:
                return jsonify({'error': 'POSTAMP obrigatório'}), 400
            body = request.get_json(silent=True) or {}
            rdata = str(body.get('RDATA') or '').strip()
            if not rdata:
                return jsonify({'error': 'RDATA obrigatória'}), 400

            try:
                d = date.fromisoformat(rdata)
            except Exception:
                return jsonify({'error': 'RDATA inválida (YYYY-MM-DD)'}), 400

            dt0 = datetime.combine(d, datetime.min.time())

            # garantir que existe
            exists = db.session.execute(
                text("SELECT 1 AS X FROM dbo.PO WHERE POSTAMP = :p"),
                {'p': postamp}
            ).mappings().first()
            if not exists:
                return jsonify({'error': 'Pagamento não encontrado.'}), 404

            po_cols = set(
                r['COLUMN_NAME'].upper()
                for r in (
                    db.session.execute(
                        text("""
                            SELECT COLUMN_NAME
                            FROM INFORMATION_SCHEMA.COLUMNS
                            WHERE TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'PO'
                        """)
                    ).mappings().all()
                    or []
                )
            )

            set_parts = ["RDATA = :d", "PROCDATA = :d"]
            if 'DVALOR' in po_cols:
                set_parts.append("DVALOR = :d")
            if 'DTVALOR' in po_cols:
                set_parts.append("DTVALOR = :d")

            db.session.execute(
                text(f"""
                    UPDATE dbo.PO
                    SET {", ".join(set_parts)}
                    WHERE POSTAMP = :p
                """),
                {'d': dt0, 'p': postamp}
            )
            db.session.execute(
                text("""
                    UPDATE dbo.PL
                    SET RDATA = :d
                    WHERE POSTAMP = :p
                """),
                {'d': dt0, 'p': postamp}
            )
            db.session.commit()
            return jsonify({'ok': True, 'POSTAMP': postamp, 'RDATA': rdata})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/pagamentos/<postamp>', methods=['DELETE'])
    @login_required
    def api_pagamentos_delete(postamp):
        """
        Elimina um pagamento (PO) e as suas linhas (PL).
        """
        try:
            postamp = (postamp or '').strip()
            if not postamp:
                return jsonify({'error': 'POSTAMP obrigatório'}), 400

            # delete linhas primeiro
            db.session.execute(text("DELETE FROM dbo.PL WHERE POSTAMP = :p"), {'p': postamp})
            res = db.session.execute(text("DELETE FROM dbo.PO WHERE POSTAMP = :p"), {'p': postamp})
            # SQLAlchemy 2: rowcount disponível
            if getattr(res, 'rowcount', 0) == 0:
                db.session.rollback()
                return jsonify({'error': 'Pagamento não encontrado.'}), 404
            db.session.commit()
            return jsonify({'ok': True, 'POSTAMP': postamp})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    # -----------------------------
    # Recebimentos (RE)
    # -----------------------------
    @app.route('/recebimentos')
    @login_required
    def recebimentos_page():
        return render_template('recebimentos.html', page_title='Recebimentos')

    # -----------------------------
    # Calendário de Reservas (Timeline/Gantt-like)
    # -----------------------------
    @app.route('/calendario_reservas')
    @login_required
    def calendario_reservas_page():
        return render_template('calendario_reservas.html', page_title='Calendário de Reservas')

    @app.route('/api/calendario_reservas')
    @login_required
    def api_calendario_reservas():
        # Janela: por defeito hoje-5 até +60 dias (inclusive)
        try:
            start_str = (request.args.get('start') or '').strip()
            end_str = (request.args.get('end') or '').strip()
            if start_str:
                start_d = datetime.strptime(start_str, '%Y-%m-%d').date()
            else:
                start_d = date.today() - timedelta(days=5)
            if end_str:
                end_d = datetime.strptime(end_str, '%Y-%m-%d').date()
            else:
                end_d = start_d + timedelta(days=119)
        except Exception:
            start_d = date.today() - timedelta(days=5)
            end_d = start_d + timedelta(days=119)

        if end_d < start_d:
            end_d = start_d

        end_excl = end_d + timedelta(days=1)

        try:
            al_rows = db.session.execute(text("""
                SELECT
                    LTRIM(RTRIM(ISNULL(NOME,''))) AS NOME,
                    ISNULL(PBASE,0) AS PBASE
                FROM AL
                WHERE ISNULL(INATIVO,0) = 0
                ORDER BY LTRIM(RTRIM(ISNULL(NOME,'')))
            """)).mappings().all()
            alojamentos = [str(r.get('NOME') or '').strip() for r in al_rows if str(r.get('NOME') or '').strip()]
            def _norm_key(v: str) -> str:
                s = str(v or '').strip()
                if not s:
                    return ''
                s = unicodedata.normalize('NFD', s)
                s = ''.join(ch for ch in s if not unicodedata.combining(ch))
                return s.upper()

            pbase_by_name = {}
            for r in al_rows:
                nome = str(r.get('NOME') or '').strip()
                if not nome:
                    continue
                pbase_by_name[_norm_key(nome)] = float(r.get('PBASE') or 0)
        except Exception as e:
            return jsonify({'error': f'Erro ao obter alojamentos: {e}'}), 500

        # detetar colunas disponíveis em RS (nome do hóspede e stamp)
        try:
            cols = db.session.execute(text("""
                SELECT UPPER(COLUMN_NAME) AS COL
                FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_NAME = 'RS'
            """)).fetchall()
            rs_cols = {str(r[0] or '').strip().upper() for r in cols if r and r[0]}
        except Exception:
            rs_cols = set()

        stamp_expr = "LEFT(NEWID(),25) AS RSSTAMP"
        if 'RSSTAMP' in rs_cols:
            stamp_expr = "RS.RSSTAMP AS RSSTAMP"

        # nome do hóspede: tentar várias colunas comuns
        guest_col = None
        for c in ('NOME', 'HOSPEDE', 'NOMEHOSPEDE', 'CLIENTE', 'NOMECLIENTE'):
            if c in rs_cols:
                guest_col = c
                break
        guest_expr = "'' AS HOSPEDE"
        if guest_col:
            guest_expr = f"ISNULL(RS.[{guest_col}], '') AS HOSPEDE"

        # Reservas que intersectam a janela (start..end)
        try:
            sql_rs = f"""
                SELECT
                    {stamp_expr},
                    {guest_expr},
                    LTRIM(RTRIM(ISNULL(RS.ALOJAMENTO,''))) AS ALOJAMENTO,
                    -- offsets dentro do dia para evitar barras "coladas" (start ~25%, end ~20%)
                    DATEADD(hour, 6, CAST(CAST(RS.DATAIN AS date) AS datetime)) AS STARTDT,
                    DATEADD(minute, 48, DATEADD(hour, 4, CAST(CAST(RS.DATAOUT AS date) AS datetime))) AS ENDDT,
                    CAST(RS.DATAIN AS date) AS DATAIN,
                    CAST(RS.DATAOUT AS date) AS DATAOUT,
                    ISNULL(RS.NOITES,0) AS NOITES,
                    ISNULL(RS.ADULTOS,0) AS ADULTOS,
                    ISNULL(RS.CRIANCAS,0) AS CRIANCAS,
                    (ISNULL(RS.ESTADIA,0) + ISNULL(RS.LIMPEZA,0) - ISNULL(RS.COMISSAO,0)) AS VALOR
                FROM RS
                WHERE RS.DATAIN IS NOT NULL
                  AND RS.DATAOUT IS NOT NULL
                  AND ISNULL(RS.CANCELADA,0) = 0
                  AND CAST(RS.DATAOUT AS date) >= :start
                  AND CAST(RS.DATAIN AS date) < :end_excl
            """
            rs_rows = db.session.execute(text(sql_rs), {'start': start_d, 'end_excl': end_excl}).mappings().all()
        except Exception as e:
            return jsonify({'error': f'Erro ao obter reservas: {e}'}), 500

        # grupos únicos (por alojamento) - evitar duplicados e normalizar por case/trim
        groups_by_key = {}
        for a in alojamentos:
            name = str(a or '').strip()
            if not name:
                continue
            k = _norm_key(name)
            if k not in groups_by_key:
                groups_by_key[k] = name

        items = []
        seen_item_ids = set()
        occupied_by_group = {}
        for r in rs_rows:
            aloj = str(r.get('ALOJAMENTO') or '').strip()
            if not aloj:
                continue

            aloj_key = _norm_key(aloj)
            if aloj_key not in groups_by_key:
                groups_by_key[aloj_key] = aloj
            group_id = groups_by_key[aloj_key]

            # Marcar noites ocupadas para esconder cadeados quando existe reserva
            try:
                din = r.get('DATAIN')
                dout = r.get('DATAOUT')
                if isinstance(din, date) and isinstance(dout, date) and din < dout:
                    occ = occupied_by_group.setdefault(group_id, set())
                    d0 = max(din, start_d)
                    d1 = min(dout, end_excl)  # exclusivo
                    dcur = d0
                    while dcur < d1:
                        occ.add(dcur.isoformat())
                        dcur = dcur + timedelta(days=1)
            except Exception:
                pass

            rsstamp_base = str(r.get('RSSTAMP') or '').strip() or f"RS-{aloj}-{r.get('DATAIN')}-{r.get('DATAOUT')}"
            rsstamp = rsstamp_base
            if rsstamp in seen_item_ids:
                # garantir unicidade caso existam duplicados na origem
                n = 2
                while f"{rsstamp_base}-{n}" in seen_item_ids:
                    n += 1
                rsstamp = f"{rsstamp_base}-{n}"
            seen_item_ids.add(rsstamp)
            startdt = r.get('STARTDT')
            enddt = r.get('ENDDT')
            if not startdt or not enddt:
                continue
            hosp = str(r.get('HOSPEDE') or '').strip()
            noites = int(r.get('NOITES') or 0)
            ad = int(r.get('ADULTOS') or 0)
            cr = int(r.get('CRIANCAS') or 0)
            valor = float(r.get('VALOR') or 0)
            title = f"{hosp or aloj} | {noites} noites | {ad+cr} hóspedes | {valor:.2f}"
            items.append({
                'id': rsstamp,
                'group': group_id,
                'start': startdt.isoformat(sep=' '),
                'end': enddt.isoformat(sep=' '),
                'hospede': hosp,
                'valor': round(valor, 2),
                'noites': noites,
                'title': title,
                'className': 'rs-item'
            })

        groups = []
        for v in sorted(groups_by_key.values(), key=lambda x: str(x).upper()):
            name = str(v or '').strip()
            groups.append({
                'id': v,
                'content': v,
                'pbase': float(pbase_by_name.get(_norm_key(name), 0) or 0),
            })

        # Noites bloqueadas (BQ) na janela (start..end) - cadeado por célula
        try:
            bq_rows = db.session.execute(text("""
                SELECT
                    LTRIM(RTRIM(ISNULL(BQ.ALOJAMENTO,''))) AS ALOJAMENTO,
                    LTRIM(RTRIM(ISNULL(BQ.NMAIRBNB,''))) AS NMAIRBNB,
                    CAST(BQ.[DATA] AS date) AS [DATA],
                    ISNULL(BQ.TRATADO,0) AS TRATADO,
                    ISNULL(BQ.DESBLOQ,0) AS DESBLOQ,
                    ISNULL(BQ.ANULADO,0) AS ANULADO
                FROM dbo.BQ AS BQ
                WHERE CAST(BQ.[DATA] AS date) >= :start
                  AND CAST(BQ.[DATA] AS date) <= :end
                  AND ISNULL(BQ.ANULADO,0) = 0
            """), {'start': start_d, 'end': end_d}).mappings().all()
        except Exception as e:
            return jsonify({'error': f'Erro ao obter BQ: {e}'}), 500

        blocked_map = {}  # (group_id, YYYY-MM-DD) -> flags
        for r in bq_rows:
            d = r.get('DATA')
            if not d:
                continue

            cand1 = _norm_key(r.get('ALOJAMENTO'))
            cand2 = _norm_key(r.get('NMAIRBNB'))
            group_id = None
            if cand1 and cand1 in groups_by_key:
                group_id = groups_by_key[cand1]
            elif cand2 and cand2 in groups_by_key:
                group_id = groups_by_key[cand2]

            if not group_id:
                continue

            key = (group_id, d.isoformat())
            tratado = 1 if int(r.get('TRATADO') or 0) else 0
            desbloq = 1 if int(r.get('DESBLOQ') or 0) else 0
            if int(r.get('ANULADO') or 0):
                continue
            flags = blocked_map.get(key, {'has_block': False, 'has_block_treated': False, 'has_unblock': False})
            if desbloq == 1 and tratado == 0:
                flags['has_unblock'] = True
            elif desbloq == 0 and tratado == 0:
                flags['has_block'] = True
            elif desbloq == 0 and tratado == 1:
                flags['has_block_treated'] = True
            blocked_map[key] = flags

        blocked = []
        for k, v in blocked_map.items():
            if v.get('has_unblock'):
                blocked.append({'group': k[0], 'date': k[1], 'tratado': 0, 'desbloq': 1})
            elif v.get('has_block'):
                blocked.append({'group': k[0], 'date': k[1], 'tratado': 0, 'desbloq': 0})
            elif v.get('has_block_treated'):
                blocked.append({'group': k[0], 'date': k[1], 'tratado': 1, 'desbloq': 0})

        # Preços por data (PRECOS) - para mostrar nas células (noites vazias)
        try:
            pr_rows = db.session.execute(text("""
                SELECT
                    LTRIM(RTRIM(ISNULL(P.ALOJAMENTO,''))) AS ALOJAMENTO,
                    CAST(P.[DATA] AS date) AS [DATA],
                    ISNULL(P.PBASE,0) AS PBASE,
                    ISNULL(P.DESCONTO,0) AS DESCONTO,
                    ISNULL(P.PRECO,0) AS PRECO,
                    ISNULL(P.TRATADO,0) AS TRATADO
                FROM dbo.PRECOS AS P
                WHERE CAST(P.[DATA] AS date) >= :start
                  AND CAST(P.[DATA] AS date) <= :end
                ORDER BY ISNULL(P.TRATADO,0) DESC, CAST(P.[DATA] AS date) DESC
            """), {'start': start_d, 'end': end_d}).mappings().all()
        except Exception as e:
            return jsonify({'error': f'Erro ao obter PRECOS: {e}'}), 500

        price_map = {}  # (group_id, YYYY-MM-DD) -> {pbase, desconto, preco, tratado}
        for r in pr_rows:
            d = r.get('DATA')
            if not d:
                continue
            aloj = _norm_key(r.get('ALOJAMENTO'))
            if not aloj or aloj not in groups_by_key:
                continue
            group_id = groups_by_key[aloj]
            key = (group_id, d.isoformat())
            if key in price_map:
                continue
            try:
                price_map[key] = {
                    'pbase': float(r.get('PBASE') or 0),
                    'desconto': float(r.get('DESCONTO') or 0),
                    'preco': float(r.get('PRECO') or 0),
                    'tratado': 1 if int(r.get('TRATADO') or 0) else 0,
                }
            except Exception:
                price_map[key] = {'pbase': 0.0, 'desconto': 0.0, 'preco': 0.0, 'tratado': 0}

        prices = [
            {'group': k[0], 'date': k[1], **v}
            for k, v in price_map.items()
        ]

        return jsonify({
            'start': start_d.isoformat(),
            'end': end_d.isoformat(),
            'groups': groups,
            'items': items,
            'blocked': blocked,
            'prices': prices,
        })

    @app.route('/api/precos_bulk', methods=['POST'])
    @login_required
    def api_precos_bulk():
        """
        Insere/atualiza preços na tabela PRECOS para um conjunto de células (ALOJAMENTO + DATA).
        Body:
          {
            "cells": [{"group":"...", "date":"YYYY-MM-DD"}, ...],
            "pbase": number,
            "desconto": number,
            "preco": number
          }
        """
        try:
            payload = request.get_json(silent=True) or {}
            cells = payload.get('cells') or []
            pbase = float(payload.get('pbase') or 0)
            desconto = float(payload.get('desconto') or 0)
            preco = float(payload.get('preco') or 0)
        except Exception:
            return jsonify({'error': 'Payload inválido'}), 400

        if not isinstance(cells, list) or not cells:
            return jsonify({'error': 'Sem células selecionadas'}), 400

        # normalizar desconto para 0..60
        if desconto not in (0, 10, 20, 30, 40, 50, 60):
            return jsonify({'error': 'Desconto inválido'}), 400

        try:
            updated = 0
            inserted = 0
            for c in cells:
                aloj = str((c or {}).get('group') or '').strip()
                d = str((c or {}).get('date') or '').strip()
                if not aloj or not d:
                    continue

                existing = db.session.execute(text("""
                    SELECT TOP 1 PRECOSSTAMP
                    FROM dbo.PRECOS
                    WHERE LTRIM(RTRIM(ALOJAMENTO)) = LTRIM(RTRIM(:aloj))
                      AND CAST([DATA] AS date) = CAST(:d AS date)
                """), {'aloj': aloj, 'd': d}).scalar()

                if existing:
                    db.session.execute(text("""
                        UPDATE dbo.PRECOS
                        SET PBASE = :pbase,
                            DESCONTO = :desconto,
                            PRECO = :preco,
                            TRATADO = 0
                        WHERE PRECOSSTAMP = :stamp
                    """), {'pbase': pbase, 'desconto': desconto, 'preco': preco, 'stamp': existing})
                    updated += 1
                else:
                    stamp = new_stamp()
                    db.session.execute(text("""
                        INSERT INTO dbo.PRECOS (PRECOSSTAMP, ALOJAMENTO, [DATA], PBASE, DESCONTO, PRECO, TRATADO)
                        VALUES (:stamp, :aloj, CAST(:d AS date), :pbase, :desconto, :preco, 0)
                    """), {'stamp': stamp, 'aloj': aloj, 'd': d, 'pbase': pbase, 'desconto': desconto, 'preco': preco})
                    inserted += 1

            db.session.commit()
            return jsonify({'ok': True, 'inserted': inserted, 'updated': updated})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/recebimentos')
    @login_required
    def api_recebimentos_list():
        """
        Lista recebimentos (RE) com filtros simples.
        Query:
          - de=YYYY-MM-DD
          - ate=YYYY-MM-DD
          - q=texto (NO, NOME, NRDOC, CDESC, NMDOC, OLCODIGO, OLLOCAL)
        """
        try:
            de = (request.args.get('de') or '').strip()
            ate = (request.args.get('ate') or '').strip()
            q = (request.args.get('q') or '').strip()

            where = []
            params = {}
            if de:
                where.append("CAST(RE.RDATA AS date) >= :de")
                params['de'] = de
            if ate:
                where.append("CAST(RE.RDATA AS date) <= :ate")
                params['ate'] = ate
            if q:
                where.append("""(
                    CAST(RE.NO AS varchar(50)) LIKE :q OR
                    RE.NOME LIKE :q OR
                    RE.NMDOC LIKE :q OR
                    RE.OLCODIGO LIKE :q OR
                    RE.OLLOCAL LIKE :q OR
                    EXISTS (
                        SELECT 1 FROM dbo.RL AS RL
                        WHERE RL.RESTAMP = RE.RESTAMP
                          AND (
                            CAST(ISNULL(RL.NRDOC,0) AS varchar(50)) LIKE :q OR
                            RL.CDESC LIKE :q
                          )
                    )
                )""")
                params['q'] = f"%{q}%"

            where_sql = ("WHERE " + " AND ".join(where)) if where else ""

            sql = text(f"""
                SELECT TOP 2000
                    RE.RESTAMP,
                    RE.RNO,
                    CAST(RE.RDATA AS date) AS RDATA,
                    ISNULL(RE.NO,0) AS NO,
                    ISNULL(RE.NOME,'') AS NOME,
                    ISNULL(RE.MOEDA,'') AS MOEDA,
                    ISNULL(RE.ETOTAL,0) AS ETOTAL,
                    ISNULL(RE.OLLOCAL,'') AS OLLOCAL,
                    ISNULL(RE.OLCODIGO,'') AS OLCODIGO,
                    ISNULL(RE.TELOCAL,'') AS TELOCAL,
                    ISNULL(RE.NMDOC,'') AS NMDOC
                FROM dbo.RE AS RE
                {where_sql}
                ORDER BY CAST(RE.RDATA AS date) DESC, RE.RNO DESC
            """)
            rows = db.session.execute(sql, params).mappings().all()

            out = []
            total = 0.0
            for r in rows:
                d = r.get('RDATA')
                if isinstance(d, (datetime, date)):
                    d = d.strftime('%Y-%m-%d')
                else:
                    d = str(d) if d is not None else ''
                et = float(r.get('ETOTAL') or 0)
                total += et
                out.append({
                    'RESTAMP': r.get('RESTAMP') or '',
                    'RNO': int(r.get('RNO') or 0),
                    'RDATA': d,
                    'NO': int(r.get('NO') or 0),
                    'NOME': r.get('NOME') or '',
                    'MOEDA': r.get('MOEDA') or '',
                    'ETOTAL': round(et, 2),
                    'OLLOCAL': r.get('OLLOCAL') or '',
                    'OLCODIGO': r.get('OLCODIGO') or '',
                    'TELOCAL': r.get('TELOCAL') or '',
                    'NMDOC': r.get('NMDOC') or ''
                })

            return jsonify({'rows': out, 'total': round(total, 2), 'count': len(out)})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/recebimentos/<restamp>')
    @login_required
    def api_recebimentos_detail(restamp):
        """
        Detalhe de recebimento: cabeçalho (RE) + linhas (RL).
        """
        try:
            restamp = (restamp or '').strip()
            if not restamp:
                return jsonify({'error': 'RESTAMP obrigatório'}), 400

            head_sql = text("""
                SELECT TOP 1
                    RE.RESTAMP,
                    RE.RNO,
                    CAST(RE.RDATA AS date) AS RDATA,
                    ISNULL(RE.NO,0) AS NO,
                    ISNULL(RE.NOME,'') AS NOME,
                    ISNULL(RE.MOEDA,'') AS MOEDA,
                    ISNULL(RE.ETOTAL,0) AS ETOTAL,
                    ISNULL(RE.OLLOCAL,'') AS OLLOCAL,
                    ISNULL(RE.OLCODIGO,'') AS OLCODIGO,
                    ISNULL(RE.TELOCAL,'') AS TELOCAL,
                    ISNULL(RE.NMDOC,'') AS NMDOC,
                    ISNULL(RE.PROCESS,0) AS PROCESS
                FROM dbo.RE AS RE
                WHERE RE.RESTAMP = :s
            """)
            head = db.session.execute(head_sql, {'s': restamp}).mappings().first()
            if not head:
                return jsonify({'error': 'Recebimento não encontrado.'}), 404

            rdata = head.get('RDATA')
            rdata_str = rdata.strftime('%Y-%m-%d') if isinstance(rdata, (datetime, date)) else (str(rdata) if rdata is not None else '')
            header = {
                'RESTAMP': head.get('RESTAMP') or '',
                'RNO': int(head.get('RNO') or 0),
                'RDATA': rdata_str,
                'NO': int(head.get('NO') or 0),
                'NOME': head.get('NOME') or '',
                'MOEDA': head.get('MOEDA') or '',
                'ETOTAL': round(float(head.get('ETOTAL') or 0), 2),
                'OLLOCAL': head.get('OLLOCAL') or '',
                'OLCODIGO': head.get('OLCODIGO') or '',
                'TELOCAL': head.get('TELOCAL') or '',
                'NMDOC': head.get('NMDOC') or '',
                'PROCESS': int(head.get('PROCESS') or 0),
            }

            rl_cols = set(
                r['COLUMN_NAME'].upper()
                for r in (
                    db.session.execute(
                        text("""
                            SELECT COLUMN_NAME
                            FROM INFORMATION_SCHEMA.COLUMNS
                            WHERE TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'RL'
                        """)
                    ).mappings().all()
                    or []
                )
            )
            value_col = 'EVAL' if 'EVAL' in rl_cols else ('VAL' if 'VAL' in rl_cols else 'EVAL')

            lines_sql = text(f"""
                SELECT
                    RL.RLSTAMP,
                    RL.RNO,
                    CAST(RL.RDATA AS date) AS RDATA,
                    ISNULL(RL.CDESC,'') AS CDESC,
                    ISNULL(RL.NRDOC,0) AS NRDOC,
                    CAST(RL.DATALC AS date) AS DATALC,
                    CAST(RL.DATAVEN AS date) AS DATAVEN,
                    ISNULL(RL.CCSTAMP,'') AS CCSTAMP,
                    ISNULL(RL.CM,0) AS CM,
                    ISNULL(RL.{value_col},0) AS EVAL,
                    ISNULL(RL.EREC,0) AS EREC,
                    ISNULL(RL.PROCESS,0) AS PROCESS,
                    ISNULL(RL.MOEDA,'') AS MOEDA
                FROM dbo.RL AS RL
                WHERE RL.RESTAMP = :s
                ORDER BY CAST(RL.DATAVEN AS date), CAST(RL.DATALC AS date), RL.RLSTAMP
            """)
            rows = db.session.execute(lines_sql, {'s': restamp}).mappings().all()

            lines = []
            for r in rows:
                dlc = r.get('DATALC')
                dven = r.get('DATAVEN')
                lines.append({
                    'RLSTAMP': r.get('RLSTAMP') or '',
                    'RNO': int(r.get('RNO') or 0),
                    'RDATA': (r.get('RDATA').strftime('%Y-%m-%d') if isinstance(r.get('RDATA'), (datetime, date)) else str(r.get('RDATA') or '')),
                    'CDESC': r.get('CDESC') or '',
                    'NRDOC': int(r.get('NRDOC') or 0),
                    'DATALC': dlc.strftime('%Y-%m-%d') if isinstance(dlc, (datetime, date)) else (str(dlc) if dlc is not None else ''),
                    'DATAVEN': dven.strftime('%Y-%m-%d') if isinstance(dven, (datetime, date)) else (str(dven) if dven is not None else ''),
                    'CCSTAMP': r.get('CCSTAMP') or '',
                    'CM': int(r.get('CM') or 0),
                    'EVAL': round(float(r.get('EVAL') or 0), 2),
                    'EREC': round(float(r.get('EREC') or 0), 2),
                    'PROCESS': int(r.get('PROCESS') or 0),
                    'MOEDA': r.get('MOEDA') or ''
                })

            return jsonify({'header': header, 'lines': lines})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/recebimentos/<restamp>', methods=['PUT'])
    @login_required
    def api_recebimentos_update(restamp):
        """
        Atualiza campos do recebimento (RE):
          - RDATA/PROCDATA (data sem hora)
          - RL.RDATA (consistência)
        Body: { "RDATA": "YYYY-MM-DD" }
        """
        try:
            restamp = (restamp or '').strip()
            if not restamp:
                return jsonify({'error': 'RESTAMP obrigatório'}), 400
            body = request.get_json(silent=True) or {}
            rdata = str(body.get('RDATA') or '').strip()
            if not rdata:
                return jsonify({'error': 'RDATA obrigatória'}), 400

            try:
                d = date.fromisoformat(rdata)
            except Exception:
                return jsonify({'error': 'RDATA inválida (YYYY-MM-DD)'}), 400

            dt0 = datetime.combine(d, datetime.min.time())

            exists = db.session.execute(
                text("SELECT 1 AS X FROM dbo.RE WHERE RESTAMP = :s"),
                {'s': restamp}
            ).mappings().first()
            if not exists:
                return jsonify({'error': 'Recebimento não encontrado.'}), 404

            db.session.execute(
                text("""
                    UPDATE dbo.RE
                    SET RDATA = :d, PROCDATA = :d
                    WHERE RESTAMP = :s
                """),
                {'d': dt0, 's': restamp}
            )
            db.session.execute(
                text("""
                    UPDATE dbo.RL
                    SET RDATA = :d
                    WHERE RESTAMP = :s
                """),
                {'d': dt0, 's': restamp}
            )
            db.session.commit()
            return jsonify({'ok': True, 'RESTAMP': restamp, 'RDATA': rdata})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/recebimentos/<restamp>', methods=['DELETE'])
    @login_required
    def api_recebimentos_delete(restamp):
        """
        Elimina um recebimento (RE) e as suas linhas (RL).
        """
        try:
            restamp = (restamp or '').strip()
            if not restamp:
                return jsonify({'error': 'RESTAMP obrigatório'}), 400

            exists = db.session.execute(
                text("SELECT 1 AS X FROM dbo.RE WHERE RESTAMP = :s"),
                {'s': restamp}
            ).mappings().first()
            if not exists:
                return jsonify({'error': 'Recebimento não encontrado.'}), 404

            db.session.execute(text("DELETE FROM dbo.RL WHERE RESTAMP = :s"), {'s': restamp})
            db.session.execute(text("DELETE FROM dbo.RE WHERE RESTAMP = :s"), {'s': restamp})
            db.session.commit()
            return jsonify({'ok': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    # -----------------------------
    # Registo de tempos (Limpeza)
    # -----------------------------
    @app.route('/tempos_limpeza')
    @login_required
    def tempos_limpeza_page():
        try:
            can_assign = bool(getattr(current_user, 'ADMIN', False) or getattr(current_user, 'LPADMIN', False))
        except Exception:
            can_assign = False
        current_login = (getattr(current_user, 'LOGIN', '') or '').strip()
        return render_template(
            'tempos_limpeza.html',
            page_title='Limpezas de Hoje',
            can_assign=can_assign,
            current_login=current_login,
        )

    @app.route('/api/tempos_limpeza/hoje')
    @login_required
    def api_tempos_limpeza_hoje():
        """
        Lista tarefas de limpeza (ORIGEM='LP') do dia de hoje para o utilizador autenticado.
        """
        try:
            user = (getattr(current_user, 'LOGIN', '') or '').strip()
            if not user:
                return jsonify({'error': 'Utilizador inválido.'}), 400

            cols = set(
                r[0] for r in db.session.execute(
                    text("SELECT UPPER(COLUMN_NAME) FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = 'TAREFAS'")
                ).fetchall()
            )
            need = {'HORAINI', 'HORAFIM'}
            missing = sorted(list(need - cols))
            if missing:
                return jsonify({'error': f"Campos em falta na TAREFAS: {', '.join(missing)}"}), 400

            sql = text("""
                SELECT
                    T.TAREFASSTAMP,
                    CAST(T.DATA AS date) AS DATA,
                    ISNULL(T.HORA,'') AS HORA,
                    ISNULL(T.ALOJAMENTO,'') AS ALOJAMENTO,
                    ISNULL(T.TAREFA,'') AS TAREFA,
                    ISNULL(T.TRATADO,0) AS TRATADO,
                    ISNULL(T.HORAINI,'') AS HORAINI,
                    ISNULL(T.HORAFIM,'') AS HORAFIM
                FROM dbo.TAREFAS AS T
                WHERE LTRIM(RTRIM(ISNULL(T.ORIGEM,''))) = 'LP'
                  AND CAST(T.DATA AS date) = CAST(GETDATE() AS date)
                  AND ISNULL(T.UTILIZADOR,'') = :u
                ORDER BY ISNULL(T.HORA,''), ISNULL(T.ALOJAMENTO,''), T.TAREFASSTAMP
            """)
            rows = db.session.execute(sql, {'u': user}).mappings().all()
            out = []
            for r in rows:
                d = r.get('DATA')
                d = d.strftime('%Y-%m-%d') if isinstance(d, (date, datetime)) else (str(d) if d is not None else '')
                out.append({
                    'TAREFASSTAMP': r.get('TAREFASSTAMP') or '',
                    'DATA': d,
                    'HORA': r.get('HORA') or '',
                    'ALOJAMENTO': r.get('ALOJAMENTO') or '',
                    'TAREFA': r.get('TAREFA') or '',
                    'TRATADO': int(r.get('TRATADO') or 0),
                    'HORAINI': (r.get('HORAINI') or '').strip(),
                    'HORAFIM': (r.get('HORAFIM') or '').strip(),
                })
            return jsonify({'rows': out, 'count': len(out), 'user': user})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/historico_reservas')
    @login_required
    def historico_reservas_page():
        current_login = (getattr(current_user, 'LOGIN', '') or '').strip()
        can_filter = bool(getattr(current_user, 'ADMIN', False) or getattr(current_user, 'LPADMIN', False))
        return render_template(
            'historico_reservas.html',
            page_title='Histórico de Tarefas',
            current_login=current_login,
            can_filter=can_filter,
        )

    @app.route('/api/historico_reservas_users')
    @login_required
    def api_historico_reservas_users():
        try:
            is_admin = bool(getattr(current_user, 'ADMIN', False) or getattr(current_user, 'LPADMIN', False))
            if not is_admin:
                return jsonify({'users': []})

            sql = text("""
                SELECT LOGIN, NOME
                FROM US
                WHERE ISNULL(INATIVO,0)=0
                  AND LTRIM(RTRIM(ISNULL(EQUIPA,''))) <> ''
                ORDER BY NOME
            """)
            rows = db.session.execute(sql).fetchall()
            users = [{'login': r[0], 'nome': r[1]} for r in rows]
            return jsonify({'users': users})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/historico_reservas')
    @login_required
    def api_historico_reservas():
        """
        Lista tarefas (ORIGEM='LP') por data para o utilizador autenticado.
        """
        try:
            user = (getattr(current_user, 'LOGIN', '') or '').strip()
            if not user:
                return jsonify({'error': 'Utilizador inválido.'}), 400
            req_user = (request.args.get('user') or '').strip()
            is_admin = bool(getattr(current_user, 'ADMIN', False) or getattr(current_user, 'LPADMIN', False))
            if is_admin and req_user:
                user = req_user

            date_str = (request.args.get('data') or '').strip()
            if not date_str:
                date_str = datetime.now().strftime('%Y-%m-%d')
            try:
                dt = datetime.strptime(date_str, '%Y-%m-%d').date()
            except Exception:
                return jsonify({'error': 'Data inválida.'}), 400

            cols = set(
                r[0] for r in db.session.execute(
                    text("SELECT UPPER(COLUMN_NAME) FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = 'TAREFAS'")
                ).fetchall()
            )
            need = {'HORAINI', 'HORAFIM'}
            missing = sorted(list(need - cols))
            if missing:
                return jsonify({'error': f"Campos em falta na TAREFAS: {', '.join(missing)}"}), 400

            sql = text("""
                SELECT
                    T.TAREFASSTAMP,
                    CAST(T.DATA AS date) AS DATA,
                    ISNULL(T.HORA,'') AS HORA,
                    ISNULL(T.ALOJAMENTO,'') AS ALOJAMENTO,
                    ISNULL(T.TAREFA,'') AS TAREFA,
                    ISNULL(T.TRATADO,0) AS TRATADO,
                    ISNULL(T.HORAINI,'') AS HORAINI,
                    ISNULL(T.HORAFIM,'') AS HORAFIM,
                    ISNULL(AL.TIPOLOGIA,'') AS TIPOLOGIA
                FROM dbo.TAREFAS AS T
                LEFT JOIN dbo.AL AS AL ON LTRIM(RTRIM(AL.NOME)) = LTRIM(RTRIM(T.ALOJAMENTO))
                WHERE LTRIM(RTRIM(ISNULL(T.ORIGEM,''))) = 'LP'
                  AND CAST(T.DATA AS date) = :d
                  AND ISNULL(T.UTILIZADOR,'') = :u
                ORDER BY ISNULL(T.HORA,''), ISNULL(T.ALOJAMENTO,''), T.TAREFASSTAMP
            """)
            rows = db.session.execute(sql, {'u': user, 'd': dt}).mappings().all()
            out = []
            for r in rows:
                d = r.get('DATA')
                d = d.strftime('%Y-%m-%d') if isinstance(d, (date, datetime)) else (str(d) if d is not None else '')
                out.append({
                    'TAREFASSTAMP': r.get('TAREFASSTAMP') or '',
                    'DATA': d,
                    'HORA': r.get('HORA') or '',
                    'ALOJAMENTO': r.get('ALOJAMENTO') or '',
                    'TAREFA': r.get('TAREFA') or '',
                    'TRATADO': int(r.get('TRATADO') or 0),
                    'HORAINI': (r.get('HORAINI') or '').strip(),
                    'HORAFIM': (r.get('HORAFIM') or '').strip(),
                    'TIPOLOGIA': (r.get('TIPOLOGIA') or '').strip(),
                })
            return jsonify({'rows': out, 'count': len(out), 'user': user, 'data': dt.strftime('%Y-%m-%d')})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/bq_toggle', methods=['POST'])
    @login_required
    def api_bq_toggle():
        """
        Bloquear/desbloquear noite (BQ) para um alojamento e data.
        Body: { "alojamento": "...", "date": "YYYY-MM-DD", "action": "block|unblock" }
        """
        try:
            data = request.get_json(silent=True) or {}
            aloj = (data.get('alojamento') or '').strip()
            date_str = (data.get('date') or '').strip()
            action = (data.get('action') or '').strip().lower()
            if not aloj or not date_str or action not in ('block', 'unblock'):
                return jsonify({'error': 'Parâmetros inválidos.'}), 400
            try:
                dt = datetime.strptime(date_str, '%Y-%m-%d').date()
            except Exception:
                return jsonify({'error': 'Data inválida.'}), 400

            # Trata bloqueio existente quando desbloqueia
            if action == 'unblock':
                db.session.execute(text("""
                    UPDATE dbo.BQ
                    SET TRATADO = 1,
                        ANULADO = 1
                    WHERE (
                        LTRIM(RTRIM(ISNULL(ALOJAMENTO,''))) COLLATE Latin1_General_CI_AI = :a
                        OR LTRIM(RTRIM(ISNULL(NMAIRBNB,''))) COLLATE Latin1_General_CI_AI = :a
                    )
                      AND CAST([DATA] AS date) = :d
                      AND ISNULL(DESBLOQ,0) = 0
                """), {'a': aloj, 'd': dt})

                db.session.execute(text("""
                    INSERT INTO dbo.BQ (BQSTAMP, ALOJAMENTO, NMAIRBNB, [DATA], TRATADO, DESBLOQ)
                    VALUES (LEFT(CONVERT(varchar(36), NEWID()),25), :a, :a, :d, 0, 1)
                """), {'a': aloj, 'd': dt})
            else:
                # Se já existir bloqueio pendente, não duplica
                exists = db.session.execute(text("""
                    SELECT TOP 1 1
                    FROM dbo.BQ
                    WHERE LTRIM(RTRIM(ISNULL(ALOJAMENTO,''))) = :a
                      AND CAST([DATA] AS date) = :d
                      AND ISNULL(DESBLOQ,0) = 0
                      AND ISNULL(TRATADO,0) = 0
                      AND ISNULL(ANULADO,0) = 0
                """), {'a': aloj, 'd': dt}).fetchone()
                if not exists:
                    db.session.execute(text("""
                        INSERT INTO dbo.BQ (BQSTAMP, ALOJAMENTO, NMAIRBNB, [DATA], TRATADO, DESBLOQ)
                        VALUES (LEFT(CONVERT(varchar(36), NEWID()),25), :a, :a, :d, 0, 0)
                    """), {'a': aloj, 'd': dt})

            db.session.commit()
            return jsonify({'ok': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    # -----------------------------
    # Fecho Mensal
    # -----------------------------
    @app.route('/fecho_mensal')
    @login_required
    def fecho_mensal_page():
        today = date.today()
        if today.month == 1:
            default_ano = today.year - 1
            default_mes = 12
        else:
            default_ano = today.year
            default_mes = today.month - 1
        return render_template(
            'fecho_mensal.html',
            page_title='Fecho Mensal',
            fecho_ano=default_ano,
            fecho_mes=default_mes
        )

    @app.route('/api/fecho_mensal')
    @login_required
    def api_fecho_mensal():
        try:
            ano = int(request.args.get('ano') or date.today().year)
            mes = int(request.args.get('mes') or date.today().month)
            if mes < 1 or mes > 12:
                return jsonify({'error': 'Mês inválido'}), 400
            def _norm_cc(value):
                try:
                    return ''.join(str(value or '').strip().upper().split())
                except Exception:
                    return ''
            def _norm_sort(value):
                s = _norm_cc(value)
                return (s
                        .replace('Á', 'A').replace('À', 'A').replace('Â', 'A').replace('Ã', 'A')
                        .replace('É', 'E').replace('Ê', 'E')
                        .replace('Í', 'I')
                        .replace('Ó', 'O').replace('Ô', 'O').replace('Õ', 'O')
                        .replace('Ú', 'U')
                        .replace('Ç', 'C'))
            excluded_cc = {'ADMIN', 'CAIXAAB'}
            base_cc_rows = db.session.execute(text("""
                SELECT DISTINCT LTRIM(RTRIM(ISNULL(CCUSTO,''))) AS CCUSTO
                FROM v_cct
                WHERE LTRIM(RTRIM(ISNULL(CCUSTO,''))) <> ''
            """)).mappings().all()
            fam_name_rows = db.session.execute(text("""
                SELECT LTRIM(RTRIM(ref)) AS REF, MAX(ISNULL(nome,'')) AS NOME
                FROM v_stfami
                WHERE LTRIM(RTRIM(ref)) IN ('1','2','3','4','9')
                GROUP BY LTRIM(RTRIM(ref))
            """)).mappings().all()
            fam_labels = {str(r.get('REF') or '').strip(): str(r.get('NOME') or '').strip() for r in fam_name_rows}

            custos_rows = db.session.execute(text("""
                SELECT
                    LTRIM(RTRIM(ISNULL(CCUSTO,''))) AS CCUSTO,
                    LEFT(LTRIM(RTRIM(ISNULL(FAMILIA,''))),1) AS FAM1,
                    SUM(ISNULL(TOTAL,0)) AS VALOR
                FROM v_custo
                WHERE YEAR([DATA]) = :ano
                  AND MONTH([DATA]) = :mes
                  AND LEFT(LTRIM(RTRIM(ISNULL(FAMILIA,''))),1) IN ('1','2','3','4','9')
                GROUP BY
                    LTRIM(RTRIM(ISNULL(CCUSTO,''))),
                    LEFT(LTRIM(RTRIM(ISNULL(FAMILIA,''))),1)
            """), {'ano': ano, 'mes': mes}).mappings().all()

            cc_map = {}
            for r in base_cc_rows:
                cc = (r.get('CCUSTO') or '').strip()
                cc_key = _norm_cc(cc)
                if (not cc_key) or cc_key in excluded_cc:
                    continue
                cc_map.setdefault(cc_key, {'CCUSTO': cc, '1': 0.0, '2': 0.0, '3': 0.0, '4': 0.0, '9': 0.0})
            for r in custos_rows:
                cc = (r.get('CCUSTO') or '').strip()
                cc_key = _norm_cc(cc)
                fam = (r.get('FAM1') or '').strip()
                if not cc_key or fam not in ('1', '2', '3', '4', '9'):
                    continue
                node = cc_map.setdefault(cc_key, {'CCUSTO': cc, '1': 0.0, '2': 0.0, '3': 0.0, '4': 0.0, '9': 0.0})
                node[fam] += float(r.get('VALOR') or 0)

            tipo_rows = db.session.execute(text("""
                SELECT
                    LTRIM(RTRIM(ISNULL(CCUSTO,''))) AS CCUSTO,
                    UPPER(LTRIM(RTRIM(ISNULL(TIPO,'')))) AS TIPO,
                    MAX(LTRIM(RTRIM(ISNULL(TIPOLOGIA,'')))) AS TIPOLOGIA
                FROM dbo.AL
                WHERE LTRIM(RTRIM(ISNULL(CCUSTO,''))) <> ''
                  AND UPPER(LTRIM(RTRIM(ISNULL(TIPO,'')))) IN ('EXPLORACAO','GESTAO')
                GROUP BY LTRIM(RTRIM(ISNULL(CCUSTO,''))), UPPER(LTRIM(RTRIM(ISNULL(TIPO,''))))
            """)).mappings().all()
            tipo_map = {_norm_cc(r.get('CCUSTO')): (r.get('TIPO') or '').strip() for r in tipo_rows}
            tipologia_map = {_norm_cc(r.get('CCUSTO')): (r.get('TIPOLOGIA') or '').strip() for r in tipo_rows}

            limpezas_rows = db.session.execute(text("""
                SELECT
                    LTRIM(RTRIM(ISNULL(a.CCUSTO,''))) AS CCUSTO,
                    COUNT(1) AS LIMPEZAS
                FROM RS rs
                JOIN AL a
                  ON LTRIM(RTRIM(ISNULL(a.NOME,''))) COLLATE SQL_Latin1_General_CP1_CI_AI
                   = LTRIM(RTRIM(ISNULL(rs.ALOJAMENTO,''))) COLLATE SQL_Latin1_General_CP1_CI_AI
                WHERE YEAR(ISNULL(rs.DATAOUT, rs.DATAIN)) = :ano
                  AND MONTH(ISNULL(rs.DATAOUT, rs.DATAIN)) = :mes
                GROUP BY LTRIM(RTRIM(ISNULL(a.CCUSTO,'')))
            """), {'ano': ano, 'mes': mes}).mappings().all()
            limpezas_map = {_norm_cc(r.get('CCUSTO')): int(r.get('LIMPEZAS') or 0) for r in limpezas_rows}

            rows = []
            for cc_key, node in cc_map.items():
                if cc_key in excluded_cc:
                    continue
                row = {
                    'CCUSTO': node.get('CCUSTO') or '',
                    'TIPO': tipo_map.get(cc_key, ''),
                    'TIPOLOGIA': tipologia_map.get(cc_key, ''),
                    'LIMPEZAS': limpezas_map.get(cc_key, 0),
                    'F1': round(float(node['1'] or 0), 2),
                    'F2': round(float(node['2'] or 0), 2),
                    'F3': round(float(node['3'] or 0), 2),
                    'F4': round(float(node['4'] or 0), 2),
                    'F9': round(float(node['9'] or 0), 2),
                }
                row['TOTAL_MES'] = round(row['F1'] + row['F2'] + row['F3'] + row['F4'] - row['F9'], 2)
                rows.append(row)

            estrutura_order = {
                'SEDE': 0,
                'LIMPEZA': 1,
                'LIMPEAZ': 1,
                'LAVANDARIA': 2,
                'HELPDESK': 3,
                'MANUTENCAO': 4,
                'MENUTENCAO': 4
            }
            rows.sort(key=lambda x: (
                0 if x.get('TIPO') not in ('EXPLORACAO', 'GESTAO') else 1,
                estrutura_order.get(_norm_sort(x.get('CCUSTO')), 999),
                (x.get('TIPO') or ''),
                (x.get('CCUSTO') or '')
            ))

            totals = {
                'F1': round(sum(float(r['F1']) for r in rows), 2),
                'F2': round(sum(float(r['F2']) for r in rows), 2),
                'F3': round(sum(float(r['F3']) for r in rows), 2),
                'F4': round(sum(float(r['F4']) for r in rows), 2),
                'F9': round(sum(float(r['F9']) for r in rows), 2),
                'TOTAL_MES': round(
                    sum(float(r['F1']) + float(r['F2']) + float(r['F3']) + float(r['F4']) - float(r['F9']) for r in rows),
                    2
                ),
            }
            imp_rows = db.session.execute(text("""
                SELECT
                    UPPER(LTRIM(RTRIM(ISNULL(CCUSTOORI,'')))) AS CCUSTOORI,
                    LTRIM(RTRIM(ISNULL(FAMILIA,''))) AS FAMILIA,
                    COUNT(1) AS N
                FROM dbo.TXADM
                WHERE ANO = :ano
                  AND MES = :mes
                  AND LTRIM(RTRIM(ISNULL(FAMILIA,''))) IN ('4.1','4.2','4.3','4.4')
                GROUP BY UPPER(LTRIM(RTRIM(ISNULL(CCUSTOORI,'')))), LTRIM(RTRIM(ISNULL(FAMILIA,'')))
            """), {'ano': ano, 'mes': mes}).mappings().all()
            imp_status = {'SEDE': False, 'LIMPEZA': False, 'LAVANDARIA': False, 'HELPDESK': False}
            family_to_cc = {'4.1': 'SEDE', '4.2': 'LIMPEZA', '4.3': 'LAVANDARIA', '4.4': 'HELPDESK'}
            for r in imp_rows:
                fam = (r.get('FAMILIA') or '').strip()
                cc = (r.get('CCUSTOORI') or '').strip().upper()
                expected_cc = family_to_cc.get(fam)
                if expected_cc and cc == expected_cc and int(r.get('N') or 0) > 0:
                    imp_status[expected_cc] = True

            return jsonify({
                'ano': ano,
                'mes': mes,
                'rows': rows,
                'totals': totals,
                'fam_labels': fam_labels,
                'imput_status': imp_status
            })
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/fecho_mensal/imputar', methods=['POST'])
    @login_required
    def api_fecho_mensal_imputar():
        try:
            payload = request.get_json(silent=True) or {}
            ano = int(payload.get('ano') or date.today().year)
            mes = int(payload.get('mes') or date.today().month)
            ccusto_ori = (payload.get('ccusto_ori') or '').strip().upper()
            familia = (payload.get('familia') or '').strip()
            if mes < 1 or mes > 12:
                return jsonify({'error': 'Mês inválido'}), 400
            if not ccusto_ori or not familia:
                return jsonify({'error': 'Parâmetros inválidos'}), 400

            ccusto_norm = ''.join(ccusto_ori.upper().split())
            ccusto_norm_sql = """
                UPPER(
                    REPLACE(
                        REPLACE(
                            REPLACE(LTRIM(RTRIM(ISNULL(CCUSTO,''))), ' ', ''),
                            CHAR(9), ''
                        ),
                        CHAR(160), ''
                    )
                )
            """
            total_origem = db.session.execute(text("""
                SELECT SUM(ISNULL(TOTAL,0))
                FROM v_custo
                WHERE YEAR([DATA]) = :ano
                  AND MONTH([DATA]) = :mes
                  AND LEFT(REPLACE(LTRIM(RTRIM(ISNULL(FAMILIA,''))), ' ', ''), 1) = '4'
                  AND """ + ccusto_norm_sql + """ = :ccusto_norm
            """), {
                'ano': ano,
                'mes': mes,
                'ccusto_norm': ccusto_norm
            }).scalar() or 0
            total_origem = float(total_origem or 0)

            # Fallback: usa o total de custos (famílias 1..4) do centro no mês.
            # Mantém alinhamento com o que o utilizador vê na grelha (saldo/colunas).
            if abs(total_origem) < 0.005:
                total_origem = db.session.execute(text("""
                    SELECT SUM(ISNULL(TOTAL,0))
                    FROM v_custo
                    WHERE YEAR([DATA]) = :ano
                      AND MONTH([DATA]) = :mes
                      AND LEFT(REPLACE(LTRIM(RTRIM(ISNULL(FAMILIA,''))), ' ', ''), 1) IN ('1','2','3','4')
                      AND """ + ccusto_norm_sql + """ = :ccusto_norm
                """), {
                    'ano': ano,
                    'mes': mes,
                    'ccusto_norm': ccusto_norm
                }).scalar() or 0
                total_origem = float(total_origem or 0)

            if abs(total_origem) < 0.005:
                return jsonify({'ok': True, 'inserted': 0, 'message': 'Sem valor a imputar no centro de custo origem.'})

            base_rows = db.session.execute(text("""
                SELECT
                    LTRIM(RTRIM(ISNULL(a.CCUSTO,''))) AS CCUSTO_DEST,
                    SUM(
                        CASE UPPER(LTRIM(RTRIM(ISNULL(a.TIPOLOGIA,''))))
                            WHEN 'T2' THEN 90
                            WHEN 'T3' THEN 120
                            WHEN 'T4' THEN 150
                            ELSE 60
                        END
                    ) AS BASE_QTD
                FROM RS rs
                JOIN AL a
                  ON LTRIM(RTRIM(ISNULL(a.NOME,''))) COLLATE SQL_Latin1_General_CP1_CI_AI
                   = LTRIM(RTRIM(ISNULL(rs.ALOJAMENTO,''))) COLLATE SQL_Latin1_General_CP1_CI_AI
                WHERE YEAR(ISNULL(rs.DATAOUT, rs.DATAIN)) = :ano
                  AND MONTH(ISNULL(rs.DATAOUT, rs.DATAIN)) = :mes
                  AND UPPER(LTRIM(RTRIM(ISNULL(a.TIPO,'')))) IN ('GESTAO','EXPLORACAO')
                GROUP BY LTRIM(RTRIM(ISNULL(a.CCUSTO,'')))
                HAVING SUM(
                    CASE UPPER(LTRIM(RTRIM(ISNULL(a.TIPOLOGIA,''))))
                        WHEN 'T2' THEN 90
                        WHEN 'T3' THEN 120
                        WHEN 'T4' THEN 150
                        ELSE 60
                    END
                ) > 0
            """), {'ano': ano, 'mes': mes}).mappings().all()

            if not base_rows:
                return jsonify({'error': 'Sem base de checkouts/tipologia para imputar.'}), 400

            base_total = float(sum(float(r.get('BASE_QTD') or 0) for r in base_rows))
            if base_total <= 0:
                return jsonify({'error': 'Base total inválida para imputação.'}), 400

            # Recriar apenas a imputação deste botão (família/centro)
            db.session.execute(text("""
                DELETE FROM dbo.TXADM
                WHERE ANO = :ano AND MES = :mes
                  AND LTRIM(RTRIM(ISNULL(FAMILIA,''))) = :familia
                  AND UPPER(REPLACE(LTRIM(RTRIM(ISNULL(CCUSTOORI,''))), ' ', '')) = :ccusto_norm
            """), {
                'ano': ano,
                'mes': mes,
                'familia': familia,
                'ccusto_norm': ccusto_norm
            })

            values = []
            running = 0.0
            for idx, r in enumerate(base_rows):
                qtd = float(r.get('BASE_QTD') or 0)
                pct = (qtd / base_total) if base_total else 0.0
                if idx < len(base_rows) - 1:
                    valor = round(total_origem * pct, 2)
                    running += valor
                else:
                    valor = round(total_origem - running, 2)
                values.append({
                    'ano': ano,
                    'mes': mes,
                    'familia': familia,
                    'ccusto_ori': ccusto_ori,
                    'ccusto_dest': (r.get('CCUSTO_DEST') or '').strip(),
                    'valor': valor,
                    'base_qtd': qtd,
                    'base_total': base_total,
                    'percentagem': round(pct * 100.0, 6),
                    'gerado_por': getattr(current_user, 'LOGIN', '') or ''
                })

            db.session.execute(text("""
                INSERT INTO dbo.TXADM
                (ANO, MES, DATA_FECHO, FAMILIA, CCUSTOORI, CCUSTODEST, VALOR, CRITERIO, BASE_QTD, BASE_TOTAL, PERCENTAGEM, GERADO_EM, GERADO_POR)
                VALUES
                (:ano, :mes, CAST(GETDATE() AS date), :familia, :ccusto_ori, :ccusto_dest, :valor, 'CHECKOUT_TIPOLOGIA', :base_qtd, :base_total, :percentagem, GETDATE(), :gerado_por)
            """), values)

            db.session.commit()
            return jsonify({'ok': True, 'inserted': len(values), 'total_origem': round(total_origem, 2)})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/fecho_mensal/imputacoes_mes', methods=['DELETE', 'POST'])
    @login_required
    def api_fecho_mensal_imputacoes_mes():
        try:
            payload = request.get_json(silent=True) or {}
            ano = int(payload.get('ano') or date.today().year)
            mes = int(payload.get('mes') or date.today().month)
            if mes < 1 or mes > 12:
                return jsonify({'error': 'Mês inválido'}), 400
            result = db.session.execute(text("""
                DELETE FROM dbo.TXADM
                WHERE ANO = :ano
                  AND MES = :mes
                  AND LTRIM(RTRIM(ISNULL(FAMILIA,''))) LIKE '4.%'
            """), {'ano': ano, 'mes': mes})
            db.session.commit()
            return jsonify({'ok': True, 'deleted': int(result.rowcount or 0)})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    # -----------------------------
    # Imputação a proprietários (FO/MN)
    # -----------------------------
    @app.route('/imputacao_proprietarios')
    @login_required
    def imputacao_proprietarios_page():
        return render_template('imputacao_proprietarios.html', page_title='Imputação a Proprietários')

    @app.route('/api/imputacao_proprietarios')
    @login_required
    def api_imputacao_proprietarios():
        try:
            # FN (linhas) com CCUSTO de gestão
            fn_rows = db.session.execute(text("""
                SELECT
                    FN.FNSTAMP,
                    FN.FOSTAMP,
                    FN.REF,
                    FN.DESIGN,
                    FN.FNCCUSTO,
                    FN.ETILIQUIDO,
                    FN.EPV,
                    FN.IMPUTAR,
                    FN.IMPUTMES,
                    FN.IMPUTANO,
                    FN.IMPUTVALOR,
                    FN.IMPUTDESIGN,
                    FN.NIMPUTAR,
                    F.DATA AS DATA,
                    F.DOCNOME,
                    F.ADOC,
                    F.NOME AS NOME_FORN
                FROM dbo.FN AS FN
                JOIN dbo.FO AS F ON F.FOSTAMP = FN.FOSTAMP
                JOIN dbo.AL AS A ON LTRIM(RTRIM(A.NOME)) = LTRIM(RTRIM(FN.FNCCUSTO))
                WHERE ISNULL(A.TIPO,'') = 'GESTAO'
            """)).mappings().all()

            fn_by_fo = {}
            for r in fn_rows:
                fn_by_fo.setdefault(r.get('FOSTAMP'), []).append(r)

            # FO: incluir cabeçalho quando o CCUSTO é gestão OU tem linhas FN gestão
            fo_params = {}
            fo_filter = "ISNULL(A.TIPO,'') = 'GESTAO'"
            if fn_by_fo:
                keys = []
                for i, k in enumerate(fn_by_fo.keys()):
                    key = f"f{i}"
                    keys.append(f":{key}")
                    fo_params[key] = k
                fo_filter = f"({fo_filter} OR F.FOSTAMP IN ({','.join(keys)}))"

            fo_sql = text(f"""
                SELECT
                    F.FOSTAMP AS STAMP,
                    CAST(F.DATA AS date) AS DATA,
                    CONCAT(ISNULL(F.DOCNOME,''), ' ', ISNULL(F.ADOC,'')) AS DOC,
                    ISNULL(F.NOME,'') AS NOME,
                    ISNULL(F.CCUSTO,'') AS ALOJAMENTO,
                    ISNULL(F.ETOTAL,0) AS TOTAL,
                    ISNULL(F.ETTILIQ,0) AS BASE,
                    ISNULL(F.IMPUTAR,0) AS IMPUTAR,
                    ISNULL(F.IMPUTMES,0) AS IMPUTMES,
                    ISNULL(F.IMPUTANO,0) AS IMPUTANO,
                    ISNULL(F.IMPUTVALOR,0) AS IMPUTVALOR,
                    ISNULL(F.IMPUTDESIGN,'') AS IMPUTDESIGN,
                    ISNULL(F.NIMPUTAR,0) AS NIMPUTAR
                FROM dbo.FO AS F
                LEFT JOIN dbo.AL AS A
                  ON LTRIM(RTRIM(A.NOME)) = LTRIM(RTRIM(F.CCUSTO))
                WHERE {fo_filter}
            """)
            fo_rows = db.session.execute(fo_sql, fo_params).mappings().all()

            out = []
            for fo in fo_rows:
                fostamp = fo.get('STAMP')
                lines = fn_by_fo.get(fostamp, [])
                distinct_cc = {str(l.get('FNCCUSTO') or '').strip() for l in lines if (l.get('FNCCUSTO') or '').strip()}
                fo_cc = str(fo.get('ALOJAMENTO') or '').strip()
                show_lines = False
                if lines:
                    if len(distinct_cc) > 1 or (distinct_cc and (fo_cc not in distinct_cc)):
                        show_lines = True
                out.append({
                    'ORIGEM': 'FO',
                    'STAMP': fostamp,
                    'DATA': fo.get('DATA'),
                    'DOC': fo.get('DOC') or '',
                    'NOME': fo.get('NOME') or '',
                    'ALOJAMENTO': fo_cc,
                    'TOTAL': float(fo.get('TOTAL') or 0),
                    'BASE': float(fo.get('BASE') or 0),
                    'IMPUTAR': int(fo.get('IMPUTAR') or 0),
                    'IMPUTMES': int(fo.get('IMPUTMES') or 0),
                    'IMPUTANO': int(fo.get('IMPUTANO') or 0),
                    'IMPUTVALOR': float(fo.get('IMPUTVALOR') or 0),
                    'IMPUTDESIGN': fo.get('IMPUTDESIGN') or '',
                    'NIMPUTAR': int(fo.get('NIMPUTAR') or 0),
                })
                if show_lines:
                    for l in lines:
                        base = l.get('EPV')
                        if base is None:
                            base = l.get('ETILIQUIDO') or 0
                        out.append({
                            'ORIGEM': 'FN',
                            'STAMP': l.get('FNSTAMP'),
                            'FOSTAMP': fostamp,
                            'DATA': l.get('DATA'),
                            'DOC': f"Linha: {l.get('REF') or ''} - {l.get('DESIGN') or ''}",
                            'NOME': l.get('NOME_FORN') or '',
                            'ALOJAMENTO': l.get('FNCCUSTO') or '',
                            'TOTAL': float(base or 0),
                            'BASE': float(base or 0),
                            'IMPUTAR': int(l.get('IMPUTAR') or 0),
                            'IMPUTMES': int(l.get('IMPUTMES') or 0),
                            'IMPUTANO': int(l.get('IMPUTANO') or 0),
                            'IMPUTVALOR': float(l.get('IMPUTVALOR') or 0),
                            'IMPUTDESIGN': l.get('IMPUTDESIGN') or '',
                            'NIMPUTAR': int(l.get('NIMPUTAR') or 0),
                        })

            # MN (manutenções) gestão
            mn_rows = db.session.execute(text("""
                SELECT
                    M.MNSTAMP AS STAMP,
                    CAST(M.DATA AS date) AS DATA,
                    ISNULL(M.INCIDENCIA,'') AS DOC,
                    ISNULL(M.NOME,'') AS NOME,
                    ISNULL(M.ALOJAMENTO,'') AS ALOJAMENTO,
                    ISNULL(M.IMPUTAR,0) AS IMPUTAR,
                    ISNULL(M.IMPUTMES,0) AS IMPUTMES,
                    ISNULL(M.IMPUTANO,0) AS IMPUTANO,
                    ISNULL(M.IMPUTVALOR,0) AS IMPUTVALOR,
                    ISNULL(M.IMPUTDESIGN,'') AS IMPUTDESIGN,
                    ISNULL(M.NIMPUTAR,0) AS NIMPUTAR
                FROM dbo.MN AS M
                JOIN dbo.AL AS A
                  ON LTRIM(RTRIM(A.NOME)) = LTRIM(RTRIM(M.ALOJAMENTO))
                WHERE ISNULL(A.TIPO,'') = 'GESTAO'
            """)).mappings().all()
            for r in mn_rows:
                out.append({
                    'ORIGEM': 'MN',
                    'STAMP': r.get('STAMP'),
                    'DATA': r.get('DATA'),
                    'DOC': r.get('DOC') or '',
                    'NOME': r.get('NOME') or '',
                    'ALOJAMENTO': r.get('ALOJAMENTO') or '',
                    'TOTAL': 0.0,
                    'BASE': 0.0,
                    'IMPUTAR': int(r.get('IMPUTAR') or 0),
                    'IMPUTMES': int(r.get('IMPUTMES') or 0),
                    'IMPUTANO': int(r.get('IMPUTANO') or 0),
                    'IMPUTVALOR': float(r.get('IMPUTVALOR') or 0),
                    'IMPUTDESIGN': r.get('IMPUTDESIGN') or '',
                    'NIMPUTAR': int(r.get('NIMPUTAR') or 0),
                })

            return jsonify({'rows': out})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/imputacao_proprietarios/save', methods=['POST'])
    @login_required
    def api_imputacao_proprietarios_save():
        try:
            payload = request.get_json(silent=True) or {}
            rows = payload.get('rows') or []
            if not isinstance(rows, list):
                return jsonify({'error': 'Formato inválido.'}), 400

            for r in rows:
                origem = (r.get('origem') or r.get('ORIGEM') or '').upper()
                stamp = (r.get('stamp') or r.get('STAMP') or '').strip()
                if not origem or not stamp:
                    continue
                data = {
                    'IMPUTAR': 1 if int(r.get('imputar') or 0) else 0,
                    'NIMPUTAR': 1 if int(r.get('nimputar') or 0) else 0,
                    'IMPUTMES': int(r.get('imputmes') or 0),
                    'IMPUTANO': int(r.get('imputano') or 0),
                    'IMPUTVALOR': float(r.get('imputvalor') or 0),
                    'IMPUTDESIGN': (r.get('imputdesign') or '').strip()[:60],
                    'STAMP': stamp
                }
                if origem == 'FO':
                    db.session.execute(text("""
                        UPDATE dbo.FO
                        SET IMPUTAR = :IMPUTAR,
                            NIMPUTAR = :NIMPUTAR,
                            IMPUTMES = :IMPUTMES,
                            IMPUTANO = :IMPUTANO,
                            IMPUTVALOR = :IMPUTVALOR,
                            IMPUTDESIGN = :IMPUTDESIGN
                        WHERE FOSTAMP = :STAMP
                    """), data)
                elif origem == 'FN':
                    db.session.execute(text("""
                        UPDATE dbo.FN
                        SET IMPUTAR = :IMPUTAR,
                            NIMPUTAR = :NIMPUTAR,
                            IMPUTMES = :IMPUTMES,
                            IMPUTANO = :IMPUTANO,
                            IMPUTVALOR = :IMPUTVALOR,
                            IMPUTDESIGN = :IMPUTDESIGN
                        WHERE FNSTAMP = :STAMP
                    """), data)
                elif origem == 'MN':
                    db.session.execute(text("""
                        UPDATE dbo.MN
                        SET IMPUTAR = :IMPUTAR,
                            NIMPUTAR = :NIMPUTAR,
                            IMPUTMES = :IMPUTMES,
                            IMPUTANO = :IMPUTANO,
                            IMPUTVALOR = :IMPUTVALOR,
                            IMPUTDESIGN = :IMPUTDESIGN
                        WHERE MNSTAMP = :STAMP
                    """), data)

            db.session.commit()
            return jsonify({'ok': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    # -----------------------------
    # Processamento Mensal (DM)
    # -----------------------------
    @app.route('/processamento_mensal')
    @login_required
    def processamento_mensal_page():
        return render_template('processamento_mensal.html', page_title='Processamento Mensal')

    @app.route('/api/processamento_mensal')
    @login_required
    def api_processamento_mensal():
        try:
            ano = int(request.args.get('ano') or datetime.now().year)
            mes = int(request.args.get('mes') or (datetime.now().month))
            dm_cols = set(r[0] for r in db.session.execute(
                text("SELECT UPPER(COLUMN_NAME) FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME='DM'")
            ).fetchall())
            has_com = 'COMISSOES' in dm_cols
            has_imp = 'IMPUTACOES' in dm_cols
            has_tot = 'TOTAL' in dm_cols
            sql = text("""
                SELECT
                    DMSTAMP, ANO, MES, NO, NOME, FATURATG, FTVALOR, FDATA,
                    DOSSIER, BOVALOR, DATAOBRA, FTFILE, ENVIADO
                    {extra_cols}
                FROM dbo.DM
                WHERE ANO = :ano AND MES = :mes
                ORDER BY NOME, NO, DOSSIER
            """.format(extra_cols=(
                (", COMISSOES" if has_com else "") +
                (", IMPUTACOES" if has_imp else "") +
                (", TOTAL" if has_tot else "")
            )))
            rows = db.session.execute(sql, {'ano': ano, 'mes': mes}).mappings().all()
            out = []
            for r in rows:
                out.append({
                    'DMSTAMP': r.get('DMSTAMP') or '',
                    'ANO': int(r.get('ANO') or 0),
                    'MES': int(r.get('MES') or 0),
                    'NO': r.get('NO') or '',
                    'NOME': r.get('NOME') or '',
                    'FATURATG': r.get('FATURATG') or '',
                    'FTVALOR': float(r.get('FTVALOR') or 0),
                    'FDATA': r.get('FDATA'),
                    'DOSSIER': r.get('DOSSIER') or '',
                    'BOVALOR': float(r.get('BOVALOR') or 0),
                    'DATAOBRA': r.get('DATAOBRA'),
                    'FTFILE': r.get('FTFILE') or '',
                    'ENVIADO': int(r.get('ENVIADO') or 0),
                    'COMISSOES': float(r.get('COMISSOES') or 0) if has_com else 0,
                    'IMPUTACOES': float(r.get('IMPUTACOES') or 0) if has_imp else 0,
                    'TOTAL': float(r.get('TOTAL') or 0) if has_tot else 0,
                })
            return jsonify({'rows': out})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/processamento_mensal/fatura_info')
    @login_required
    def api_processamento_mensal_fatura_info():
        try:
            stamp = (request.args.get('stamp') or '').strip()
            if not stamp:
                return jsonify({'error': 'DMSTAMP obrigatório'}), 400
            dm = db.session.execute(text("""
                SELECT DMSTAMP, ANO, MES, NO, FATURATG, FTVALOR
                FROM dbo.DM
                WHERE DMSTAMP = :s
            """), {'s': stamp}).mappings().first()
            if not dm:
                return jsonify({'error': 'Registo DM não encontrado'}), 404

            targets = db.session.execute(text("""
                SELECT DMSTAMP, ANO, MES
                FROM dbo.DM
                WHERE NO = :no
                  AND (LTRIM(RTRIM(ISNULL(FATURATG,''))) = '')
                  AND DMSTAMP <> :s
                ORDER BY ANO, MES
            """), {'no': dm['NO'], 's': stamp}).mappings().all()
            t = [{'stamp': r['DMSTAMP'], 'label': f"{int(r['MES']):02d}/{int(r['ANO'])}"} for r in targets]
            periodo = f"{int(dm['MES']):02d}/{int(dm['ANO'])}"
            return jsonify({
                'fatura': dm.get('FATURATG') or '',
                'valor': float(dm.get('FTVALOR') or 0),
                'periodo': periodo,
                'targets': t
            })
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/processamento_mensal/mover_fatura', methods=['POST'])
    @login_required
    def api_processamento_mensal_mover_fatura():
        try:
            payload = request.get_json(silent=True) or {}
            from_stamp = (payload.get('from_stamp') or '').strip()
            to_stamp = (payload.get('to_stamp') or '').strip()
            if not from_stamp or not to_stamp:
                return jsonify({'error': 'Parâmetros inválidos'}), 400

            src = db.session.execute(text("""
                SELECT DMSTAMP, FATURATG, FTVALOR, FDATA, FTFILE, ENVIADO
                FROM dbo.DM
                WHERE DMSTAMP = :s
            """), {'s': from_stamp}).mappings().first()
            if not src:
                return jsonify({'error': 'Registo origem não encontrado'}), 404

            db.session.execute(text("""
                UPDATE dbo.DM
                SET FATURATG = :fat, FTVALOR = :val, FDATA = :fdata, FTFILE = :file, ENVIADO = :env
                WHERE DMSTAMP = :s
            """), {
                'fat': src.get('FATURATG') or '',
                'val': float(src.get('FTVALOR') or 0),
                'fdata': src.get('FDATA') or datetime(1900, 1, 1),
                'file': src.get('FTFILE') or '',
                'env': int(src.get('ENVIADO') or 0),
                's': to_stamp
            })

            db.session.execute(text("""
                UPDATE dbo.DM
                SET FATURATG = '', FTVALOR = 0, FDATA = :fdata, FTFILE = '', ENVIADO = 0
                WHERE DMSTAMP = :s
            """), {'fdata': datetime(1900, 1, 1), 's': from_stamp})

            db.session.commit()
            return jsonify({'ok': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/processamento_mensal/clientes_disponiveis')
    @login_required
    def api_processamento_mensal_clientes_disponiveis():
        try:
            ano = int(request.args.get('ano') or datetime.now().year)
            mes = int(request.args.get('mes') or datetime.now().month)
            rows = db.session.execute(text("""
                SELECT CL.NO, CL.NOME
                FROM dbo.CL AS CL
                WHERE ISNULL(CL.INATIVO,0)=0
                  AND NOT EXISTS (
                    SELECT 1 FROM dbo.DM
                    WHERE ANO = :ano AND MES = :mes AND NO = CL.NO
                  )
                ORDER BY CL.NOME
            """), {'ano': ano, 'mes': mes}).mappings().all()
            out = [{'NO': r.get('NO'), 'NOME': r.get('NOME') or ''} for r in rows]
            return jsonify({'rows': out})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/processamento_mensal/add_clientes', methods=['POST'])
    @login_required
    def api_processamento_mensal_add_clientes():
        try:
            payload = request.get_json(silent=True) or {}
            ano = int(payload.get('ano') or datetime.now().year)
            mes = int(payload.get('mes') or datetime.now().month)
            items = payload.get('items') or []
            if not isinstance(items, list):
                return jsonify({'error': 'Formato inválido'}), 400

            added = []
            for it in items:
                no = int(it.get('no') or 0)
                nome = (it.get('nome') or '').strip()[:60]
                if not no or not nome:
                    continue
                exists = db.session.execute(text("""
                    SELECT 1 FROM dbo.DM WHERE ANO = :ano AND MES = :mes AND NO = :no
                """), {'ano': ano, 'mes': mes, 'no': no}).fetchone()
                if exists:
                    continue
                stamp = db.session.execute(text("SELECT LEFT(CONVERT(varchar(36), NEWID()),25)")).scalar()
                db.session.execute(text("""
                    INSERT INTO dbo.DM
                    (DMSTAMP, ANO, MES, NO, NOME, FATURATG, FTVALOR, FDATA, DOSSIER, BOVALOR, DATAOBRA, FTFILE, ENVIADO)
                    VALUES
                    (:stamp, :ano, :mes, :no, :nome, '', 0, :fdata, '', 0, :fdata, '', 0)
                """), {'stamp': stamp, 'ano': ano, 'mes': mes, 'no': no, 'nome': nome, 'fdata': datetime(1900,1,1)})
                added.append({
                    'DMSTAMP': stamp,
                    'ANO': ano,
                    'MES': mes,
                    'NO': no,
                    'NOME': nome,
                    'FATURATG': '',
                    'FTVALOR': 0,
                    'FDATA': datetime(1900,1,1),
                    'DOSSIER': '',
                    'BOVALOR': 0,
                    'DATAOBRA': datetime(1900,1,1),
                    'FTFILE': '',
                    'ENVIADO': 0
                })
            db.session.commit()
            return jsonify({'ok': True, 'rows': added})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/processamento_mensal/delete', methods=['POST'])
    @login_required
    def api_processamento_mensal_delete():
        try:
            payload = request.get_json(silent=True) or {}
            stamp = (payload.get('stamp') or '').strip()
            if not stamp:
                return jsonify({'error': 'DMSTAMP obrigatório'}), 400
            row = db.session.execute(text("""
                SELECT DMSTAMP, FATURATG, DOSSIER
                FROM dbo.DM
                WHERE DMSTAMP = :s
            """), {'s': stamp}).mappings().first()
            if not row:
                return jsonify({'error': 'Registo não encontrado'}), 404
            if (row.get('FATURATG') or '').strip() or (row.get('DOSSIER') or '').strip():
                return jsonify({'error': 'Só é possível eliminar registos sem fatura e sem dossier.'}), 400
            db.session.execute(text("DELETE FROM dbo.DM WHERE DMSTAMP = :s"), {'s': stamp})
            db.session.commit()
            return jsonify({'ok': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/processamento_mensal/drilldown')
    @login_required
    def api_processamento_mensal_drilldown():
        try:
            stamp = (request.args.get('stamp') or '').strip()
            dtype = (request.args.get('type') or '').strip().lower()
            if not stamp or dtype not in ('comissoes', 'imputacoes'):
                return jsonify({'error': 'Parâmetros inválidos'}), 400
            dm = db.session.execute(text("""
                SELECT ANO, MES, NO, NOME
                FROM dbo.DM
                WHERE DMSTAMP = :s
            """), {'s': stamp}).mappings().first()
            if not dm:
                return jsonify({'error': 'Registo DM não encontrado'}), 404

            ano = int(dm['ANO'])
            mes = int(dm['MES'])
            no = int(dm['NO'])

            if dtype == 'comissoes':
                rows = db.session.execute(text("""
                    SELECT
                        RS.ALOJAMENTO AS ALOJAMENTO,
                        CAST(RS.DATAOUT AS date) AS DATAOUT,
                        CASE WHEN ISNULL(RS.CANCELADA,0) = 1 THEN 'Cancelada' ELSE '' END AS ESTADO,
                        ISNULL(RS.ESTADIA,0) AS ESTADIA,
                        ISNULL(RS.LIMPEZA,0) AS LIMPEZA,
                        ISNULL(RS.COMISSAO,0) AS COMISSAO,
                        ISNULL(RS.PCANCEL,0) AS PCANCEL,
                        ISNULL(AL.COMISSAO,0) AS COMISSAO_PERC,
                        ROUND(CASE WHEN ISNULL(AL.FTLIMPEZA,0) = 0 THEN
                          CASE WHEN ISNULL(RS.CANCELADA,0) = 1
                            THEN ISNULL(RS.PCANCEL,0)
                            ELSE ISNULL(RS.ESTADIA,0) + ISNULL(RS.LIMPEZA,0) - ISNULL(RS.COMISSAO,0)
                          END * (ISNULL(AL.COMISSAO,0) / 100.0)
                        ELSE
                          CASE WHEN ISNULL(RS.CANCELADA,0) = 1
                            THEN ISNULL(RS.PCANCEL,0)
                            ELSE ISNULL(RS.ESTADIA,0) - ISNULL(RS.COMISSAO,0)
                          END * (ISNULL(AL.COMISSAO,0) / 100.0) + ISNULL(RS.LIMPEZA,0)
                        END, 2) AS VALOR
                    FROM dbo.CL
                    JOIN dbo.AL ON LTRIM(RTRIM(ISNULL(AL.CLIENTE,''))) = LTRIM(RTRIM(ISNULL(CL.NOME,'')))
                    JOIN dbo.RS ON LTRIM(RTRIM(ISNULL(RS.ALOJAMENTO,''))) = LTRIM(RTRIM(ISNULL(AL.NOME,'')))
                    WHERE CL.NO = :no
                      AND YEAR(RS.DATAOUT) = :ano AND MONTH(RS.DATAOUT) = :mes
                    ORDER BY RS.DATAOUT, RS.ALOJAMENTO
                """), {'no': no, 'ano': ano, 'mes': mes}).mappings().all()
                cols = ['ALOJAMENTO', 'DATAOUT', 'ESTADO', 'ESTADIA', 'LIMPEZA', 'COMISSAO', 'PCANCEL', 'COMISSAO_PERC', 'VALOR']
                out = [{c: r.get(c) for c in cols} for r in rows]
                return jsonify({'columns': cols, 'rows': out})

            # imputações
            im_cols = set(r[0] for r in db.session.execute(
                text("SELECT UPPER(COLUMN_NAME) FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME='IM'")
            ).fetchall())
            value_col = 'VALOR' if 'VALOR' in im_cols else ('IMPUTVALOR' if 'IMPUTVALOR' in im_cols else ('TOTAL' if 'TOTAL' in im_cols else None))
            ano_col = 'ANO' if 'ANO' in im_cols else None
            mes_col = 'MES' if 'MES' in im_cols else None
            no_col = 'NO' if 'NO' in im_cols else ('CLIENTE' if 'CLIENTE' in im_cols else None)

            if not (value_col and ano_col and mes_col and no_col):
                return jsonify({'columns': [], 'rows': []})

            if no_col == 'NO':
                sql = text(f"""
                    SELECT {no_col} AS NO, {value_col} AS VALOR
                    FROM dbo.IM
                    WHERE {ano_col} = :ano AND {mes_col} = :mes AND {no_col} = :no
                    ORDER BY {no_col}
                """)
                rows = db.session.execute(sql, {'ano': ano, 'mes': mes, 'no': no}).mappings().all()
                cols = ['NO', 'VALOR']
                out = [{c: r.get(c) for c in cols} for r in rows]
                return jsonify({'columns': cols, 'rows': out})
            else:
                sql = text(f"""
                    SELECT IM.{no_col} AS CLIENTE, IM.{value_col} AS VALOR
                    FROM dbo.IM AS IM
                    JOIN dbo.CL AS CL ON LTRIM(RTRIM(ISNULL(CL.NOME,''))) = LTRIM(RTRIM(ISNULL(IM.{no_col},'')))
                    WHERE IM.{ano_col} = :ano AND IM.{mes_col} = :mes AND CL.NO = :no
                    ORDER BY IM.{no_col}
                """)
                rows = db.session.execute(sql, {'ano': ano, 'mes': mes, 'no': no}).mappings().all()
                cols = ['CLIENTE', 'VALOR']
                out = [{c: r.get(c) for c in cols} for r in rows]
                return jsonify({'columns': cols, 'rows': out})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/processamento_mensal/calcular', methods=['POST'])
    @login_required
    def api_processamento_mensal_calcular():
        try:
            payload = request.get_json(silent=True) or {}
            ano = int(payload.get('ano') or datetime.now().year)
            mes = int(payload.get('mes') or datetime.now().month)

            # Comissões por cliente
            com_rows = db.session.execute(text("""
                SELECT DM.NO,
                       SUM(
                         CASE WHEN ISNULL(AL.FTLIMPEZA,0) = 0 THEN
                           CASE WHEN ISNULL(RS.CANCELADA,0) = 1
                             THEN ISNULL(RS.PCANCEL,0)
                             ELSE ISNULL(RS.ESTADIA,0) + ISNULL(RS.LIMPEZA,0) - ISNULL(RS.COMISSAO,0)
                           END * (ISNULL(AL.COMISSAO,0) / 100.0)
                         ELSE
                           CASE WHEN ISNULL(RS.CANCELADA,0) = 1
                             THEN ISNULL(RS.PCANCEL,0)
                             ELSE ISNULL(RS.ESTADIA,0) - ISNULL(RS.COMISSAO,0)
                           END * (ISNULL(AL.COMISSAO,0) / 100.0) + ISNULL(RS.LIMPEZA,0)
                         END
                       ) AS COMISSOES
                FROM dbo.DM
                JOIN dbo.CL ON CL.NO = DM.NO
                JOIN dbo.AL ON LTRIM(RTRIM(ISNULL(AL.CLIENTE,''))) = LTRIM(RTRIM(ISNULL(CL.NOME,'')))
                JOIN dbo.RS ON LTRIM(RTRIM(ISNULL(RS.ALOJAMENTO,''))) = LTRIM(RTRIM(ISNULL(AL.NOME,'')))
                WHERE DM.ANO = :ano AND DM.MES = :mes
                  AND YEAR(RS.DATAOUT) = :ano AND MONTH(RS.DATAOUT) = :mes
                GROUP BY DM.NO
            """), {'ano': ano, 'mes': mes}).mappings().all()
            com_map = {int(r['NO']): float(r.get('COMISSOES') or 0) for r in com_rows if r.get('NO') is not None}

            # Imputações por cliente (tabela IM)
            im_cols = set(r[0] for r in db.session.execute(
                text("SELECT UPPER(COLUMN_NAME) FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME='IM'")
            ).fetchall())
            value_col = 'VALOR' if 'VALOR' in im_cols else ('IMPUTVALOR' if 'IMPUTVALOR' in im_cols else ('TOTAL' if 'TOTAL' in im_cols else None))
            ano_col = 'ANO' if 'ANO' in im_cols else None
            mes_col = 'MES' if 'MES' in im_cols else None
            no_col = 'NO' if 'NO' in im_cols else ('CLIENTE' if 'CLIENTE' in im_cols else None)

            imp_map = {}
            if value_col and ano_col and mes_col and no_col:
                if no_col == 'NO':
                    sql_im = text(f"""
                        SELECT NO, SUM(ISNULL({value_col},0)) AS IMPUTACOES
                        FROM dbo.IM
                        WHERE {ano_col} = :ano AND {mes_col} = :mes
                        GROUP BY NO
                    """)
                    im_rows = db.session.execute(sql_im, {'ano': ano, 'mes': mes}).mappings().all()
                    imp_map = {int(r['NO']): float(r.get('IMPUTACOES') or 0) for r in im_rows if r.get('NO') is not None}
                else:
                    sql_im = text(f"""
                        SELECT CL.NO AS NO, SUM(ISNULL(IM.{value_col},0)) AS IMPUTACOES
                        FROM dbo.IM AS IM
                        JOIN dbo.CL AS CL ON LTRIM(RTRIM(ISNULL(CL.NOME,''))) = LTRIM(RTRIM(ISNULL(IM.{no_col},'')))
                        WHERE IM.{ano_col} = :ano AND IM.{mes_col} = :mes
                        GROUP BY CL.NO
                    """)
                    im_rows = db.session.execute(sql_im, {'ano': ano, 'mes': mes}).mappings().all()
                    imp_map = {int(r['NO']): float(r.get('IMPUTACOES') or 0) for r in im_rows if r.get('NO') is not None}

            rows = db.session.execute(text("""
                SELECT DMSTAMP, NO FROM dbo.DM WHERE ANO = :ano AND MES = :mes
            """), {'ano': ano, 'mes': mes}).mappings().all()

            dm_cols = set(r[0] for r in db.session.execute(
                text("SELECT UPPER(COLUMN_NAME) FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME='DM'")
            ).fetchall())
            has_com = 'COMISSOES' in dm_cols
            has_imp = 'IMPUTACOES' in dm_cols
            has_tot = 'TOTAL' in dm_cols

            out = []
            for r in rows:
                no = int(r.get('NO') or 0)
                com = float(com_map.get(no, 0))
                imp = float(imp_map.get(no, 0))
                total = com + imp
                if has_com or has_imp or has_tot:
                    db.session.execute(text("""
                        UPDATE dbo.DM
                        SET {set_cols}
                        WHERE DMSTAMP = :s
                    """.format(set_cols=",".join(
                        [c for c in [
                            ("COMISSOES = :com" if has_com else None),
                            ("IMPUTACOES = :imp" if has_imp else None),
                            ("TOTAL = :tot" if has_tot else None)
                        ] if c]
                    ))), {'com': com, 'imp': imp, 'tot': total, 's': r['DMSTAMP']})
                out.append({
                    'DMSTAMP': r['DMSTAMP'],
                    'COMISSOES': com,
                    'IMPUTACOES': imp,
                    'TOTAL': total
                })
            db.session.commit()
            return jsonify({'rows': out})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500
    @app.route('/api/processamento_mensal/fetch_dossier', methods=['POST'])
    @login_required
    def api_processamento_mensal_fetch_dossier():
        try:
            payload = request.get_json(silent=True) or {}
            ano = int(payload.get('ano') or datetime.now().year)
            mes = int(payload.get('mes') or datetime.now().month)

            dms = db.session.execute(text("""
                SELECT DMSTAMP, ANO, MES, NO
                FROM dbo.DM
                WHERE ANO = :ano AND MES = :mes
            """), {'ano': ano, 'mes': mes}).mappings().all()

            updated = 0
            for dm in dms:
                row = db.session.execute(text("""
                    SELECT TOP 1
                        NMDOS, OBRANO, DATAOBRA,
                        ROUND(ETOTALDEB * 1.23, 2) AS VALOR
                    FROM guest_spa_tur..BO
                    WHERE YEAR(DATAOBRA) = :ano
                      AND MONTH(DATAOBRA) = :mes
                      AND NDOS = 15
                      AND NO = :no
                    ORDER BY DATAOBRA DESC, OBRANO DESC
                """), {'ano': dm['ANO'], 'mes': dm['MES'], 'no': dm['NO']}).mappings().first()
                if not row:
                    continue
                dossier = f"{row.get('NMDOS') or ''} nº {int(row.get('OBRANO') or 0)}"
                db.session.execute(text("""
                    UPDATE dbo.DM
                    SET DOSSIER = :dos,
                        DATAOBRA = :dataobra,
                        BOVALOR = :valor
                    WHERE DMSTAMP = :s
                """), {
                    'dos': dossier,
                    'dataobra': row.get('DATAOBRA'),
                    'valor': float(row.get('VALOR') or 0),
                    's': dm['DMSTAMP']
                })
                updated += 1

            db.session.commit()
            return jsonify({'updated': updated})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/tempos_limpeza/start', methods=['POST'])
    @login_required
    def api_tempos_limpeza_start():
        """
        Define TAREFAS.HORAINI (HH:MM) para a tarefa, se ainda não estiver definida.
        Body: { "id": "<TAREFASSTAMP>" }
        """
        try:
            body = request.get_json(silent=True) or {}
            tid = (body.get('id') or '').strip()
            if not tid:
                return jsonify({'error': 'id obrigatório'}), 400

            user = (getattr(current_user, 'LOGIN', '') or '').strip()
            if not user:
                return jsonify({'error': 'Utilizador inválido.'}), 400

            # Só permite iniciar tarefas do próprio utilizador e de hoje
            exists = db.session.execute(text("""
                SELECT 1 AS X
                FROM dbo.TAREFAS
                WHERE TAREFASSTAMP = :id
                  AND LTRIM(RTRIM(ISNULL(ORIGEM,''))) = 'LP'
                  AND CAST(DATA AS date) = CAST(GETDATE() AS date)
                  AND ISNULL(UTILIZADOR,'') = :u
            """), {'id': tid, 'u': user}).fetchone()
            if not exists:
                return jsonify({'error': 'Tarefa não encontrada.'}), 404

            # Hora atual HH:MM, sem segundos
            now_hm = db.session.execute(text("SELECT CONVERT(varchar(5), GETDATE(), 108)")).fetchone()[0]

            # Não sobrescreve se já existir
            db.session.execute(text("""
                UPDATE dbo.TAREFAS
                SET HORAINI = CASE WHEN LTRIM(RTRIM(ISNULL(HORAINI,''))) = '' THEN :hm ELSE HORAINI END
                WHERE TAREFASSTAMP = :id
            """), {'hm': now_hm, 'id': tid})
            db.session.commit()
            return jsonify({'ok': True, 'id': tid, 'HORAINI': now_hm})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/tempos_limpeza/stop', methods=['POST'])
    @login_required
    def api_tempos_limpeza_stop():
        """
        Define TAREFAS.HORAFIM (HH:MM) e marca como tratada.
        Também garante HORAINI se estiver vazio.
        Body: { "id": "<TAREFASSTAMP>" }
        """
        try:
            body = request.get_json(silent=True) or {}
            tid = (body.get('id') or '').strip()
            if not tid:
                return jsonify({'error': 'id obrigatório'}), 400

            user = (getattr(current_user, 'LOGIN', '') or '').strip()
            if not user:
                return jsonify({'error': 'Utilizador inválido.'}), 400

            exists = db.session.execute(text("""
                SELECT 1 AS X
                FROM dbo.TAREFAS
                WHERE TAREFASSTAMP = :id
                  AND LTRIM(RTRIM(ISNULL(ORIGEM,''))) = 'LP'
                  AND CAST(DATA AS date) = CAST(GETDATE() AS date)
                  AND ISNULL(UTILIZADOR,'') = :u
            """), {'id': tid, 'u': user}).fetchone()
            if not exists:
                return jsonify({'error': 'Tarefa não encontrada.'}), 404

            now_hm = db.session.execute(text("SELECT CONVERT(varchar(5), GETDATE(), 108)")).fetchone()[0]

            db.session.execute(text("""
                UPDATE dbo.TAREFAS
                SET
                  HORAINI = CASE WHEN LTRIM(RTRIM(ISNULL(HORAINI,''))) = '' THEN :hm ELSE HORAINI END,
                  HORAFIM = CASE WHEN LTRIM(RTRIM(ISNULL(HORAFIM,''))) = '' THEN :hm ELSE HORAFIM END,
                  TRATADO = 1,
                  NMTRATADO = :u,
                  DTTRATADO = CAST(GETDATE() AS date)
                WHERE TAREFASSTAMP = :id
            """), {'hm': now_hm, 'u': user, 'id': tid})
            db.session.commit()
            return jsonify({'ok': True, 'id': tid, 'HORAFIM': now_hm})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    # -----------------------------
    # Configuração de Escalas (ES/ESL)
    # -----------------------------
    @app.route('/escalas')
    @login_required
    def escalas_page():
        return render_template('escalas.html', page_title='Configuração de Escalas')

    @app.route('/api/escalas', methods=['GET', 'POST'])
    @login_required
    def api_escalas():
        try:
            if request.method == 'GET':
                sql = text(
                    """
                    SELECT E.ESSTAMP, E.ESCALA, ISNULL(L.DIA, 0) AS DIA, ISNULL(L.FOLGA,0) AS FOLGA
                    FROM ES AS E
                    LEFT JOIN ESL AS L ON L.ESSTAMP = E.ESSTAMP
                    ORDER BY E.ESCALA, E.ESSTAMP, ISNULL(L.DIA, 0)
                    """
                )
                rows = db.session.execute(sql).fetchall()
                escalas = {}
                for r in rows:
                    es_id, nome, dia, folga = r[0], r[1], int(r[2] or 0), int(r[3] or 0)
                    if es_id not in escalas:
                        escalas[es_id] = { 'id': es_id, 'escala': nome, 'dias': {}, 'max_dia': 0 }
                    if dia:
                        escalas[es_id]['dias'][dia] = folga
                        if dia > escalas[es_id]['max_dia']:
                            escalas[es_id]['max_dia'] = dia
                items = []
                max_dias = 0
                for es in escalas.values():
                    max_dias = max(max_dias, es['max_dia'])
                    items.append({ 'id': es['id'], 'escala': es['escala'], 'dias': es['dias'], 'total_dias': es['max_dia'] })
                return jsonify({ 'rows': items, 'max_dias': max_dias })

            # POST (create): { escala, dias }
            body = request.get_json(silent=True) or {}
            nome = (body.get('escala') or '').strip()
            try:
                dias = int(body.get('dias') or 0)
            except Exception:
                dias = 0
            if not nome:
                return jsonify({ 'error': 'Nome da escala obrigatório' }), 400
            if dias <= 0 or dias > 366:
                return jsonify({ 'error': 'Número de dias inválido' }), 400
            new_id_row = db.session.execute(text("SELECT LEFT(CONVERT(varchar(36), NEWID()), 25)")).fetchone()
            es_id = new_id_row[0]
            db.session.execute(text("INSERT INTO ES (ESSTAMP, ESCALA) VALUES (:id, :nome)"), { 'id': es_id, 'nome': nome })
            # Preenche dias com FOLGA=0
            ins = text("INSERT INTO ESL (ESLSTAMP, ESSTAMP, DIA, FOLGA) VALUES (LEFT(CONVERT(varchar(36), NEWID()), 25), :es, :dia, 0)")
            for d in range(1, dias+1):
                db.session.execute(ins, { 'es': es_id, 'dia': d })
            db.session.commit()
            return jsonify({ 'ok': True, 'id': es_id })
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/escalas/<es_id>', methods=['DELETE'])
    @login_required
    def api_escalas_delete(es_id):
        try:
            db.session.execute(text("DELETE FROM ESL WHERE ESSTAMP = :id"), { 'id': es_id })
            res = db.session.execute(text("DELETE FROM ES WHERE ESSTAMP = :id"), { 'id': es_id })
            db.session.commit()
            return jsonify({ 'ok': True, 'deleted': res.rowcount })
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/escalas/<es_id>/toggle', methods=['POST'])
    @login_required
    def api_escalas_toggle(es_id):
        try:
            body = request.get_json(silent=True) or {}
            try:
                dia = int(body.get('dia') or 0)
            except Exception:
                dia = 0
            if dia <= 0:
                return jsonify({ 'error': 'Dia inválido' }), 400
            # Tenta atualizar; se não existir, cria com FOLGA=1
            upd = text("UPDATE ESL SET FOLGA = CASE WHEN ISNULL(FOLGA,0)=1 THEN 0 ELSE 1 END WHERE ESSTAMP = :es AND DIA = :dia")
            res = db.session.execute(upd, { 'es': es_id, 'dia': dia })
            if res.rowcount == 0:
                db.session.execute(text("INSERT INTO ESL (ESLSTAMP, ESSTAMP, DIA, FOLGA) VALUES (LEFT(CONVERT(varchar(36), NEWID()), 25), :es, :dia, 1)"), { 'es': es_id, 'dia': dia })
                new_val = 1
            else:
                # fetch new value
                cur = db.session.execute(text("SELECT ISNULL(FOLGA,0) FROM ESL WHERE ESSTAMP=:es AND DIA=:dia"), { 'es': es_id, 'dia': dia }).fetchone()
                new_val = int(cur[0]) if cur else 0
            db.session.commit()
            return jsonify({ 'ok': True, 'folga': new_val })
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    # API: Detalhe de faturação por reserva para um alojamento
    @app.route('/api/performance/detalhe')
    @login_required
    def api_performance_detalhe():
        try:
            alojamento = (request.args.get('alojamento') or '').strip()
            data_inicio = (request.args.get('data_inicio') or '').strip()
            data_fim = (request.args.get('data_fim') or '').strip()
            if not alojamento or not data_inicio or not data_fim:
                return jsonify({'error': 'Parâmetros em falta'}), 400

            # Query: agrupa por reserva, soma valor, mostra uma linha por reserva
            sql = text(
                """
                SELECT
                    CAST(MIN(V.data) AS date) AS data_reserva,
                    MAX(ISNULL(V.nome, ''))    AS hospede,
                    ISNULL(V.reserva, '')      AS reserva,
                    SUM(ISNULL(V.valor, 0))    AS total
                FROM v_diario_all AS V
                WHERE V.CCUSTO = :aloj
                  AND ISNULL(V.valor, 0) <> 0
                  AND V.data BETWEEN :data_inicio AND :data_fim
                GROUP BY ISNULL(V.reserva, '')
                ORDER BY data_reserva, reserva
                """
            )

            rows = db.session.execute(sql, {
                'aloj': alojamento,
                'data_inicio': data_inicio,
                'data_fim': data_fim
            }).mappings().all()

            out = []
            total_sum = 0.0
            for r in rows:
                data_res = r.get('data_reserva')
                if isinstance(data_res, (datetime, date)):
                    data_str = data_res.strftime('%Y-%m-%d')
                else:
                    data_str = str(data_res) if data_res is not None else ''
                val = float(r.get('total') or 0)
                total_sum += val
                out.append({
                    'data_reserva': data_str,
                    'hospede': r.get('hospede') or '',
                    'reserva': r.get('reserva') or '',
                    'valor': round(val, 2)
                })

            # Custos: agrupar por referência (REF), somando o valor no período
            sql_c = text(
                """
                SELECT ISNULL(C.ref,'') AS ref, SUM(ISNULL(C.valor,0)) AS valor
                FROM v_custo AS C
                WHERE C.ccusto = :aloj
                  AND C.data BETWEEN :data_inicio AND :data_fim
                GROUP BY ISNULL(C.ref,'')
                ORDER BY ISNULL(C.ref,'')
                """
            )
            c_rows_db = db.session.execute(sql_c, {
                'aloj': alojamento,
                'data_inicio': data_inicio,
                'data_fim': data_fim
            }).mappings().all()

            custos_rows = []
            custos_total = 0.0
            for r in c_rows_db:
                v = float(r.get('valor') or 0)
                custos_total += v
                custos_rows.append({
                    'ref': r.get('ref') or '',
                    'valor': round(v, 2)
                })

            return jsonify({
                'rows': out,
                'total': round(total_sum, 2),
                'custos': custos_rows,
                'total_custos': round(custos_total, 2),
                'resultado': round(total_sum - custos_total, 2)
            })
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    # API: Ocupação por dia (datas e totais) para um alojamento num período
    @app.route('/api/performance/ocupacao')
    @login_required
    def api_performance_ocupacao():
        try:
            alojamento = (request.args.get('alojamento') or '').strip()
            data_inicio = (request.args.get('data_inicio') or '').strip()
            data_fim = (request.args.get('data_fim') or '').strip()
            if not alojamento or not data_inicio or not data_fim:
                return jsonify({'error': 'Parâmetros em falta'}), 400

            sql = text(
                """
                SELECT CAST(V.data AS date) AS dia, SUM(ISNULL(V.valor,0)) AS total
                FROM v_diario_all V
                WHERE V.CCUSTO = :aloj
                  AND ISNULL(V.valor,0) <> 0
                  AND V.data BETWEEN :data_inicio AND :data_fim
                GROUP BY CAST(V.data AS date)
                ORDER BY dia
                """
            )
            rows = db.session.execute(sql, {
                'aloj': alojamento,
                'data_inicio': data_inicio,
                'data_fim': data_fim
            }).fetchall()
            dias = []
            mapa = {}
            for r in rows:
                d = r[0]
                tot = float(r[1] or 0)
                if isinstance(d, (date, datetime)):
                    key = d.strftime('%Y-%m-%d')
                else:
                    key = str(d)
                dias.append(key)
                mapa[key] = round(tot, 2)
            return jsonify({'dias': dias, 'mapa': mapa})
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    from sqlalchemy import text

    from flask_login import current_user

    @app.route('/newmn')
    @login_required
    def newmn():
        alojamentos = [row[0] for row in db.session.execute(text("SELECT NOME FROM AL ORDER BY 1")).fetchall()]
        users = [row[0] for row in db.session.execute(text("SELECT LOGIN FROM US ORDER BY 1")).fetchall()]
        utilizador = current_user.LOGIN
        return render_template('newmn.html', alojamentos=alojamentos, users=users, utilizador=utilizador, page_title='ManutenÃ§Ã£o')
    
    from sqlalchemy import text

    @app.route('/newfs')
    @login_required
    def newfs():
        # Ajusta os SELECTs para as tuas tabelas reais de alojamentos e utilizadores
        alojamentos = [row[0] for row in db.session.execute(text("SELECT NOME FROM AL ORDER BY 1")).fetchall()]
        users = [row[0] for row in db.session.execute(text("SELECT LOGIN FROM US ORDER BY 1")).fetchall()]
        utilizador = getattr(current_user, 'LOGIN', '')  # default seguro
        return render_template('newfs.html',
                            alojamentos=alojamentos,
                            users=users,
                            utilizador=utilizador,
                            page_title='Faltas')

    
    @app.route('/newanexo')
    @login_required
    def newanexo():
        table = request.args.get('table')
        rec = request.args.get('rec')
        if not table or not rec:
            abort(400, "Faltam parÃ¢metros")
        return render_template('newanexo.html', table=table, rec=rec, page_title='Anexos')

    from sqlalchemy import text

    @app.route('/planner/api/imprimir_etiquetas', methods=['POST'])
    @login_required
    def imprimir_etiquetas():
        data = request.args.get('date')
        if not data:
            return jsonify({'error': 'Data em falta'}), 400

        # Inserir todos os LP com DATA = data na tabela ET
        sql = text("""
            INSERT INTO ET (LPSTAMP, TRATADO)
            SELECT LPSTAMP, 0 FROM LP WHERE DATA = :data
        """)
        try:
            db.session.execute(sql, {'data': data})
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500


    @app.route('/api/etiqueta_rapida', methods=['POST'])
    @login_required
    def criar_etiqueta_rapida():
        lpstamp = request.json.get('lpstamp')
        if not lpstamp:
            return jsonify({'error': 'LPSTAMP em falta'}), 400
        try:
            # sÃ³ insere se nÃ£o existir ainda
            sql = text('INSERT INTO ET (LPSTAMP, TRATADO) VALUES (:lpstamp, 0)')
            db.session.execute(sql, {'lpstamp': lpstamp})
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500


    from flask import render_template, redirect, url_for

    @app.route('/profile')
    @login_required
    def profile():
        return render_template('profile.html', user=current_user, page_title='Perfil')


    @app.route('/api/profile/change_password', methods=['POST'])
    @login_required
    def change_password():
        data = request.json
        new_pwd = data.get('password', '').strip()
        if not new_pwd or len(new_pwd) < 4:
            return {'error': 'Password demasiado curta'}, 400

        from app import db  # ou usa db conforme jÃ¡ tens no teu projeto
        user = db.session.query(US).get(current_user.USSTAMP)
        if not user:
            return {'error': 'Utilizador nÃ£o encontrado'}, 404
        user.PASSWORD = new_pwd
        db.session.commit()
        return {'success': True}

    @app.route('/api/profile/save', methods=['POST'])
    @login_required
    def save_profile():
        try:
            data = request.get_json(force=True) or {}
            # Campos permitidos a serem atualizados no perfil
            allowed_fields = {'EMAIL', 'COR'}

            user = db.session.query(US).get(current_user.USSTAMP)
            if not user:
                return {'error': 'Utilizador nÃ£o encontrado'}, 404

            updated = {}
            for field in allowed_fields:
                if field in data:
                    setattr(user, field, data[field])
                    updated[field] = data[field]

            if not updated:
                return {'error': 'Sem alteraÃ§Ãµes vÃ¡lidas'}, 400

            db.session.commit()
            return jsonify({'success': True, 'updated': updated})
        except Exception as e:
            db.session.rollback()
            return {'error': str(e)}, 500

    @app.route('/api/profile/upload_photo', methods=['POST'])
    @login_required
    def upload_photo():
        try:
            from werkzeug.utils import secure_filename
            import uuid as _uuid
            photo = request.files.get('photo')
            if not photo or photo.filename == '':
                return {'error': 'Ficheiro em falta'}, 400
            fname = secure_filename(photo.filename)
            ext = os.path.splitext(fname)[1].lower()
            if ext not in ('.jpg', '.jpeg', '.png', '.gif', '.webp'):
                return {'error': 'Formato invÃ¡lido'}, 400

            target_dir = os.path.join(app.static_folder, 'images', 'profile')
            os.makedirs(target_dir, exist_ok=True)
            new_name = f"{_uuid.uuid4().hex}{ext}"
            save_path = os.path.join(target_dir, new_name)
            photo.save(save_path)

            # Atualiza caminho relativo sob /static
            rel_path = f"images/profile/{new_name}"
            user = db.session.query(US).get(current_user.USSTAMP)
            if not user:
                return {'error': 'Utilizador nÃ£o encontrado'}, 404
            user.FOTO = rel_path
            db.session.commit()

            return jsonify({'success': True, 'path': rel_path})
        except Exception as e:
            db.session.rollback()
            return {'error': str(e)}, 500

    # Monitor: lista de tarefas unificada com filtros simples
    @app.route('/generic/api/monitor_tasks')
    @app.route('/generic/api/monitor_tasks_filtered')
    @login_required
    def api_monitor_tasks():
        try:
            params = request.args or {}
            only_mine = params.get('only_mine') in ('1', 'true', 'True')
            start = params.get('start')
            end = params.get('end')
            # aceita tambÃ©m CSV alternativo (users, aloj, origins)
            utilizadores = params.getlist('UTILIZADOR') or ([] if params.get('UTILIZADOR') is None else [params.get('UTILIZADOR')])
            if not utilizadores and params.get('users'):
                utilizadores = [u.strip() for u in params.get('users').split(',') if u.strip() != '']
            alojamentos = params.getlist('ALOJAMENTO') or ([] if params.get('ALOJAMENTO') is None else [params.get('ALOJAMENTO')])
            if not alojamentos and params.get('aloj'):
                alojamentos = [a.strip() for a in params.get('aloj').split(',')]
            origens = params.getlist('ORIGEM') or ([] if params.get('ORIGEM') is None else [params.get('ORIGEM')])
            if not origens and params.get('origins'):
                origens = [o.strip() for o in params.get('origins').split(',')]
            # Regra: LPADMIN vê sempre FS. Não remove filtros existentes; apenas acrescenta FS.
            try:
                is_lp_admin = bool(getattr(current_user, 'LPADMIN', 0))
            except Exception:
                is_lp_admin = False
            if is_lp_admin and origens:
                if 'FS' not in {str(o).upper() for o in origens}:
                    origens.append('FS')

            # Defaults date window
            if not start:
                start = (date.today()).toordinal() - 30
                start = date.fromordinal(start).isoformat()
            if not end:
                end = (date.today()).toordinal() + 60
                end = date.fromordinal(end).isoformat()

            # Helper to get columns present in a table
            def get_columns(table_name: str):
                try:
                    rows = db.session.execute(text(
                        "SELECT UPPER(COLUMN_NAME) AS CN FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = :t"
                    ), { 't': table_name }).fetchall()
                    return { r[0] for r in rows }
                except Exception:
                    return set()

            def exists_table(table_name: str) -> bool:
                try:
                    r = db.session.execute(text(
                        "SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = :t"
                    ), { 't': table_name }).fetchone()
                    return r is not None
                except Exception:
                    return False

            # Build dynamic SELECT for each table, normalizing fields
            def build_select(table: str, origem_code: str):
                cols = get_columns(table)
                if not cols:
                    return None, {}
                # Choose STAMP
                stamp_col = next((c for c in cols if c.endswith('STAMP')), None)
                data_col = 'DATA' if 'DATA' in cols else None
                hora_col = 'HORA' if 'HORA' in cols else None
                aloj_col = 'ALOJAMENTO' if 'ALOJAMENTO' in cols else None
                user_col = 'UTILIZADOR' if 'UTILIZADOR' in cols else None
                tarefa_col = next((c for c in ('TAREFA','TITULO','ASSUNTO','DESCRICAO','DESCR') if c in cols), None)
                tratado_col = 'TRATADO' if 'TRATADO' in cols else None

                # SELECT list
                sel = [
                    f"CAST({stamp_col} AS varchar(50)) AS TAREFASSTAMP" if stamp_col else "NULL AS TAREFASSTAMP",
                    f"CAST({data_col} AS date) AS DATA" if data_col else "CAST(NULL AS date) AS DATA",
                    (f"CAST({hora_col} AS varchar(8)) AS HORA" if hora_col else "'' AS HORA"),
                    (f"ISNULL({aloj_col}, '') AS ALOJAMENTO" if aloj_col else "'' AS ALOJAMENTO"),
                    ("ISNULL(T.ORIGEM,'') AS ORIGEM" if table == 'TAREFAS' and 'ORIGEM' in cols else f"'{origem_code}' AS ORIGEM"),
                    (f"{tarefa_col} AS TAREFA" if tarefa_col else f"'{origem_code if origem_code else 'Tarefa'}' AS TAREFA"),
                    (f"ISNULL({tratado_col}, 0) AS TRATADO" if tratado_col else "0 AS TRATADO"),
                    (f"ISNULL({user_col}, '') AS UTILIZADOR" if user_col else "'' AS UTILIZADOR")
                ]
                sel_sql = ",\n                    ".join(sel)
                base = f"SELECT\n                    {sel_sql}\n                FROM {table}"
                binds = {}
                where = []
                if data_col:
                    where.append(f"{data_col} BETWEEN :start AND :end")
                    binds.update({ 'start': start, 'end': end })
                if only_mine and getattr(current_user, 'LOGIN', None) and user_col:
                    where.append(f"{user_col} = :onlymine")
                    binds['onlymine'] = current_user.LOGIN
                if where:
                    base += "\nWHERE " + " AND ".join(where)
                return base, binds

            selects = []
            binds_union = {}
            # Always try TAREFAS
            if exists_table('TAREFAS'):
                s, b = build_select('TAREFAS', '')
                if s:
                    selects.append(s)
                    binds_union.update({ f"t_{k}": v for k, v in b.items() })
                    # rename binds in SQL to unique names
                    for k in list(b.keys()):
                        s = s.replace(f":{k}", f":t_{k}")
                    selects[-1] = s
            # Optional MN/LP/FS
            for tab, code, prefix in (( 'MN','MN','m'), ('LP','LP','l'), ('FS','FS','f')):
                if exists_table(tab):
                    s, b = build_select(tab, code)
                    if s:
                        selects.append(s)
                        binds_union.update({ f"{prefix}_{k}": v for k, v in b.items() })
                        for k in list(b.keys()):
                            s = s.replace(f":{k}", f":{prefix}_{k}")
                        selects[-1] = s

            # LPADMIN: garantir FS visível, ignorando only_mine
            try:
                _lpadmin_flag = bool(getattr(current_user, 'LPADMIN', 0))
            except Exception:
                _lpadmin_flag = False
            if _lpadmin_flag:
                try:
                    _only_backup = only_mine
                except Exception:
                    _only_backup = False
                try:
                    only_mine = False
                    if exists_table('FS'):
                        s2, b2 = build_select('FS', 'FS')
                        if s2:
                            # Renomear binds para evitar colisões
                            selects.append(s2)
                            binds_union.update({ f"fsall_{k}": v for k, v in b2.items() })
                            for k in list(b2.keys()):
                                s2 = s2.replace(f":{k}", f":fsall_{k}")
                            selects[-1] = s2
                    elif exists_table('TAREFAS') and 'ORIGEM' in get_columns('TAREFAS'):
                        cols = get_columns('TAREFAS')
                        data_col = 'DATA' if 'DATA' in cols else None
                        hora_col = 'HORA' if 'HORA' in cols else None
                        aloj_col = 'ALOJAMENTO' if 'ALOJAMENTO' in cols else None
                        user_col = 'UTILIZADOR' if 'UTILIZADOR' in cols else None
                        tarefa_col = next((c for c in ('TAREFA','TITULO','ASSUNTO','DESCRICAO','DESCR') if c in cols), None)
                        tratado_col = 'TRATADO' if 'TRATADO' in cols else None
                        if data_col:
                            sel = [
                                "NULL AS TAREFASSTAMP",
                                f"CAST({data_col} AS date) AS DATA",
                                (f"CAST({hora_col} AS varchar(8)) AS HORA" if hora_col else "'' AS HORA"),
                                (f"ISNULL({aloj_col}, '') AS ALOJAMENTO" if aloj_col else "'' AS ALOJAMENTO"),
                                "'FS' AS ORIGEM",
                                (f"{tarefa_col} AS TAREFA" if tarefa_col else "'Tarefa' AS TAREFA"),
                                (f"ISNULL({tratado_col}, 0) AS TRATADO" if tratado_col else "0 AS TRATADO"),
                                (f"ISNULL({user_col}, '') AS UTILIZADOR" if user_col else "'' AS UTILIZADOR"),
                            ]
                            sel_sql = ",\n                    ".join(sel)
                            fs_sql = f"SELECT\n                    {sel_sql}\n                FROM TAREFAS\n                WHERE {data_col} BETWEEN :fs_start AND :fs_end AND ISNULL(ORIGEM,'') = 'FS'"
                            selects.append(fs_sql)
                            binds_union.update({ 'fs_start': start, 'fs_end': end })
                finally:
                    only_mine = _only_backup

            if not selects:
                return jsonify([])

            union_sql = "\nUNION ALL\n".join(selects)
            outer = f"SELECT X.*, ISNULL(U.NOME, X.UTILIZADOR) AS UTILIZADOR_NOME, '#6c757d' AS UTILIZADOR_COR\nFROM (\n{union_sql}\n) X\nLEFT JOIN US U ON U.LOGIN = X.UTILIZADOR\n"

            # Outer filters (IN lists)
            where_outer = []
            outer_bind = dict(binds_union)
            if utilizadores:
                keys = []
                for i, u in enumerate(utilizadores):
                    k = f"ou{i}"
                    outer_bind[k] = u
                    keys.append(f":{k}")
                where_outer.append("X.UTILIZADOR IN (" + ",".join(keys) + ")")
            if alojamentos:
                keys = []
                for i, a in enumerate(alojamentos):
                    k = f"oa{i}"
                    outer_bind[k] = a
                    keys.append(f":{k}")
                where_outer.append("X.ALOJAMENTO IN (" + ",".join(keys) + ")")
            if origens:
                keys = []
                for i, o in enumerate(origens):
                    k = f"oo{i}"
                    outer_bind[k] = o
                    keys.append(f":{k}")
                where_outer.append("X.ORIGEM IN (" + ",".join(keys) + ")")
            if where_outer:
                outer += "WHERE " + " AND ".join(where_outer) + "\n"
            outer += "ORDER BY X.DATA, X.HORA\n"

            base_rows = [ dict(r) for r in db.session.execute(text(outer), outer_bind).mappings().all() ]

            # Regra adicional: Se o utilizador tiver LP num alojamento num dia,
            # deve ver MN/FS desse alojamento nesse dia, mesmo de outros utilizadores.
            # 1) recolher pares (DATA, ALOJAMENTO) de LP para utilizadores selecionados (ou only_mine)
            lp_pairs = set()
            target_users = set(utilizadores)
            if only_mine and getattr(current_user, 'LOGIN', None):
                target_users.add(current_user.LOGIN)

            def add_lp_pairs_from_table(tab: str):
                cols = get_columns(tab)
                if not cols or 'DATA' not in cols or 'ALOJAMENTO' not in cols:
                    return
                where = ["DATA BETWEEN :lp_start AND :lp_end"]
                bind = { 'lp_start': start, 'lp_end': end }
                if target_users and 'UTILIZADOR' in cols:
                    ukeys = []
                    for i, u in enumerate(target_users):
                        k = f"lpu{i}"
                        bind[k] = u
                        ukeys.append(f":{k}")
                    where.append("UTILIZADOR IN (" + ",".join(ukeys) + ")")
                if alojamentos:
                    akeys = []
                    for i, a in enumerate(alojamentos):
                        k = f"lpa{i}"
                        bind[k] = a
                        akeys.append(f":{k}")
                    where.append("ISNULL(ALOJAMENTO,'') IN (" + ",".join(akeys) + ")")
                sql_lp = text(f"SELECT CAST(DATA AS date) AS DATA, ISNULL(ALOJAMENTO,'') AS ALOJAMENTO FROM {tab} WHERE " + " AND ".join(where))
                for r in db.session.execute(sql_lp, bind).fetchall():
                    lp_pairs.add((r[0].isoformat() if hasattr(r[0], 'isoformat') else str(r[0]), r[1]))

            # Prefer LP table if exists, else TAREFAS where ORIGEM='LP'
            if exists_table('LP'):
                add_lp_pairs_from_table('LP')
            elif exists_table('TAREFAS') and 'ORIGEM' in get_columns('TAREFAS'):
                # Fallback via TAREFAS
                cols = get_columns('TAREFAS')
                if 'DATA' in cols and 'ALOJAMENTO' in cols:
                    where = ["DATA BETWEEN :lp_start AND :lp_end", "ISNULL(ORIGEM,'') = 'LP'"]
                    bind = { 'lp_start': start, 'lp_end': end }
                    if target_users and 'UTILIZADOR' in cols:
                        ukeys = []
                        for i, u in enumerate(target_users):
                            k = f"lptu{i}"
                            bind[k] = u
                            ukeys.append(f":{k}")
                        where.append("UTILIZADOR IN (" + ",".join(ukeys) + ")")
                    if alojamentos:
                        akeys = []
                        for i, a in enumerate(alojamentos):
                            k = f"lpta{i}"
                            bind[k] = a
                            akeys.append(f":{k}")
                        where.append("ISNULL(ALOJAMENTO,'') IN (" + ",".join(akeys) + ")")
                    sql_lp = text("SELECT CAST(DATA AS date) AS DATA, ISNULL(ALOJAMENTO,'') AS ALOJAMENTO FROM TAREFAS WHERE " + " AND ".join(where))
                    for r in db.session.execute(sql_lp, bind).fetchall():
                        lp_pairs.add((r[0].isoformat() if hasattr(r[0], 'isoformat') else str(r[0]), r[1]))

            # 2) se houver pares LP, buscar MN e FS para esses pares (independente de utilizador)
            extra_rows = []
            if lp_pairs:
                # limita por origens selecionadas, se fornecidas; caso vazio, considera MN e FS
                target_origens = set([o for o in origens if o in ('MN','FS')]) if origens else {'MN','FS'}

                def fetch_extra_from_table(tab: str, origem_code: str):
                    if origem_code not in target_origens:
                        return
                    cols = get_columns(tab)
                    if not cols:
                        return
                    stamp_col = next((c for c in cols if c.endswith('STAMP')), None)
                    data_col = 'DATA' if 'DATA' in cols else None
                    hora_col = 'HORA' if 'HORA' in cols else None
                    aloj_col = 'ALOJAMENTO' if 'ALOJAMENTO' in cols else None
                    user_col = 'UTILIZADOR' if 'UTILIZADOR' in cols else None
                    tarefa_col = next((c for c in ('TAREFA','TITULO','ASSUNTO','DESCRICAO','DESCR') if c in cols), None)
                    if not data_col or not aloj_col:
                        return
                    pair_conditions = []
                    bind = { 'ex_start': start, 'ex_end': end }
                    idx = 0
                    for (d,a) in lp_pairs:
                        kd = f"d{idx}"; ka = f"a{idx}"; idx += 1
                        bind[kd] = d; bind[ka] = a
                        pair_conditions.append(f"({data_col} = :{kd} AND ISNULL({aloj_col},'') = :{ka})")
                    if not pair_conditions:
                        return
                    where = [f"{data_col} BETWEEN :ex_start AND :ex_end", "(" + " OR ".join(pair_conditions) + ")"]
                    sql = f"SELECT {stamp_col or 'NULL'} AS STAMP, CAST({data_col} AS date) AS DATA, " \
                          f"{('CAST('+hora_col+' AS varchar(8))' if hora_col else "''")} AS HORA, " \
                          f"ISNULL({aloj_col}, '') AS ALOJAMENTO, '{origem_code}' AS ORIGEM, " \
                          f"{(tarefa_col or "''")} AS TAREFA, {('ISNULL('+user_col+', \'\')' if user_col else "''")} AS UTILIZADOR " \
                          f"FROM {tab} WHERE " + " AND ".join(where)
                    for r in db.session.execute(text(sql), bind).mappings().all():
                        extra_rows.append({
                            'TAREFASSTAMP': r.get('STAMP'),
                            'DATA': r.get('DATA'),
                            'HORA': r.get('HORA') or '',
                            'ALOJAMENTO': r.get('ALOJAMENTO') or '',
                            'ORIGEM': origem_code,
                            'TAREFA': r.get('TAREFA') or (origem_code or 'Tarefa'),
                            'TRATADO': 0,
                            'UTILIZADOR': r.get('UTILIZADOR') or ''
                        })

                # Prefer MN/FS tables
                if exists_table('MN'):
                    fetch_extra_from_table('MN', 'MN')
                if exists_table('FS'):
                    fetch_extra_from_table('FS', 'FS')
                # Fallback: TAREFAS com ORIGEM MN/FS
                if exists_table('TAREFAS') and 'ORIGEM' in get_columns('TAREFAS'):
                    cols = get_columns('TAREFAS')
                    data_col = 'DATA' if 'DATA' in cols else None
                    hora_col = 'HORA' if 'HORA' in cols else None
                    aloj_col = 'ALOJAMENTO' if 'ALOJAMENTO' in cols else None
                    user_col = 'UTILIZADOR' if 'UTILIZADOR' in cols else None
                    tarefa_col = next((c for c in ('TAREFA','TITULO','ASSUNTO','DESCRICAO','DESCR') if c in cols), None)
                    if data_col and aloj_col:
                        pair_conditions = []
                        bind = { 'ex_start': start, 'ex_end': end }
                        idx = 0
                        for (d,a) in lp_pairs:
                            kd = f"td{idx}"; ka = f"ta{idx}"; idx += 1
                            bind[kd] = d; bind[ka] = a
                            pair_conditions.append(f"({data_col} = :{kd} AND ISNULL({aloj_col},'') = :{ka})")
                        if pair_conditions:
                            org_filter = "('MN','FS')" if not origens else "(" + ",".join([f"'{o}'" for o in target_origens]) + ")"
                            sql = f"SELECT CAST({data_col} AS date) AS DATA, {('CAST('+hora_col+' AS varchar(8))' if hora_col else "''")} AS HORA, " \
                                  f"ISNULL({aloj_col}, '') AS ALOJAMENTO, ISNULL(ORIGEM,'') AS ORIGEM, " \
                                  f"{(tarefa_col or "''")} AS TAREFA, {('ISNULL('+user_col+', \'\')' if user_col else "''")} AS UTILIZADOR " \
                                  f"FROM TAREFAS WHERE {data_col} BETWEEN :ex_start AND :ex_end AND ISNULL(ORIGEM,'') IN {org_filter} " \
                                  f"AND (" + " OR ".join(pair_conditions) + ")"
                            for r in db.session.execute(text(sql), bind).mappings().all():
                                extra_rows.append({
                                    'TAREFASSTAMP': None,
                                    'DATA': r.get('DATA'),
                                    'HORA': r.get('HORA') or '',
                                    'ALOJAMENTO': r.get('ALOJAMENTO') or '',
                                    'ORIGEM': r.get('ORIGEM') or '',
                                    'TAREFA': r.get('TAREFA') or 'Tarefa',
                                    'TRATADO': 0,
                                    'UTILIZADOR': r.get('UTILIZADOR') or ''
                                })

            # 3) juntar e eliminar duplicados por chave lÃ³gica
            seen = set()
            out = []
            def key_of(x):
                return f"{x.get('ORIGEM','')}|{x.get('DATA','')}|{x.get('HORA','')}|{x.get('ALOJAMENTO','')}|{x.get('TAREFA','')}|{x.get('UTILIZADOR','')}"
            for r in base_rows + extra_rows:
                k = key_of(r)
                if k in seen:
                    continue
                seen.add(k)
                out.append(r)
            # ordenar por data/hora
            out.sort(key=lambda r: (str(r.get('DATA') or ''), str(r.get('HORA') or '')))
            return jsonify(out)
        except Exception as e:
            return jsonify({ 'error': str(e), 'rows': [] })

    @app.route("/api/tarefa_info/<stamp>")
    @login_required
    def tarefa_info(stamp):
        result = db.session.execute(
            db.text("SELECT dbo.info_tarefa(:stamp) AS info"),
            {"stamp": stamp}
        ).fetchone()
        
        return jsonify({"info": result.info if result else ""})

    @app.route('/api/alojamentos')
    @login_required
    def api_alojamentos():
        try:
            # lightweight mode: just names, no extra queries
            basic_param = (request.args.get('basic') or '').strip().lower()
            basic = basic_param in ('1', 'true', 'yes', 'y')

            can_open = False
            try:
                q = db.session.execute(text("""
                    SELECT CONSULTAR FROM ACESSOS
                    WHERE UTILIZADOR = :u AND TABELA = 'AL'
                """), { 'u': current_user.LOGIN }).fetchone()
                can_open = bool(q[0]) if q is not None else False
            except Exception:
                can_open = False

            if basic:
                rows = db.session.execute(text(
                    """
                    SELECT NOME
                    FROM AL
                    WHERE ISNULL(INATIVO,0) = 0
                    ORDER BY NOME
                    """
                )).fetchall()
                nomes = [r[0] for r in rows if r and r[0]]
                return jsonify({ 'rows': [ { 'NOME': n } for n in nomes ], 'can_open': can_open })

            rows = db.session.execute(text(
                """
                SELECT ALSTAMP, NOME, MORADA, CODPOST, LOCAL, TIPOLOGIA, LOTACAO
                FROM AL
                WHERE ISNULL(INATIVO,0) = 0
                ORDER BY NOME
                """
            )).mappings().all()

            result = []
            for r in rows:
                d = dict(r)
                # carregar codigos (ALC) por alojamento
                try:
                    cods = db.session.execute(text(
                        """
                        SELECT CARACTERISTICA, VALOR
                        FROM ALC
                        WHERE ALOJAMENTO = :nome AND GRUPO = 'CODIGOS'
                        ORDER BY CARACTERISTICA
                        """
                    ), { 'nome': r['NOME'] }).mappings().all()
                    d['CODIGOS'] = [ dict(c) for c in cods ]
                except Exception:
                    d['CODIGOS'] = []
                result.append(d)

            return jsonify({ 'rows': result, 'can_open': can_open })
        except Exception as e:
            return jsonify({ 'error': str(e) }), 500


    @app.route('/alojamentos/geo')
    @app.route('/alojamentos_geo')
    @login_required
    def alojamentos_geo_page():
        return render_template('alojamentos_geo.html', page_title='Geolocalização de Alojamentos')


    @app.route('/api/alojamentos_geo', methods=['GET'])
    @login_required
    def api_alojamentos_geo_list():
        rows = db.session.execute(text("""
            SELECT ALSTAMP, NOME, MORADA, CODPOST, LOCAL, LAT, LON
            FROM AL
            ORDER BY NOME
        """)).mappings().all()
        para_rows = db.session.execute(text("""
            SELECT PARAMETRO, CVALOR, NVALOR, DVALOR, LVALOR, TIPO
            FROM dbo.PARA
            WHERE UPPER(LTRIM(RTRIM(PARAMETRO))) IN ('EMP_LAT','EMP_LON','EMP_MORADA','EMP_CODPOST','EMP_LOCAL')
        """)).mappings().all()
        para = {}
        for p in para_rows:
            k = (p.get('PARAMETRO') or '').strip().upper()
            t = (p.get('TIPO') or '').strip().upper()
            if t == 'N':
                try:
                    para[k] = float(p.get('NVALOR') or 0)
                except Exception:
                    para[k] = 0.0
            else:
                para[k] = (p.get('CVALOR') or '').strip()

        out = []
        emp_lat = para.get('EMP_LAT')
        emp_lon = para.get('EMP_LON')
        try:
            if emp_lat is not None and emp_lon is not None:
                out.append({
                    'ALSTAMP': '__SEDE__',
                    'NOME': 'SEDE',
                    'MORADA': para.get('EMP_MORADA', ''),
                    'CODPOST': para.get('EMP_CODPOST', ''),
                    'LOCAL': para.get('EMP_LOCAL', ''),
                    'LAT': float(emp_lat),
                    'LON': float(emp_lon),
                    'IS_SEDE': 1
                })
        except Exception:
            pass

        out.extend([dict(r) for r in rows])
        return jsonify(out)


    @app.route('/api/alojamentos_geo/<alstamp>', methods=['GET'])
    @login_required
    def api_alojamentos_geo_detail(alstamp):
        if str(alstamp or '').strip() == '__SEDE__':
            para_rows = db.session.execute(text("""
                SELECT PARAMETRO, CVALOR, NVALOR, TIPO
                FROM dbo.PARA
                WHERE UPPER(LTRIM(RTRIM(PARAMETRO))) IN ('EMP_LAT','EMP_LON','EMP_MORADA','EMP_CODPOST','EMP_LOCAL')
            """)).mappings().all()
            para = {}
            for p in para_rows:
                k = (p.get('PARAMETRO') or '').strip().upper()
                t = (p.get('TIPO') or '').strip().upper()
                if t == 'N':
                    try:
                        para[k] = float(p.get('NVALOR') or 0)
                    except Exception:
                        para[k] = 0.0
                else:
                    para[k] = (p.get('CVALOR') or '').strip()
            return jsonify({
                'ALSTAMP': '__SEDE__',
                'NOME': 'SEDE',
                'MORADA': para.get('EMP_MORADA', ''),
                'CODPOST': para.get('EMP_CODPOST', ''),
                'LOCAL': para.get('EMP_LOCAL', ''),
                'LAT': para.get('EMP_LAT', None),
                'LON': para.get('EMP_LON', None),
                'IS_SEDE': 1
            })
        row = db.session.execute(text("""
            SELECT ALSTAMP, NOME, MORADA, CODPOST, LOCAL, LAT, LON
            FROM AL
            WHERE ALSTAMP = :alstamp
        """), {'alstamp': alstamp}).mappings().first()
        if not row:
            return jsonify({'error': 'Alojamento não encontrado'}), 404
        return jsonify(dict(row))


    @app.route('/api/alojamentos_geo/<alstamp>/coords', methods=['PUT'])
    @login_required
    def api_alojamentos_geo_update_coords(alstamp):
        if str(alstamp or '').strip() == '__SEDE__':
            return jsonify({'error': 'A geolocalização da sede não é editável neste ecrã.'}), 400
        data = request.get_json() or {}
        try:
            lat = data.get('lat', None)
            lon = data.get('lon', None)
            if lat is None or lon is None:
                return jsonify({'error': 'Latitude/Longitude em falta'}), 400
            lat = round(float(lat), 6)
            lon = round(float(lon), 6)
        except Exception:
            return jsonify({'error': 'Latitude/Longitude inválidas'}), 400

        current = db.session.execute(text("""
            SELECT MORADA, CODPOST, LOCAL
            FROM AL
            WHERE ALSTAMP = :alstamp
        """), {'alstamp': alstamp}).mappings().first()
        if not current:
            return jsonify({'error': 'Alojamento não encontrado'}), 404

        morada = data.get('morada', current.get('MORADA'))
        codpost = data.get('codpost', current.get('CODPOST'))
        local = data.get('local', current.get('LOCAL'))

        db.session.execute(text("""
            UPDATE AL
            SET LAT = :lat,
                LON = :lon,
                MORADA = :morada,
                CODPOST = :codpost,
                LOCAL = :local
            WHERE ALSTAMP = :alstamp
        """), {
            'lat': lat,
            'lon': lon,
            'morada': morada or '',
            'codpost': codpost or '',
            'local': local or '',
            'alstamp': alstamp
        })
        db.session.commit()
        return jsonify({'success': True})


    @app.route('/api/alojamentos_geo/geocode', methods=['POST'])
    @login_required
    def api_alojamentos_geo_geocode():
        payload = request.get_json() or {}
        morada = (payload.get('morada') or '').strip()
        codpost = (payload.get('codpost') or '').strip()
        local = (payload.get('local') or '').strip()
        limit = int(payload.get('limit') or 5)
        query = _geo_build_query(morada, codpost, local)
        if not query:
            return jsonify({'error': 'Morada em falta'}), 400
        url = f"https://nominatim.openstreetmap.org/search?format=json&limit={limit}&countrycodes=pt&q={quote(query)}"
        try:
            data = _geo_fetch_json(url)
        except Exception:
            return jsonify({'error': 'Falha ao contactar o serviço de geocoding'}), 502
        results = []
        for item in data or []:
            try:
                results.append({
                    'display_name': item.get('display_name', ''),
                    'lat': float(item.get('lat')),
                    'lon': float(item.get('lon'))
                })
            except Exception:
                continue
        return jsonify({'results': results})

    @app.route('/api/alojamentos_geo/rota', methods=['POST'])
    @login_required
    def api_alojamentos_geo_rota():
        from itertools import permutations
        payload = request.get_json(silent=True) or {}
        from_stamp = (payload.get('from_stamp') or '').strip()
        to_stamp = (payload.get('to_stamp') or '').strip()
        to_stamps = payload.get('to_stamps') or []
        return_to_origin = bool(payload.get('return_to_origin') or False)
        if not isinstance(to_stamps, list):
            to_stamps = []
        if to_stamp and to_stamp not in to_stamps:
            to_stamps.append(to_stamp)
        to_stamps = [str(x).strip() for x in to_stamps if str(x).strip()]
        if not from_stamp or not to_stamps:
            return jsonify({'error': 'Origem e destinos são obrigatórios.'}), 400

        def _loc_by_stamp(stamp):
            if stamp == '__SEDE__':
                para_rows = db.session.execute(text("""
                    SELECT PARAMETRO, CVALOR, NVALOR, TIPO
                    FROM dbo.PARA
                    WHERE UPPER(LTRIM(RTRIM(PARAMETRO))) IN ('EMP_LAT','EMP_LON')
                """)).mappings().all()
                vals = {}
                for r in para_rows:
                    k = (r.get('PARAMETRO') or '').strip().upper()
                    if (r.get('TIPO') or '').strip().upper() == 'N':
                        vals[k] = float(r.get('NVALOR') or 0)
                    else:
                        vals[k] = float(r.get('CVALOR') or 0)
                return {'NOME': 'SEDE', 'LAT': vals.get('EMP_LAT'), 'LON': vals.get('EMP_LON')}
            row = db.session.execute(text("""
                SELECT NOME, LAT, LON
                FROM AL
                WHERE ALSTAMP = :s
            """), {'s': stamp}).mappings().first()
            return dict(row) if row else None

        requested = [from_stamp] + to_stamps
        requested_unique = []
        seen = set()
        for s in requested:
            if s in seen:
                continue
            seen.add(s)
            requested_unique.append(s)
        if len(requested_unique) < 2:
            return jsonify({'error': 'Seleciona pelo menos dois locais diferentes.'}), 400

        locs = {}
        for s in requested_unique:
            loc = _loc_by_stamp(s)
            if not loc:
                return jsonify({'error': f'Local não encontrado: {s}'}), 404
            try:
                loc['LAT'] = float(loc.get('LAT'))
                loc['LON'] = float(loc.get('LON'))
            except Exception:
                return jsonify({'error': f'O local "{loc.get("NOME") or s}" não tem coordenadas válidas.'}), 400
            locs[s] = loc

        dests = [s for s in requested_unique if s != from_stamp]

        all_for_matrix = [from_stamp] + dests
        matrix_idx = {s: i for i, s in enumerate(all_for_matrix)}
        matrix_dur = []
        matrix_dist = []
        try:
            matrix_coords = ";".join([f"{locs[x]['LON']},{locs[x]['LAT']}" for x in all_for_matrix])
            table_url = f"https://router.project-osrm.org/table/v1/driving/{matrix_coords}?annotations=duration,distance"
            table_data = _geo_fetch_json(table_url)
            matrix_dur = table_data.get('durations') or []
            matrix_dist = table_data.get('distances') or []
        except Exception:
            matrix_dur = []
            matrix_dist = []

        def _leg_duration(a_stamp, b_stamp):
            try:
                ai = matrix_idx[a_stamp]
                bi = matrix_idx[b_stamp]
                val = float(matrix_dur[ai][bi])
                return val if val > 0 else None
            except Exception:
                return None

        def _leg_distance(a_stamp, b_stamp):
            try:
                ai = matrix_idx[a_stamp]
                bi = matrix_idx[b_stamp]
                val = float(matrix_dist[ai][bi])
                return val if val >= 0 else None
            except Exception:
                return None

        def _coord_key(stamp):
            loc = locs.get(stamp) or {}
            try:
                lat = round(float(loc.get('LAT')), 5)
                lon = round(float(loc.get('LON')), 5)
                return (lat, lon)
            except Exception:
                return ('STAMP', stamp)

        def _coord_split_penalty(order):
            # Penaliza quando o mesmo ponto (coordenadas) aparece em blocos separados:
            # ex.: A(1) -> B -> A(2). O ideal é A(1),A(2) juntos.
            seq = [_coord_key(s) for s in order[1:]]
            pos = {}
            for i, key in enumerate(seq):
                pos.setdefault(key, []).append(i)
            penalty = 0
            for _, idxs in pos.items():
                if len(idxs) <= 1:
                    continue
                span = (idxs[-1] - idxs[0] + 1)
                gaps = span - len(idxs)
                if gaps > 0:
                    penalty += gaps + 1
            return penalty

        best_order = None
        best_duration = None
        best_distance = None
        best_penalty = None
        combinations = []
        if len(dests) <= 8 and matrix_dur:
            for perm in permutations(dests):
                order = [from_stamp] + list(perm)
                total_duration = 0.0
                total_distance = 0.0
                valid = True
                for i in range(len(order) - 1):
                    d = _leg_duration(order[i], order[i + 1])
                    km = _leg_distance(order[i], order[i + 1])
                    if d is None:
                        valid = False
                        break
                    total_duration += d
                    if km is not None:
                        total_distance += km
                if valid and return_to_origin:
                    d_back = _leg_duration(order[-1], from_stamp)
                    km_back = _leg_distance(order[-1], from_stamp)
                    if d_back is None:
                        valid = False
                    else:
                        total_duration += d_back
                        if km_back is not None:
                            total_distance += km_back
                combo_order = list(order)
                if return_to_origin and combo_order[-1] != from_stamp:
                    combo_order.append(from_stamp)
                combo_names = [(locs[s].get('NOME') or s) for s in combo_order]
                if not valid:
                    combinations.append({
                        'route_names': combo_names,
                        'distance_km': None,
                        'duration_min': None,
                        'status': 'sem_rota'
                    })
                    continue
                penalty = _coord_split_penalty(order)
                if penalty > 0:
                    # Regra rígida: locais com mesmas coordenadas têm de ficar contíguos.
                    combinations.append({
                        'route_names': combo_names,
                        'distance_km': round(total_distance / 1000.0, 2),
                        'duration_min': int(round(total_duration / 60.0)),
                        'status': 'coords_separadas'
                    })
                    continue
                combinations.append({
                    'route_names': combo_names,
                    'distance_km': round(total_distance / 1000.0, 2),
                    'duration_min': int(round(total_duration / 60.0)),
                    'status': 'ok'
                })
                better = False
                if best_duration is None:
                    better = True
                else:
                    # 1) menor duração; 2) menor distância
                    if total_duration < (best_duration - 0.5):
                        better = True
                    elif abs(total_duration - best_duration) <= 0.5:
                        if (best_distance is None) or (total_distance < best_distance):
                            better = True
                if better:
                    best_duration = total_duration
                    best_distance = total_distance
                    best_penalty = penalty
                    best_order = order
        if not best_order:
            best_order = [from_stamp] + dests

        full_order = list(best_order)
        if return_to_origin and full_order[-1] != from_stamp:
            full_order.append(from_stamp)

        legs = []
        polyline_coords = []
        total_distance_m = 0.0
        total_duration_s = 0.0
        for i in range(len(full_order) - 1):
            a = full_order[i]
            b = full_order[i + 1]
            la = locs[a]
            lb = locs[b]
            leg_url = f"https://router.project-osrm.org/route/v1/driving/{la['LON']},{la['LAT']};{lb['LON']},{lb['LAT']}?overview=full&geometries=geojson"
            try:
                leg_data = _geo_fetch_json(leg_url)
            except Exception:
                return jsonify({'error': 'Falha ao contactar o serviço de rotas.'}), 502
            if leg_data.get('code') != 'Ok' or not leg_data.get('routes'):
                return jsonify({'error': f'Não foi possível calcular o percurso entre {la.get("NOME")} e {lb.get("NOME")}.'}), 400
            route = leg_data['routes'][0]
            total_distance_m += float(route.get('distance') or 0)
            total_duration_s += float(route.get('duration') or 0)
            geom = (route.get('geometry') or {}).get('coordinates') or []
            if geom:
                if polyline_coords and geom[0] == polyline_coords[-1]:
                    polyline_coords.extend(geom[1:])
                else:
                    polyline_coords.extend(geom)
            legs.append({
                'from_stamp': a,
                'to_stamp': b,
                'from_name': la.get('NOME') or a,
                'to_name': lb.get('NOME') or b,
                'distance_km': round(float(route.get('distance') or 0) / 1000.0, 2),
                'duration_min': int(round(float(route.get('duration') or 0) / 60.0))
            })

        return jsonify({
            'ok': True,
            'from_name': locs[from_stamp].get('NOME') or 'Origem',
            'to_name': locs[full_order[-1]].get('NOME') or 'Destino',
            'route_stamps': full_order,
            'route_names': [(locs[s].get('NOME') or s) for s in full_order],
            'route_points': [{
                'stamp': s,
                'name': locs[s].get('NOME') or s,
                'lat': locs[s].get('LAT'),
                'lon': locs[s].get('LON')
            } for s in full_order],
            'distance_km': round(total_distance_m / 1000.0, 2),
            'duration_min': int(round(total_duration_s / 60.0)),
            'geometry': {'type': 'LineString', 'coordinates': polyline_coords},
            'legs': legs,
            'combinations': combinations
        })


    rotas_workers = {}

    def _rotas_worker_alive(job_id: str) -> bool:
        t = rotas_workers.get(str(job_id))
        return bool(t and t.is_alive())

    def ensure_rotas_job_table():
        db.session.execute(text("""
            IF OBJECT_ID('dbo.ROTAS_JOB', 'U') IS NULL
            BEGIN
                CREATE TABLE dbo.ROTAS_JOB(
                    JobId UNIQUEIDENTIFIER NOT NULL PRIMARY KEY,
                    State VARCHAR(20) NOT NULL,
                    Stage VARCHAR(20) NOT NULL,
                    Total INT NOT NULL,
                    Processed INT NOT NULL,
                    Ok INT NOT NULL,
                    Errors INT NOT NULL,
                    Pending INT NOT NULL,
                    StartedAt DATETIME NULL,
                    UpdatedAt DATETIME NULL,
                    FinishedAt DATETIME NULL,
                    Message VARCHAR(255) NULL,
                    RequestedBy VARCHAR(50) NULL
                )
            END
        """))
        db.session.commit()


    def rotas_job_update(job_id, **fields):
        if not fields:
            return
        sets = []
        params = { 'JobId': job_id }
        for key, value in fields.items():
            sets.append(f"{key} = :{key}")
            params[key] = value
        sql = f"UPDATE dbo.ROTAS_JOB SET {', '.join(sets)} WHERE JobId = :JobId"
        db.session.execute(text(sql), params)
        db.session.commit()


    def rotas_job_get(job_id):
        row = db.session.execute(text("""
            SELECT *
            FROM dbo.ROTAS_JOB
            WHERE JobId = :JobId
        """), {'JobId': job_id}).mappings().first()
        return dict(row) if row else None


    def rotas_job_get_running():
        row = db.session.execute(text("""
            SELECT TOP 1 *
            FROM dbo.ROTAS_JOB
            WHERE State IN ('running','stopping')
            ORDER BY StartedAt DESC
        """)).mappings().first()
        return dict(row) if row else None


    def _rotas_job_is_stale(job_row, seconds=300):
        if not job_row:
            return False
        ts = job_row.get('UpdatedAt') or job_row.get('StartedAt')
        if not ts:
            return False
        if isinstance(ts, str):
            try:
                ts = datetime.fromisoformat(ts)
            except Exception:
                return False
        return (datetime.now() - ts).total_seconds() > seconds


    def run_rotas_rebuild(job_id, full_rebuild=True):
        with app.app_context():
            try:
                if full_rebuild:
                    rotas_job_update(
                        job_id,
                        State='running',
                        Stage='truncate',
                        Message='A limpar...',
                        UpdatedAt=datetime.now()
                    )
                    try:
                        db.session.execute(text("TRUNCATE TABLE dbo.ROTAS"))
                    except Exception:
                        db.session.execute(text("DELETE FROM dbo.ROTAS"))
                    db.session.commit()

                    rotas_job_update(
                        job_id,
                        Stage='generate',
                        Message='A gerar pares...',
                        UpdatedAt=datetime.now()
                    )
                aloj_rows = db.session.execute(text("""
                    SELECT ALSTAMP, LAT, LON
                    FROM AL
                    WHERE LAT IS NOT NULL AND LON IS NOT NULL
                """)).mappings().all()
                aloj = []
                coords = {}
                for r in aloj_rows:
                    try:
                        lat = float(r['LAT'])
                        lon = float(r['LON'])
                    except Exception:
                        continue
                    coords[str(r['ALSTAMP'])] = (lat, lon)
                    aloj.append(str(r['ALSTAMP']))

                aloj = sorted(set(aloj))
                if full_rebuild:
                    total = int(len(aloj) * (len(aloj) - 1) / 2)
                    rotas_job_update(
                        job_id,
                        Total=total,
                        Processed=0,
                        Ok=0,
                        Errors=0,
                        Pending=total,
                        UpdatedAt=datetime.now()
                    )

                    batch = []
                    for i in range(len(aloj)):
                        for j in range(i + 1, len(aloj)):
                            batch.append({
                                'orig': aloj[i],
                                'dest': aloj[j]
                            })
                            if len(batch) >= 1000:
                                db.session.execute(text("""
                                    INSERT INTO dbo.ROTAS
                                    (OrigemStamp, DestinoStamp, Km, Segundos, Perfil, Provider, Status, Tentativas, Erro, UpdatedAt)
                                    VALUES (:orig, :dest, 0, 0, 'driving', 'OSRM', 0, 0, '', GETDATE())
                                """), batch)
                                db.session.commit()
                                batch = []
                    if batch:
                        db.session.execute(text("""
                            INSERT INTO dbo.ROTAS
                            (OrigemStamp, DestinoStamp, Km, Segundos, Perfil, Provider, Status, Tentativas, Erro, UpdatedAt)
                            VALUES (:orig, :dest, 0, 0, 'driving', 'OSRM', 0, 0, '', GETDATE())
                        """), batch)
                        db.session.commit()

                    rotas_job_update(
                        job_id,
                        Stage='compute',
                        Message='A calcular rotas...',
                        UpdatedAt=datetime.now()
                    )

                    processed = 0
                    ok = 0
                    errors = 0
                    pending = total
                else:
                    total = db.session.execute(text("""
                        SELECT COUNT(*) AS C
                        FROM dbo.ROTAS
                        WHERE Status = 0
                    """)).scalar() or 0
                    rotas_job_update(
                        job_id,
                        Stage='compute',
                        Message='A retomar pendentes...',
                        Total=total,
                        Processed=0,
                        Ok=0,
                        Errors=0,
                        Pending=total,
                        UpdatedAt=datetime.now()
                    )
                    processed = 0
                    ok = 0
                    errors = 0
                    pending = total
                    if total == 0:
                        rotas_job_update(
                            job_id,
                            State='done',
                            Message='Nada pendente',
                            FinishedAt=datetime.now(),
                            UpdatedAt=datetime.now()
                        )
                        return

                last_job_update = datetime.now()
                while True:
                    state_row = rotas_job_get(job_id)
                    if not state_row:
                        return
                    if state_row.get('State') == 'stopping':
                        rotas_job_update(
                            job_id,
                            State='stopped',
                            Message='Parado',
                            FinishedAt=datetime.now(),
                            UpdatedAt=datetime.now()
                        )
                        return

                    rows = db.session.execute(text("""
                        SELECT TOP 200 OrigemStamp, DestinoStamp, Tentativas
                        FROM dbo.ROTAS
                        WHERE Status = 0
                        ORDER BY OrigemStamp, DestinoStamp
                    """)).mappings().all()
                    if not rows:
                        break

                    stop_now = False
                    for r in rows:
                        state_row = rotas_job_get(job_id)
                        if state_row and state_row.get('State') == 'stopping':
                            stop_now = True
                            break
                        orig = str(r['OrigemStamp'])
                        dest = str(r['DestinoStamp'])
                        lat1, lon1 = coords.get(orig, (None, None))
                        lat2, lon2 = coords.get(dest, (None, None))
                        status = 2
                        km = 0
                        segundos = 0
                        err = ''
                        if lat1 is None or lat2 is None:
                            status = 2
                            err = 'Sem coordenadas'
                        else:
                            url = f"https://router.project-osrm.org/route/v1/driving/{lon1},{lat1};{lon2},{lat2}?overview=false"
                            try:
                                data = _geo_fetch_json(url)
                                if data.get('code') == 'Ok' and data.get('routes'):
                                    route = data['routes'][0]
                                    km = round(float(route.get('distance', 0)) / 1000.0, 2)
                                    segundos = int(route.get('duration', 0))
                                    status = 1
                                elif data.get('code') == 'NoRoute':
                                    status = 3
                                    err = 'Sem rota'
                                else:
                                    status = 2
                                    err = str(data.get('code') or 'Erro')
                            except Exception as e:
                                status = 2
                                err = str(e)[:255]

                        db.session.execute(text("""
                            UPDATE dbo.ROTAS
                            SET Km = :km,
                                Segundos = :segundos,
                                Status = :status,
                                Tentativas = :tent,
                                Erro = :erro,
                                UpdatedAt = GETDATE()
                            WHERE OrigemStamp = :orig AND DestinoStamp = :dest
                        """), {
                            'km': km,
                            'segundos': segundos,
                            'status': status,
                            'tent': int(r['Tentativas'] or 0) + 1,
                            'erro': err[:255],
                            'orig': orig,
                            'dest': dest
                        })
                        processed += 1
                        pending = max(0, pending - 1)
                        if status == 1:
                            ok += 1
                        else:
                            errors += 1
                        now_dt = datetime.now()
                        if processed % 10 == 0 or (now_dt - last_job_update).total_seconds() >= 2:
                            rotas_job_update(
                                job_id,
                                Processed=processed,
                                Ok=ok,
                                Errors=errors,
                                Pending=pending,
                                UpdatedAt=now_dt
                            )
                            last_job_update = now_dt

                    if stop_now:
                        db.session.commit()
                        rotas_job_update(
                            job_id,
                            State='stopped',
                            Message='Parado',
                            FinishedAt=datetime.now(),
                            UpdatedAt=datetime.now()
                        )
                        return

                    db.session.commit()
                    rotas_job_update(
                        job_id,
                        Processed=processed,
                        Ok=ok,
                        Errors=errors,
                        Pending=pending,
                        UpdatedAt=datetime.now()
                    )

                rotas_job_update(
                    job_id,
                    State='done',
                    Message='Concluído',
                    FinishedAt=datetime.now(),
                    UpdatedAt=datetime.now()
                )
            except Exception as e:
                rotas_job_update(
                    job_id,
                    State='error',
                    Message=str(e)[:255],
                    UpdatedAt=datetime.now(),
                    FinishedAt=datetime.now()
                )
            finally:
                rotas_workers.pop(str(job_id), None)


    @app.route('/api/rotas/rebuild/start', methods=['POST'])
    @login_required
    def api_rotas_rebuild_start():
        ensure_rotas_job_table()
        running = rotas_job_get_running()
        if running and not _rotas_worker_alive(str(running['JobId'])):
            rotas_job_update(
                running['JobId'],
                State='stopped',
                Message='Parado (sem worker)',
                FinishedAt=datetime.now(),
                UpdatedAt=datetime.now()
            )
            running = None
        if running and _rotas_job_is_stale(running, seconds=300):
            rotas_job_update(
                running['JobId'],
                State='stopped',
                Message='Parado (servidor reiniciado)',
                FinishedAt=datetime.now(),
                UpdatedAt=datetime.now()
            )
            running = None
        if running:
            return jsonify({'job_id': str(running['JobId']), 'state': running['State']}), 409
        job_id = str(uuid.uuid4())
        db.session.execute(text("""
            INSERT INTO dbo.ROTAS_JOB
            (JobId, State, Stage, Total, Processed, Ok, Errors, Pending, StartedAt, UpdatedAt, Message, RequestedBy)
            VALUES (:job_id, 'running', 'truncate', 0, 0, 0, 0, 0, GETDATE(), GETDATE(), 'A iniciar...', :user)
        """), { 'job_id': job_id, 'user': getattr(current_user, 'LOGIN', '') })
        db.session.commit()
        t = threading.Thread(target=run_rotas_rebuild, args=(job_id, True), daemon=True)
        rotas_workers[str(job_id)] = t
        t.start()
        return jsonify({'job_id': job_id})


    @app.route('/api/rotas/rebuild/stop', methods=['POST'])
    @login_required
    def api_rotas_rebuild_stop():
        data = request.get_json() or {}
        job_id = data.get('job_id')
        if not job_id:
            running = rotas_job_get_running()
            if not running:
                return jsonify({'error': 'Sem job ativo'}), 404
            job_id = running['JobId']
        rotas_job_update(job_id, State='stopping', Message='A parar...', UpdatedAt=datetime.now())
        return jsonify({'ok': True})


    @app.route('/api/rotas/rebuild/resume', methods=['POST'])
    @login_required
    def api_rotas_rebuild_resume():
        ensure_rotas_job_table()
        running = rotas_job_get_running()
        if running and not _rotas_worker_alive(str(running['JobId'])):
            rotas_job_update(
                running['JobId'],
                State='stopped',
                Message='Parado (sem worker)',
                FinishedAt=datetime.now(),
                UpdatedAt=datetime.now()
            )
            running = None
        if running and _rotas_job_is_stale(running, seconds=300):
            rotas_job_update(
                running['JobId'],
                State='stopped',
                Message='Parado (servidor reiniciado)',
                FinishedAt=datetime.now(),
                UpdatedAt=datetime.now()
            )
            running = None
        if running:
            return jsonify({'job_id': str(running['JobId']), 'state': running['State']}), 409
        job_id = str(uuid.uuid4())
        db.session.execute(text("""
            INSERT INTO dbo.ROTAS_JOB
            (JobId, State, Stage, Total, Processed, Ok, Errors, Pending, StartedAt, UpdatedAt, Message, RequestedBy)
            VALUES (:job_id, 'running', 'compute', 0, 0, 0, 0, 0, GETDATE(), GETDATE(), 'A retomar...', :user)
        """), { 'job_id': job_id, 'user': getattr(current_user, 'LOGIN', '') })
        db.session.commit()
        t = threading.Thread(target=run_rotas_rebuild, args=(job_id, False), daemon=True)
        rotas_workers[str(job_id)] = t
        t.start()
        return jsonify({'job_id': job_id})


    @app.route('/api/rotas/rebuild/generate_missing', methods=['POST'])
    @login_required
    def api_rotas_rebuild_generate_missing():
        inserted = db.session.execute(text("""
            ;WITH coords AS (
                SELECT ALSTAMP
                FROM AL
                WHERE LAT IS NOT NULL AND LON IS NOT NULL
            )
            INSERT INTO dbo.ROTAS
            (OrigemStamp, DestinoStamp, Km, Segundos, Perfil, Provider, Status, Tentativas, Erro, UpdatedAt)
            SELECT a.ALSTAMP, b.ALSTAMP, 0, 0, 'driving', 'OSRM', 0, 0, '', GETDATE()
            FROM coords a
            JOIN coords b ON a.ALSTAMP < b.ALSTAMP
            WHERE NOT EXISTS (
                SELECT 1
                FROM dbo.ROTAS r
                WHERE (r.OrigemStamp = a.ALSTAMP AND r.DestinoStamp = b.ALSTAMP)
                   OR (r.OrigemStamp = b.ALSTAMP AND r.DestinoStamp = a.ALSTAMP)
            )
        """)).rowcount
        db.session.commit()
        return jsonify({'ok': True, 'inserted': inserted})


    @app.route('/api/rotas/rebuild/status', methods=['GET'])
    @login_required
    def api_rotas_rebuild_status():
        job_id = request.args.get('job_id')
        if not job_id:
            running = rotas_job_get_running()
            if not running:
                return jsonify({'state': 'idle'})
            job_id = running['JobId']
        job = rotas_job_get(job_id)
        if not job:
            return jsonify({'state': 'idle'})
        if job.get('State') in ('running', 'stopping'):
            if not _rotas_worker_alive(str(job_id)):
                rotas_job_update(
                    job_id,
                    State='stopped',
                    Message='Parado (sem worker)',
                    FinishedAt=datetime.now(),
                    UpdatedAt=datetime.now()
                )
                job = rotas_job_get(job_id) or job
            elif _rotas_job_is_stale(job, seconds=300):
                rotas_job_update(
                    job_id,
                    State='stopped',
                    Message='Parado (timeout)',
                    FinishedAt=datetime.now(),
                    UpdatedAt=datetime.now()
                )
                job = rotas_job_get(job_id) or job
        job['JobId'] = str(job['JobId'])
        return jsonify(job)


    # Report per client/month (HTML)
    @app.route('/report/<cliente_id>/<int:ano>/<int:mes>')
    @login_required
    def report_cliente_mes(cliente_id, ano, mes):
        # Resolve month range
        try:
            start = date(ano, mes, 1)
            if mes == 12:
                end = date(ano + 1, 1, 1)
            else:
                end = date(ano, mes + 1, 1)
            end = end.replace(day=1)
            # last day of month
            from datetime import timedelta
            end = end - timedelta(days=1)
        except Exception:
            # fallback: current month
            today = date.today()
            start = date(today.year, today.month, 1)
            if today.month == 12:
                end = date(today.year + 1, 1, 1) - timedelta(days=1)
            else:
                end = date(today.year, today.month + 1, 1) - timedelta(days=1)

        # Fetch client's alojamentos; prefer CLIENTID when numeric
        alojamentos = []
        monthly_map = {}
        try:
            is_num = False
            try:
                _cid_int = int(cliente_id)
                is_num = True
            except Exception:
                is_num = False

            if is_num:
                rows = db.session.execute(text("""
                    SELECT NOME, TIPOLOGIA, EMPRESA, PROPRIETARIO, CCUSTO, TIPO
                    FROM AL
                    WHERE ISNULL(INATIVO,0)=0 AND CLIENTID = :cid
                    ORDER BY NOME
                """), { 'cid': _cid_int }).mappings().all()
            else:
                rows = db.session.execute(text("""
                    SELECT NOME, TIPOLOGIA, EMPRESA, PROPRIETARIO, CCUSTO, TIPO
                    FROM AL
                    WHERE ISNULL(INATIVO,0)=0 AND (PROPRIETARIO = :cid OR EMPRESA = :cid)
                    ORDER BY NOME
                """), { 'cid': cliente_id }).mappings().all()
            alojamentos = [ dict(r) for r in rows ]

            if is_num:
                sql_aggr = text("""
                    SELECT A.NOME AS ALOJ, MONTH(RS.DATAOUT) AS MES,
                           SUM(ISNULL(RS.ESTADIA,0)+ISNULL(RS.LIMPEZA,0)) AS VENDAS,
                           SUM(ISNULL(RS.COMISSAO,0)) AS COM_CANAL
                    FROM AL A
                    LEFT JOIN RS ON ISNULL(RS.ALOJAMENTO, RS.CCUSTO) = A.NOME AND YEAR(RS.DATAOUT) = :ano
                    WHERE ISNULL(A.INATIVO,0)=0 AND A.CLIENTID = :cid
                    GROUP BY A.NOME, MONTH(RS.DATAOUT)
                    ORDER BY A.NOME, MES
                """)
                ag = db.session.execute(sql_aggr, { 'cid': _cid_int, 'ano': ano }).all()
                for aloj, mesn, vendas, com_canal in ag:
                    monthly_map.setdefault(aloj, {})[int(mesn or 0)] = {
                        'vendas': float(vendas or 0),
                        'com_canal': float(com_canal or 0)
                    }
        except Exception:
            alojamentos = []
            monthly_map = {}

        # Render standalone printable HTML
        return render_template('report.html', cliente_id=cliente_id, ano=ano, mes=mes, inicio=start, fim=end, alojamentos=alojamentos, monthly_map=monthly_map)# Report per client/month (PDF)
    @app.route('/report/<cliente_id>/<int:ano>/<int:mes>/pdf')
    @login_required
    def report_cliente_mes_pdf(cliente_id, ano, mes):
        # Reuse same data gathering as HTML
        try:
            start = date(ano, mes, 1)
            if mes == 12:
                next_month_first = date(ano + 1, 1, 1)
            else:
                next_month_first = date(ano, mes + 1, 1)
            from datetime import timedelta
            end = next_month_first - timedelta(days=1)
        except Exception:
            today = date.today()
            start = date(today.year, today.month, 1)
            if today.month == 12:
                end = date(today.year + 1, 1, 1) - timedelta(days=1)
            else:
                end = date(today.year, today.month + 1, 1) - timedelta(days=1)

        alojamentos = []
        try:
            query = text(
                """
                SELECT NOME, TIPOLOGIA, EMPRESA, PROPRIETARIO, CCUSTO, TIPO
                FROM AL
                WHERE ISNULL(INATIVO,0)=0
                  AND (
                        PROPRIETARIO = :cid
                     OR EMPRESA = :cid
                     OR (CASE WHEN TRY_CAST(:cid AS int) IS NOT NULL THEN IDPLABS ELSE NULL END) = TRY_CAST(:cid AS int)
                  )
                ORDER BY NOME
                """
            )
            rows = db.session.execute(query, { 'cid': cliente_id }).mappings().all()
            alojamentos = [ dict(r) for r in rows ]
        except Exception:
            alojamentos = []

        # Render HTML
        from flask import render_template_string, current_app, Response
        html = render_template('report.html', cliente_id=cliente_id, ano=ano, mes=mes, inicio=start, fim=end, alojamentos=alojamentos, monthly_map={})

        # Try WeasyPrint / wkhtmltopdf / Headless Chrome
        pdf_bytes = None
        try:
            from weasyprint import HTML
            pdf_bytes = HTML(string=html, base_url=current_app.root_path).write_pdf()
        except Exception:
            try:
                import pdfkit  # requires wkhtmltopdf installed
                pdf_bytes = pdfkit.from_string(html, False)
            except Exception:
                # Try Headless Chrome/Edge as a last resort
                try:
                    chrome_paths = [
                        os.environ.get('CHROME_PATH') or '',
                        shutil.which('chrome') or '',
                        shutil.which('google-chrome') or '',
                        shutil.which('msedge') or '',
                        r"C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe",
                        r"C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe",
                        r"C:\\Program Files\\Microsoft\\Edge\\Application\\msedge.exe",
                        r"/usr/bin/google-chrome",
                        r"/usr/bin/chromium",
                    ]
                    chrome = next((p for p in chrome_paths if p and os.path.exists(p)), None)
                    if chrome:
                        with tempfile.TemporaryDirectory() as td:
                            html_path = os.path.join(td, 'report.html')
                            pdf_path = os.path.join(td, 'report.pdf')
                            with open(html_path, 'w', encoding='utf-8') as f:
                                f.write(html)
                            uri = 'file:///' + html_path.replace('\\', '/')
                            cmd = [
                                chrome,
                                '--headless=new',
                                '--disable-gpu',
                                f'--print-to-pdf={pdf_path}',
                                '--no-sandbox',
                                uri
                            ]
                            subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                            with open(pdf_path, 'rb') as pf:
                                pdf_bytes = pf.read()
                except Exception:
                    pdf_bytes = None

        if not pdf_bytes:
            # Fallback: return HTML with hint
            return html + "<!-- PDF generator not available. Install WeasyPrint (pip install weasyprint + deps), wkhtmltopdf/pdfkit, or set CHROME_PATH to a Chrome/Edge binary for headless printing. -->"

        headers = {
            'Content-Type': 'application/pdf',
            'Content-Disposition': f'inline; filename="report_{cliente_id}_{ano:04d}-{mes:02d}.pdf"'
        }
        return Response(pdf_bytes, headers=headers)


    return app


def _geo_fetch_json(url: str):
    req = Request(
        url,
        headers={
            "User-Agent": "APP_WEB/1.0 (geo)"
        }
    )
    with urlopen(req, timeout=12) as resp:
        return json.loads(resp.read().decode("utf-8"))


def _geo_build_query(morada: str, codpost: str, local: str) -> str:
    parts = [morada, codpost, local]
    parts = [p.strip() for p in parts if p and str(p).strip()]
    base = ", ".join(parts)
    if not base:
        return ""
    if "portugal" not in base.lower():
        base = f"{base}, Portugal"
    return base




app = create_app()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
